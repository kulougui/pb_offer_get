"""
广告联盟Offer管理工具
支持从PartnerBoost获取offer信息，并与Google Ads和飞书表格联动管理
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import requests
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import threading
import os
import re
import time
from collections import Counter
from requests.exceptions import ReadTimeout, Timeout
from google.ads.googleads.client import GoogleAdsClient
from google.oauth2 import service_account


# 硬编码API基础URL
PB_API_BASE_URL = "https://app.partnerboost.com"
YP_API_BASE_URL = "https://yeahpromos.com"
FEISHU_API_BASE_URL = "https://open.feishu.cn/open-apis"
SUMMARY_STATUS_TEXT = "相同offer统计行"
SUMMARY_PLACEHOLDER_TEXT = "——" * 10
CAMPAIGNS_SHEET_ID = "XrkOF7"
ADS_BRAND_SHEET_TITLE = "ads | 品牌"

# 配置文件路径
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")

# 日志文件路径
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
LOG_FILE = os.path.join(LOG_DIR, "app.log")
DEBUG_LOG_FILE = os.path.join(LOG_DIR, "debug.log")  # 调试日志文件
LOG_MAX_SIZE_MB = 10  # 日志文件最大大小（MB）
LOG_KEEP_DAYS = 7     # 保留最近几天的日志

# 旧广告系列花费记录文件
OLD_CAMPAIGN_CONSUME_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "old_campaign_consume.xlsx")

# 广告系列花费快照文件（用于追踪每个广告系列的花费变化）
CAMPAIGN_COST_SNAPSHOT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "campaign_cost_snapshot.xlsx")


class OfferToolApp:
    def __init__(self, root):
        self.root = root
        self.root.title("广告联盟Offer管理工具")
        self.root.geometry("900x750")
        self.root.resizable(True, True)
        
        # 初始化日志系统
        self.init_log_system()
        
        # 设置样式
        style = ttk.Style()
        style.configure('TLabel', font=('Microsoft YaHei', 10))
        style.configure('TButton', font=('Microsoft YaHei', 10))
        style.configure('Header.TLabel', font=('Microsoft YaHei', 12, 'bold'))
        
        # 加载配置
        self.config = self.load_config()
        
        # 创建Notebook（选项卡）
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 创建各个选项卡
        self.create_offer_get_tab()
        self.create_offer_manage_tab()
        self.create_config_tab()
    
    def init_log_system(self):
        """初始化日志系统"""
        # 创建日志目录
        if not os.path.exists(LOG_DIR):
            os.makedirs(LOG_DIR)
        
        # 清理旧日志
        self.clean_old_logs()
        
        # 清空调试日志（每次启动时重新开始）
        try:
            with open(DEBUG_LOG_FILE, 'w', encoding='utf-8') as f:
                f.write(f"=== 调试日志 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
        except Exception:
            pass
        
        # 写入启动日志
        self.write_log_to_file(f"\n{'='*60}")
        self.write_log_to_file(f"程序启动 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.write_log_to_file(f"{'='*60}")
    
    def write_log_to_file(self, message):
        """写入日志到文件"""
        try:
            # 检查文件大小，如果超过限制则轮转
            if os.path.exists(LOG_FILE):
                size_mb = os.path.getsize(LOG_FILE) / (1024 * 1024)
                if size_mb > LOG_MAX_SIZE_MB:
                    self.rotate_log_file()
            
            # 写入日志
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            with open(LOG_FILE, 'a', encoding='utf-8') as f:
                f.write(f"[{timestamp}] {message}\n")
        except Exception as e:
            print(f"写入日志失败: {e}")
    
    def rotate_log_file(self):
        """轮转日志文件"""
        try:
            if os.path.exists(LOG_FILE):
                # 重命名为带时间戳的备份文件
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_file = os.path.join(LOG_DIR, f"app_{timestamp}.log")
                os.rename(LOG_FILE, backup_file)
        except Exception as e:
            print(f"轮转日志失败: {e}")
    
    def clean_old_logs(self):
        """清理旧日志文件"""
        try:
            if not os.path.exists(LOG_DIR):
                return
            
            cutoff_date = datetime.now() - timedelta(days=LOG_KEEP_DAYS)
            
            for filename in os.listdir(LOG_DIR):
                if filename.endswith('.log'):
                    filepath = os.path.join(LOG_DIR, filename)
                    # 获取文件修改时间
                    file_time = datetime.fromtimestamp(os.path.getmtime(filepath))
                    if file_time < cutoff_date:
                        os.remove(filepath)
                        print(f"已清理旧日志: {filename}")
        except Exception as e:
            print(f"清理日志失败: {e}")
        
    def load_config(self):
        """加载配置文件"""
        default_config = {
            # PartnerBoost配置
            "pb_token": "5wXyGERfQ3rEQTdI",
            # YeahPromos配置
            "yp_token": "",
            "yp_site_id": "",
            # 飞书配置
            "feishu_app_id": "cli_a8517363cf3bd013",
            "feishu_app_secret": "O4Sm3UNHjpykF9OZq3LroblsrCVYyQEp",
            "feishu_spreadsheet_token": "KnJ1wphpBiVMrGkWl5ncUkMGnfe",
            "feishu_sheet_id": "kPlW5z",
            # Google Ads配置
            "google_developer_token": "1YsRjWGxV6XUdxtX8MiT3Q",
            "google_mcc_id": "6885177935",
            "google_mcc_ids": [],
            "google_mcc_accounts": [],
            "google_service_account_file": "credentials/google_ads_service_account.json",
            "google_service_account_key": "",
            "stats_start_date": "2026-01-01",
            "stats_increment_days": "7",
            # 保存路径
            "save_path": os.path.join(os.path.expanduser("~"), "Desktop", "offers.xlsx")
        }
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    saved_config = json.load(f)
                    default_config.update(saved_config)
        except Exception:
            pass
        if not default_config.get("google_service_account_key"):
            default_config["google_service_account_key"] = self.load_google_service_account_key(
                default_config.get("google_service_account_file", "")
            )
        default_config["google_mcc_accounts"] = self.normalize_google_mcc_accounts(default_config)
        return default_config

    def normalize_google_mcc_accounts(self, config):
        """标准化多MCC配置，兼容旧的单密钥/多ID配置。"""
        accounts = []
        seen = set()

        for item in config.get("google_mcc_accounts", []) or []:
            if not isinstance(item, dict):
                continue
            mcc_id = self.parse_google_mcc_ids(item.get("mcc_id", ""))
            mcc_id = mcc_id[0] if mcc_id else ""
            if not mcc_id or mcc_id in seen:
                continue
            seen.add(mcc_id)
            accounts.append({
                "name": str(item.get("name", "") or "").strip(),
                "mcc_id": mcc_id,
                "developer_token": str(item.get("developer_token", "") or config.get("google_developer_token", "") or "").strip(),
                "service_account_key": str(item.get("service_account_key", "") or "").strip(),
            })

        if accounts:
            return accounts

        legacy_key = str(config.get("google_service_account_key", "") or "").strip()
        legacy_token = str(config.get("google_developer_token", "") or "").strip()
        for mcc_id in self.parse_google_mcc_ids(config.get("google_mcc_ids") or config.get("google_mcc_id", "")):
            if mcc_id in seen:
                continue
            seen.add(mcc_id)
            accounts.append({
                "name": "",
                "mcc_id": mcc_id,
                "developer_token": legacy_token,
                "service_account_key": legacy_key,
            })
        return accounts

    def load_google_service_account_key(self, sa_file):
        """读取服务账号密钥文本"""
        try:
            if sa_file and os.path.exists(sa_file):
                with open(sa_file, 'r', encoding='utf-8') as f:
                    return f.read()
        except Exception:
            pass
        return ""

    def save_google_service_account_key(self, sa_file, key_text):
        """保存服务账号密钥文本"""
        if not sa_file:
            return

        sa_dir = os.path.dirname(sa_file)
        if sa_dir and not os.path.exists(sa_dir):
            os.makedirs(sa_dir, exist_ok=True)
        with open(sa_file, 'w', encoding='utf-8') as f:
            f.write(key_text.strip())
    
    def save_config(self):
        """保存配置文件"""
        try:
            google_mcc_accounts = self.get_google_mcc_accounts()
            google_mcc_ids = [account["mcc_id"] for account in google_mcc_accounts]
            first_google_account = google_mcc_accounts[0] if google_mcc_accounts else {}
            key_text = first_google_account.get("service_account_key", "")
            developer_token = first_google_account.get("developer_token", "")
            config_to_save = {
                "pb_token": self.pb_token_var.get().strip(),
                "yp_token": self.yp_token_var.get().strip(),
                "yp_site_id": self.yp_site_id_var.get().strip(),
                "feishu_app_id": self.feishu_app_id_var.get().strip(),
                "feishu_app_secret": self.feishu_app_secret_var.get().strip(),
                "feishu_spreadsheet_token": self.feishu_spreadsheet_var.get().strip(),
                "feishu_sheet_id": self.feishu_sheet_id_var.get().strip(),
                "google_developer_token": developer_token,
                "google_mcc_id": google_mcc_ids[0] if google_mcc_ids else "",
                "google_mcc_ids": google_mcc_ids,
                "google_mcc_accounts": google_mcc_accounts,
                "google_service_account_file": self.google_sa_file_var.get().strip(),
                "google_service_account_key": key_text,
                "stats_start_date": self.stats_start_date_var.get().strip() if hasattr(self, 'stats_start_date_var') else self.config.get("stats_start_date", "2026-01-01"),
                "stats_increment_days": self.stats_increment_days_var.get().strip() if hasattr(self, 'stats_increment_days_var') else self.config.get("stats_increment_days", "7"),
                "save_path": self.save_path_var.get()
            }
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config_to_save, f, ensure_ascii=False, indent=2)
            if key_text:
                self.save_google_service_account_key(config_to_save["google_service_account_file"], key_text)
        except Exception as e:
            print(f"保存配置失败: {e}")

    # ==================== Offer获取选项卡 ====================
    def create_offer_get_tab(self):
        """创建Offer获取选项卡"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="  Offer获取  ")
        
        main_frame = ttk.Frame(tab, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # API参数配置区域
        api_frame = ttk.LabelFrame(main_frame, text="API参数配置", padding="10")
        api_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 使用Grid布局，2列
        ttk.Label(api_frame, text="国家代码:", width=15, anchor=tk.W).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.country_code_var = tk.StringVar(value="")
        ttk.Entry(api_frame, textvariable=self.country_code_var, width=20).grid(row=0, column=1, sticky=tk.W, pady=5, padx=(0, 30))
        
        ttk.Label(api_frame, text="品牌ID:", width=15, anchor=tk.W).grid(row=0, column=2, sticky=tk.W, pady=5)
        self.brand_id_var = tk.StringVar(value="")
        ttk.Entry(api_frame, textvariable=self.brand_id_var, width=20).grid(row=0, column=3, sticky=tk.W, pady=5)
        
        ttk.Label(api_frame, text="ASIN(逗号分隔):", width=15, anchor=tk.W).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.asins_var = tk.StringVar(value="")
        ttk.Entry(api_frame, textvariable=self.asins_var, width=20).grid(row=1, column=1, sticky=tk.W, pady=5, padx=(0, 30))
        
        ttk.Label(api_frame, text="排序方式:", width=15, anchor=tk.W).grid(row=1, column=2, sticky=tk.W, pady=5)
        self.sort_var = tk.StringVar(value="")
        ttk.Entry(api_frame, textvariable=self.sort_var, width=20).grid(row=1, column=3, sticky=tk.W, pady=5)
        
        ttk.Label(api_frame, text="Relationship:", width=15, anchor=tk.W).grid(row=2, column=0, sticky=tk.W, pady=5)
        self.relationship_var = tk.StringVar(value="1")
        ttk.Combobox(api_frame, textvariable=self.relationship_var, values=["0", "1"], width=17).grid(row=2, column=1, sticky=tk.W, pady=5, padx=(0, 30))
        
        # 复选框
        self.default_filter_var = tk.IntVar(value=0)
        ttk.Checkbutton(api_frame, text="默认筛选", variable=self.default_filter_var).grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=5)
        
        self.is_original_currency_var = tk.IntVar(value=0)
        ttk.Checkbutton(api_frame, text="原始货币", variable=self.is_original_currency_var).grid(row=3, column=2, columnspan=2, sticky=tk.W, pady=5)
        
        self.has_promo_code_var = tk.IntVar(value=0)
        ttk.Checkbutton(api_frame, text="有促销码", variable=self.has_promo_code_var).grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=5)
        
        self.has_acc_var = tk.IntVar(value=0)
        ttk.Checkbutton(api_frame, text="有ACC佣金", variable=self.has_acc_var).grid(row=4, column=2, columnspan=2, sticky=tk.W, pady=5)
        
        self.filter_sexual_wellness_var = tk.IntVar(value=0)
        ttk.Checkbutton(api_frame, text="过滤成人用品", variable=self.filter_sexual_wellness_var).grid(row=5, column=0, columnspan=2, sticky=tk.W, pady=5)
        
        # UID追踪参数
        ttk.Label(api_frame, text="UID模式:", width=15, anchor=tk.W).grid(row=6, column=0, sticky=tk.W, pady=5)
        self.uid_mode_var = tk.StringVar(value="随机")
        uid_mode_combo = ttk.Combobox(api_frame, textvariable=self.uid_mode_var, values=["无", "随机", "固定"], width=17, state='readonly')
        uid_mode_combo.grid(row=6, column=1, sticky=tk.W, pady=5, padx=(0, 30))
        uid_mode_combo.bind('<<ComboboxSelected>>', self._on_uid_mode_changed)
        
        ttk.Label(api_frame, text="固定UID:", width=15, anchor=tk.W).grid(row=6, column=2, sticky=tk.W, pady=5)
        self.fixed_uid_var = tk.StringVar(value="")
        self.fixed_uid_entry = ttk.Entry(api_frame, textvariable=self.fixed_uid_var, width=20, state='disabled')
        self.fixed_uid_entry.grid(row=6, column=3, sticky=tk.W, pady=5)
        
        # 保存配置区域
        save_frame = ttk.LabelFrame(main_frame, text="保存配置", padding="10")
        save_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(save_frame, text="保存路径:").pack(side=tk.LEFT)
        self.save_path_var = tk.StringVar(value=self.config.get("save_path", ""))
        save_entry = ttk.Entry(save_frame, textvariable=self.save_path_var, width=50)
        save_entry.pack(side=tk.LEFT, padx=(10, 10))
        ttk.Button(save_frame, text="浏览...", command=self.browse_save_path).pack(side=tk.LEFT)
        
        # 操作按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.extract_btn = ttk.Button(button_frame, text="提取PB Offer", command=self.extract_offers)
        self.extract_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.update_yp_links_btn = ttk.Button(button_frame, text="更新YP链接", command=self.update_yp_links)
        self.update_yp_links_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.copy_offer_btn = ttk.Button(button_frame, text="复制Offer", command=self.copy_offers)
        self.copy_offer_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.brand_search_btn = ttk.Button(button_frame, text="获取品牌搜索量报告", command=self.get_brand_search_volume)
        self.brand_search_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.brand_search_scope_var = tk.StringVar(value="全量")
        ttk.Combobox(
            button_frame,
            textvariable=self.brand_search_scope_var,
            values=["50个", "全量"],
            width=8,
            state='readonly'
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        self.progress_label_get = ttk.Label(button_frame, text="")
        self.progress_label_get.pack(side=tk.RIGHT, padx=(10, 0))
        
        self.progress_get = ttk.Progressbar(button_frame, mode='indeterminate', length=150)
        self.progress_get.pack(side=tk.RIGHT)
        
        # 日志区域
        log_frame = ttk.LabelFrame(main_frame, text="运行日志", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text_get = scrolledtext.ScrolledText(log_frame, height=12, font=('Consolas', 9))
        self.log_text_get.pack(fill=tk.BOTH, expand=True)

    # ==================== Offer管理选项卡 ====================
    def create_offer_manage_tab(self):
        """创建Offer管理选项卡"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="  Offer管理  ")
        
        main_frame = ttk.Frame(tab, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 说明区域
        info_frame = ttk.LabelFrame(main_frame, text="功能说明", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        info_text = """• "开始统计"：获取Google Ads/PB数据，更新offer状态、广告系列数量、花费、佣金
• "更新 PB offer"：按品牌批量读取PB offer，更新已有PB offer信息并新增PB offer
• "offer顺序整理"：对Offer表和广告系列表排序（按状态分组+总佣金降序）"""
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT).pack(anchor=tk.W)
        
        # 统计参数区域
        date_frame = ttk.Frame(main_frame)
        date_frame.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(date_frame, text="开始日期:").pack(side=tk.LEFT)
        self.stats_start_date_var = tk.StringVar(value=self.config.get("stats_start_date", "2026-01-01"))
        ttk.Entry(date_frame, textvariable=self.stats_start_date_var, width=12).pack(side=tk.LEFT, padx=(5, 10))

        ttk.Label(date_frame, text="结束日期:").pack(side=tk.LEFT)
        self.stats_end_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        self.stats_end_date_entry = ttk.Entry(date_frame, textvariable=self.stats_end_date_var, width=12)
        self.stats_end_date_entry.pack(side=tk.LEFT, padx=(5, 10))

        self.stats_to_today_var = tk.IntVar(value=1)
        self.stats_to_today_check = ttk.Checkbutton(
            date_frame, text="至今", variable=self.stats_to_today_var, command=self._on_stats_to_today_changed
        )
        self.stats_to_today_check.pack(side=tk.LEFT, padx=(0, 15))
        self._on_stats_to_today_changed()

        ttk.Label(date_frame, text="新增花费/佣金统计天数:").pack(side=tk.LEFT)
        self.stats_increment_days_var = tk.StringVar(value=self.config.get("stats_increment_days", "7"))
        ttk.Entry(date_frame, textvariable=self.stats_increment_days_var, width=6).pack(side=tk.LEFT, padx=(5, 10))

        # 操作按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.start_stats_btn = ttk.Button(button_frame, text="开始统计", command=self.start_statistics)
        self.start_stats_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.stop_btn = ttk.Button(button_frame, text="停止", command=self.stop_statistics, state='disabled')
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.update_offers_btn = ttk.Button(button_frame, text="更新 PB offer", command=self.start_update_offers)
        self.update_offers_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.sort_tables_btn = ttk.Button(button_frame, text="offer顺序整理", command=self.start_sort_tables)
        self.sort_tables_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.check_links_btn = ttk.Button(button_frame, text="检查链接健康", command=self.start_check_link_health)
        self.check_links_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.progress_label_manage = ttk.Label(button_frame, text="")
        self.progress_label_manage.pack(side=tk.RIGHT, padx=(10, 0))
        
        self.progress_manage = ttk.Progressbar(button_frame, mode='indeterminate', length=200)
        self.progress_manage.pack(side=tk.RIGHT)
        
        # 日志区域
        log_frame = ttk.LabelFrame(main_frame, text="运行日志", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text_manage = scrolledtext.ScrolledText(log_frame, height=20, font=('Consolas', 9))
        self.log_text_manage.pack(fill=tk.BOTH, expand=True)
        
        # 停止标志
        self.stop_flag = False

    def _on_stats_to_today_changed(self):
        """切换统计结束日期是否固定为今天"""
        if self.stats_to_today_var.get():
            self.stats_end_date_entry.config(state='disabled')
            self.stats_end_date_var.set(datetime.now().strftime('%Y-%m-%d'))
        else:
            self.stats_end_date_entry.config(state='normal')

    # ==================== 配置选项卡 ====================
    def create_config_tab(self):
        """创建配置选项卡"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="  API配置  ")
        
        main_frame = ttk.Frame(tab, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # PartnerBoost配置
        pb_frame = ttk.LabelFrame(main_frame, text="PartnerBoost配置", padding="10")
        pb_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(pb_frame, text="API Token:", width=20, anchor=tk.W).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.pb_token_var = tk.StringVar(value=self.config.get("pb_token", ""))
        ttk.Entry(pb_frame, textvariable=self.pb_token_var, width=50).grid(row=0, column=1, sticky=tk.W, pady=5)

        ttk.Label(pb_frame, text="YP Token:", width=20, anchor=tk.W).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.yp_token_var = tk.StringVar(value=self.config.get("yp_token", ""))
        ttk.Entry(pb_frame, textvariable=self.yp_token_var, width=50).grid(row=1, column=1, sticky=tk.W, pady=5)

        ttk.Label(pb_frame, text="YP Site ID:", width=20, anchor=tk.W).grid(row=2, column=0, sticky=tk.W, pady=5)
        self.yp_site_id_var = tk.StringVar(value=self.config.get("yp_site_id", ""))
        ttk.Entry(pb_frame, textvariable=self.yp_site_id_var, width=50).grid(row=2, column=1, sticky=tk.W, pady=5)
        
        # 飞书配置
        feishu_frame = ttk.LabelFrame(main_frame, text="飞书配置", padding="10")
        feishu_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(feishu_frame, text="App ID:", width=20, anchor=tk.W).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.feishu_app_id_var = tk.StringVar(value=self.config.get("feishu_app_id", ""))
        ttk.Entry(feishu_frame, textvariable=self.feishu_app_id_var, width=50).grid(row=0, column=1, sticky=tk.W, pady=5)
        
        ttk.Label(feishu_frame, text="App Secret:", width=20, anchor=tk.W).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.feishu_app_secret_var = tk.StringVar(value=self.config.get("feishu_app_secret", ""))
        ttk.Entry(feishu_frame, textvariable=self.feishu_app_secret_var, width=50, show="*").grid(row=1, column=1, sticky=tk.W, pady=5)
        
        ttk.Label(feishu_frame, text="电子表格Token:", width=20, anchor=tk.W).grid(row=2, column=0, sticky=tk.W, pady=5)
        self.feishu_spreadsheet_var = tk.StringVar(value=self.config.get("feishu_spreadsheet_token", ""))
        ttk.Entry(feishu_frame, textvariable=self.feishu_spreadsheet_var, width=50).grid(row=2, column=1, sticky=tk.W, pady=5)
        
        ttk.Label(feishu_frame, text="工作表ID:", width=20, anchor=tk.W).grid(row=3, column=0, sticky=tk.W, pady=5)
        self.feishu_sheet_id_var = tk.StringVar(value=self.config.get("feishu_sheet_id", ""))
        ttk.Entry(feishu_frame, textvariable=self.feishu_sheet_id_var, width=50).grid(row=3, column=1, sticky=tk.W, pady=5)
        
        # Google Ads配置
        google_frame = ttk.LabelFrame(main_frame, text="Google Ads配置", padding="10")
        google_frame.pack(fill=tk.X, pady=(0, 10))
        google_frame.columnconfigure(1, weight=1)
        google_frame.columnconfigure(3, weight=1)
        
        self.google_dev_token_var = tk.StringVar(value=self.config.get("google_developer_token", ""))
        self.google_sa_file_var = tk.StringVar(value=self.config.get("google_service_account_file", ""))
        self.google_mcc_accounts = [
            dict(account) for account in self.normalize_google_mcc_accounts(self.config)
        ]
        self.current_google_mcc_index = None

        list_frame = ttk.Frame(google_frame)
        list_frame.grid(row=0, column=0, rowspan=6, sticky=tk.NS, padx=(0, 12))
        ttk.Label(list_frame, text="MCC账号", anchor=tk.W).pack(fill=tk.X)
        self.google_mcc_listbox = tk.Listbox(list_frame, width=28, height=12, exportselection=False)
        self.google_mcc_listbox.pack(fill=tk.BOTH, expand=True, pady=(5, 6))
        self.google_mcc_listbox.bind("<<ListboxSelect>>", self.on_google_mcc_select)
        list_btn_frame = ttk.Frame(list_frame)
        list_btn_frame.pack(fill=tk.X)
        ttk.Button(list_btn_frame, text="新增", command=self.add_google_mcc_account).pack(side=tk.LEFT)
        ttk.Button(list_btn_frame, text="删除", command=self.delete_google_mcc_account).pack(side=tk.LEFT, padx=(6, 0))

        ttk.Label(google_frame, text="名称:", width=16, anchor=tk.W).grid(row=0, column=1, sticky=tk.W, pady=5)
        self.google_mcc_name_var = tk.StringVar()
        ttk.Entry(google_frame, textvariable=self.google_mcc_name_var, width=42).grid(row=0, column=2, sticky=tk.EW, pady=5)

        ttk.Label(google_frame, text="MCC ID:", width=16, anchor=tk.W).grid(row=1, column=1, sticky=tk.W, pady=5)
        self.google_mcc_id_var = tk.StringVar()
        ttk.Entry(google_frame, textvariable=self.google_mcc_id_var, width=42).grid(row=1, column=2, sticky=tk.EW, pady=5)

        ttk.Label(google_frame, text="Developer Token:", width=16, anchor=tk.W).grid(row=2, column=1, sticky=tk.W, pady=5)
        ttk.Entry(google_frame, textvariable=self.google_dev_token_var, width=42).grid(row=2, column=2, sticky=tk.EW, pady=5)

        ttk.Label(google_frame, text="服务账号密钥:", width=16, anchor=tk.NW).grid(row=3, column=1, sticky=tk.NW, pady=5)
        self.google_sa_key_var = scrolledtext.ScrolledText(google_frame, width=58, height=10, font=('Consolas', 9))
        self.google_sa_key_var.grid(row=3, column=2, sticky=tk.EW, pady=5)

        guide_text = (
            "MCC账号更换流程\n"
            "①准备一个有Basic Access的mcc账号\n"
            "②去一个已经开通google ads api的google cloud账号（比如mcc1）\n"
            "③创建新项目→创建新服务账号→创建服务账号的密钥\n"
            "④新mcc账号后台，拉新服务账号，给予权限\n"
            "⑤该卡片填入所有信息\n\n"
            "多MCC配置说明\n"
            "• 每个MCC填写自己的Developer Token、MCC ID和服务账号密钥\n"
            "• “开始统计”会依次读取所有MCC下的直属Ads账号\n"
            "• 切换左侧MCC前会自动暂存当前编辑内容\n"
            "• 旧的单MCC配置会自动迁移为第一条记录"
        )
        ttk.Label(
            google_frame,
            text=guide_text,
            justify=tk.LEFT,
            foreground="#808080"
        ).grid(row=0, column=3, rowspan=4, sticky=tk.NW, padx=(20, 0), pady=5)
        self.refresh_google_mcc_listbox()
        if self.google_mcc_accounts:
            self.select_google_mcc_account(0)
        else:
            self.add_google_mcc_account()
        
        # 保存按钮
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        ttk.Button(btn_frame, text="保存配置", command=self.save_and_notify).pack(side=tk.LEFT)
        self.test_conn_btn = ttk.Button(btn_frame, text="测试连接", command=self.test_connections)
        self.test_conn_btn.pack(side=tk.LEFT, padx=10)

    # ==================== 通用方法 ====================
    def browse_save_path(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            title="选择保存位置"
        )
        if file_path:
            self.save_path_var.set(file_path)
    
    def save_and_notify(self):
        self.save_config()
        messagebox.showinfo("成功", "配置已保存！")

    def format_google_mcc_list_label(self, account):
        name = str(account.get("name", "") or "").strip()
        mcc_id = str(account.get("mcc_id", "") or "").strip()
        if name and mcc_id:
            return f"{name} ({mcc_id})"
        return mcc_id or name or "未命名MCC"

    def refresh_google_mcc_listbox(self):
        if not hasattr(self, 'google_mcc_listbox'):
            return
        self.google_mcc_listbox.delete(0, tk.END)
        for account in self.google_mcc_accounts:
            self.google_mcc_listbox.insert(tk.END, self.format_google_mcc_list_label(account))

    def save_current_google_mcc_editor(self):
        if not hasattr(self, 'google_mcc_accounts'):
            return
        index = getattr(self, 'current_google_mcc_index', None)
        if index is None or index < 0 or index >= len(self.google_mcc_accounts):
            return
        mcc_ids = self.parse_google_mcc_ids(self.google_mcc_id_var.get())
        self.google_mcc_accounts[index] = {
            "name": self.google_mcc_name_var.get().strip(),
            "mcc_id": mcc_ids[0] if mcc_ids else "",
            "developer_token": self.google_dev_token_var.get().strip(),
            "service_account_key": self.google_sa_key_var.get("1.0", tk.END).strip(),
        }

    def select_google_mcc_account(self, index):
        if not hasattr(self, 'google_mcc_accounts') or not self.google_mcc_accounts:
            return
        index = max(0, min(index, len(self.google_mcc_accounts) - 1))
        self.current_google_mcc_index = index
        account = self.google_mcc_accounts[index]
        self.google_mcc_name_var.set(account.get("name", ""))
        self.google_mcc_id_var.set(account.get("mcc_id", ""))
        self.google_dev_token_var.set(account.get("developer_token", ""))
        self.google_sa_key_var.delete("1.0", tk.END)
        self.google_sa_key_var.insert("1.0", account.get("service_account_key", ""))
        self.google_mcc_listbox.selection_clear(0, tk.END)
        self.google_mcc_listbox.selection_set(index)
        self.google_mcc_listbox.activate(index)

    def on_google_mcc_select(self, event=None):
        if not hasattr(self, 'google_mcc_listbox'):
            return
        selection = self.google_mcc_listbox.curselection()
        if not selection:
            return
        new_index = selection[0]
        if new_index == getattr(self, 'current_google_mcc_index', None):
            return
        self.save_current_google_mcc_editor()
        self.refresh_google_mcc_listbox()
        self.select_google_mcc_account(new_index)

    def add_google_mcc_account(self):
        self.save_current_google_mcc_editor()
        default_token = self.google_dev_token_var.get().strip() if hasattr(self, 'google_dev_token_var') else self.config.get("google_developer_token", "")
        self.google_mcc_accounts.append({
            "name": "",
            "mcc_id": "",
            "developer_token": default_token,
            "service_account_key": "",
        })
        self.refresh_google_mcc_listbox()
        self.select_google_mcc_account(len(self.google_mcc_accounts) - 1)

    def delete_google_mcc_account(self):
        if not getattr(self, 'google_mcc_accounts', None):
            return
        selection = self.google_mcc_listbox.curselection()
        index = selection[0] if selection else getattr(self, 'current_google_mcc_index', 0)
        if index is None or index < 0 or index >= len(self.google_mcc_accounts):
            return
        del self.google_mcc_accounts[index]
        self.current_google_mcc_index = None
        self.refresh_google_mcc_listbox()
        if self.google_mcc_accounts:
            self.select_google_mcc_account(min(index, len(self.google_mcc_accounts) - 1))
        else:
            self.add_google_mcc_account()

    def _run_on_ui_thread(self, func):
        """确保Tkinter控件只在主线程更新"""
        try:
            self.root.after(0, func)
        except Exception:
            pass

    def _append_log_text(self, text_widget, message, timestamp_fmt):
        """线程安全地向日志框追加内容"""
        def _append():
            timestamp = datetime.now().strftime(timestamp_fmt)
            text_widget.insert(tk.END, f"[{timestamp}] {message}\n")
            text_widget.see(tk.END)
            self.root.update_idletasks()
        self._run_on_ui_thread(_append)
    
    def log_get(self, message):
        self._append_log_text(self.log_text_get, message, "%Y-%m-%d %H:%M:%S")
        # 同时写入日志文件
        self.write_log_to_file(f"[Offer获取] {message}")
    
    def log_manage(self, message):
        self._append_log_text(self.log_text_manage, message, "%H:%M:%S")
        # 同时写入日志文件
        self.write_log_to_file(message)

    def log_config(self, message):
        """API配置页日志"""
        self.write_log_to_file(f"[API配置] {message}")
    
    def log_debug(self, message):
        """写入调试日志（仅写入文件，不显示在GUI中）"""
        try:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            with open(DEBUG_LOG_FILE, 'a', encoding='utf-8') as f:
                f.write(f"[{timestamp}] {message}\n")
        except Exception:
            pass  # 调试日志写入失败不影响主程序
    
    def update_progress_get(self, text):
        self._run_on_ui_thread(lambda: (
            self.progress_label_get.config(text=text),
            self.root.update_idletasks()
        ))
    
    def update_progress_manage(self, text):
        self._run_on_ui_thread(lambda: (
            self.progress_label_manage.config(text=text),
            self.root.update_idletasks()
        ))

    # ==================== Offer获取功能 ====================
    def build_request_body(self, page):
        body = {
            "token": self.pb_token_var.get().strip(),
            "page_size": 100,
            "page": page,
            "default_filter": self.default_filter_var.get(),
            "country_code": self.country_code_var.get().strip(),
            "brand_id": self.brand_id_var.get().strip() if self.brand_id_var.get().strip() else None,
            "sort": self.sort_var.get().strip(),
            "asins": self.asins_var.get().strip(),
            "relationship": int(self.relationship_var.get()),
            "is_original_currency": self.is_original_currency_var.get(),
            "has_promo_code": self.has_promo_code_var.get(),
            "has_acc": self.has_acc_var.get(),
            "filter_sexual_wellness": self.filter_sexual_wellness_var.get()
        }
        return body
    
    def generate_random_uid(self):
        """生成7位随机UID（避免重复）"""
        import secrets
        import string
        
        # 加载已使用的UID
        used_uids_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "used_uids.txt")
        used_uids = set()
        if os.path.exists(used_uids_file):
            with open(used_uids_file, 'r', encoding='utf-8') as f:
                for line in f:
                    uid = line.strip()
                    if uid:
                        used_uids.add(uid)
        
        # 生成不重复的UID
        alphabet = string.ascii_lowercase + string.digits
        max_attempts = 1000
        for _ in range(max_attempts):
            uid = ''.join(secrets.choice(alphabet) for _ in range(7))
            if uid not in used_uids:
                # 保存新UID
                used_uids.add(uid)
                with open(used_uids_file, 'a', encoding='utf-8') as f:
                    f.write(uid + '\n')
                return uid
        
        # 如果多次尝试都重复，使用时间戳
        import time
        uid = f"T{int(time.time()) % 10000000}"
        with open(used_uids_file, 'a', encoding='utf-8') as f:
            f.write(uid + '\n')
        return uid
    
    def _on_uid_mode_changed(self, event=None):
        """UID模式改变时的回调函数"""
        mode = self.uid_mode_var.get()
        if mode == "固定":
            self.fixed_uid_entry.config(state='normal')
        else:
            self.fixed_uid_entry.config(state='disabled')
            self.fixed_uid_var.set("")
    
    def get_uid_for_offer(self):
        """根据UID模式获取用于投放链接的UID"""
        mode = self.uid_mode_var.get()
        if mode == "无":
            return ""
        elif mode == "随机":
            return self.generate_random_uid()
        elif mode == "固定":
            uid = self.fixed_uid_var.get().strip()
            if not uid:
                self.log_get("警告: 固定UID模式但未输入UID，将使用随机UID")
                return self.generate_random_uid()
            return uid
        return ""
    
    def resolve_redirect_url(self, tracking_link):
        """访问投放链接并解析最终产品链接。"""
        if not tracking_link:
            return ""
        try:
            resp = requests.get(tracking_link, timeout=15, allow_redirects=True)
            final_url = str(getattr(resp, 'url', '') or '').strip()
            if final_url and final_url != tracking_link:
                return final_url
            refresh_header = str(resp.headers.get('refresh') or resp.headers.get('Refresh') or '').strip()
            if refresh_header:
                match = re.search(r'url\s*=\s*(.+)$', refresh_header, re.IGNORECASE)
                if match:
                    return match.group(1).strip().strip('"\'')
            if resp.status_code == 200:
                html = resp.text
                match = re.search(r'var\s+u\s*=\s*"([^"]+)";\s*\n\s*location\.replace', html)
                if match:
                    return match.group(1)
                match = re.search(r'<meta\s+http-equiv="refresh"\s+content="[^;]*;\s*url=([^"]+)"', html)
                if match:
                    return match.group(1)
            return ""
        except Exception as e:
            self.log_get(f"    [警告] 解析重定向失败: {str(e)[:80]}")
            return ""
    
    def get_partnerboost_link(self, asin, country_code, uid=""):
        try:
            url = f"{PB_API_BASE_URL}/api/datafeed/get_amazon_link_by_asin"
            body = {
                "token": self.pb_token_var.get().strip(),
                "asins": asin,
                "country_code": country_code,
                "return_partnerboost_link": 1
            }
            if uid:
                body["uid"] = uid
            response = requests.post(url, json=body, timeout=30)
            if response.status_code == 200:
                data = response.json()
                status_code = data.get("status", {}).get("code")
                if status_code == 0:
                    link_data = data.get("data", [])
                    if link_data:
                        link = link_data[0].get("partnerboost_link", "")
                        if link:
                            return link
                        else:
                            # 数据中没有partnerboost_link字段
                            self.log_get(f"    [警告] ASIN={asin} 返回数据中无投放链接")
                    else:
                        # data为空列表
                        self.log_get(f"    [警告] ASIN={asin} API返回空数据")
                else:
                    # API返回错误码
                    error_msg = data.get("status", {}).get("message", "未知错误")
                    self.log_get(f"    [错误] ASIN={asin} API错误: {error_msg}")
            else:
                self.log_get(f"    [错误] ASIN={asin} HTTP状态码: {response.status_code}")
            return ""
        except Exception as e:
            self.log_get(f"    [异常] ASIN={asin} 获取链接失败: {str(e)}")
            return ""
    
    def extract_offers(self):
        if not self.pb_token_var.get().strip():
            messagebox.showerror("错误", "请在API配置中输入PartnerBoost Offer Token!")
            return
        
        self.extract_btn.config(state='disabled')
        self.progress_get.start()
        self.log_text_get.delete(1.0, tk.END)
        
        thread = threading.Thread(target=self._do_extract)
        thread.daemon = True
        thread.start()
    
    def _do_extract(self):
        all_offers = []
        page = 1
        has_more = True
        
        try:
            endpoint = "/api/datafeed/get_fba_products"
            full_url = PB_API_BASE_URL + endpoint
            self.log_get(f"开始获取Offer数据...")
            
            headers = {"Content-Type": "application/json", "Accept": "application/json"}
            
            while has_more:
                request_body = self.build_request_body(page)
                self.root.after(0, lambda p=page, t=len(all_offers): self.update_progress_get(f"第{p}页 | 已获取{t}条"))
                
                response = requests.post(full_url, json=request_body, headers=headers, timeout=60)
                
                if response.status_code == 200:
                    data = response.json()
                    if data.get("status", {}).get("code") == 0:
                        offers = data.get("data", {}).get("list", [])
                        has_more = data.get("data", {}).get("has_more", False)
                        all_offers.extend(offers)
                        self.log_get(f"第{page}页: 获取 {len(offers)} 个offer，累计 {len(all_offers)} 个")
                        page += 1
                    else:
                        self.log_get(f"API返回错误: {data.get('status', {}).get('msg')}")
                        break
                else:
                    self.log_get(f"HTTP请求失败: {response.status_code}")
                    break
            
            self.log_get(f"Offer数据获取完成，共 {len(all_offers)} 个")
            
            if all_offers:
                self.log_get(f"开始获取原始投放链接...")
                
                total = len(all_offers)
                for idx, offer in enumerate(all_offers):
                    asin = offer.get("asin", "")
                    country_code = offer.get("country_code", "")
                    if asin and country_code:
                        self.root.after(0, lambda i=idx+1, t=total: self.update_progress_get(f"获取链接 {i}/{t}"))
                        link = self.get_partnerboost_link(asin, country_code)
                        offer["partnerboost_link"] = link
                        
                        # 访问投放链接获取重定向后的最终URL，替换产品链接
                        if link:
                            redirect_url = self.resolve_redirect_url(link)
                            if redirect_url:
                                offer["url"] = redirect_url
                        
                        if (idx + 1) % 10 == 0:
                            self.log_get(f"已获取 {idx + 1}/{total} 个投放链接")
                        time.sleep(0.1)
                    else:
                        offer["partnerboost_link"] = ""
                
                self.save_to_xlsx(all_offers)
            else:
                self.log_get("没有获取到任何offer数据")
                
        except Exception as e:
            self.log_get(f"错误: {str(e)}")
        finally:
            self.root.after(0, self._restore_get_ui)
    
    def _restore_get_ui(self):
        self.extract_btn.config(state='normal')
        self.update_yp_links_btn.config(state='normal')
        self.copy_offer_btn.config(state='normal')
        self.brand_search_btn.config(state='normal')
        self.progress_get.stop()
        self.update_progress_get("")

    def update_yp_links(self):
        """回填YP投放链接对应的产品链接。"""
        if not self.feishu_app_id_var.get().strip() or not self.feishu_app_secret_var.get().strip():
            messagebox.showerror("错误", "请在API配置中输入飞书App ID和App Secret!")
            return

        self.extract_btn.config(state='disabled')
        self.update_yp_links_btn.config(state='disabled')
        self.copy_offer_btn.config(state='disabled')
        self.brand_search_btn.config(state='disabled')
        self.progress_get.start()
        self.log_text_get.delete(1.0, tk.END)

        thread = threading.Thread(target=self._do_update_yp_links)
        thread.daemon = True
        thread.start()

    def _is_yp_tracking_link(self, link):
        link = str(link or '').strip()
        return ('https://yeahpromos.com/' in link) or ('https://www.yeahpromos.com/' in link)

    def _is_pb_tracking_link(self, link):
        """仅识别明确的PB投放链接，空链接和非PB链接都不算PB。"""
        normalized_link = str(link or '').strip().lower()
        if not normalized_link:
            return False
        return (
            'https://pboost.me/' in normalized_link or
            'http://pboost.me/' in normalized_link or
            'https://www.pboost.me/' in normalized_link or
            'http://www.pboost.me/' in normalized_link or
            'https://partnerboost.com/' in normalized_link or
            'http://partnerboost.com/' in normalized_link or
            'https://www.partnerboost.com/' in normalized_link or
            'http://www.partnerboost.com/' in normalized_link
        )

    def build_copied_offer_tracking_link(self, asin, country, source_tracking_link):
        """为复制offer决定投放链接。YP保留原链接，PB生成新UID链接。"""
        normalized_source_link = str(source_tracking_link or '').strip()
        if self._is_yp_tracking_link(normalized_source_link):
            return normalized_source_link, False, ''

        uid = self.generate_random_uid()
        new_link = self.get_partnerboost_link(asin, country, uid)
        return new_link, True, uid

    def is_pb_offer_row(self, row, col_indices):
        """判断飞书offer行是否属于PB且不是统计行。"""
        status_idx = (col_indices or {}).get('状态')
        if status_idx is not None and status_idx < len(row):
            status = str(self.normalize_sheet_cell_value(row[status_idx]) or '').strip()
            if status == SUMMARY_STATUS_TEXT:
                return False

        tracking_link_idx = (col_indices or {}).get('投放链接')
        if tracking_link_idx is None:
            return False
        if tracking_link_idx >= len(row):
            return False
        tracking_link = str(self.normalize_sheet_cell_value(row[tracking_link_idx]) or '').strip()
        return self._is_pb_tracking_link(tracking_link)

    def build_copy_offer_match_key(self, row, col_indices):
        """为复制offer功能构建源offer匹配键：ASIN + 国家代码 + 品牌ID。"""
        asin_idx = (col_indices or {}).get('ASIN')
        country_idx = (col_indices or {}).get('国家代码')
        brand_id_idx = (col_indices or {}).get('品牌ID')

        asin = str(row[asin_idx]).strip() if asin_idx is not None and asin_idx < len(row) and row[asin_idx] else ''
        country = str(row[country_idx]).strip().upper() if country_idx is not None and country_idx < len(row) and row[country_idx] else ''
        brand_id = str(row[brand_id_idx]).strip() if brand_id_idx is not None and brand_id_idx < len(row) and row[brand_id_idx] else ''

        if not asin or not country or not brand_id:
            return None
        return asin, country, brand_id

    def copy_offer_has_non_key_data(self, row, col_indices, copy_fields):
        """判断复制offer行是否包含匹配键之外的有效数据。"""
        identity_fields = {'品牌ID'}
        for field in copy_fields or []:
            if field in identity_fields:
                continue
            if field in col_indices:
                field_col = col_indices[field]
                if field_col < len(row) and row[field_col]:
                    cell_value = str(row[field_col]).strip()
                    if cell_value and cell_value.lower() not in ['none', '']:
                        return True
        return False

    def extract_tracking_uid_from_link(self, link):
        """从投放链接提取UID，兼容PB的uid参数和YP的 <uid>={tag1} 格式。"""
        if isinstance(link, list) and len(link) > 0 and isinstance(link[0], dict):
            link = link[0].get('link', '') or link[0].get('text', '')
        link = str(link or '').strip()
        if not link:
            return ''

        uid_match = re.search(r'[?&]uid=([^&]+)', link, re.IGNORECASE)
        if uid_match:
            uid_value = uid_match.group(1)
            if uid_value.lower() not in ('{tag1}', '%7btag1%7d'):
                return uid_value

        yp_match = re.search(r'[?&]([^=&?#]+)=\{tag1\}(?:[&#]|$)', link, re.IGNORECASE)
        if yp_match:
            return yp_match.group(1)

        yp_encoded_match = re.search(r'[?&]([^=&?#]+)=%7Btag1%7D(?:[&#]|$)', link, re.IGNORECASE)
        if yp_encoded_match:
            return yp_encoded_match.group(1)

        return ''

    def normalize_sheet_cell_value(self, value):
        """标准化飞书单元格为简单字符串值。"""
        if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
            return value[0].get('link', '') or value[0].get('text', '')
        return value if value is not None else ''

    def build_yp_brand_key(self, brand_id='', brand_name=''):
        """YP没有品牌ID时，用品牌名生成稳定分组键。"""
        brand_id = str(brand_id or '').strip()
        if brand_id and brand_id.lower() not in ('none', 'null'):
            return brand_id
        brand_key = self.normalize_brand_key(brand_name or '')
        return f"name:{brand_key}" if brand_key else ''

    def build_yp_row_brand_key(self, row):
        row = row or {}
        return self.build_yp_brand_key(row.get('品牌ID', ''), row.get('品牌名称', ''))

    def build_yp_transaction_brand_key(self, trans, fallback_entry=None):
        trans = trans or {}
        fallback_entry = fallback_entry or {}
        brand_ids = sorted(fallback_entry.get('brand_ids', set()) or [])
        return self.build_yp_brand_key(
            trans.get('brand_id', '') or trans.get('bid', '') or (brand_ids[0] if brand_ids else ''),
            trans.get('advert_name', '') or trans.get('merchant_name', '') or trans.get('brand_name', '') or fallback_entry.get('brand', '')
        )

    def display_offer_brand_id(self, brand_id):
        """内部YP品牌名键不写入用户可见的品牌ID位置。"""
        brand_id = str(brand_id or '').strip()
        return '' if brand_id.startswith('name:') else brand_id

    def get_column_index_by_header(self, header):
        """根据表头名返回0-based列索引。"""
        col_letter = getattr(self, 'feishu_column_map', {}).get(header)
        if not col_letter:
            return None
        return self.column_letter_to_index(col_letter)

    def column_letter_to_index(self, col_letter):
        """将列字母转换为0-based列索引。"""
        if not col_letter:
            return None
        index = 0
        for ch in str(col_letter).strip():
            if not ('A' <= ch.upper() <= 'Z'):
                return None
            index = index * 26 + (ord(ch.upper()) - ord('A') + 1)
        return index - 1

    def extract_yp_offer_marker(self, tracking_link='', product_link=''):
        """提取用于区分YP复制offer的稳定标识，优先使用产品链接中的aa_adgroupid。"""
        normalized_product_link = str(self.normalize_sheet_cell_value(product_link) or '').strip()
        if normalized_product_link:
            adgroupid_match = re.search(r'[?&]aa_adgroupid=([^&]+)', normalized_product_link, re.IGNORECASE)
            if adgroupid_match:
                return f"adg:{adgroupid_match.group(1)}"

        normalized_tracking_link = str(self.normalize_sheet_cell_value(tracking_link) or '').strip()
        if not normalized_tracking_link:
            return ''

        uid = self.extract_tracking_uid_from_link(normalized_tracking_link)
        if uid:
            return f"uid:{uid}"

        pid_match = re.search(r'[?&]pid=([^&]+)', normalized_tracking_link, re.IGNORECASE)
        if pid_match:
            return f"pid:{pid_match.group(1)}"

        return ''

    def get_yp_campaign_group_key(self, offer_key, offer_summary=None):
        """广告系列表按同一个YP投放链接聚合；Offer表仍可用aa_adgroupid区分行。"""
        if not offer_key:
            return None
        return (offer_key[0], offer_key[1], offer_key[2])

    def build_offer_row_groups(self, feishu_data):
        """构建offer表中有效数据行的分组信息。"""
        asin_country_rows = {}
        yp_asin_brand_rows = {}
        row_by_index = {}
        row_sequence = []

        for row in feishu_data:
            row_index = row.get('row_index')
            if not row_index or row_index <= 2:
                continue

            status = str(row.get('状态', '') or '').strip()
            asin = str(row.get('ASIN', '') or '').strip()
            country = self.normalize_country_code(row.get('国家代码', '') or '')
            brand_id = self.build_yp_row_brand_key(row)
            tracking_link = str(self.normalize_sheet_cell_value(row.get('投放链接', '')) or '').strip()

            if not asin or status == SUMMARY_STATUS_TEXT:
                continue

            row_by_index[row_index] = row
            row_sequence.append(row_index)

            if country:
                asin_country_rows.setdefault((asin, country), []).append(row_index)

            if self._is_yp_tracking_link(tracking_link) and brand_id:
                yp_marker = self.extract_yp_offer_marker(tracking_link, row.get('产品链接', ''))
                yp_asin_brand_rows.setdefault((asin, brand_id, country, yp_marker), []).append(row_index)

        return {
            'asin_country_rows': asin_country_rows,
            'yp_asin_brand_rows': yp_asin_brand_rows,
            'row_by_index': row_by_index,
            'row_sequence': row_sequence,
        }

    def build_yp_offer_group_context(self, feishu_data, offer_row_commissions=None):
        """构建YP offer分组上下文，供广告系列表分组与统计行复用。"""
        context = {
            'offer_group_summary_by_key': {},
            'campaign_name_to_offer_keys': {},
            'campaign_id_to_offer_keys': {},
            'tracking_link_to_offer_keys': {},
        }

        def ensure_group(asin, brand_id, country, yp_marker='', brand_name=''):
            asin = str(asin or '').strip()
            brand_id = self.build_yp_brand_key(brand_id, brand_name)
            country = self.normalize_country_code(country or '')
            yp_marker = str(yp_marker or '').strip()
            brand_name = str(brand_name or '').strip()
            if not asin or not brand_id or not country:
                return None

            offer_key = (asin, brand_id, country, yp_marker)
            group_summary = context['offer_group_summary_by_key'].setdefault(offer_key, {
                'asin': asin,
                'brand_id': brand_id,
                'brand_name': brand_name,
                'country': country,
                'yp_marker': yp_marker,
                'tracking_link': '',
                'is_yp': True,
                'commission': 0.0,
            })
            if brand_name and not group_summary.get('brand_name'):
                group_summary['brand_name'] = brand_name
            return offer_key, group_summary

        def remember_tracking_link(link, offer_key):
            normalized_link = str(self.normalize_sheet_cell_value(link) or '').strip()
            if not normalized_link:
                return
            context['tracking_link_to_offer_keys'].setdefault(normalized_link, set()).add(offer_key)
            group_summary = context['offer_group_summary_by_key'].get(offer_key)
            if group_summary is not None and not group_summary.get('tracking_link'):
                group_summary['tracking_link'] = normalized_link

        def remember_campaign_names(campaign_names, offer_key):
            for campaign_name in campaign_names or []:
                campaign_name = str(campaign_name or '').strip()
                if campaign_name:
                    context['campaign_name_to_offer_keys'].setdefault(campaign_name, set()).add(offer_key)

        for row in feishu_data or []:
            if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                continue

            tracking_link = str(self.normalize_sheet_cell_value(row.get('投放链接', '')) or '').strip()
            if not self._is_yp_tracking_link(tracking_link):
                continue

            group_result = ensure_group(
                row.get('ASIN', ''),
                row.get('品牌ID', ''),
                row.get('国家代码', ''),
                self.extract_yp_offer_marker(tracking_link, row.get('产品链接', '')),
                row.get('品牌名称', '')
            )
            if not group_result:
                continue

            offer_key, _ = group_result
            remember_tracking_link(tracking_link, offer_key)
            remember_campaign_names(
                [name.strip() for name in str(row.get('广告系列名称', '') or '').split(',') if name.strip()],
                offer_key
            )

        for row_info in (offer_row_commissions or {}).values():
            if not row_info.get('is_yp'):
                continue

            group_result = ensure_group(
                row_info.get('asin', ''),
                row_info.get('brand_id', ''),
                row_info.get('country', ''),
                self.extract_yp_offer_marker(row_info.get('tracking_link', ''), row_info.get('product_link', '')),
                row_info.get('brand_name', '')
            )
            if not group_result:
                continue

            offer_key, group_summary = group_result
            group_summary['commission'] += float(row_info.get('commission', 0.0) or 0.0)

            remember_tracking_link(row_info.get('tracking_link', ''), offer_key)
            remember_campaign_names(row_info.get('campaign_names', []), offer_key)
            for campaign_id in row_info.get('campaign_ids', []) or []:
                campaign_id = str(campaign_id or '').strip()
                if campaign_id:
                    context['campaign_id_to_offer_keys'].setdefault(campaign_id, set()).add(offer_key)

        return context

    def resolve_yp_campaign_offer_group(self, campaign_name, campaign_info, offer_context, tracking_link=''):
        """解析YP广告系列所属的offer分组。"""
        offer_context = offer_context or {}
        summary_by_key = offer_context.get('offer_group_summary_by_key', {})
        campaign_name_to_offer_keys = offer_context.get('campaign_name_to_offer_keys', {})
        campaign_id_to_offer_keys = offer_context.get('campaign_id_to_offer_keys', {})
        tracking_link_to_offer_keys = offer_context.get('tracking_link_to_offer_keys', {})

        campaign_name = str(campaign_name or '').strip()
        campaign_info = campaign_info or {}
        campaign_id = str(campaign_info.get('campaign_id', '') or '').strip()
        asin = str(campaign_info.get('asin', '') or '').strip()
        country = self.normalize_country_code(campaign_info.get('country', '') or '')
        campaign_marker = self.extract_yp_offer_marker(
            tracking_link or campaign_info.get('tracking_link', ''),
            campaign_info.get('product_link', '')
        )

        candidate_keys = set()
        if campaign_name:
            candidate_keys.update(campaign_name_to_offer_keys.get(campaign_name, set()))
        if campaign_id:
            candidate_keys.update(campaign_id_to_offer_keys.get(campaign_id, set()))

        normalized_links = []
        for value in (tracking_link, campaign_info.get('tracking_link', '')):
            normalized_link = str(self.normalize_sheet_cell_value(value) or '').strip()
            if normalized_link and normalized_link not in normalized_links:
                normalized_links.append(normalized_link)
                candidate_keys.update(tracking_link_to_offer_keys.get(normalized_link, set()))

        has_explicit_yp_signal = bool(candidate_keys) or any(
            self._is_yp_tracking_link(link) for link in normalized_links
        )

        if len(candidate_keys) == 1:
            offer_key = next(iter(candidate_keys))
            offer_summary = summary_by_key.get(offer_key, {})
            return self.get_yp_campaign_group_key(offer_key, offer_summary), offer_summary

        if asin and country and candidate_keys:
            matched_keys = [key for key in candidate_keys if key[0] == asin and key[2] == country]
            if campaign_marker:
                marker_matched_keys = [key for key in matched_keys if (len(key) > 3 and key[3] == campaign_marker)]
                if len(marker_matched_keys) == 1:
                    offer_key = marker_matched_keys[0]
                    offer_summary = summary_by_key.get(offer_key, {})
                    return self.get_yp_campaign_group_key(offer_key, offer_summary), offer_summary
            if len(matched_keys) == 1:
                offer_key = matched_keys[0]
                offer_summary = summary_by_key.get(offer_key, {})
                return self.get_yp_campaign_group_key(offer_key, offer_summary), offer_summary

        if has_explicit_yp_signal and asin and country:
            matched_keys = [key for key in summary_by_key.keys() if key[0] == asin and key[2] == country]
            if campaign_marker:
                marker_matched_keys = [key for key in matched_keys if (len(key) > 3 and key[3] == campaign_marker)]
                if len(marker_matched_keys) == 1:
                    offer_key = marker_matched_keys[0]
                    offer_summary = summary_by_key.get(offer_key, {})
                    return self.get_yp_campaign_group_key(offer_key, offer_summary), offer_summary
            if len(matched_keys) == 1:
                offer_key = matched_keys[0]
                offer_summary = summary_by_key.get(offer_key, {})
                return self.get_yp_campaign_group_key(offer_key, offer_summary), offer_summary

        return None, {}

    def summarize_campaign_group_rows(self, row_by_index, row_indices):
        """汇总广告系列表中同一offer分组的展示数据。"""
        metrics = {
            'total_cost': 0.0,
            'total_commission': 0.0,
            'total_clicks': 0,
            'commission_asins': [],
            'brand_break_even_cpc': '',
            'increment_cost': 0.0,
            'increment_commission': 0.0,
        }

        commission_asin_seen = set()
        break_even_seen = set()

        for row_index in row_indices or []:
            row_data = row_by_index.get(row_index, {}) or {}
            status_text = str(row_data.get('状态', '') or '').strip()
            total_commission_raw = str(row_data.get('总佣金', '') or '').strip()
            metrics['total_cost'] += self.parse_commission_value(row_data.get('广告系列总花费', ''))
            metrics['total_commission'] += self.parse_commission_value(total_commission_raw)
            metrics['total_clicks'] += int(self.parse_commission_value(row_data.get('总点击数', '')) or 0)
            metrics['increment_cost'] += self.parse_commission_value(row_data.get('新增广告系列花费', ''))
            if total_commission_raw != '↑':
                metrics['increment_commission'] += self.parse_commission_value(row_data.get('新增佣金', ''))

            commission_asins_value = str(row_data.get('佣金ASIN', '') or '').strip()
            if commission_asins_value:
                for asin_item in commission_asins_value.split(','):
                    asin_item = asin_item.strip()
                    if asin_item and asin_item not in commission_asin_seen:
                        commission_asin_seen.add(asin_item)
                        metrics['commission_asins'].append(asin_item)

            if status_text.startswith('投放已结束'):
                continue

            break_even_value = str(row_data.get('品牌收支平衡CPC', '') or '').strip()
            if break_even_value and break_even_value not in break_even_seen:
                break_even_seen.add(break_even_value)
                if not metrics['brand_break_even_cpc']:
                    metrics['brand_break_even_cpc'] = break_even_value

        return metrics

    def parse_campaign_summary_label(self, value):
        """解析广告系列统计行标题，兼容旧格式和新格式。"""
        parts = [part.strip() for part in str(value or '').split('|')]
        if len(parts) >= 9:
            return {
                'brand_name': parts[0],
                'asin': parts[1],
                'brand_id': parts[2],
                'country': self.normalize_country_code(parts[3]),
                'platform': parts[4],
                'status_counts': ' | '.join(parts[5:8]),
                'marker': parts[8] if len(parts) >= 9 else '',
            }
        if len(parts) >= 3:
            return {
                'brand_name': '',
                'asin': parts[0],
                'brand_id': parts[1],
                'country': self.normalize_country_code(parts[2]),
                'platform': '',
                'status_counts': '',
                'marker': parts[3] if len(parts) >= 4 else '',
            }
        return {}

    def normalize_campaign_summary_key(self, key):
        """规范化广告系列统计行分组key。"""
        if not key or len(key) < 3:
            return None
        asin = str(key[0] or '').strip().upper()
        brand_id = str(key[1] or '').strip()
        country = self.normalize_country_code(key[2] or '')
        marker = ''
        if len(key) > 3:
            marker = str(key[3] or '').strip()
            if marker == '-':
                marker = ''
        if not asin or not country:
            return None
        return (asin, brand_id, country, marker)

    def build_campaign_commission_baseline_snapshot(self, rows):
        """读取运行前广告系列表佣金快照，用于新增佣金差值口径。"""
        snapshot = {
            'campaign_commission': {},
            'summary_commission': {},
            'summary_children': {},
        }
        row_by_index = {}
        summary_rows = []

        for row in rows or []:
            row_index = row.get('row_index')
            if row_index:
                row_by_index[row_index] = row

            status = str(row.get('状态', '') or '').strip()
            campaign_name = str(row.get('广告系列名称', '') or '').strip()
            if status == SUMMARY_STATUS_TEXT:
                parsed = self.parse_campaign_summary_label(campaign_name)
                summary_key = self.normalize_campaign_summary_key((
                    parsed.get('asin', ''),
                    parsed.get('brand_id', ''),
                    parsed.get('country', ''),
                    parsed.get('marker', ''),
                )) if parsed else None
                if summary_key:
                    summary_rows.append((row_index, summary_key))
                    snapshot['summary_commission'][summary_key] = self.parse_commission_value(row.get('总佣金', ''))
                continue

            if campaign_name and campaign_name != SUMMARY_PLACEHOLDER_TEXT:
                snapshot['campaign_commission'][campaign_name] = self.parse_commission_value(row.get('总佣金', ''))

        for summary_row_index, summary_key in summary_rows:
            children = []
            for row_index in sorted(row_by_index):
                if row_index <= summary_row_index:
                    continue
                row_data = row_by_index.get(row_index, {}) or {}
                status = str(row_data.get('状态', '') or '').strip()
                if status == SUMMARY_STATUS_TEXT:
                    break
                campaign_name = str(row_data.get('广告系列名称', '') or '').strip()
                if campaign_name and campaign_name != SUMMARY_PLACEHOLDER_TEXT:
                    children.append(campaign_name)
            snapshot['summary_children'][summary_key] = children

        return snapshot

    def apply_campaign_increment_commission_delta(self, rows, baseline_snapshot, grouped_offer_keys=None):
        """按“本次总佣金 - 运行前总佣金”覆盖广告系列行新增佣金。"""
        previous_by_campaign = (baseline_snapshot or {}).get('campaign_commission', {})
        grouped_offer_keys = {
            self.get_yp_campaign_group_key(key, {}) or key
            for key in (grouped_offer_keys or [])
            if key
        }
        for item in rows or []:
            campaign_name = str(item.get('campaign_name') or item.get('广告系列名称') or '').strip()
            if not campaign_name:
                continue
            item_group_key = self.get_yp_campaign_group_key(item.get('_offer_group_key'), {}) or item.get('_offer_group_key')
            if item_group_key in grouped_offer_keys:
                item['新增佣金'] = '↑'
                item['总佣金'] = '↑'
                continue
            current_total = self.parse_commission_value(item.get('总佣金', ''))
            previous_total = float(previous_by_campaign.get(campaign_name, 0.0) or 0.0)
            item['新增佣金'] = f"${current_total - previous_total:.2f}"

    def get_campaign_summary_previous_commission(self, offer_key, child_campaign_names, baseline_snapshot):
        """计算统计行差值基准，兼容单行升级为新统计行和统计行新增子广告系列。"""
        normalized_key = self.normalize_campaign_summary_key(offer_key)
        if not normalized_key:
            return 0.0

        summary_commission = (baseline_snapshot or {}).get('summary_commission', {})
        summary_children = (baseline_snapshot or {}).get('summary_children', {})
        campaign_commission = (baseline_snapshot or {}).get('campaign_commission', {})

        child_names = []
        for name in child_campaign_names or []:
            name = str(name or '').strip()
            if name and name not in child_names:
                child_names.append(name)

        if normalized_key in summary_commission:
            previous_total = float(summary_commission.get(normalized_key, 0.0) or 0.0)
            old_children = set(summary_children.get(normalized_key, []) or [])
            for campaign_name in child_names:
                if campaign_name not in old_children:
                    previous_total += float(campaign_commission.get(campaign_name, 0.0) or 0.0)
            return previous_total

        return sum(float(campaign_commission.get(name, 0.0) or 0.0) for name in child_names)

    def build_campaign_summary_label(self, brand_name, asin, brand_id, country, platforms, status_counts, marker=''):
        platform_text = ','.join(platforms or [])
        parts = [
            str(brand_name or '').strip() or '-',
            str(asin or '').strip(),
            str(brand_id or '').strip(),
            self.normalize_country_code(country or ''),
            platform_text or '-',
            status_counts or '0 | 0 | 0',
        ]
        marker = str(marker or '').strip() or '-'
        parts.append(marker)
        return ' | '.join(parts)

    def apply_yp_group_commissions_to_campaign_rows(self, updates, new_rows, offer_group_summary_by_key, existing_campaign_commission=None, existing_campaign_cost=None, increment_yp_commission_by_group=None):
        """把已落到offer分组的YP佣金补写到广告系列行。

        规则：
        - 若某个YP offer分组只对应 1 个广告系列行，则把该组佣金直接写入该广告系列行。
        - 若对应多个广告系列行，则佣金保留给后续“相同offer统计行”展示，避免重复落到多个广告系列。
        """
        existing_campaign_cost = existing_campaign_cost or {}

        offer_group_to_rows = {}
        for item in (updates or []) + (new_rows or []):
            offer_key = item.get('_offer_group_key')
            campaign_name = str(item.get('campaign_name', '') or '').strip()
            if not offer_key or not campaign_name:
                continue
            offer_group_to_rows.setdefault(offer_key, [])
            if item not in offer_group_to_rows[offer_key]:
                offer_group_to_rows[offer_key].append(item)

        applied_rows = 0
        skipped_groups = 0

        campaign_summary_by_key = {}
        for offer_key, group_summary in (offer_group_summary_by_key or {}).items():
            campaign_key = self.get_yp_campaign_group_key(offer_key, group_summary)
            if not campaign_key:
                continue
            if campaign_key not in campaign_summary_by_key:
                campaign_summary_by_key[campaign_key] = {
                    **(group_summary or {}),
                    'commission': float((group_summary or {}).get('commission', 0.0) or 0.0),
                }
            else:
                campaign_summary = campaign_summary_by_key[campaign_key]
                campaign_summary['commission'] = (
                    float(campaign_summary.get('commission', 0.0) or 0.0)
                    + float((group_summary or {}).get('commission', 0.0) or 0.0)
                )

        for offer_key, group_summary in campaign_summary_by_key.items():
            yp_commission = float(group_summary.get('commission', 0.0) or 0.0)
            if yp_commission <= 0:
                continue

            targets = offer_group_to_rows.get(offer_key, [])
            if len(targets) != 1:
                if len(targets) > 1:
                    skipped_groups += 1
                continue

            target = targets[0]
            campaign_name = str(target.get('campaign_name', '') or '').strip()
            current_commission = self.parse_commission_value(target.get('总佣金', ''))
            total_commission = current_commission + yp_commission
            target['总佣金'] = f"${total_commission:.2f}"
            if not str(target.get('新增佣金', '') or '').strip():
                target['新增佣金'] = f"${yp_commission:.2f}"

            cost_value = self.parse_commission_value(target.get('广告系列总花费', ''))
            if cost_value <= 0:
                cost_value = float(existing_campaign_cost.get(campaign_name, 0.0) or 0.0)
            roi_value = round(total_commission / cost_value, 1) if cost_value > 0 else 0
            target['ROI'] = f"{roi_value}"

            asin_country_value = f"{offer_key[0]}_{offer_key[2]}"
            existing_asins = [
                item.strip() for item in str(target.get('佣金ASIN', '') or '').split(',')
                if item.strip()
            ]
            if asin_country_value not in existing_asins:
                existing_asins.append(asin_country_value)
            target['佣金ASIN'] = ', '.join(existing_asins)

            applied_rows += 1

        return {
            'applied_rows': applied_rows,
            'skipped_groups': skipped_groups,
        }

    def get_pb_offer_row_campaign_commission(self, row_info):
        """取PB Offer行可转给广告系列表的佣金，排除YP和已按UID精确归因的佣金。"""
        if not row_info or row_info.get('is_yp'):
            return 0.0

        commission = (
            float(row_info.get('non_uid_commission', 0.0) or 0.0)
            + float(row_info.get('asin_only_commission', 0.0) or 0.0)
        )
        for item in row_info.get('uid_allocations', []) or []:
            if item.get('match_type') == 'asin_country_fallback':
                commission += float(item.get('commission', 0.0) or 0.0)
        return commission

    def build_pb_no_uid_commission_map(self, commission_data):
        """按(ASIN, 国家)汇总PB佣金，用于无UID场景复用Offer表归因。"""
        result = {}
        for trans in commission_data or []:
            if str(trans.get('status', '') or '').strip() == 'Rejected':
                continue
            asin = str(trans.get('prod_id', '') or trans.get('asin', '') or '').strip().upper()
            if not asin:
                continue
            country = self.normalize_country_code(
                trans.get('customer_country', '') or trans.get('geo', '') or trans.get('country', '')
            )
            if not country and trans.get('merchant_name', ''):
                country = self.normalize_country_code(self.extract_country_from_merchant_name(trans.get('merchant_name', '')))
            if not country and trans.get('mcid', ''):
                country = self.normalize_country_code(self.extract_country_from_mcid(trans.get('mcid', '')))
            if not country:
                country = 'US'
            key = (asin, country)
            result[key] = result.get(key, 0.0) + float(trans.get('sale_comm', 0.0) or 0.0)
        return result

    def apply_pb_offer_commissions_to_campaign_rows(self, updates, new_rows, offer_row_commissions, existing_campaign_cost=None, increment_pb_commission_by_campaign=None):
        """把无UID PB佣金从Offer行归因到广告系列表。

        单广告系列Offer直接写到该广告系列；多广告系列Offer保留给相同offer统计行汇总。
        """
        existing_campaign_cost = existing_campaign_cost or {}
        row_items = {}
        for item in (updates or []) + (new_rows or []):
            campaign_name = str(item.get('campaign_name', '') or '').strip()
            if campaign_name:
                row_items[campaign_name] = item

        applied_rows = 0
        skipped_groups = 0
        for row_info in (offer_row_commissions or {}).values():
            commission = self.get_pb_offer_row_campaign_commission(row_info)
            if commission <= 0:
                continue

            campaign_names = []
            for campaign_name in row_info.get('campaign_names', []) or []:
                campaign_name = str(campaign_name or '').strip()
                if campaign_name and campaign_name not in campaign_names:
                    campaign_names.append(campaign_name)

            if len(campaign_names) != 1:
                if len(campaign_names) > 1:
                    skipped_groups += 1
                continue

            campaign_name = campaign_names[0]
            target = row_items.get(campaign_name)
            if not target:
                continue

            total_commission = self.parse_commission_value(target.get('总佣金', '')) + commission
            target['总佣金'] = f"${total_commission:.2f}"

            cost_value = self.parse_commission_value(target.get('广告系列总花费', ''))
            if cost_value <= 0:
                cost_value = float(existing_campaign_cost.get(campaign_name, 0.0) or 0.0)
            target['ROI'] = f"{round(total_commission / cost_value, 1) if cost_value > 0 else 0}"
            increment_commission = (increment_pb_commission_by_campaign or {}).get(campaign_name)
            if increment_commission is not None:
                target['新增佣金'] = f"${float(increment_commission or 0.0):.2f}"

            asin_country_value = f"{row_info.get('asin', '')}_{row_info.get('country', '')}"
            existing_asins = [
                item.strip() for item in str(target.get('佣金ASIN', '') or '').split(',')
                if item.strip()
            ]
            if asin_country_value.strip('_') and asin_country_value not in existing_asins:
                existing_asins.append(asin_country_value)
            target['佣金ASIN'] = ', '.join(existing_asins)
            applied_rows += 1

        return {'applied_rows': applied_rows, 'skipped_groups': skipped_groups}

    def build_break_even_brand_country_totals(self, offer_row_commissions, offer_group_summary_by_key, mcc_campaigns, removed_campaign_metrics=None, removed_campaign_sources=None):
        """按品牌+国家汇总佣金与点击。"""
        brand_country_totals = {}

        def ensure_aggregate(brand_name, country):
            brand_name = str(brand_name or '').strip()
            country = self.normalize_country_code(country or '')
            if not brand_name or not country:
                return None

            key = (self.normalize_brand_key(brand_name), country)
            aggregate = brand_country_totals.setdefault(key, {
                'brand': brand_name,
                'country': country,
                'commission': 0.0,
                'clicks': 0,
            })
            if not aggregate.get('brand'):
                aggregate['brand'] = brand_name
            return aggregate

        for row_info in (offer_row_commissions or {}).values():
            aggregate = ensure_aggregate(
                row_info.get('brand_name', ''),
                row_info.get('country', '')
            )
            if not aggregate:
                continue
            aggregate['commission'] += float(row_info.get('commission', 0.0) or 0.0)

        for campaign_name, campaign_info in (mcc_campaigns or {}).items():
            brand, country = self.extract_brand_and_country_from_campaign_name(campaign_name)
            aggregate = ensure_aggregate(brand, country)
            if not aggregate:
                continue
            aggregate['clicks'] += int(campaign_info.get('clicks', 0) or 0)

        for campaign_name, metrics in (removed_campaign_metrics or {}).items():
            brand, country = self.extract_brand_and_country_from_campaign_name(campaign_name)
            aggregate = ensure_aggregate(brand, country)
            if not aggregate:
                continue
            aggregate['clicks'] += int(metrics.get('clicks', 0) or 0)

        return brand_country_totals

    def normalize_brand_key(self, brand_name):
        """标准化品牌名，便于跨来源匹配。"""
        text = str(brand_name or '').strip()
        if not text:
            return ''
        text = re.sub(r'\s+', ' ', text)

        country_tokens = {
            'us', 'uk', 'gb', 'de', 'fr', 'it', 'es', 'ca', 'jp', 'au', 'nl', 'be',
            'mx', 'br', 'in', 'sg', 'ae', 'sa', 'pl', 'se', 'tr', 'eg',
            'united states', 'united kingdom', 'great britain', 'germany', 'france',
            'italy', 'spain', 'canada', 'japan', 'australia', 'netherlands', 'belgium',
            'mexico', 'brazil', 'india', 'singapore', 'uae', 'saudi arabia', 'poland',
            'sweden', 'turkey', 'egypt',
        }

        lower_text = text.lower()
        changed = True
        while changed and lower_text:
            changed = False
            for token in sorted(country_tokens, key=len, reverse=True):
                escaped = re.escape(token)
                patterns = [
                    rf'^{escaped}[\s\-_]+',
                    rf'[\s\-_]+{escaped}$',
                ]
                for pattern in patterns:
                    candidate = re.sub(pattern, '', lower_text, count=1).strip(' _-')
                    if candidate != lower_text:
                        lower_text = candidate
                        changed = True
                        break
                if changed:
                    break

        normalized = re.sub(r'\s+', ' ', lower_text).strip()
        brand_aliases = {
            'ogery-outdoors': 'ogery',
            'ogery outdoors': 'ogery',
        }
        return brand_aliases.get(normalized, normalized)

    def build_break_even_brand_country_offer_index(self, feishu_data, offer_row_commissions=None):
        """基于offer表构建品牌+国家映射，供品牌级佣金与点击汇总复用。"""
        offer_index = {}
        row_commissions = offer_row_commissions or {}
        row_by_index = {row.get('row_index'): row for row in (feishu_data or []) if row.get('row_index')}

        def ensure_entry(brand_name, brand_id, country, asin=''):
            normalized_brand = self.normalize_brand_key(brand_name)
            normalized_country = self.normalize_country_code(country or '')
            normalized_brand_id = str(brand_id or '').strip()
            normalized_asin = str(asin or '').strip().upper()
            if not normalized_country or (not normalized_brand and not normalized_brand_id):
                return None

            key = (normalized_brand, normalized_country)
            entry = offer_index.setdefault(key, {
                'brand': str(brand_name or '').strip(),
                'brand_lower': normalized_brand,
                'country': normalized_country,
                'brand_ids': set(),
                'asins': set(),
            })
            if brand_name and not entry.get('brand'):
                entry['brand'] = str(brand_name).strip()
            if normalized_brand_id:
                entry['brand_ids'].add(normalized_brand_id)
            if normalized_asin:
                entry['asins'].add(normalized_asin)
            return entry

        for row in (feishu_data or []):
            if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                continue
            ensure_entry(
                row.get('品牌名称', ''),
                row.get('品牌ID', ''),
                row.get('国家代码', ''),
                row.get('ASIN', '')
            )

        for row_info in row_commissions.values():
            row_data = row_by_index.get(row_info.get('row_index'))
            brand_name = row_info.get('brand_name', '')
            brand_id = row_info.get('brand_id', '')
            country = row_info.get('country', '')
            asin = row_info.get('asin', '')
            if row_data:
                brand_name = brand_name or row_data.get('品牌名称', '')
                brand_id = brand_id or row_data.get('品牌ID', '')
                country = country or row_data.get('国家代码', '')
                asin = asin or row_data.get('ASIN', '')
            ensure_entry(brand_name, brand_id, country, asin)

        return offer_index

    def build_brand_country_lookup_maps(self, offer_index):
        """把offer索引拆成多种查找维度。"""
        lookup = {
            'by_brand_id_country': {},
            'by_asin_brand_id_candidates': {},
            'by_asin_brand_id_country': {},
            'by_asin_country': {},
            'by_asin_brand': {},
            'by_brand': {},
            'by_brand_country': {},
            'by_country_brand': {},
        }

        for key, entry in (offer_index or {}).items():
            brand_key, country = key
            if not country:
                continue
            lookup['by_brand_country'][key] = entry
            if brand_key:
                lookup['by_country_brand'].setdefault(country, []).append(entry)
                lookup['by_brand'].setdefault(brand_key, [])
                if entry not in lookup['by_brand'][brand_key]:
                    lookup['by_brand'][brand_key].append(entry)
            for asin in entry.get('asins', set()):
                lookup['by_asin_country'].setdefault((asin, country), [])
                if entry not in lookup['by_asin_country'][(asin, country)]:
                    lookup['by_asin_country'][(asin, country)].append(entry)
                if brand_key:
                    lookup['by_asin_brand'].setdefault((asin, brand_key), [])
                    if entry not in lookup['by_asin_brand'][(asin, brand_key)]:
                        lookup['by_asin_brand'][(asin, brand_key)].append(entry)
            for brand_id in entry.get('brand_ids', set()):
                lookup['by_brand_id_country'][(brand_id, country)] = entry
                for asin in entry.get('asins', set()):
                    lookup['by_asin_brand_id_candidates'].setdefault((asin, brand_id), [])
                    if entry not in lookup['by_asin_brand_id_candidates'][(asin, brand_id)]:
                        lookup['by_asin_brand_id_candidates'][(asin, brand_id)].append(entry)
                    lookup['by_asin_brand_id_country'][(asin, brand_id, country)] = entry

        return lookup

    def resolve_yp_brand_country_entry(self, trans, offer_lookup):
        """根据YP交易和offer映射解析品牌+国家归属。"""
        offer_lookup = offer_lookup or {}
        by_asin_brand_id_country = offer_lookup.get('by_asin_brand_id_country', {})
        by_asin_brand_id_candidates = offer_lookup.get('by_asin_brand_id_candidates', {})
        by_brand_id_country = offer_lookup.get('by_brand_id_country', {})
        by_asin_country = offer_lookup.get('by_asin_country', {})
        by_asin_brand = offer_lookup.get('by_asin_brand', {})
        by_brand = offer_lookup.get('by_brand', {})
        by_country_brand = offer_lookup.get('by_country_brand', {})

        asin = str(trans.get('asin', '') or trans.get('prod_id', '') or '').strip().upper()
        if not asin:
            asin = self.extract_asin_from_yp_order_id(trans.get('id', ''))
        brand_id = str(trans.get('brand_id', '') or trans.get('bid', '')).strip()
        advert_name_key = self.normalize_brand_key(
            trans.get('advert_name', '') or trans.get('merchant_name', '') or trans.get('brand_name', '')
        )

        country_candidates = []
        for field in ('customer_country', 'geo', 'country', 'country_code'):
            candidate = self.normalize_country_code(trans.get(field, '') or '')
            if candidate and candidate not in country_candidates:
                country_candidates.append(candidate)

        merchant_country = self.normalize_country_code(
            self.extract_country_from_merchant_name(trans.get('merchant_name', ''))
        )
        if merchant_country and merchant_country not in country_candidates:
            country_candidates.append(merchant_country)

        for country in country_candidates:
            if asin and brand_id:
                entry = by_asin_brand_id_country.get((asin, brand_id, country))
                if entry:
                    return entry
            if brand_id:
                entry = by_brand_id_country.get((brand_id, country))
                if entry:
                    return entry
            if asin and advert_name_key:
                candidates = [
                    entry for entry in by_asin_country.get((asin, country), [])
                    if entry.get('brand_lower') == advert_name_key
                ]
                if len(candidates) == 1:
                    return candidates[0]

        if asin and brand_id:
            candidates = by_asin_brand_id_candidates.get((asin, brand_id), [])
            if len(candidates) == 1:
                return candidates[0]

        if asin and advert_name_key:
            candidates = []
            for country in country_candidates:
                for entry in by_asin_country.get((asin, country), []):
                    if entry.get('brand_lower') == advert_name_key and entry not in candidates:
                        candidates.append(entry)
            if len(candidates) == 1:
                return candidates[0]

            candidates = by_asin_brand.get((asin, advert_name_key), [])
            if len(candidates) == 1:
                return candidates[0]

        if advert_name_key:
            candidates = by_brand.get(advert_name_key, [])
            if len(candidates) == 1:
                return candidates[0]

        for country in country_candidates:
            country_entries = by_country_brand.get(country, [])
            if len(country_entries) == 1:
                return country_entries[0]

        return None

    def resolve_pb_brand_country_entry(self, trans, offer_lookup):
        """根据PB交易和offer映射解析品牌+国家归属。"""
        offer_lookup = offer_lookup or {}
        by_asin_brand_id_country = offer_lookup.get('by_asin_brand_id_country', {})
        by_brand_id_country = offer_lookup.get('by_brand_id_country', {})
        by_asin_brand_id_candidates = offer_lookup.get('by_asin_brand_id_candidates', {})
        by_brand_country = offer_lookup.get('by_brand_country', {})

        asin = str(trans.get('prod_id', '') or trans.get('asin', '') or '').strip().upper()
        brand_id = str(trans.get('brand_id', '') or '').strip()
        merchant_name = str(trans.get('merchant_name', '') or '').strip()
        parsed_brand_name = ''
        if merchant_name:
            parsed_brand_name = str(self._parse_brand_name_country(merchant_name)[0] or '').strip()

        country_candidates = []
        for field in ('customer_country', 'geo', 'country', 'country_code'):
            candidate = self.normalize_country_code(trans.get(field, '') or '')
            if candidate and candidate not in country_candidates:
                country_candidates.append(candidate)

        merchant_country = self.normalize_country_code(self.extract_country_from_merchant_name(merchant_name))
        if merchant_country and merchant_country not in country_candidates:
            country_candidates.append(merchant_country)

        mcid_country = self.normalize_country_code(self.extract_country_from_mcid(trans.get('mcid', '') or ''))
        if mcid_country and mcid_country not in country_candidates:
            country_candidates.append(mcid_country)

        if not country_candidates:
            country_candidates.append('US')

        for country in country_candidates:
            if asin and brand_id:
                entry = by_asin_brand_id_country.get((asin, brand_id, country))
                if entry:
                    return entry
            if brand_id:
                entry = by_brand_id_country.get((brand_id, country))
                if entry:
                    return entry
            if parsed_brand_name:
                entry = by_brand_country.get((self.normalize_brand_key(parsed_brand_name), country))
                if entry:
                    return entry

        if asin and brand_id:
            candidates = by_asin_brand_id_candidates.get((asin, brand_id), [])
            if len(candidates) == 1:
                return candidates[0]

        return None

    def calculate_break_even_brand_country_commissions(self, feishu_data, commission_data=None, yp_commission_data=None):
        """按品牌+国家汇总统计窗口内 PB+YP 总佣金。"""
        totals = {}
        offer_index = self.build_break_even_brand_country_offer_index(feishu_data)
        offer_lookup = self.build_brand_country_lookup_maps(offer_index)

        def ensure_total(entry):
            if not entry:
                return None
            key = (entry.get('brand_lower', ''), entry.get('country', ''))
            if not key[0] or not key[1]:
                return None
            return totals.setdefault(key, {
                'brand': entry.get('brand', ''),
                'country': entry.get('country', ''),
                'commission': 0.0,
                'clicks': 0,
            })

        for trans in (commission_data or []):
            if str(trans.get('status', '') or '').strip() == 'Rejected':
                continue
            entry = self.resolve_pb_brand_country_entry(trans, offer_lookup)
            aggregate = ensure_total(entry)
            if not aggregate:
                continue
            aggregate['commission'] += float(trans.get('sale_comm', 0.0) or 0.0)

        for trans in (yp_commission_data or []):
            status = str(trans.get('status', '') or '').strip().lower()
            if status == 'rejected':
                continue
            entry = self.resolve_yp_brand_country_entry(trans, offer_lookup)
            aggregate = ensure_total(entry)
            if not aggregate:
                continue
            aggregate['commission'] += float(trans.get('sale_comm', 0.0) or 0.0)

        return totals, offer_index

    def campaign_name_matches_brand_country(self, campaign_name, brand_name, country):
        """判断广告系列名称是否包含品牌名和国家代码。"""
        name = str(campaign_name or '').strip()
        brand = str(brand_name or '').strip()
        normalized_country = self.normalize_country_code(country or '')
        if not name or not brand or not normalized_country:
            return False

        campaign_country = self.extract_country_from_campaign_name(name)
        if campaign_country != normalized_country:
            return False

        tokens = [token.strip().lower() for token in re.split(r'[-_]', name) if token.strip()]
        normalized_brand = self.normalize_brand_key(brand)
        brand_tokens = [token.strip().lower() for token in re.split(r'[-_\s]+', normalized_brand) if token.strip()]
        if not brand_tokens:
            return False
        return all(token in tokens for token in brand_tokens)

    def get_brand_country_clicks(self, start_date_str, end_date_str, offer_index):
        """直接调用Google Ads API，汇总名称包含品牌名和国家代码的广告系列点击。"""
        brand_country_clicks = {
            key: {
                'brand': entry.get('brand', ''),
                'country': entry.get('country', ''),
                'clicks': 0,
            }
            for key, entry in (offer_index or {}).items()
        }
        if not brand_country_clicks:
            return brand_country_clicks

        try:
            sub_accounts = self.iter_google_ads_mcc_accounts()
            matched_campaigns = 0
            client_by_mcc = {}
            for account in sub_accounts:
                if self.stop_flag:
                    break
                try:
                    mcc_id = account.get('mcc_id', '')
                    client = client_by_mcc.get(mcc_id)
                    if client is None:
                        client = self.get_google_ads_client(mcc_id)
                        client_by_mcc[mcc_id] = client
                    if not client:
                        continue
                    ga_service = client.get_service('GoogleAdsService')
                    query = """
                        SELECT
                            campaign.name,
                            metrics.clicks
                        FROM campaign
                        WHERE campaign.name IS NOT NULL
                    """
                    if start_date_str and end_date_str:
                        query += f"""
                            AND segments.date >= '{start_date_str}'
                            AND segments.date <= '{end_date_str}'
                        """

                    campaign_response = ga_service.search(customer_id=account['id'], query=query)
                    for row in campaign_response:
                        campaign_name = str(row.campaign.name or '').strip()
                        if not campaign_name:
                            continue
                        clicks = int(row.metrics.clicks or 0)
                        if clicks <= 0:
                            continue
                        for key, entry in brand_country_clicks.items():
                            if self.campaign_name_matches_brand_country(
                                campaign_name,
                                entry.get('brand', ''),
                                entry.get('country', '')
                            ):
                                entry['clicks'] += clicks
                                matched_campaigns += 1
                except Exception as e:
                    self.log_manage(
                        f"  ⚠ 账户 {account['name']}({account['id']}) 汇总品牌点击失败: {str(e)[:100]}"
                    )

            self.log_manage(f"  品牌国家点击汇总完成: 匹配到 {matched_campaigns} 个广告系列")
        except Exception as e:
            self.log_manage(f"  汇总品牌国家点击失败: {e}")

        return brand_country_clicks

    def format_brand_break_even_cpc(self, total_commission, total_clicks):
        """格式化品牌收支平衡CPC展示：$2.00（$100/50）。"""
        commission = float(total_commission or 0.0)
        clicks = int(total_clicks or 0)
        break_even_cpc = (commission / clicks) if clicks > 0 else 0.0
        return f"${break_even_cpc:.2f}（${commission:.2f}/{clicks}）"

    def apply_brand_break_even_cpc(self, updates, new_rows, mcc_campaigns, brand_country_totals):
        """把品牌收支平衡CPC写入广告系列更新结果。"""
        break_even_updates = 0

        def apply_to_rows(rows):
            nonlocal break_even_updates
            for item in rows:
                status = str(item.get('状态', '') or '').strip()
                if status != '投放中' and not status.startswith('广告系列暂停中'):
                    continue

                campaign_name = str(item.get('campaign_name', '') or '').strip()
                brand, country = self.extract_brand_and_country_from_campaign_name(campaign_name)
                if not brand or not country:
                    continue

                aggregate = brand_country_totals.get((self.normalize_brand_key(brand), country))
                if not aggregate:
                    continue

                total_clicks = int(aggregate.get('clicks', 0) or 0)
                total_commission = float(aggregate.get('commission', 0.0) or 0.0)
                item['品牌收支平衡CPC'] = self.format_brand_break_even_cpc(total_commission, total_clicks)
                break_even_updates += 1

        apply_to_rows(updates or [])
        apply_to_rows(new_rows or [])
        return break_even_updates

    def apply_brand_break_even_cpc_to_ended_rows(self, updates, brand_country_totals):
        """把品牌收支平衡CPC写入“投放已结束”广告系列更新结果。"""
        break_even_updates = 0

        for item in updates or []:
            status = str(item.get('状态', '') or '').strip()
            if status != '投放已结束':
                continue

            campaign_name = str(item.get('campaign_name', '') or '').strip()
            brand, country = self.extract_brand_and_country_from_campaign_name(campaign_name)
            if not brand or not country:
                continue

            aggregate = brand_country_totals.get((self.normalize_brand_key(brand), country))
            if not aggregate:
                continue

            total_clicks = int(aggregate.get('clicks', 0) or 0)
            total_commission = float(aggregate.get('commission', 0.0) or 0.0)
            item['品牌收支平衡CPC'] = self.format_brand_break_even_cpc(total_commission, total_clicks)
            break_even_updates += 1

        return break_even_updates

    def extract_asin_from_yp_order_id(self, value):
        """从YP返回的id字段末尾提取ASIN。"""
        text = str(value or '').strip()
        if not text:
            return ''
        match = re.search(r'_([A-Z0-9]{10})(?:$|_)', text, re.IGNORECASE)
        if match:
            return match.group(1).upper()
        return ''

    def _batch_update_sheet_cells(self, token, spreadsheet_token, updates, sheet_id=None):
        """批量更新飞书单元格值。"""
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }

        target_sheet_id = sheet_id or self.feishu_sheet_id_var.get().strip()

        value_ranges = []
        for row_idx, col_idx, value in updates:
            col_letter = self.index_to_column_letter(col_idx)
            value_ranges.append({
                'range': f"{target_sheet_id}!{col_letter}{row_idx}:{col_letter}{row_idx}",
                'values': [[value]]
            })

        batch_size = 100
        success = True
        for i in range(0, len(value_ranges), batch_size):
            batch = value_ranges[i:i+batch_size]
            url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
            body = {"valueRanges": batch}

            try:
                response = requests.post(url, headers=headers, json=body, timeout=30)
                data = response.json()
                batch_no = i // batch_size + 1
                if data.get('code') == 0:
                    self.log_get(f"  批次 {batch_no}: 更新成功 {len(batch)} 个单元格")
                else:
                    success = False
                    self.log_get(f"  批次 {batch_no}: 更新失败 - {data.get('msg', 'Unknown error')}")
            except Exception as e:
                success = False
                self.log_get(f"  批次 {i // batch_size + 1}: 更新异常 - {str(e)}")

        return success

    def insert_sheet_rows(self, token, spreadsheet_token, sheet_id, start_index_0based, count):
        """在飞书工作表插入空行。"""
        if count <= 0:
            return True

        url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/insert_dimension_range"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        body = {
            "dimension": {
                "sheetId": sheet_id,
                "majorDimension": "ROWS",
                "startIndex": start_index_0based,
                "endIndex": start_index_0based + count
            },
            "inheritStyle": "BEFORE"
        }
        try:
            response = requests.post(url, headers=headers, json=body, timeout=30)
            result = response.json()
            if result.get('code') != 0:
                self.log_manage(f"    插入行失败: code={result.get('code')}, {result.get('msg', '')}")
                return False
            return True
        except Exception as e:
            self.log_manage(f"    插入行异常: {e}")
            return False

    def delete_sheet_rows(self, token, spreadsheet_token, sheet_id, row_indices):
        """删除飞书工作表中的指定行，row_indices 为 1-based 行号。"""
        rows = sorted({int(row) for row in row_indices if row and int(row) > 1}, reverse=True)
        if not rows:
            return True

        url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/dimension_range"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }

        for row_index in rows:
            start_index = row_index - 1
            body = {
                "dimension": {
                    "sheetId": sheet_id,
                    "majorDimension": "ROWS",
                    "startIndex": start_index,
                    "endIndex": start_index + 1
                }
            }
            try:
                response = requests.delete(url, headers=headers, json=body, timeout=30)
                result = response.json()
                if result.get('code') != 0:
                    self.log_manage(f"    删除行失败 row={row_index}: code={result.get('code')}, {result.get('msg', '')}")
                    return False
            except Exception as e:
                self.log_manage(f"    删除行异常 row={row_index}: {e}")
                return False

        return True

    def _do_update_yp_links(self):
        try:
            self.log_get("=" * 50)
            self.log_get("开始回填YP产品链接...")
            self.log_get("=" * 50)

            token = self.get_feishu_token()
            if not token:
                self.log_get("获取飞书访问令牌失败")
                return

            spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
            if not spreadsheet_token or not self.feishu_sheet_id_var.get().strip():
                self.log_get("飞书电子表格配置缺失")
                return

            self.log_get("\n【步骤1】读取飞书Offer表格数据...")
            self.update_progress_get("读取飞书数据...")
            feishu_data = self.get_feishu_sheet_data(token)
            tracking_link_col = self.feishu_column_map.get('投放链接')
            product_link_col = self.feishu_column_map.get('产品链接')
            if tracking_link_col is None or product_link_col is None:
                self.log_get("未找到“投放链接”列或“产品链接”列")
                return

            product_link_col_idx = ord(product_link_col) - ord('A')
            yp_rows = []
            empty_link_count = 0
            existing_product_link_count = 0

            for row in feishu_data:
                link = str(self.normalize_sheet_cell_value(row.get('投放链接', '')) or '').strip()
                if not link:
                    empty_link_count += 1
                    continue
                if not self._is_yp_tracking_link(link):
                    continue

                product_link = str(self.normalize_sheet_cell_value(row.get('产品链接', '')) or '').strip()
                if product_link:
                    existing_product_link_count += 1
                    continue

                row_index = row.get('row_index')
                if not row_index:
                    continue
                yp_rows.append((row_index, link))

            self.log_get(f"  扫描到 {len(feishu_data)} 行Offer")
            self.log_get(f"  符合条件的YP行: {len(yp_rows)} 行")
            self.log_get(f"  已有产品链接的YP行: {existing_product_link_count} 行")
            if empty_link_count:
                self.log_get(f"  空投放链接: {empty_link_count} 行")

            if not yp_rows:
                self.log_get("没有需要回填产品链接的YP行")
                return

            self.log_get("\n【步骤2】访问YP投放链接并解析产品链接...")
            self.update_progress_get("解析产品链接...")
            updates = []
            failed_rows = []
            for idx, (row_idx, link) in enumerate(yp_rows, start=1):
                self.root.after(0, lambda i=idx, t=len(yp_rows): self.update_progress_get(f"解析链接 {i}/{t}"))
                product_link = self.resolve_redirect_url(link)
                if product_link:
                    updates.append((row_idx, product_link_col_idx, product_link))
                    self.log_get(f"  第{row_idx}行: 已获取产品链接")
                else:
                    failed_rows.append(row_idx)
                    self.log_get(f"  第{row_idx}行: 未获取到产品链接")
                time.sleep(0.1)

            if not updates:
                self.log_get("未解析到任何可写回的产品链接")
                return

            self.log_get("\n【步骤3】写回产品链接到飞书表格...")
            self.update_progress_get("写入飞书...")
            success = self._batch_update_sheet_cells(token, spreadsheet_token, updates)

            self.log_get("\n" + "=" * 50)
            self.log_get("更新完成！统计信息：")
            self.log_get(f"  成功写回产品链接: {len(updates)} 行")
            self.log_get(f"  解析失败: {len(failed_rows)} 行")
            self.log_get(f"  飞书写入结果: {'成功' if success else '部分失败'}")
            self.log_get("=" * 50)
        except Exception as e:
            self.log_get(f"回填YP产品链接时发生错误: {str(e)}")
            import traceback
            self.log_get(traceback.format_exc())
        finally:
            self.root.after(0, self._restore_get_ui)

    def copy_offers(self):
        """复制offer功能：识别只有国家代码和ASIN的行，复制完整信息并生成新投放链接"""
        if not self.feishu_app_id_var.get().strip() or not self.feishu_app_secret_var.get().strip():
            messagebox.showerror("错误", "请在API配置中输入飞书App ID和App Secret!")
            return
        
        if not self.pb_token_var.get().strip():
            messagebox.showerror("错误", "请在API配置中输入PartnerBoost Offer Token!")
            return
        
        self.extract_btn.config(state='disabled')
        self.update_yp_links_btn.config(state='disabled')
        self.copy_offer_btn.config(state='disabled')
        self.progress_get.start()
        self.log_text_get.delete(1.0, tk.END)
        
        thread = threading.Thread(target=self._do_copy_offers)
        thread.daemon = True
        thread.start()
    
    def _do_copy_offers(self):
        """执行复制offer操作"""
        try:
            self.log_get("=" * 50)
            self.log_get("开始复制offer...")
            self.log_get("=" * 50)
            
            # 获取飞书访问令牌
            token = self.get_feishu_token()
            if not token:
                self.log_get("获取飞书访问令牌失败")
                return
            
            spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
            sheet_id = self.feishu_sheet_id_var.get().strip()
            
            if not spreadsheet_token or not sheet_id:
                self.log_get("飞书电子表格配置缺失")
                return
            
            # 读取飞书数据
            self.log_get("\n【步骤1】读取飞书offer表格数据...")
            self.update_progress_get("读取飞书数据...")
            
            url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values/{sheet_id}!A1:Z2000"
            headers = {"Authorization": f"Bearer {token}"}
            resp = requests.get(url, headers=headers)
            data = resp.json()
            
            if data.get('code') != 0:
                self.log_get(f"读取飞书数据失败: {data.get('msg')}")
                return
            
            values = data.get('data', {}).get('valueRange', {}).get('values', [])
            if len(values) < 2:
                self.log_get("飞书表格数据不足")
                return
            
            headers_row = values[0]
            
            # 找到所有列的索引
            col_indices = {}
            for i, h in enumerate(headers_row):
                h_str = str(h).strip() if h else ''
                if h_str:
                    col_indices[h_str] = i
            
            self.log_get(f"  检测到的列: {list(col_indices.keys())}")
            
            # 必需的列
            if 'ASIN' not in col_indices or '国家代码' not in col_indices or '品牌ID' not in col_indices:
                self.log_get("未找到ASIN、国家代码或品牌ID列")
                return
            
            asin_col = col_indices['ASIN']
            country_col = col_indices['国家代码']
            brand_id_col = col_indices['品牌ID']
            tracking_link_col = col_indices.get('投放链接')
            product_link_col = col_indices.get('产品链接')
            
            # 定义需要复制的字段（除投放链接和公式列外）
            # 注意："每单佣金"是公式列，不需要复制，会自动计算
            copy_fields = ['品牌名称', '品牌ID', '折扣价', '佣金', '产品名称', '产品链接', 
                          '库存状态', '原价', '货币', '更新时间', '图片URL', '分类', '子分类',
                          '评分', '评论数', '折扣', '折扣码', '优惠券']
            
            # 第一遍扫描：识别复制源和复制对象
            self.log_get("\n【步骤2】识别复制源和复制对象...")
            
            copy_sources = []  # [(row_idx, asin, country, brand_id), ...]
            complete_offers = {}  # {(asin, country, brand_id): (row_idx, row_data), ...} 第一个完整offer
            
            for row_idx, row in enumerate(values[1:], start=2):
                key = self.build_copy_offer_match_key(row, col_indices)
                if not key:
                    continue
                asin, country, brand_id = key
                
                # 检查是否只有ASIN、国家代码和品牌ID有值（复制源）
                has_other_data = self.copy_offer_has_non_key_data(row, col_indices, copy_fields)
                
                # 检查投放链接
                has_tracking_link = False
                if tracking_link_col is not None and tracking_link_col < len(row) and row[tracking_link_col]:
                    link_value = row[tracking_link_col]
                    if isinstance(link_value, list) and len(link_value) > 0 and isinstance(link_value[0], dict):
                        link_value = link_value[0].get('link', '') or link_value[0].get('text', '')
                    if str(link_value).strip():
                        has_tracking_link = True
                
                if not has_other_data and not has_tracking_link:
                    # 这是复制源（只有ASIN、国家代码和品牌ID）
                    copy_sources.append((row_idx, asin, country, brand_id))
                elif has_other_data:
                    # 这是有完整信息的offer，记录第一个作为复制对象
                    if key not in complete_offers:
                        complete_offers[key] = (row_idx, row)
            
            self.log_get(f"  找到 {len(copy_sources)} 个复制源（只有ASIN、国家代码和品牌ID）")
            self.log_get(f"  找到 {len(complete_offers)} 个有完整信息的offer")
            
            if not copy_sources:
                self.log_get("没有找到需要复制的offer（复制源）")
                return
            
            # 第二遍处理：执行复制
            self.log_get("\n【步骤3】执行复制操作...")
            self.update_progress_get("复制offer...")
            
            updates = []  # [(row_idx, col_idx, value), ...]
            style_updates = []  # [(row_idx, col_idx, style_type), ...] 样式更新
            success_count = 0
            skip_count = 0
            
            for row_idx, asin, country, brand_id in copy_sources:
                key = (asin, country, brand_id)
                
                if key not in complete_offers:
                    self.log_get(f"  [跳过] {asin}_{country}_{brand_id}: 未找到同ASIN+国家+品牌ID的完整offer")
                    skip_count += 1
                    continue
                
                source_row_idx, source_row = complete_offers[key]
                self.log_get(f"  [复制] {asin}_{country}_{brand_id}: 从第{source_row_idx}行复制到第{row_idx}行")
                
                # 复制字段（除投放链接外）
                for field in copy_fields:
                    if field in col_indices:
                        field_col = col_indices[field]
                        if field_col < len(source_row) and source_row[field_col]:
                            source_value = source_row[field_col]
                            # 处理URL类型的单元格
                            if isinstance(source_value, list) and len(source_value) > 0 and isinstance(source_value[0], dict):
                                source_value = source_value[0].get('link', '') or source_value[0].get('text', '')
                            if source_value and str(source_value).strip():
                                updates.append((row_idx, field_col, source_value))
                
                # 为"每单佣金"列写入公式（公式格式: =L{row}*M{row}）
                if '每单佣金' in col_indices:
                    commission_per_order_col = col_indices['每单佣金']
                    formula = f"=L{row_idx}*M{row_idx}"
                    updates.append((row_idx, commission_per_order_col, formula))
                    self.log_get(f"    写入每单佣金公式: {formula}")
                
                # 设置"状态"列为"新复制"（蓝色加粗）
                if '状态' in col_indices:
                    status_col = col_indices['状态']
                    updates.append((row_idx, status_col, '新复制'))
                    style_updates.append((row_idx, status_col, 'blue_bold'))
                    self.log_get(f"    设置状态: 新复制 (蓝色加粗)")
                
                source_tracking_link = ''
                if tracking_link_col is not None and tracking_link_col < len(source_row) and source_row[tracking_link_col]:
                    source_tracking_link = source_row[tracking_link_col]
                    if isinstance(source_tracking_link, list) and len(source_tracking_link) > 0 and isinstance(source_tracking_link[0], dict):
                        source_tracking_link = source_tracking_link[0].get('link', '') or source_tracking_link[0].get('text', '')
                    source_tracking_link = str(source_tracking_link).strip()

                new_link, generated_new_link, uid = self.build_copied_offer_tracking_link(asin, country, source_tracking_link)
                if not generated_new_link and self._is_yp_tracking_link(source_tracking_link):
                    self.log_get("    识别为YP链接，保留原投放链接")
                else:
                    self.log_get(f"    获取新投放链接...")

                if new_link and tracking_link_col is not None:
                    updates.append((row_idx, tracking_link_col, new_link))
                    if not generated_new_link and self._is_yp_tracking_link(new_link):
                        self.log_get(f"    投放链接已保留: {new_link[-40:]}")
                    else:
                        self.log_get(f"    新投放链接: {new_link[-40:]} (uid={uid})")

                    # 仅PB链接需要解析重定向并更新产品链接；YP链接保留复制来的产品链接
                    if product_link_col is not None and not self._is_yp_tracking_link(new_link):
                        redirect_url = self.resolve_redirect_url(new_link)
                        if redirect_url:
                            updates.append((row_idx, product_link_col, redirect_url))
                            self.log_get(f"    产品链接已更新: {redirect_url[:50]}...")

                    success_count += 1
                else:
                    self.log_get(f"    [警告] 获取投放链接失败")
                    success_count += 1  # 即使没有链接，其他字段也复制了
                
                time.sleep(0.3)  # API调用间隔
            
            # 应用更新到飞书
            if updates:
                self.log_get(f"\n【步骤4】写入飞书表格... ({len(updates)} 个单元格)")
                self.update_progress_get("写入飞书...")
                self._apply_copy_updates(token, spreadsheet_token, sheet_id, updates, style_updates, headers_row)
            
            # 输出统计
            self.log_get("\n" + "=" * 50)
            self.log_get("复制完成！统计信息：")
            self.log_get(f"  成功复制: {success_count} 个offer")
            self.log_get(f"  跳过（无复制对象）: {skip_count} 个")
            self.log_get("=" * 50)
            
        except Exception as e:
            self.log_get(f"复制offer时发生错误: {str(e)}")
            import traceback
            self.log_get(traceback.format_exc())
        finally:
            self.root.after(0, self._restore_get_ui)
    
    def _apply_copy_updates(self, token, spreadsheet_token, sheet_id, updates, style_updates, headers_row):
        """应用复制offer的更新到飞书表格"""
        self._batch_update_sheet_cells(token, spreadsheet_token, updates)
        
        # 应用样式更新（蓝色加粗）
        if style_updates:
            self._apply_copy_style_updates(token, spreadsheet_token, sheet_id, style_updates)
    
    def _apply_copy_style_updates(self, token, spreadsheet_token, sheet_id, style_updates):
        """应用复制offer的样式更新（蓝色加粗）"""
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        for row_idx, col_idx, style_type in style_updates:
            col_letter = self.index_to_column_letter(col_idx)
            range_str = f"{sheet_id}!{col_letter}{row_idx}:{col_letter}{row_idx}"
            
            if style_type == 'blue_bold':
                style = {
                    "font": {
                        "bold": True
                    },
                    "foreColor": "#0000FF"  # 蓝色
                }
            else:
                style = {
                    "font": {
                        "bold": False
                    },
                    "foreColor": "#000000"  # 黑色
                }
            
            url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/style"
            body = {
                "appendStyle": {
                    "range": range_str,
                    "style": style
                }
            }
            
            try:
                resp = requests.put(url, headers=headers, json=body, timeout=10)
                result = resp.json()
                if result.get('code') != 0:
                    self.log_get(f"  样式更新失败: {result.get('msg', 'Unknown error')}")
            except Exception as e:
                self.log_get(f"  样式更新异常: {str(e)}")
    
    # ==================== 品牌搜索量报告 ====================
    def get_brand_search_volume(self):
        """获取品牌搜索量报告"""
        if not self.pb_token_var.get().strip():
            messagebox.showerror("错误", "请在API配置中输入PartnerBoost Offer Token!")
            return

        self.extract_btn.config(state='disabled')
        self.update_yp_links_btn.config(state='disabled')
        self.copy_offer_btn.config(state='disabled')
        self.brand_search_btn.config(state='disabled')
        self.progress_get.start()
        self.log_text_get.delete(1.0, tk.END)
        
        thread = threading.Thread(target=self._do_brand_search_volume)
        thread.daemon = True
        thread.start()
    
    def _do_brand_search_volume(self):
        """执行品牌搜索量报告生成"""
        try:
            self.log_get("=" * 50)
            self.log_get("开始获取品牌搜索量报告...")
            self.log_get("=" * 50)
            
            # 步骤1：获取PB partnered品牌列表
            self.log_get("\n【步骤1】获取PartnerBoost已合作品牌列表...")
            self.update_progress_get("获取品牌列表...")
            
            brands = self._get_partnered_brands()
            if not brands:
                self.log_get("未获取到任何品牌数据")
                return
            
            self.log_get(f"  共获取到 {len(brands)} 个品牌")

            scope = self.brand_search_scope_var.get().strip() if hasattr(self, 'brand_search_scope_var') else "全量"
            if scope == "50个":
                brands = brands[:50]
                self.log_get("  已选择测试模式：仅处理前50个品牌")
            else:
                self.log_get("  已选择全量模式：处理全部品牌")
            
            # 步骤2：解析品牌名和国家代码
            self.log_get("\n【步骤2】解析品牌名和国家代码...")
            self.update_progress_get("解析品牌信息...")
            
            parsed_brands = []
            for brand in brands:
                bid = brand.get('bid', '')
                raw_name = brand.get('brand_name', '')
                brand_name, country_code = self._parse_brand_name_country(raw_name)
                parsed_brands.append({
                    'bid': bid,
                    'raw_name': raw_name,
                    'brand_name': brand_name,
                    'country_code': country_code
                })
            
            # 统计国家分布
            country_stats = {}
            for b in parsed_brands:
                cc = b['country_code']
                country_stats[cc] = country_stats.get(cc, 0) + 1
            self.log_get(f"  国家分布: {country_stats}")
            self.log_get(f"  前5个品牌示例:")
            for b in parsed_brands[:5]:
                self.log_get(f"    • {b['raw_name']} -> 品牌名={b['brand_name']}, 国家={b['country_code']}, BID={b['bid']}")
            
            # 步骤3：调用Google Ads API获取搜索量
            self.log_get("\n【步骤3】调用Google Ads API获取关键词搜索量...")
            self.update_progress_get("获取搜索量...")
            
            client = self.get_google_ads_client()
            if not client:
                self.log_get("  ✗ 无法创建Google Ads客户端，请检查配置")
                return
            
            mcc_ids = self.get_google_mcc_ids()
            mcc_id = mcc_ids[0] if mcc_ids else ''
            if not mcc_id:
                self.log_get("  ✗ 未配置MCC ID")
                return
            
            # 按国家代码分组，批量查询
            country_brand_groups = {}  # {country_code: [brand_info, ...]}
            for b in parsed_brands:
                cc = b['country_code']
                if cc not in country_brand_groups:
                    country_brand_groups[cc] = []
                country_brand_groups[cc].append(b)
            
            results = []
            total_brands = len(parsed_brands)
            processed = 0
            
            for country_code, brand_list in country_brand_groups.items():
                self.log_get(f"\n  查询国家 {country_code} 的品牌搜索量（{len(brand_list)} 个品牌）...")
                
                # 分批查询，每批最多20个关键词
                batch_size = 20
                max_retries = 5
                for i in range(0, len(brand_list), batch_size):
                    batch = brand_list[i:i+batch_size]
                    keywords = [b['brand_name'] for b in batch]
                    
                    success = False
                    for attempt in range(max_retries):
                        try:
                            metrics = self._get_keyword_historical_metrics(client, mcc_id, keywords, country_code)
                            
                            for b in batch:
                                keyword = b['brand_name']
                                metric = metrics.get(keyword.lower(), {})
                                b['avg_monthly_searches'] = metric.get('avg_monthly_searches', 0)
                                b['competition'] = metric.get('competition', 'N/A')
                                b['competition_index'] = metric.get('competition_index', 'N/A')
                                b['low_top_of_page_bid'] = metric.get('low_top_of_page_bid', 0)
                                b['high_top_of_page_bid'] = metric.get('high_top_of_page_bid', 0)
                                results.append(b)
                            
                            processed += len(batch)
                            self.update_progress_get(f"搜索量 {processed}/{total_brands}...")
                            success = True
                            break
                            
                        except Exception as e:
                            error_str = str(e)
                            # 解析429限流错误中的等待时间
                            is_rate_limit = '429' in error_str or 'Resource has been exhausted' in error_str
                            if is_rate_limit and attempt < max_retries - 1:
                                # 尝试从错误信息中提取等待秒数
                                wait_seconds = 5  # 默认等待时间
                                retry_match = re.search(r'Retry in (\d+) seconds', error_str)
                                if retry_match:
                                    wait_seconds = int(retry_match.group(1)) + 1
                                else:
                                    wait_seconds = min(5 * (2 ** attempt), 60)  # 指数退避，最长60秒
                                self.log_get(f"    [限流] 第{attempt+1}次重试，等待{wait_seconds}秒...")
                                self.update_progress_get(f"限流等待{wait_seconds}秒... ({processed}/{total_brands})")
                                time.sleep(wait_seconds)
                            else:
                                self.log_get(f"    [错误] 批次查询失败 (已重试{attempt}次): {error_str[:200]}")
                                break
                    
                    if not success:
                        # 所有重试都失败，记录错误结果
                        for b in batch:
                            b['avg_monthly_searches'] = 'ERROR'
                            b['competition'] = 'ERROR'
                            b['competition_index'] = 'ERROR'
                            b['low_top_of_page_bid'] = 'ERROR'
                            b['high_top_of_page_bid'] = 'ERROR'
                            results.append(b)
                        processed += len(batch)
                    
                    time.sleep(1)  # API调用间隔
            
            self.log_get(f"\n  搜索量查询完成，共 {len(results)} 个品牌")
            
            # 步骤4：查询品牌Storefront链接
            self.log_get("\n【步骤4】查询品牌Storefront链接...")
            self.update_progress_get("查询Storefront链接...")
            
            all_bids = [r['bid'] for r in results if r.get('bid')]
            storefront_map = self._get_brand_storefront_links(all_bids)
            
            for r in results:
                r['storefront_link'] = storefront_map.get(str(r.get('bid', '')), '')
            
            found_count = sum(1 for v in storefront_map.values() if v)
            self.log_get(f"  Storefront链接查询完成: {found_count}/{len(all_bids)} 个品牌有链接")
            
            # 步骤5：查询各品牌产品数量
            self.log_get("\n【步骤5】查询各品牌产品数量...")
            self.update_progress_get("查询产品数量...")
            
            product_count_map, commission_raw_map = self._get_brand_product_counts(all_bids)
            
            for r in results:
                r['product_count'] = product_count_map.get(str(r.get('bid', '')), 0)
                r['commission_raw'] = commission_raw_map.get(str(r.get('bid', '')), '')
            
            has_products = sum(1 for v in product_count_map.values() if v > 0)
            total_products = sum(product_count_map.values())
            self.log_get(f"  产品数量查询完成: {has_products}/{len(all_bids)} 个品牌有产品，共 {total_products} 个产品")
            
            # 步骤6：生成报告
            self.log_get("\n【步骤6】生成Excel报告...")
            self.update_progress_get("生成报告...")
            
            self._save_brand_search_report(results)
            
            self.log_get("\n" + "=" * 50)
            self.log_get("品牌搜索量报告生成完成！")
            self.log_get("=" * 50)
            
        except Exception as e:
            self.log_get(f"生成品牌搜索量报告时发生错误: {str(e)}")
            import traceback
            self.log_get(traceback.format_exc())
        finally:
            self.root.after(0, self._restore_get_ui)
    
    def _get_partnered_brands(self):
        """获取PB已合作品牌列表（自动翻页）"""
        all_brands = []
        page = 1
        page_size = 50
        
        token = self.pb_token_var.get().strip()
        url = f"{PB_API_BASE_URL}/api/datafeed/get_amazon_joined_brands"
        
        while True:
            body = {
                "token": token,
                "bids": "",
                "page_size": page_size,
                "page": page
            }
            
            try:
                response = requests.post(url, json=body, timeout=30)
                data = response.json()
                
                if data.get('status', {}).get('code') != 0:
                    self.log_get(f"  获取品牌列表失败 (页{page}): {data.get('status', {}).get('msg', 'Unknown error')}")
                    break
                
                brand_list = data.get('data', {}).get('list', [])
                if not brand_list:
                    break
                
                all_brands.extend(brand_list)
                has_more = data.get('data', {}).get('hasMore', False)
                
                self.log_get(f"  第{page}页: 获取 {len(brand_list)} 个品牌 (累计 {len(all_brands)})")
                
                if not has_more:
                    break
                
                page += 1
                time.sleep(0.3)
                
            except Exception as e:
                self.log_get(f"  获取品牌列表异常 (页{page}): {str(e)}")
                break
        
        return all_brands
    
    def _get_brand_storefront_links(self, bids):
        """批量查询品牌storefront链接，返回 {bid: storefront_link} 映射"""
        storefront_map = {}
        token = self.pb_token_var.get().strip()
        url = f"{PB_API_BASE_URL}/api/datafeed/get_fba_brand_link"
        
        batch_size = 50
        for i in range(0, len(bids), batch_size):
            batch = bids[i:i+batch_size]
            bids_str = ','.join(str(b) for b in batch)
            
            try:
                response = requests.post(url, json={"token": token, "bids": bids_str}, timeout=30)
                data = response.json()
                
                if data.get('status', {}).get('code') == 0 and data.get('data'):
                    for item in data['data']:
                        bid = item.get('bid', '')
                        link = item.get('storefront_link', '')
                        if link:
                            storefront_map[str(bid)] = link
                
                self.log_get(f"    Storefront查询批次 {i//batch_size+1}: 成功 {len([b for b in batch if str(b) in storefront_map])}/{len(batch)}")
                
            except Exception as e:
                self.log_get(f"    Storefront查询批次 {i//batch_size+1} 失败: {str(e)[:100]}")
            
            if i + batch_size < len(bids):
                time.sleep(0.5)
        
        return storefront_map
    
    def _get_brand_product_counts(self, bids):
        """查询各品牌的产品数量和佣金字段，返回 ({bid: count}, {bid: commission_raw})"""
        count_map = {}
        commission_raw_map = {}
        token = self.pb_token_var.get().strip()
        url = f"{PB_API_BASE_URL}/api/datafeed/get_fba_products"
        headers = {"Content-Type": "application/json", "Accept": "application/json"}
        
        total = len(bids)
        for idx, bid in enumerate(bids):
            bid_str = str(bid).strip()
            if not bid_str:
                continue
            
            product_count = 0
            commission_raw = ''
            page = 1
            has_more = True
            
            while has_more:
                try:
                    body = {
                        "token": token,
                        "page_size": 100,
                        "page": page,
                        "default_filter": 0,
                        "country_code": "",
                        "brand_id": bid_str,
                        "sort": "",
                        "asins": "",
                        "relationship": 1,
                        "is_original_currency": 0,
                        "has_promo_code": 0,
                        "has_acc": 0,
                        "filter_sexual_wellness": 0
                    }
                    resp = requests.post(url, json=body, headers=headers, timeout=60)
                    data = resp.json()
                    
                    if data.get("status", {}).get("code") == 0:
                        products = data.get("data", {}).get("list", [])
                        product_count += len(products)
                        if not commission_raw:
                            for product in products:
                                raw_commission = str(product.get('commission', '')).strip()
                                if raw_commission:
                                    commission_raw = raw_commission
                                    break
                        has_more = data.get("data", {}).get("has_more", False)
                        page += 1
                    else:
                        break
                except Exception:
                    break
                
                time.sleep(0.1)
            
            count_map[bid_str] = product_count
            commission_raw_map[bid_str] = commission_raw
            
            if (idx + 1) % 10 == 0 or idx == total - 1:
                self.log_get(f"    产品数量查询进度: {idx+1}/{total}")
                self.update_progress_get(f"产品数量 {idx+1}/{total}...")
        
        return count_map, commission_raw_map
    
    def _parse_brand_name_country(self, raw_name):
        """从品牌名中智能解析品牌名和国家代码
        
        示例：
            "Odoland_UN-US" -> ("Odoland", "US")
            "PARWIN PRO BEAUTY--DE" -> ("PARWIN PRO BEAUTY", "DE")
            "DE-Anker" -> ("Anker", "DE")
            "Brand_FR" -> ("Brand", "FR")
            "SimpleBrand" -> ("SimpleBrand", "US")  # 默认US
        """
        if not raw_name:
            return (raw_name, 'US')
        
        country_codes = ['US', 'UK', 'DE', 'FR', 'IT', 'ES', 'CA', 'JP', 'AU', 'NL', 'BE', 'MX', 'BR', 'IN', 'SG', 'AE', 'SA', 'PL', 'SE', 'TR', 'EG', 'GB']
        
        name = raw_name.strip()
        
        # 模式1: 结尾有国家代码，用各种分隔符 如 "Brand_UN-US", "Brand--DE", "Brand_FR", "Brand-US"
        for code in country_codes:
            # 结尾匹配：各种分隔符+国家代码
            patterns = [
                (f'[-_]+UN[-_]+{code}$', code),        # _UN-US, -UN_US
                (f'[-_]{{2,}}{code}$', code),           # --DE, __DE
                (f'[-_]{code}$', code),                 # -US, _FR
            ]
            for pattern, matched_code in patterns:
                match = re.search(pattern, name, re.IGNORECASE)
                if match:
                    brand_name = name[:match.start()].strip()
                    if brand_name:
                        return (brand_name, matched_code)
        
        # 模式2: 开头有国家代码 如 "DE-Anker"
        for code in country_codes:
            match = re.match(rf'^({code})[-_]+(.+)$', name, re.IGNORECASE)
            if match:
                return (match.group(2).strip(), code)
        
        # 没有找到国家代码，默认US
        return (name, 'US')
    
    # 国家代码到Google Ads Geo Target Constant ID的映射
    GEO_TARGET_MAP = {
        'US': '2840', 'UK': '2826', 'GB': '2826', 'DE': '2276', 'FR': '2250',
        'IT': '2380', 'ES': '2724', 'CA': '2124', 'JP': '2392', 'AU': '2036',
        'NL': '2528', 'BE': '2056', 'MX': '2484', 'BR': '2076', 'IN': '2356',
        'SG': '2702', 'AE': '2784', 'SA': '2682', 'PL': '2616', 'SE': '2752',
        'TR': '2792', 'EG': '2818',
    }
    
    # 国家代码到Google Ads Language Constant ID的映射
    LANGUAGE_MAP = {
        'US': '1000', 'UK': '1000', 'GB': '1000', 'CA': '1000', 'AU': '1000',  # English
        'DE': '1001',  # German
        'FR': '1002',  # French
        'IT': '1004',  # Italian
        'ES': '1003',  # Spanish
        'JP': '1005',  # Japanese
        'NL': '1010',  # Dutch
        'BE': '1010',  # Dutch (Belgium)
        'MX': '1003',  # Spanish
        'BR': '1014',  # Portuguese
        'IN': '1000',  # English (India)
        'SG': '1000',  # English (Singapore)
        'AE': '1019',  # Arabic
        'SA': '1019',  # Arabic
        'PL': '1030',  # Polish
        'SE': '1015',  # Swedish
        'TR': '1037',  # Turkish
        'EG': '1019',  # Arabic
    }

    COUNTRY_CODE_ALIASES = {
        'GB': 'UK',
    }

    def normalize_country_code(self, country_code):
        """标准化国家代码，统一历史别名，便于跨表匹配。"""
        code = str(country_code or '').upper().strip()
        if not code:
            return ''
        return self.COUNTRY_CODE_ALIASES.get(code, code)
    
    def _get_keyword_historical_metrics(self, client, customer_id, keywords, country_code):
        """调用Google Ads API获取关键词历史搜索量指标
        
        Args:
            client: GoogleAdsClient实例
            customer_id: MCC客户ID
            keywords: 关键词列表
            country_code: 国家代码
            
        Returns:
            {keyword_lower: {avg_monthly_searches, competition, ...}}
        """
        googleads_service = client.get_service("GoogleAdsService")
        keyword_plan_idea_service = client.get_service("KeywordPlanIdeaService")
        
        request = client.get_type("GenerateKeywordHistoricalMetricsRequest")
        request.customer_id = customer_id
        request.keywords.extend(keywords)
        
        # 设置地理位置
        normalized_country = self.normalize_country_code(country_code)
        geo_id = self.GEO_TARGET_MAP.get(normalized_country, '2840')  # 默认US
        request.geo_target_constants.append(
            googleads_service.geo_target_constant_path(geo_id)
        )
        
        # 设置搜索网络
        request.keyword_plan_network = client.enums.KeywordPlanNetworkEnum.GOOGLE_SEARCH
        
        # 设置语言
        lang_id = self.LANGUAGE_MAP.get(normalized_country, '1000')  # 默认English
        request.language = googleads_service.language_constant_path(lang_id)
        
        response = keyword_plan_idea_service.generate_keyword_historical_metrics(
            request=request
        )
        
        metrics_map = {}
        for result in response.results:
            keyword = result.text.lower() if result.text else ''
            m = result.keyword_metrics
            metrics_map[keyword] = {
                'avg_monthly_searches': m.avg_monthly_searches if m.avg_monthly_searches else 0,
                'competition': m.competition.name if m.competition else 'N/A',
                'competition_index': m.competition_index if m.competition_index else 0,
                'low_top_of_page_bid': round(m.low_top_of_page_bid_micros / 1_000_000, 2) if m.low_top_of_page_bid_micros else 0,
                'high_top_of_page_bid': round(m.high_top_of_page_bid_micros / 1_000_000, 2) if m.high_top_of_page_bid_micros else 0,
            }
        
        return metrics_map
    
    def _save_brand_search_report(self, results):
        """保存品牌搜索量报告为xlsx"""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = "品牌搜索量报告"
        
        # 表头
        headers = ['Brand ID', 'Storefront Link', '产品数量', '佣金比例', '原始品牌名', '品牌名', '国家代码', '月均搜索量', 
                   '竞争程度', '竞争指数', '页首低价出价($)', '页首高价出价($)']
        ws.append(headers)
        
        # 数据行
        for r in results:
            ws.append([
                r.get('bid', ''),
                r.get('storefront_link', ''),
                r.get('product_count', 0),
                r.get('commission_raw', ''),
                r.get('raw_name', ''),
                r.get('brand_name', ''),
                r.get('country_code', ''),
                r.get('avg_monthly_searches', 0),
                r.get('competition', 'N/A'),
                r.get('competition_index', 0),
                r.get('low_top_of_page_bid', 0),
                r.get('high_top_of_page_bid', 0),
            ])
        
        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(headers[col_idx - 1]))
            for row_idx in range(2, min(len(results) + 2, 100)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, min(len(str(cell_value)), 40))
            ws.column_dimensions[col_letter].width = max_length + 2
        
        # 按月均搜索量降序排序（表头行之后）
        data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
        data_rows.sort(key=lambda x: (x[7] if isinstance(x[7], (int, float)) else 0), reverse=True)
        
        # 清除数据行，重新写入排序后的数据
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).value = None
        
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx).value = value
        
        # 保存文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_dir = os.path.dirname(self.save_path_var.get())
        if not save_dir:
            save_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        save_path = os.path.join(save_dir, f"品牌搜索量报告_{timestamp}.xlsx")
        
        wb.save(save_path)
        self.log_get(f"  报告已保存到: {save_path}")
        self.root.after(0, lambda: self._ask_open_file(save_path))
    
    def save_to_xlsx(self, offers):
        try:
            save_path = self.save_path_var.get()
            wb = Workbook()
            ws = wb.active
            ws.title = "Offers"
            
            columns_config = [
                ("投放链接", "partnerboost_link", False),
                ("国家代码", "country_code", False),
                ("品牌名称", "brand_name", False),
                ("折扣价", "discount_price", False),
                ("佣金", "commission", False),
                ("Reviews", "reviews", False),
                ("产品ID", "product_id", True),
                ("产品名称", "product_name", False),
                ("图片URL", "image", True),
                ("ASIN", "asin", False),
                ("折扣", "discount", True),
                ("折扣码", "discount_code", True),
                ("优惠券", "coupon", True),
                ("分类", "category", True),
                ("子分类", "subcategory", True),
                ("父ASIN", "parent_asin", True),
                ("变体ASIN", "variant_asin", True),
                ("库存状态", "availability", False),
                ("评分", "rating", True),
                ("产品链接", "url", False),
                ("品牌ID", "brand_id", False),
                ("更新时间", "update_time", False),
                ("Relationship", "relationship", True),
                ("原价", "original_price", False),
                ("货币", "currency", False),
                ("ACC佣金", "acc_commission", True),
                ("ACC开始日期", "acc_start_date", True),
                ("ACC结束日期", "acc_end_date", True),
                ("促销码列表", "promo_code_list", True),
            ]
            
            headers = [col[0] for col in columns_config]
            ws.append(headers)
            
            for offer in offers:
                row = []
                for header, field, hidden in columns_config:
                    if field == "promo_code_list":
                        promo_codes = offer.get("promo_code_list", [])
                        if promo_codes:
                            promo_code_str = "; ".join([f"{p.get('promo_code', '')}({p.get('promotional_price', '')})" for p in promo_codes])
                        else:
                            promo_code_str = ""
                        row.append(promo_code_str)
                    else:
                        row.append(offer.get(field, ""))
                ws.append(row)
            
            for col_idx, (header, field, hidden) in enumerate(columns_config, start=1):
                col_letter = get_column_letter(col_idx)
                if hidden:
                    ws.column_dimensions[col_letter].hidden = True
                else:
                    max_length = len(header)
                    for row_idx in range(1, min(len(offers) + 2, 100)):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            max_length = max(max_length, min(len(str(cell_value)), 50))
                    ws.column_dimensions[col_letter].width = max_length + 2
            
            wb.save(save_path)
            self.log_get(f"数据已保存到: {save_path}")
            self.root.after(0, lambda: self._ask_open_file(save_path))
            
        except Exception as e:
            self.log_get(f"保存文件失败: {str(e)}")
    
    def _ask_open_file(self, path):
        if messagebox.askyesno("保存成功", f"文件已保存到:\n{path}\n\n是否打开文件?"):
            os.startfile(path)

    # ==================== Offer管理功能 ====================
    def test_connections(self):
        """测试所有API连接"""
        if hasattr(self, 'test_conn_btn'):
            self.test_conn_btn.config(state='disabled')
        self.log_text_manage.delete(1.0, tk.END)
        thread = threading.Thread(target=self._do_test_connections)
        thread.daemon = True
        thread.start()
    
    def _do_test_connections(self):
        try:
            results = []

            self.log_manage("开始测试API连接...")
            self.log_config("开始测试API连接...")
            
            # 测试飞书
            self.log_manage("\n[飞书API]")
            self.log_config("\n[飞书API]")
            try:
                token = self.get_feishu_token()
                if token:
                    self.log_manage("  ✓ 飞书Token获取成功")
                    self.log_config("  ✓ 飞书Token获取成功")
                    results.append("飞书: 成功")
                else:
                    self.log_manage("  ✗ 飞书Token获取失败")
                    self.log_config("  ✗ 飞书Token获取失败")
                    results.append("飞书: 失败")
            except Exception as e:
                self.log_manage(f"  ✗ 飞书连接失败: {e}")
                self.log_config(f"  ✗ 飞书连接失败: {e}")
                results.append("飞书: 失败")
            
            # 测试Google Ads
            self.log_manage("\n[Google Ads API]")
            self.log_config("\n[Google Ads API]")
            try:
                google_accounts = self.get_google_mcc_accounts()
                google_success = 0
                for account in google_accounts:
                    mcc_id = account.get("mcc_id", "")
                    label = self.format_google_mcc_list_label(account)
                    try:
                        client = self.get_google_ads_client(mcc_id)
                        if client:
                            customer_service = client.get_service('CustomerService')
                            accessible_customers = customer_service.list_accessible_customers()
                            self.log_manage(f"  ✓ {label} 连接成功，可访问 {len(accessible_customers.resource_names)} 个账户")
                            self.log_config(f"  ✓ {label} 连接成功，可访问 {len(accessible_customers.resource_names)} 个账户")
                            google_success += 1
                        else:
                            self.log_manage(f"  ✗ {label} 客户端创建失败")
                            self.log_config(f"  ✗ {label} 客户端创建失败")
                    except Exception as e:
                        self.log_manage(f"  ✗ {label} 连接失败: {e}")
                        self.log_config(f"  ✗ {label} 连接失败: {e}")
                if google_accounts and google_success == len(google_accounts):
                    results.append("Google Ads: 成功")
                elif google_success:
                    results.append(f"Google Ads: 部分成功 ({google_success}/{len(google_accounts)})")
                else:
                    if not google_accounts:
                        self.log_manage("  ✗ 未配置Google Ads MCC")
                        self.log_config("  ✗ 未配置Google Ads MCC")
                    results.append("Google Ads: 失败")
            except Exception as e:
                self.log_manage(f"  ✗ Google Ads连接失败: {e}")
                self.log_config(f"  ✗ Google Ads连接失败: {e}")
                results.append("Google Ads: 失败")
            
            # 测试PartnerBoost
            self.log_manage("\n[PartnerBoost API]")
            self.log_config("\n[PartnerBoost API]")
            try:
                url = f"{PB_API_BASE_URL}/api.php?mod=medium&op=transaction"
                body = {
                    "token": self.pb_token_var.get().strip(),
                    "begin_date": (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'),
                    "end_date": datetime.now().strftime('%Y-%m-%d'),
                    "type": "json",
                    "status": "All",
                    "limit": 1,
                    "page": 1
                }
                response = requests.post(url, json=body, timeout=30)
                data = response.json()
                if data.get("status", {}).get("code") == 0:
                    self.log_manage("  ✓ PartnerBoost连接成功")
                    self.log_config("  ✓ PartnerBoost连接成功")
                    results.append("PartnerBoost: 成功")
                else:
                    self.log_manage(f"  ✗ PartnerBoost返回错误: {data.get('status', {}).get('msg')}")
                    self.log_config(f"  ✗ PartnerBoost返回错误: {data.get('status', {}).get('msg')}")
                    results.append("PartnerBoost: 失败")
            except Exception as e:
                self.log_manage(f"  ✗ PartnerBoost连接失败: {e}")
                self.log_config(f"  ✗ PartnerBoost连接失败: {e}")
                results.append("PartnerBoost: 失败")

            # 测试YeahPromos
            self.log_manage("\n[YeahPromos API]")
            self.log_config("\n[YeahPromos API]")
            try:
                yp_token = self.yp_token_var.get().strip()
                yp_site_id = self.yp_site_id_var.get().strip()
                if yp_token and yp_site_id:
                    url = f"{YP_API_BASE_URL}/index/Getorder/getorder"
                    headers = {"token": yp_token}
                    params = {
                        "site_id": yp_site_id,
                        "startDate": (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'),
                        "endDate": datetime.now().strftime('%Y-%m-%d'),
                        "is_amazon": 1,
                        "limit": 1,
                        "page": 1,
                    }
                    response = requests.get(url, headers=headers, params=params, timeout=30)
                    data = response.json()
                    if str(data.get("code", "")) == "100000":
                        self.log_manage("  ✓ YeahPromos连接成功")
                        self.log_config("  ✓ YeahPromos连接成功")
                        results.append("YeahPromos: 成功")
                    else:
                        msg = data.get("msg") or data.get("message") or data.get("status") or "Unknown error"
                        self.log_manage(f"  ✗ YeahPromos返回错误: {msg}")
                        self.log_config(f"  ✗ YeahPromos返回错误: {msg}")
                        results.append("YeahPromos: 失败")
                else:
                    self.log_manage("  - 未配置YP Token或Site ID，跳过测试")
                    self.log_config("  - 未配置YP Token或Site ID，跳过测试")
                    results.append("YeahPromos: 跳过")
            except Exception as e:
                self.log_manage(f"  ✗ YeahPromos连接失败: {e}")
                self.log_config(f"  ✗ YeahPromos连接失败: {e}")
                results.append("YeahPromos: 失败")
            
            self.log_manage("\n测试完成！")
            self.log_config("\n测试完成！")
            summary = "\n".join(results)
            self._run_on_ui_thread(lambda: messagebox.showinfo("测试完成", summary))
        finally:
            self._run_on_ui_thread(lambda: self.test_conn_btn.config(state='normal'))
    
    def get_feishu_token(self):
        """获取飞书tenant_access_token"""
        url = f"{FEISHU_API_BASE_URL}/auth/v3/tenant_access_token/internal"
        body = {
            "app_id": self.feishu_app_id_var.get().strip(),
            "app_secret": self.feishu_app_secret_var.get().strip()
        }
        response = requests.post(url, json=body, timeout=30)
        data = response.json()
        if data.get("code") == 0:
            return data.get("tenant_access_token")
        return None

    def get_feishu_sheet_id_by_title(self, token, spreadsheet_token, title):
        """按工作表标题查询sheet ID。"""
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        url = f"{FEISHU_API_BASE_URL}/sheets/v3/spreadsheets/{spreadsheet_token}/sheets/query"
        response = requests.get(url, headers=headers, timeout=30)
        data = response.json()
        if data.get('code') != 0:
            self.log_manage(f"  查询工作表列表失败: {data.get('msg', 'Unknown error')}")
            return None

        for sheet in data.get('data', {}).get('sheets', []) or []:
            if str(sheet.get('title', '') or '').strip() == title:
                return sheet.get('sheet_id')
        return None
    
    def parse_google_mcc_ids(self, value):
        """解析MCC ID列表，兼容旧的单值配置、逗号分隔和多行输入。"""
        if isinstance(value, (list, tuple, set)):
            raw_values = value
        else:
            raw_values = re.split(r'[\s,;，；]+', str(value or ''))

        mcc_ids = []
        seen = set()
        for raw in raw_values:
            mcc_id = re.sub(r'\D', '', str(raw or '').strip())
            if not mcc_id or mcc_id in seen:
                continue
            seen.add(mcc_id)
            mcc_ids.append(mcc_id)
        return mcc_ids

    def get_google_mcc_ids(self):
        """获取当前配置的MCC ID列表。"""
        accounts = self.get_google_mcc_accounts()
        if accounts:
            return [account["mcc_id"] for account in accounts]
        return self.parse_google_mcc_ids(self.config.get("google_mcc_ids") or self.config.get("google_mcc_id", ""))

    def get_google_mcc_accounts(self):
        """获取当前界面的多MCC账号配置。"""
        if hasattr(self, 'google_mcc_accounts'):
            self.save_current_google_mcc_editor()
            normalized = self.normalize_google_mcc_accounts({
                "google_mcc_accounts": self.google_mcc_accounts,
                "google_developer_token": self.config.get("google_developer_token", ""),
            })
            self.google_mcc_accounts = [dict(account) for account in normalized]
            if self.google_mcc_accounts:
                current = getattr(self, 'current_google_mcc_index', 0) or 0
                self.current_google_mcc_index = min(current, len(self.google_mcc_accounts) - 1)
            return normalized
        return self.normalize_google_mcc_accounts(self.config)

    def get_google_mcc_account_config(self, login_customer_id=None):
        """按MCC ID查找对应的Developer Token和服务账号密钥。"""
        accounts = self.get_google_mcc_accounts()
        if not accounts:
            return None
        login_customer_id = re.sub(r'\D', '', str(login_customer_id or '').strip())
        if not login_customer_id:
            return accounts[0]
        for account in accounts:
            if account.get("mcc_id") == login_customer_id:
                return account
        return None

    def get_google_ads_client(self, login_customer_id=None):
        """创建Google Ads客户端。login_customer_id为空时使用第一个MCC。"""
        account_config = self.get_google_mcc_account_config(login_customer_id)
        if not account_config:
            return None
        key_text = str(account_config.get("service_account_key", "") or "").strip()
        if not key_text:
            return None

        login_customer_id = re.sub(r'\D', '', str(login_customer_id or account_config.get("mcc_id", "")).strip())
        if not login_customer_id:
            return None

        sa_info = json.loads(key_text)
        credentials = service_account.Credentials.from_service_account_info(
            sa_info,
            scopes=['https://www.googleapis.com/auth/adwords']
        )
        
        client = GoogleAdsClient(
            credentials=credentials,
            developer_token=str(account_config.get("developer_token", "") or "").strip(),
            login_customer_id=login_customer_id,
            use_proto_plus=True
        )
        return client

    def iter_google_ads_mcc_accounts(self, include_mcc_name=False):
        """按配置的多个MCC遍历其直属子账户。"""
        accounts = []
        for mcc_config in self.get_google_mcc_accounts():
            mcc_id = mcc_config.get("mcc_id", "")
            configured_mcc_name = str(mcc_config.get("name", "") or "").strip()
            if self.stop_flag:
                break
            try:
                client = self.get_google_ads_client(mcc_id)
                if not client:
                    self.log_manage(f"  ✗ 无法创建Google Ads客户端: MCC {mcc_id}")
                    continue
                ga_service = client.get_service('GoogleAdsService')
                query = """
                    SELECT customer_client.id, customer_client.descriptive_name, customer_client.manager, customer_client.currency_code, customer_client.level
                    FROM customer_client
                    WHERE customer_client.level <= 1
                """
                response = ga_service.search(customer_id=mcc_id, query=query)

                mcc_name = configured_mcc_name
                mcc_accounts = []
                for row in response:
                    customer_id = str(row.customer_client.id)
                    if customer_id == mcc_id and row.customer_client.manager:
                        mcc_name = row.customer_client.descriptive_name or configured_mcc_name
                        for account in mcc_accounts:
                            account['mcc_name'] = mcc_name
                        continue
                    if not row.customer_client.manager:
                        mcc_accounts.append({
                            'id': customer_id,
                            'name': row.customer_client.descriptive_name,
                            'currency': row.customer_client.currency_code,
                            'mcc_id': mcc_id,
                            'mcc_name': mcc_name,
                        })
                if not mcc_name:
                    mcc_name = 'mcc11' if mcc_id == '2160853519' else mcc_id
                for account in mcc_accounts:
                    account['mcc_name'] = mcc_name
                self.log_manage(f"  MCC {mcc_id}: 找到 {len(mcc_accounts)} 个子账户")
                accounts.extend(mcc_accounts)
            except Exception as e:
                self.log_manage(f"  ⚠ MCC {mcc_id} 子账户查询失败: {str(e)[:120]}")
        return accounts
    
    def start_statistics(self):
        """开始统计"""
        start_date_str = self.stats_start_date_var.get().strip()
        if not start_date_str:
            messagebox.showerror("错误", "请输入统计开始日期，格式为 YYYY-MM-DD")
            return

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("错误", "开始日期格式错误，请使用 YYYY-MM-DD")
            return

        if self.stats_to_today_var.get():
            end_date_str = datetime.now().strftime('%Y-%m-%d')
            self.stats_end_date_var.set(end_date_str)
        else:
            end_date_str = self.stats_end_date_var.get().strip()
            if not end_date_str:
                messagebox.showerror("错误", "请输入统计结束日期，格式为 YYYY-MM-DD，或勾选“至今”")
                return

        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("错误", "结束日期格式错误，请使用 YYYY-MM-DD")
            return

        if start_date > end_date:
            messagebox.showerror("错误", "开始日期不能晚于结束日期")
            return

        increment_days_str = self.stats_increment_days_var.get().strip()
        if not increment_days_str:
            messagebox.showerror("错误", "请输入新增花费/佣金统计天数")
            return
        try:
            increment_days = int(increment_days_str)
        except ValueError:
            messagebox.showerror("错误", "新增花费/佣金统计天数必须是正整数")
            return
        if increment_days <= 0:
            messagebox.showerror("错误", "新增花费/佣金统计天数必须大于0")
            return

        self.stop_flag = False
        self.start_stats_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.update_offers_btn.config(state='disabled')
        self.sort_tables_btn.config(state='disabled')
        self.check_links_btn.config(state='disabled')
        self.progress_manage.start()
        self.log_text_manage.delete(1.0, tk.END)
        
        thread = threading.Thread(target=self._do_statistics, args=(start_date_str, end_date_str, increment_days))
        thread.daemon = True
        thread.start()
    
    def stop_statistics(self):
        """停止统计"""
        self.stop_flag = True
        self.log_manage("正在停止...")
    
    def _restore_manage_ui(self):
        self.start_stats_btn.config(state='normal')
        self.stop_btn.config(state='disabled')
        self.update_offers_btn.config(state='normal')
        self.sort_tables_btn.config(state='normal')
        self.check_links_btn.config(state='normal')
        self.progress_manage.stop()
        self.update_progress_manage("")
    
    def start_sort_tables(self):
        """开始offer顺序整理"""
        self.stop_flag = False
        self.start_stats_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.update_offers_btn.config(state='disabled')
        self.sort_tables_btn.config(state='disabled')
        self.check_links_btn.config(state='disabled')
        self.progress_manage.start()
        self.log_text_manage.delete(1.0, tk.END)
        
        thread = threading.Thread(target=self._do_sort_tables)
        thread.daemon = True
        thread.start()
    
    def _do_sort_tables(self):
        """执行offer顺序整理操作"""
        try:
            self.log_manage("=" * 50)
            self.log_manage("开始Offer顺序整理...")
            self.log_manage("=" * 50)
            
            # 获取飞书Token
            feishu_token = self.get_feishu_token()
            if not feishu_token:
                self.log_manage("✗ 获取飞书Token失败")
                return
            
            # 排序Offer表格
            self.log_manage("\n【1/2】对Offer表格行排序")
            self.update_progress_manage("排序Offer表...")
            try:
                self.sort_offer_table(feishu_token)
            except Exception as e:
                self.log_manage(f"  Offer表格排序出错: {str(e)}")
                import traceback
                self.log_manage(traceback.format_exc())
            
            if self.stop_flag:
                self.log_manage("\n已停止")
                return
            
            # 排序广告系列表格
            self.log_manage("\n【2/2】对广告系列表格行排序")
            self.update_progress_manage("排序广告系列表...")
            try:
                self.sort_campaigns_table(feishu_token)
            except Exception as e:
                self.log_manage(f"  广告系列表格排序出错: {str(e)}")
                import traceback
                self.log_manage(traceback.format_exc())
            
            self.log_manage("\n" + "=" * 50)
            self.log_manage("✅ Offer顺序整理完成！")
            self.log_manage("=" * 50)
            
        except Exception as e:
            self.log_manage(f"\n错误: {str(e)}")
            import traceback
            self.log_manage(traceback.format_exc())
        finally:
            self.root.after(0, self._restore_manage_ui)

    def start_check_link_health(self):
        """开始检查投放链接解析出的产品链接是否与表格记录一致。"""
        self.stop_flag = False
        self.start_stats_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.update_offers_btn.config(state='disabled')
        self.sort_tables_btn.config(state='disabled')
        self.check_links_btn.config(state='disabled')
        self.progress_manage.start()
        self.log_text_manage.delete(1.0, tk.END)

        thread = threading.Thread(target=self._do_check_link_health)
        thread.daemon = True
        thread.start()

    def normalize_url_for_compare(self, value):
        """标准化URL用于健康检查对比，保留查询参数语义。"""
        url = str(self.normalize_sheet_cell_value(value) or '').strip()
        if not url:
            return ''
        return url.rstrip('/')

    def normalize_product_link_key(self, value):
        """标准化产品链接，用于Offer表和广告系列表之间做精确映射。"""
        url = str(self.normalize_sheet_cell_value(value) or '').strip()
        if not url:
            return ''
        return url.rstrip('/')

    def is_ended_status(self, status):
        return str(status or '').strip().startswith('投放已结束')

    def is_link_health_target_status(self, status):
        """链接健康检查只采集投放中和暂停中的行。"""
        status_text = str(status or '').strip()
        return (
            status_text.startswith('投放中') or
            status_text.startswith('暂停中') or
            status_text.startswith('暂停') or
            status_text.startswith('广告系列暂停中')
        )

    def _add_link_health_item(self, grouped_items, tracking_link, product_link, source, row_index, label):
        tracking_link = self.normalize_url_for_compare(tracking_link)
        product_link = self.normalize_url_for_compare(product_link)
        if not tracking_link:
            return False

        key = tracking_link
        item = grouped_items.setdefault(key, {
            'tracking_link': tracking_link,
            'old_product_link': product_link,
            'sources': []
        })
        item['sources'].append({
            'source': source,
            'row_index': row_index,
            'label': label,
            'old_product_link': product_link
        })
        return True

    def collect_link_health_items(self, token, spreadsheet_token):
        """从Offer表和广告系列表采集非投放已结束行的投放链接/产品链接组。"""
        grouped_items = {}

        offer_rows = self.get_feishu_sheet_data(token)
        offer_count = 0
        for row in offer_rows or []:
            status = str(row.get('状态', '') or '').strip()
            if not self.is_link_health_target_status(status) or status == SUMMARY_STATUS_TEXT:
                continue
            tracking_link = row.get('投放链接', '')
            product_link = row.get('产品链接', '')
            asin = str(row.get('ASIN', '') or '').strip()
            brand = str(row.get('品牌名称', '') or '').strip()
            country = self.normalize_country_code(row.get('国家代码', '') or '')
            label_parts = [part for part in [f"ASIN={asin}" if asin else '', f"品牌={brand}" if brand else '', f"国家={country}" if country else ''] if part]
            label = ', '.join(label_parts) if label_parts else '未命名Offer'
            if self._add_link_health_item(grouped_items, tracking_link, product_link, 'Offer表', row.get('row_index'), label):
                offer_count += 1

        campaigns_sheet_id = CAMPAIGNS_SHEET_ID
        campaigns_data = self.read_campaigns_sheet(token, spreadsheet_token, campaigns_sheet_id)
        campaign_count = 0
        if campaigns_data:
            campaign_rows, _, _ = campaigns_data
            for row in campaign_rows or []:
                status = str(row.get('状态', '') or '').strip()
                if not self.is_link_health_target_status(status) or status == SUMMARY_STATUS_TEXT:
                    continue
                campaign_name = str(row.get('广告系列名称', '') or '').strip()
                label = campaign_name or '未命名广告系列'
                if self._add_link_health_item(grouped_items, row.get('投放链接', ''), row.get('产品链接', ''), '广告系列表', row.get('row_index'), label):
                    campaign_count += 1

        sources_by_tracking_link = {}
        for item in grouped_items.values():
            tracking_link = item.get('tracking_link', '')
            if not tracking_link:
                continue
            sources_by_tracking_link.setdefault(tracking_link, [])
            sources_by_tracking_link[tracking_link].extend(item.get('sources', []))

        for item in grouped_items.values():
            item['all_tracking_sources'] = sources_by_tracking_link.get(item.get('tracking_link', ''), [])

        return list(grouped_items.values()), offer_count, campaign_count

    def _do_check_link_health(self):
        """检查投放链接当前解析出的产品链接是否与表格旧产品链接一致。"""
        try:
            self.log_manage("=" * 50)
            self.log_manage("开始检查链接健康...")
            self.log_manage("=" * 50)

            token = self.get_feishu_token()
            if not token:
                self.log_manage("获取飞书访问令牌失败")
                return

            spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
            if not spreadsheet_token:
                self.log_manage("飞书电子表格配置缺失")
                return

            self.log_manage("\n【步骤1】采集Offer表和广告系列表链接...")
            self.update_progress_manage("采集链接...")
            items, offer_count, campaign_count = self.collect_link_health_items(token, spreadsheet_token)
            self.log_manage(f"  Offer表采集链接组: {offer_count} 组")
            self.log_manage(f"  广告系列表采集链接组: {campaign_count} 组")
            self.log_manage(f"  去重后待检查链接组: {len(items)} 组")

            if not items:
                self.log_manage("没有需要检查的链接")
                return

            self.log_manage("\n【步骤2】解析投放链接并对比产品链接...")
            self.update_progress_manage("检查链接...")
            mismatch_count = 0
            failed_count = 0
            checked_count = 0

            for idx, item in enumerate(items, start=1):
                if self.stop_flag:
                    self.log_manage("已停止")
                    return
                tracking_link = item.get('tracking_link', '')
                old_product_link = item.get('old_product_link', '')
                self.update_progress_manage(f"检查链接 {idx}/{len(items)}")
                new_product_link = self.normalize_url_for_compare(self.resolve_redirect_url(tracking_link))
                checked_count += 1

                if not new_product_link:
                    failed_count += 1
                    self.log_manage(f"  [解析失败] 投放链接: {tracking_link}")
                    continue

                if self.normalize_url_for_compare(old_product_link) == new_product_link:
                    continue

                mismatch_count += 1
                self.log_manage("\n  [产品链接不一致]")
                self.log_manage(f"    投放链接: {tracking_link}")
                self.log_manage(f"    表格旧产品链接: {old_product_link or '(空)'}")
                self.log_manage(f"    当前解析产品链接: {new_product_link}")
                self.log_manage("    本组关联位置:")
                for source in item.get('sources', []):
                    self.log_manage(
                        f"      - {source.get('source')} 第{source.get('row_index')}行: {source.get('label')}"
                    )
                all_sources = item.get('all_tracking_sources', [])
                if len(all_sources) > len(item.get('sources', [])):
                    self.log_manage("    该投放链接的全部关联位置:")
                    for source in all_sources:
                        source_old_link = source.get('old_product_link') or '(空)'
                        self.log_manage(
                            f"      - {source.get('source')} 第{source.get('row_index')}行: {source.get('label')} | 旧产品链接={source_old_link}"
                        )
                self.log_manage("    提醒: 请检查并更新该投放链接对应的产品链接")
                time.sleep(0.1)

            self.log_manage("\n" + "=" * 50)
            self.log_manage("链接健康检查完成！统计信息：")
            self.log_manage(f"  已检查: {checked_count} 组")
            self.log_manage(f"  产品链接不一致: {mismatch_count} 组")
            self.log_manage(f"  解析失败: {failed_count} 组")
            self.log_manage("=" * 50)
        except Exception as e:
            self.log_manage(f"检查链接健康时发生错误: {str(e)}")
            import traceback
            self.log_manage(traceback.format_exc())
        finally:
            self.root.after(0, self._restore_manage_ui)
    
    def start_update_offers(self):
        """开始更新PB offer"""
        self.stop_flag = False
        self.start_stats_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.update_offers_btn.config(state='disabled')
        self.sort_tables_btn.config(state='disabled')
        self.check_links_btn.config(state='disabled')
        self.progress_manage.start()
        
        thread = threading.Thread(target=self._do_update_offers)
        thread.daemon = True
        thread.start()
    
    def _do_update_offers(self):
        """执行更新PB offer操作（按品牌批量获取，仅处理PB行）"""
        try:
            self.log_manage("=" * 50)
            self.log_manage("开始更新PB offer（按品牌批量获取，仅处理PB行）...")
            self.log_manage("=" * 50)
            
            # 获取飞书访问令牌
            token = self.get_feishu_token()
            if not token:
                self.log_manage("获取飞书访问令牌失败")
                return
            
            spreadsheet_token = self.config.get('feishu_spreadsheet_token', '')
            sheet_id = self.config.get('feishu_sheet_id', '')
            
            if not spreadsheet_token or not sheet_id:
                self.log_manage("飞书电子表格配置缺失")
                return
            
            # 读取飞书数据
            self.log_manage("\n【步骤1】读取飞书现有offer数据...")
            self.update_progress_manage("读取飞书数据...")
            
            url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values/{sheet_id}!A1:Z2000"
            headers = {"Authorization": f"Bearer {token}"}
            resp = requests.get(url, headers=headers)
            data = resp.json()
            
            if data.get('code') != 0:
                self.log_manage(f"读取飞书数据失败: {data.get('msg')}")
                return
            
            values = data.get('data', {}).get('valueRange', {}).get('values', [])
            if len(values) < 2:
                self.log_manage("飞书表格数据不足")
                return
            
            headers_row = values[0]
            
            # 找到关键列的索引
            col_indices = {}
            target_cols = ['状态', 'ASIN', '国家代码', '品牌名称', '品牌ID', '佣金', '库存状态', '更新时间', 
                          '投放链接', '折扣价', '产品名称', '产品链接', '原价', '货币', '评论数', 'Reviews']
            for i, h in enumerate(headers_row):
                h_str = str(h).strip() if h else ''
                if h_str in target_cols:
                    col_indices[h_str] = i
            
            # 输出检测到的列索引（调试用）
            self.log_manage(f"  检测到的列: {col_indices}")
            
            if 'ASIN' not in col_indices:
                self.log_manage("未找到ASIN列")
                return
            
            # 获取表格实际的列数（用于验证）
            max_col_idx = len(headers_row) - 1
            
            # 收集飞书现有PB offer和品牌ID
            # 支持同一个(asin, country)有多行（复制的offer）
            existing_offers = {}  # (asin, country) -> [(row_idx, row, brand_id), ...]
            existing_asins = {}   # asin -> (row_idx, country)  用于ASIN回退匹配
            brand_ids = set()
            skipped_non_pb_rows = 0
            
            for row_idx, row in enumerate(values[1:], start=2):
                asin_idx = col_indices.get('ASIN')
                country_idx = col_indices.get('国家代码')
                brand_id_idx = col_indices.get('品牌ID')
                asin = str(row[asin_idx]).strip() if asin_idx is not None and asin_idx < len(row) else ''
                country = str(row[country_idx]).strip().upper() if country_idx is not None and country_idx < len(row) else ''
                brand_id = str(row[brand_id_idx]).strip() if brand_id_idx is not None and brand_id_idx < len(row) else ''
                
                # 跳过"总计"行或空ASIN行
                if not asin or asin.lower() in ['none', '总计', 'total']:
                    continue
                if not self.is_pb_offer_row(row, col_indices):
                    skipped_non_pb_rows += 1
                    continue
                
                key = (asin, country)
                if key not in existing_offers:
                    existing_offers[key] = []
                existing_offers[key].append((row_idx, row, brand_id))
                existing_asins[asin] = (row_idx, country)
                # 只添加有效的brand_id（过滤掉空值和None）
                if brand_id and brand_id.lower() not in ['none', '']:
                    brand_ids.add(brand_id)
            
            # 统计唯一的(asin, country)组合数和总行数
            total_rows = sum(len(rows) for rows in existing_offers.values())
            self.log_manage(f"  飞书现有PB offer: {total_rows} 行（{len(existing_offers)} 个唯一组合）")
            self.log_manage(f"  跳过非PB行（统计行/YP行/空链接行）: {skipped_non_pb_rows} 行")
            self.log_manage(f"  涉及品牌ID: {len(brand_ids)} 个")
            
            if self.stop_flag:
                self.log_manage("已停止")
                return
            
            # 按品牌ID获取PB的offer数据
            self.log_manage("\n【步骤2】按品牌ID获取PB offer数据...")
            self.update_progress_manage("获取PB数据...")
            
            all_pb_offers = {}  # (asin, country) -> offer_data
            failed_brand_requests = {}
            successful_brand_ids = set()
            
            for i, brand_id in enumerate(brand_ids):
                if self.stop_flag:
                    self.log_manage("已停止")
                    return
                
                self.log_manage(f"  获取品牌 {brand_id} 的offer... ({i+1}/{len(brand_ids)})")
                result = self.get_offers_by_brand(brand_id)
                offers = result.get('offers', [])
                if result.get('success'):
                    successful_brand_ids.add(str(brand_id).strip())
                else:
                    failed_brand_requests[str(brand_id).strip()] = result.get('error', '未知错误')
                self.log_manage(f"    品牌 {brand_id}: PB返回 {len(offers)} 个offer")
                
                for offer in offers:
                    asin = str(offer.get('asin', '')).strip()
                    pb_country = str(offer.get('country_code', '')).strip().upper()
                    brand_name = str(offer.get('brand_name', '')).strip()
                    
                    if not asin:
                        continue
                    
                    # 原始国家代码（用于调试）
                    original_pb_country = pb_country
                    
                    # 如果PB的country_code为空，尝试从品牌名称中提取国家代码
                    if not pb_country and brand_name:
                        extracted_country = self.extract_country_from_merchant_name(brand_name)
                        if extracted_country:
                            pb_country = extracted_country
                    
                    # 尝试匹配：只做精确匹配，不做ASIN回退（避免国家混淆）
                    matched_country = pb_country  # 默认用PB返回的（或从品牌名提取的）国家代码
                    
                    # 如果PB有明确的国家代码，检查是否存在精确匹配
                    if pb_country:
                        if (asin, pb_country) in existing_offers:
                            matched_country = pb_country
                        elif asin in existing_asins:
                            # PB有国家但飞书里是另一个国家，不强制回退
                            feishu_country = existing_asins[asin][1]
                            if feishu_country != pb_country:
                                # 跳过，避免用错误国家的数据覆盖
                                continue
                    else:
                        # PB没有国家代码，尝试用飞书里该ASIN的国家
                        if asin in existing_asins:
                            matched_country = existing_asins[asin][1]
                        else:
                            matched_country = 'US'  # 默认US
                    
                    offer['_matched_country'] = matched_country
                    offer['_original_pb_country'] = original_pb_country
                    all_pb_offers[(asin, matched_country)] = offer
                
                time.sleep(0.2)  # 品牌间间隔
            
            self.log_manage(f"  PB总共返回 {len(all_pb_offers)} 个有效offer")
            if failed_brand_requests:
                self.log_manage(f"  PB请求失败品牌: {len(failed_brand_requests)} 个")
                for failed_brand_id, failed_reason in failed_brand_requests.items():
                    self.log_manage(f"    失败品牌 {failed_brand_id}: {failed_reason}")
            
            if self.stop_flag:
                self.log_manage("已停止")
                return
            
            # 对比并更新PB offer
            self.log_manage("\n【步骤3】对比并更新PB offer...")
            self.update_progress_manage("对比更新...")
            
            updates = []  # (row_idx, col_idx, new_value)
            style_updates = []  # (row_idx, col_idx, style_type)  style_type: 'red_bold' or 'black_normal'
            updated_count = 0
            unchanged_count = 0
            new_offers = []
            
            # 检查现有offer的更新
            # 修改：existing_offers现在是 {(asin, country): [(row_idx, row), ...]}
            for (asin, country), rows_list in existing_offers.items():
                pb_offer = all_pb_offers.get((asin, country))
                
                # 对该(asin, country)的所有行进行相同的更新
                for row_idx, row, row_brand_id in rows_list:
                    # 调试前几行
                    if row_idx <= 5:
                        self.log_manage(f"    调试: 第{row_idx}行 ASIN={asin} 国家={country} PB匹配={'是' if pb_offer else '否'} 行数据长度={len(row)}")
                    
                    if not pb_offer:
                        normalized_row_brand_id = str(row_brand_id or '').strip()
                        if normalized_row_brand_id and normalized_row_brand_id not in successful_brand_ids:
                            self.log_manage(
                                f"  [跳过下架] {asin}_{country} (行{row_idx}): 品牌 {normalized_row_brand_id} 本次请求失败，禁止标记OUT_OF_STOCK"
                            )
                            unchanged_count += 1
                            continue
                        # PB未返回此offer，可能已下架，标记为OUT_OF_STOCK
                        current_stock = str(row[col_indices.get('库存状态', -1)]).strip() if col_indices.get('库存状态', -1) < len(row) else ''
                        if current_stock != 'OUT_OF_STOCK':
                            stock_col = col_indices.get('库存状态')
                            if stock_col is not None:
                                updates.append((row_idx, stock_col, 'OUT_OF_STOCK'))
                                style_updates.append((row_idx, stock_col, 'red_bold'))
                                self.log_manage(f"  [下架] {asin}_{country} (行{row_idx}): PB未返回，标记为OUT_OF_STOCK")
                                updated_count += 1
                        else:
                            unchanged_count += 1
                        continue
                    
                    changes = []

                    commission_col = col_indices.get('佣金')
                    if commission_col is not None:
                        old_commission = str(row[commission_col]).strip() if commission_col < len(row) else ''
                        new_commission = str(pb_offer.get('commission', '')).strip()

                        old_val = self.normalize_commission(old_commission)
                        new_val = self.normalize_commission(new_commission)

                        if abs(old_val - new_val) > 0.0001:
                            updates.append((row_idx, commission_col, new_commission))
                            changes.append(f"佣金: {old_commission} → {new_commission}")
                            if new_val < old_val:
                                style_updates.append((row_idx, commission_col, 'red_bold'))

                    reviews_col = col_indices.get('评论数')
                    if reviews_col is None:
                        reviews_col = col_indices.get('Reviews')
                    if reviews_col is not None:
                        old_reviews = str(row[reviews_col]).strip() if reviews_col < len(row) and row[reviews_col] is not None else ''
                        new_reviews = str(pb_offer.get('reviews', '')).strip()
                        if old_reviews != new_reviews and new_reviews:
                            updates.append((row_idx, reviews_col, new_reviews))
                            changes.append(f"评论数: {old_reviews} → {new_reviews}")

                    stock_col = col_indices.get('库存状态')
                    if stock_col is not None:
                        old_stock = str(row[stock_col]).strip() if stock_col < len(row) else ''
                        new_stock = str(pb_offer.get('availability', '')).strip()
                        
                        if old_stock != new_stock:
                            updates.append((row_idx, stock_col, new_stock))
                            changes.append(f"库存: {old_stock} → {new_stock}")
                            if new_stock == 'OUT_OF_STOCK':
                                style_updates.append((row_idx, stock_col, 'red_bold'))
                            elif old_stock == 'OUT_OF_STOCK' and new_stock == 'IN_STOCK':
                                style_updates.append((row_idx, stock_col, 'black_normal'))

                    update_time_col = col_indices.get('更新时间')
                    if update_time_col is not None:
                        old_time = str(row[update_time_col]).strip() if update_time_col < len(row) else ''
                        new_time = str(pb_offer.get('update_time', '')).strip()
                        if old_time != new_time and new_time:
                            updates.append((row_idx, update_time_col, new_time))
                            changes.append(f"更新时间: {old_time} → {new_time}")

                    if changes:
                        brand_name = pb_offer.get('brand_name', '')
                        pb_country_info = pb_offer.get('_original_pb_country', '')
                        rows_count = len(rows_list)
                        if rows_count > 1:
                            self.log_manage(f"  [更新] {asin}_{country} (行{row_idx}, 共{rows_count}行): {', '.join(changes)} [PB品牌:{brand_name}, PB国家:{pb_country_info}]")
                        else:
                            self.log_manage(f"  [更新] {asin}_{country}: {', '.join(changes)} [PB品牌:{brand_name}, PB国家:{pb_country_info}]")
                        updated_count += 1
                    else:
                        unchanged_count += 1

            # 检查新PB offer
            for (asin, country), pb_offer in all_pb_offers.items():
                if asin not in existing_asins:
                    commission_val = self.normalize_commission(pb_offer.get('commission', '0'))
                    if commission_val <= 0.0001:
                        continue
                    new_offers.append(pb_offer)
            
            if self.stop_flag:
                self.log_manage("已停止")
                return
            
            # 应用更新
            if updates:
                self.log_manage(f"\n【步骤4】应用更新到飞书... ({len(updates)} 个单元格)")
                self.update_progress_manage("写入飞书...")
                self._apply_offer_updates(token, updates, style_updates)

            if new_offers:
                self.log_manage(f"\n【步骤5】添加新PB offer到飞书... ({len(new_offers)} 个)")
                self.update_progress_manage("添加新offer...")
                self._append_new_offers(token, new_offers, col_indices)
            
            # 输出统计
            self.log_manage("\n" + "=" * 50)
            self.log_manage("PB offer更新完成！统计信息：")
            self.log_manage(f"  已更新offer: {updated_count} 个")
            self.log_manage(f"  无变化offer: {unchanged_count} 个")
            self.log_manage(f"  新增offer: {len(new_offers)} 个")
            self.log_manage("=" * 50)
            
        except Exception as e:
            self.log_manage(f"更新PB offer时发生错误: {str(e)}")
            import traceback
            self.log_manage(traceback.format_exc())
        finally:
            self._restore_manage_ui()
    
    def get_offers_by_brand(self, brand_id):
        """按品牌ID获取offer，带超时/502重试，并返回成功状态。"""
        all_offers = []
        page = 1
        has_more = True
        
        # 跳过无效的brand_id
        if not brand_id or str(brand_id).strip() in ['', 'None']:
            return {"success": False, "offers": [], "error": "brand_id为空"}
        
        url = f"{PB_API_BASE_URL}/api/datafeed/get_fba_products"
        headers = {"Content-Type": "application/json", "Accept": "application/json"}
        brand_id_str = str(brand_id).strip()
        
        while has_more:
            page_success = False
            last_error = ""
            for attempt in range(1, 4):
                try:
                    # 使用与获取offer功能相同的参数格式
                    request_body = {
                        "token": self.pb_token_var.get().strip(),
                        "page_size": 100,
                        "page": page,
                        "default_filter": 0,
                        "country_code": "",
                        "brand_id": brand_id_str,
                        "sort": "",
                        "asins": "",
                        "relationship": 1,
                        "is_original_currency": 0,
                        "has_promo_code": 0,
                        "has_acc": 0,
                        "filter_sexual_wellness": 0
                    }
                    
                    resp = requests.post(url, json=request_body, headers=headers, timeout=60)
                    
                    if resp.status_code == 200:
                        data = resp.json()
                        
                        if data.get("status", {}).get("code") == 0:
                            offers = data.get("data", {}).get("list", [])
                            has_more = data.get("data", {}).get("has_more", False)
                            all_offers.extend(offers)
                            page += 1
                            page_success = True
                            break
                        error_msg = str(data.get("status", {}).get("msg") or data.get("status", {}).get("message") or "未知错误")
                        self.log_manage(f"    API错误: {error_msg}")
                        return {"success": False, "offers": all_offers, "error": f"API错误: {error_msg}"}

                    if resp.status_code == 502:
                        last_error = "HTTP错误: 502"
                        self.log_manage(f"    HTTP错误: 502 (重试 {attempt}/3)")
                        if attempt < 3:
                            time.sleep(1.5 * attempt)
                            continue
                        break

                    last_error = f"HTTP错误: {resp.status_code}"
                    self.log_manage(f"    HTTP错误: {resp.status_code}")
                    return {"success": False, "offers": all_offers, "error": last_error}
                except (ReadTimeout, Timeout) as e:
                    last_error = f"超时: {str(e)}"
                    self.log_manage(f"    超时: {str(e)} (重试 {attempt}/3)")
                    if attempt < 3:
                        time.sleep(1.5 * attempt)
                        continue
                    break
                except Exception as e:
                    last_error = f"异常: {str(e)}"
                    self.log_manage(f"    异常: {str(e)}")
                    return {"success": False, "offers": all_offers, "error": last_error}

            if not page_success:
                if not last_error:
                    last_error = "未知错误"
                return {"success": False, "offers": all_offers, "error": last_error}

            time.sleep(0.1)
        
        return {"success": True, "offers": all_offers, "error": ""}
    
    def _apply_offer_updates(self, token, updates, style_updates):
        """应用offer更新到飞书"""
        spreadsheet_token = self.config.get('feishu_spreadsheet_token', '')
        sheet_id = self.config.get('feishu_sheet_id', '')
        
        # 列索引转换为列字母的辅助函数（支持超过26列）
        def col_idx_to_letter(idx):
            result = ""
            while idx >= 0:
                result = chr(ord('A') + (idx % 26)) + result
                idx = idx // 26 - 1
            return result
        
        # 过滤掉超出合理范围的更新（假设表格最多26列A-Z）
        max_col = 25  # 最大列索引（Z列）
        valid_updates = [(r, c, v) for r, c, v in updates if c <= max_col]
        
        if len(valid_updates) < len(updates):
            self.log_manage(f"    警告: 跳过 {len(updates) - len(valid_updates)} 个超出列范围的更新")
        
        # 批量更新值
        batch_size = 50
        for i in range(0, len(valid_updates), batch_size):
            batch = valid_updates[i:i+batch_size]
            value_ranges = []
            
            for row_idx, col_idx, value in batch:
                col_letter = col_idx_to_letter(col_idx)
                # 飞书API要求range格式为 sheetId!开始位置:结束位置
                range_str = f"{sheet_id}!{col_letter}{row_idx}:{col_letter}{row_idx}"
                value_ranges.append({
                    "range": range_str,
                    "values": [[value]]
                })
            
            url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }
            body = {"valueRanges": value_ranges}
            
            resp = requests.post(url, headers=headers, json=body)
            result = resp.json()
            
            if result.get('code') != 0:
                self.log_manage(f"    批量更新失败: {result.get('msg')}")
            
            time.sleep(0.2)
        
        # 应用样式更新（也过滤）
        if style_updates:
            valid_style_updates = [(r, c, s) for r, c, s in style_updates if c <= max_col]
            self._apply_style_updates(token, valid_style_updates)
    
    def _apply_style_updates(self, token, style_updates):
        """应用样式更新（红色加粗或黑色普通）"""
        spreadsheet_token = self.config.get('feishu_spreadsheet_token', '')
        sheet_id = self.config.get('feishu_sheet_id', '')
        
        # 列索引转换为列字母
        def col_idx_to_letter(idx):
            result = ""
            while idx >= 0:
                result = chr(ord('A') + (idx % 26)) + result
                idx = idx // 26 - 1
            return result
        
        for row_idx, col_idx, style_type in style_updates:
            col_letter = col_idx_to_letter(col_idx)
            range_str = f"{sheet_id}!{col_letter}{row_idx}:{col_letter}{row_idx}"
            
            if style_type == 'red_bold':
                style = {
                    "font": {
                        "bold": True
                    },
                    "foreColor": "#FF0000"  # 红色字体
                }
            elif style_type == 'blue_bold':
                style = {
                    "font": {
                        "bold": True
                    },
                    "foreColor": "#049FD7"
                }
            else:  # black_normal
                style = {
                    "font": {
                        "bold": False
                    },
                    "foreColor": "#000000"  # 黑色字体
                }
            
            url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/style"
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }
            body = {
                "appendStyle": {
                    "range": range_str,
                    "style": style
                }
            }
            
            resp = requests.put(url, headers=headers, json=body)
            result = resp.json()
            if result.get('code') != 0:
                self.log_manage(f"    样式更新失败 {col_letter}{row_idx}: {result.get('msg')}")
            time.sleep(0.1)
    
    def _append_new_offers(self, token, new_offers, col_indices):
        """添加新offer到飞书表格末尾（使用动态列索引）"""
        spreadsheet_token = self.config.get('feishu_spreadsheet_token', '')
        sheet_id = self.config.get('feishu_sheet_id', '')
        
        # 过滤掉佣金过低的offer
        valid_offers = []
        for offer in new_offers:
            commission_val = self.normalize_commission(offer.get('commission', '0'))
            if commission_val > 0.0001:
                valid_offers.append(offer)
        
        if not valid_offers:
            self.log_manage("  没有有效的新offer需要添加")
            return
        
        # 找到第一个空白行（使用ASIN列的索引）
        asin_col_idx = col_indices.get('ASIN', 0)
        first_empty_row = self._find_first_empty_row(token, spreadsheet_token, sheet_id, asin_col_idx)
        if first_empty_row is None:
            self.log_manage("  无法确定空白行位置")
            return
        
        self.log_manage(f"  从第 {first_empty_row} 行开始添加新offer")
        
        # 确保有足够的行
        needed_rows = first_empty_row + len(valid_offers)
        self._ensure_sheet_rows(token, spreadsheet_token, sheet_id, needed_rows)
        
        # 列索引转换为列字母的辅助函数
        def col_idx_to_letter(idx):
            result = ""
            while idx >= 0:
                result = chr(ord('A') + (idx % 26)) + result
                idx = idx // 26 - 1
            return result
        
        # 要写入的字段映射：列名 -> PB字段名（根据PB API返回的实际字段名）
        # 注意：产品链接不从PB直接获取(PB返回的是简单URL)，而是通过投放链接重定向获取完整版
        field_mapping = {
            '状态': '_status',  # 新增PB offer固定标记为“新增”
            '投放链接': '_tracking_link',  # 特殊处理
            '国家代码': '_country',  # 特殊处理
            '品牌名称': 'brand_name',
            '折扣价': 'discount_price',  # PB返回的是discount_price
            '佣金': 'commission',
            '评论数': 'reviews',
            'Reviews': 'reviews',
            '产品名称': 'product_name',
            'ASIN': 'asin',
            '库存状态': 'availability',
            '产品链接': '_product_link',  # 特殊处理：通过投放链接重定向获取
            '品牌ID': 'brand_id',
            '更新时间': 'update_time',  # PB返回的是update_time
            '原价': 'original_price',  # PB返回的是original_price
            '货币': 'currency'
        }
        
        # 批量写入
        batch_size = 50
        current_row = first_empty_row
        new_status_style_updates = []
        
        for i in range(0, len(valid_offers), batch_size):
            batch = valid_offers[i:i+batch_size]
            value_ranges = []
            
            for offer in batch:
                asin = offer.get('asin', '')
                
                # 确定国家代码
                country = offer.get('country_code', '').strip().upper()
                if not country:
                    brand_name = offer.get('brand_name', '')
                    country = self.extract_country_from_merchant_name(brand_name)
                if not country:
                    country = 'US'
                
                # 获取原始投放链接，不再附加UID
                tracking_link = self.get_partnerboost_link(asin, country)
                
                if not tracking_link:
                    self.log_manage(f"    警告: {asin}_{country} 无法获取投放链接")
                
                # 解析投放链接重定向，获取完整产品链接
                product_link = ''
                if tracking_link:
                    product_link = self.resolve_redirect_url(tracking_link)
                
                # 为每个需要写入的列创建单独的range
                for col_name, pb_field in field_mapping.items():
                    col_idx = col_indices.get(col_name)
                    if col_idx is None:
                        continue  # 该列不存在于表格中
                    
                    # 获取值
                    if col_name == '投放链接':
                        value = tracking_link or ''
                    elif col_name == '状态':
                        value = '新增'
                    elif col_name == '国家代码':
                        value = country
                    elif col_name == 'ASIN':
                        value = asin
                    elif col_name == '产品链接':
                        value = product_link or ''
                    else:
                        value = offer.get(pb_field, '')
                    
                    col_letter = col_idx_to_letter(col_idx)
                    range_str = f"{sheet_id}!{col_letter}{current_row}:{col_letter}{current_row}"
                    value_ranges.append({
                        "range": range_str,
                        "values": [[value]]
                    })
                    if col_name == '状态':
                        new_status_style_updates.append((current_row, col_idx, 'blue_bold'))
                
                self.log_manage(f"    新增: {asin}_{country} - {offer.get('brand_name', '')}")
                current_row += 1
            
            # 批量写入
            url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }
            body = {"valueRanges": value_ranges}
            
            resp = requests.post(url, headers=headers, json=body)
            result = resp.json()
            
            if result.get('code') != 0:
                self.log_manage(f"    批量写入失败: {result.get('msg')}")
            else:
                self.log_manage(f"    成功写入 {len(batch)} 个新offer")
            
            time.sleep(0.3)

        if new_status_style_updates:
            self._apply_style_updates(token, new_status_style_updates)
    
    def _find_first_empty_row(self, token, spreadsheet_token, sheet_id, asin_col_idx):
        """找到第一个空白行（基于ASIN列，只要文本为空就算空白）"""
        # 列索引转换为列字母
        def col_idx_to_letter(idx):
            result = ""
            while idx >= 0:
                result = chr(ord('A') + (idx % 26)) + result
                idx = idx // 26 - 1
            return result
        
        col_letter = col_idx_to_letter(asin_col_idx)
        self.log_manage(f"    查找空白行: 检查ASIN列 {col_letter} (索引{asin_col_idx})")
        
        url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values/{sheet_id}!{col_letter}1:{col_letter}3000"
        headers = {"Authorization": f"Bearer {token}"}
        
        resp = requests.get(url, headers=headers)
        data = resp.json()
        
        if data.get('code') != 0:
            self.log_manage(f"    读取失败: {data.get('msg')}")
            return None
        
        values = data.get('data', {}).get('valueRange', {}).get('values', [])
        self.log_manage(f"    ASIN列读取到 {len(values)} 行")
        
        # 显示前几行数据用于调试
        if values and len(values) > 0:
            self.log_manage(f"    前5行ASIN数据: {[str(v[0] if v else '') for v in values[:5]]}")
        
        # 找最后一个有数据的行
        last_data_row = 0
        for i, row in enumerate(values):
            cell_value = str(row[0]).strip() if row and len(row) > 0 else ''
            # 跳过表头、空值、None
            if cell_value and cell_value.lower() not in ['none', '', 'asin']:
                last_data_row = i + 1  # 1-based
        
        # 第一个空白行 = 最后有数据的行 + 1
        first_empty = last_data_row + 1
        self.log_manage(f"    最后有数据行: {last_data_row}, 第一个空白行: {first_empty}")
        return first_empty
    
    def _ensure_sheet_rows(self, token, spreadsheet_token, sheet_id, min_rows):
        """确保工作表有足够的行数"""
        # 获取当前工作表信息
        url = f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{spreadsheet_token}/sheets/{sheet_id}"
        headers = {"Authorization": f"Bearer {token}"}
        
        resp = requests.get(url, headers=headers)
        data = resp.json()
        
        current_rows = data.get('data', {}).get('sheet', {}).get('grid_properties', {}).get('row_count', 1000)
        
        if current_rows < min_rows:
            # 需要扩展行数
            rows_to_add = min_rows - current_rows + 100  # 多加100行buffer
            
            url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/dimension_range"
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }
            body = {
                "dimension": {
                    "sheetId": sheet_id,
                    "majorDimension": "ROWS",
                    "startIndex": current_rows,
                    "endIndex": current_rows + rows_to_add
                }
            }
            
            resp = requests.post(url, headers=headers, json=body)
            result = resp.json()
            
            if result.get('code') == 0:
                self.log_manage(f"    已扩展工作表行数: {current_rows} → {current_rows + rows_to_add}")
            else:
                self.log_manage(f"    扩展行数失败: {result.get('msg')}")
    
    def normalize_commission(self, val):
        """标准化佣金值为小数形式"""
        if not val:
            return 0.0
        val_str = str(val).strip()
        if '%' in val_str:
            try:
                return float(val_str.replace('%', '')) / 100
            except:
                return 0.0
        else:
            try:
                f = float(val_str)
                if f > 1:
                    return f / 100
                return f
            except:
                return 0.0
    
    def _do_statistics(self, start_date_str, end_date_str, increment_days):
        """执行统计操作"""
        try:
            self.log_manage("=" * 50)
            self.log_manage("开始Offer统计...")
            self.log_manage(f"统计日期范围: {start_date_str} 至 {end_date_str}")
            self.log_manage(f"新增花费/佣金统计天数: {increment_days}")
            self.log_manage("=" * 50)
            
            # 步骤1：获取Google Ads广告系列信息
            self.log_manage("\n【步骤1】获取Google Ads广告系列信息")
            self.update_progress_manage("获取广告系列...")
            
            campaign_data, succeeded_account_ids = self.get_all_campaigns_with_asin(start_date_str, end_date_str)
            if self.stop_flag:
                self.log_manage("已停止")
                return
            
            self.log_manage(f"  获取到 {len(campaign_data)} 个广告系列（来自 {len(succeeded_account_ids)} 个账户）")
            
            # 调试：显示前5个广告系列的信息
            for i, c in enumerate(campaign_data[:5]):
                name = c.get('campaign_name', '')[:40]
                self.log_manage(f"    [{i+1}] {name}...")
                urls = c.get('final_urls', [])
                self.log_manage(f"        Final URL: {urls[0][:80] if urls else 'None'}...")
                self.log_manage(f"        提取: ASIN={c.get('asin', 'None')}, 国家={c.get('country', 'None')}, UIDs={c.get('uids', [])}, LinkIDs={len(c.get('link_ids', []))}")
            
            # 构建(ASIN, 国家) -> 广告系列映射
            asin_country_campaigns = {}  # {(asin, country): [campaign_info, ...]}
            campaigns_without_asin = 0
            campaigns_without_country = 0
            for campaign in campaign_data:
                asin = campaign.get('asin')
                country = self.normalize_country_code(campaign.get('country'))
                if asin and country:
                    key = (asin, country)
                    if key not in asin_country_campaigns:
                        asin_country_campaigns[key] = []
                    asin_country_campaigns[key].append(campaign)
                elif not asin:
                    campaigns_without_asin += 1
                else:
                    campaigns_without_country += 1
            
            self.log_manage(f"  关联到 {len(asin_country_campaigns)} 个(ASIN+国家)组合")
            self.log_manage(f"  无法提取ASIN的广告系列: {campaigns_without_asin} 个")
            self.log_manage(f"  无法提取国家的广告系列: {campaigns_without_country} 个")
            if asin_country_campaigns:
                sample_keys = list(asin_country_campaigns.keys())[:5]
                self.log_manage(f"  样例: {sample_keys}...")
            
            # 详细显示每个广告系列的ASIN和国家提取情况（写入调试日志）
            self.log_debug(f"=== 广告系列详情（共{len(campaign_data)}个）===")
            for i, campaign in enumerate(campaign_data):
                c_name = campaign.get('campaign_name', '')
                c_asin = campaign.get('asin', 'None')
                c_country = campaign.get('country', 'None')
                c_status = campaign.get('status', 'Unknown')
                c_cost = campaign.get('cost_usd', 0)
                self.log_debug(f"  [{i+1}] {c_name} | ASIN={c_asin} | 国家={c_country} | 状态={c_status} | 花费=${c_cost:.2f}")
            
            # 步骤2：获取飞书表格数据
            self.log_manage("\n【步骤2】获取飞书表格数据")
            self.update_progress_manage("获取飞书数据...")
            
            feishu_token = self.get_feishu_token()
            if not feishu_token:
                self.log_manage("  ✗ 获取飞书Token失败")
                return
            
            feishu_data = self.get_feishu_sheet_data(feishu_token)
            if self.stop_flag:
                return
            
            self.log_manage(f"  获取到 {len(feishu_data)} 行数据")
            
            # 调试：显示表头和前3行数据
            if feishu_data:
                first_row = feishu_data[0]
                self.log_manage(f"  表头字段: {list(first_row.keys())}")
                for i, row in enumerate(feishu_data[:3]):
                    asin_val = row.get('ASIN', 'N/A')
                    status_val = row.get('状态', 'N/A')
                    self.log_manage(f"    [{i+1}] ASIN={asin_val}, 状态={status_val}")
            
            # 步骤3：获取PartnerBoost佣金数据
            self.log_manage("\n【步骤3】获取PartnerBoost佣金数据")
            self.update_progress_manage("获取佣金数据...")
            
            commission_data = self.get_all_commissions(start_date_str, end_date_str)
            if self.stop_flag:
                return
            
            # 构建(ASIN, 国家) -> 佣金映射（排除Rejected）
            asin_country_commission = {}  # {(asin, country): total_commission}
            
            # 调试：显示前几条交易数据的完整内容
            if commission_data:
                self.log_manage(f"  交易数据字段: {list(commission_data[0].keys()) if commission_data else '无'}")
                self.log_manage(f"  === 前3条交易完整数据 ===")
                for i, trans in enumerate(commission_data[:3]):
                    self.log_manage(f"    [{i+1}] 完整数据:")
                    for k, v in trans.items():
                        if v:  # 只显示非空字段
                            self.log_manage(f"        {k}: {v}")
            
            skipped_no_asin = 0
            skipped_rejected = 0
            rejected_commission_total = 0  # Rejected状态的佣金总额
            non_rejected_commission_total = 0  # 非Rejected佣金总额
            gross_commission_total = 0     # 所有状态的佣金总额（含Rejected）
            country_from_customer = 0      # 从 customer_country 字段获取
            country_from_merchant = 0      # 从 merchant_name 提取
            country_from_mcid = 0          # 从 mcid 提取
            country_default_us = 0         # 默认为 US
            
            # 构建四种佣金映射（严格区分uid和非uid，避免重复计算）：
            # 1. (ASIN, 国家) -> 佣金（所有交易的总和，仅用于总计计算）
            # 2. (ASIN, 国家, uid) -> 佣金（仅有uid的交易，用于uid精确匹配）
            # 3. (ASIN, 国家) -> 佣金（仅无uid的交易，用于ASIN+国家匹配）
            # 4. ASIN -> 佣金（仅无uid且国家默认US的交易，用于ASIN-only备选匹配）
            asin_country_uid_commission = {}  # {(asin, country, uid): total_commission} 有uid的交易
            asin_country_no_uid_commission = {}  # {(asin, country): total_commission} 无uid的交易
            asin_only_commission = {}  # {asin: total_commission} 无uid且国家默认US
            
            for trans in commission_data:
                comm = float(trans.get('sale_comm', 0) or 0)
                gross_commission_total += comm

                if trans.get('status') == 'Rejected':
                    skipped_rejected += 1
                    # 累加Rejected佣金
                    rejected_commission_total += comm
                    continue
                else:
                    non_rejected_commission_total += comm
                    
                asin = trans.get('prod_id', '')
                merchant_name = trans.get('merchant_name', '')
                mcid = trans.get('mcid', '')
                uid = trans.get('uid', '')  # 获取uid
                
                # 优先从 customer_country 字段获取国家代码
                country = trans.get('customer_country', '') or trans.get('geo', '') or trans.get('country', '')
                country_source = 'customer_country'
                
                if not country and merchant_name:
                    # 尝试从 merchant_name 提取国家代码（如 "DE-Anker" -> "DE"）
                    country = self.extract_country_from_merchant_name(merchant_name)
                    if country:
                        country_source = 'merchant_name'
                
                if not country and mcid:
                    # 尝试从 mcid 提取国家代码（如 "amzdeanker" -> "DE"）
                    country = self.extract_country_from_mcid(mcid)
                    if country:
                        country_source = 'mcid'
                
                if not country:
                    # 品牌名中无国家代码，默认为 US
                    country = 'US'
                    country_source = 'default_us'
                
                if not asin:
                    skipped_no_asin += 1
                    continue
                
                # 统计国家代码来源
                if country_source == 'customer_country':
                    country_from_customer += 1
                elif country_source == 'merchant_name':
                    country_from_merchant += 1
                elif country_source == 'mcid':
                    country_from_mcid += 1
                else:
                    country_default_us += 1
                
                country = self.normalize_country_code(country)
                key = (asin, country)
                
                # 总映射（所有交易，用于总计计算）
                asin_country_commission[key] = asin_country_commission.get(key, 0) + comm
                
                # 严格按uid分池
                if uid:
                    # 有uid的交易 → uid池（按uid精确匹配到offer行）
                    uid_key = (asin, country, uid)
                    asin_country_uid_commission[uid_key] = asin_country_uid_commission.get(uid_key, 0) + comm
                else:
                    # 无uid的交易 → 非uid池（按ASIN+国家匹配，每个组合只分配一次）
                    asin_country_no_uid_commission[key] = asin_country_no_uid_commission.get(key, 0) + comm
                    # 对于国家默认US的情况，额外加入ASIN-only备选映射
                    if country_source == 'default_us':
                        asin_only_commission[asin] = asin_only_commission.get(asin, 0) + comm
            
            self.log_manage(f"  获取到 {len(commission_data)} 条交易")
            self.log_manage(f"  佣金池分布:")
            self.log_manage(f"    • UID池(按uid匹配): {len(asin_country_uid_commission)} 个uid组合, 金额=${sum(asin_country_uid_commission.values()):.2f}")
            self.log_manage(f"    • 非UID池(按ASIN+国家匹配): {len(asin_country_no_uid_commission)} 个组合, 金额=${sum(asin_country_no_uid_commission.values()):.2f}")
            self.log_manage(f"    • ASIN-only备选池: {len(asin_only_commission)} 个ASIN")
            self.log_manage(f"    • 合计(非Rejected): {len(asin_country_commission)} 个(ASIN+国家)组合, 金额=${sum(asin_country_commission.values()):.2f}")
            self.log_manage(f"  国家代码来源统计:")
            self.log_manage(f"    • 从customer_country字段: {country_from_customer} 条")
            self.log_manage(f"    • 从merchant_name提取: {country_from_merchant} 条")
            self.log_manage(f"    • 从mcid提取: {country_from_mcid} 条")
            self.log_manage(f"    • 默认为US: {country_default_us} 条")
            if skipped_rejected > 0 or skipped_no_asin > 0:
                self.log_manage(f"  跳过: Rejected={skipped_rejected}条(佣金${rejected_commission_total:.2f}), 无ASIN={skipped_no_asin}")
            if asin_only_commission:
                sample_asins = list(asin_only_commission.keys())[:5]
                self.log_manage(f"  佣金ASIN样例: {sample_asins}")
            # 输出带uid佣金的详细信息
            if asin_country_uid_commission:
                self.log_manage(f"  带UID佣金详情（共{len(asin_country_uid_commission)}条）:")
                for uid_key, comm_value in asin_country_uid_commission.items():
                    asin, country, uid = uid_key
                    self.log_manage(f"    • UID={uid}, ASIN={asin}, 国家={country}, 佣金=${comm_value:.2f}")
            
            # 步骤3.5：获取YeahPromos佣金数据
            self.log_manage("\n【步骤3.5】获取YeahPromos佣金数据")
            yp_commission_data = self.get_all_yp_commissions(start_date_str, end_date_str)
            if self.stop_flag:
                return

            yp_asin_brand_commission = {}
            yp_rejected_commission_total = 0.0
            yp_non_rejected_commission_total = 0.0
            yp_gross_commission_total = 0.0
            yp_missing_asin_brand = 0
            yp_missing_examples = []

            yp_pid_brand_to_offer_keys = {}
            yp_offer_index = self.build_break_even_brand_country_offer_index(feishu_data)
            yp_offer_lookup = self.build_brand_country_lookup_maps(yp_offer_index)
            for row in feishu_data:
                if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                    continue
                link = str(self.normalize_sheet_cell_value(row.get('投放链接', '')) or '').strip()
                if not self._is_yp_tracking_link(link):
                    continue
                asin = str(row.get('ASIN', '') or '').strip()
                brand_id = self.build_yp_row_brand_key(row)
                country = self.normalize_country_code(row.get('国家代码', '') or '')
                if not asin or not brand_id:
                    continue
                yp_marker = self.extract_yp_offer_marker(link, row.get('产品链接', ''))
                pid_match = re.search(r'[?&]pid=([^&]+)', link, re.IGNORECASE)
                if pid_match:
                    yp_pid_brand_to_offer_keys.setdefault((pid_match.group(1), brand_id), []).append(
                        (asin, brand_id, country, yp_marker)
                    )

            for trans in yp_commission_data:
                comm = float(trans.get('sale_comm', 0) or 0)
                yp_gross_commission_total += comm
                status = str(trans.get('status', '') or '').strip().lower()
                advert_id = str(trans.get('advert_id', '') or '').strip()
                asin = str(trans.get('asin', '') or trans.get('prod_id', '') or '').strip().upper()
                if not asin:
                    asin = self.extract_asin_from_yp_order_id(trans.get('id', ''))
                entry = self.resolve_yp_brand_country_entry(trans, yp_offer_lookup)
                brand_id = self.build_yp_transaction_brand_key(trans, entry)
                country = entry.get('country', '') if entry else ''

                if not asin or not brand_id:
                    yp_missing_asin_brand += 1
                    if len(yp_missing_examples) < 10:
                        yp_missing_examples.append({
                            'asin': asin,
                            'advert_name': trans.get('advert_name', ''),
                            'advert_id': advert_id,
                            'order_id': trans.get('id', ''),
                            'sale_comm': comm,
                        })
                    continue

                pid_key = (advert_id, brand_id)
                candidate_offer_keys = yp_pid_brand_to_offer_keys.get(pid_key, [])
                if len(candidate_offer_keys) == 1:
                    key = candidate_offer_keys[0]
                else:
                    key = (asin, brand_id, country, '')
                if key not in yp_asin_brand_commission:
                    yp_asin_brand_commission[key] = {'non_rejected': 0.0, 'gross': 0.0, 'rejected': 0.0}
                yp_asin_brand_commission[key]['gross'] += comm

                if status == 'rejected':
                    yp_asin_brand_commission[key]['rejected'] += comm
                    yp_rejected_commission_total += comm
                else:
                    yp_asin_brand_commission[key]['non_rejected'] += comm
                    yp_non_rejected_commission_total += comm

            self.log_manage(
                f"  YP佣金汇总: 非Rejected=${yp_non_rejected_commission_total:.2f}, "
                f"总佣金=${yp_gross_commission_total:.2f}, Rejected=${yp_rejected_commission_total:.2f}, "
                f"匹配组数={len(yp_asin_brand_commission)}, 缺少ASIN/品牌ID={yp_missing_asin_brand}"
            )
            if yp_missing_examples:
                self.log_manage("  YP未匹配样例（多为飞书Offer表缺少对应ASIN/品牌）:")
                for item in yp_missing_examples:
                    self.log_manage(
                        f"    - advert={item.get('advert_name', '')}({item.get('advert_id', '')}), "
                        f"ASIN={item.get('asin', '') or '-'}, 佣金=${item.get('sale_comm', 0.0):.2f}, "
                        f"订单ID={item.get('order_id', '')}"
                    )

            # 步骤4：更新飞书表格
            self.log_manage("\n【步骤4】更新飞书表格")
            self.update_progress_manage("更新飞书...")
            
            
            
            # 步骤4.5：通过link_id精确分配广告系列到具体的offer行
            self.log_manage("\n  🔗 通过link_id精确匹配广告系列到offer行...")
            self.update_progress_manage("精确匹配广告系列...")
            
            row_campaigns = self._resolve_campaign_to_row_mapping(
                feishu_data, asin_country_campaigns, campaign_data
            )
            if row_campaigns:
                self.log_manage(f"  已精确分配 {len(row_campaigns)} 行的广告系列")
            
            updates, offer_commission_context = self.calculate_updates(
                feishu_data,
                asin_country_campaigns,
                asin_country_commission,
                asin_only_commission,
                asin_country_uid_commission,
                asin_country_no_uid_commission,
                row_campaigns,
                yp_asin_brand_commission
            )
            offer_commission_context['yp_matched_duplicate_keys'] = {
                key for key, value in yp_asin_brand_commission.items()
                if value.get('non_rejected', 0) > 0
            }
            
            self.log_manage(f"  计算得到 {len(updates)} 个需要更新的offer")
            
            if updates:
                # 调试：显示前5个更新
                for i, u in enumerate(updates[:5]):
                    self.log_manage(f"    [{i+1}] ASIN={u.get('asin')}, 国家={u.get('country')}, 状态={u.get('status')}")
                self.apply_feishu_updates(feishu_token, updates)
                self.consolidate_yp_duplicate_offer_rows(
                    feishu_token,
                    matched_keys=offer_commission_context.get('yp_matched_duplicate_keys')
                )
            else:
                self.log_manage("  没有需要更新的内容")
                self.log_manage("  可能原因：")
                self.log_manage("    1. 广告系列中提取不到ASIN")
                self.log_manage("    2. 飞书表格中的ASIN与广告系列不匹配")
                self.log_manage("    3. 所有offer状态都是'未测试'且没有广告系列")
            
            # 生成统计报告
            self.log_manage("\n" + "=" * 50)
            self.log_manage("【统计报告】")
            self.log_manage("=" * 50)
            
            # 构建飞书表格中所有的 (ASIN, 国家) 组合
            feishu_asin_country_set = set()
            feishu_asin_set = set()
            for row in feishu_data:
                asin = row.get('ASIN', '')
                country = row.get('国家代码', '')
                if asin:
                    feishu_asin_set.add(asin)
                    if country:
                        feishu_asin_country_set.add((asin, self.normalize_country_code(country)))
            
            # 1. 未匹配到offer的广告系列统计
            unmatched_campaigns = []
            for key, campaigns_list in asin_country_campaigns.items():
                if key not in feishu_asin_country_set:
                    for c in campaigns_list:
                        unmatched_campaigns.append({
                            'asin': key[0],
                            'country': key[1],
                            'campaign_name': c.get('campaign_name', ''),
                            'cost_usd': c.get('cost_usd', 0)
                        })
            
            # 2. 未匹配到offer的PB佣金订单统计
            unmatched_commission_asin_country = {}  # {(asin, country): commission}
            unmatched_commission_asin_only = {}  # {asin: commission} - ASIN-only匹配也失败的
            
            for key, comm in asin_country_commission.items():
                if key not in feishu_asin_country_set:
                    unmatched_commission_asin_country[key] = comm
            
            for asin, comm in asin_only_commission.items():
                if asin not in feishu_asin_set:
                    unmatched_commission_asin_only[asin] = comm
            
            # 3. 更新统计（按状态分类）
            status_counts = {}
            commission_with_asterisk = 0  # 带"*"标记的佣金数量
            total_cost_updated = 0
            total_commission_updated = 0
            
            for u in updates:
                status = u.get('status', '仅更新佣金')
                status_counts[status] = status_counts.get(status, 0) + 1
                
                if u.get('total_cost'):
                    total_cost_updated += u.get('total_cost', 0)
                
                commission = u.get('commission')
                if commission:
                    if isinstance(commission, str) and '*' in str(commission):
                        commission_with_asterisk += 1
                        # 提取数值部分
                        try:
                            total_commission_updated += float(str(commission).replace('$', '').replace('*', ''))
                        except:
                            pass
                    else:
                        total_commission_updated += float(commission) if commission else 0
            
            # 输出报告
            self.log_manage("\n📊 数据概览:")
            self.log_manage(f"  • Google Ads 广告系列总数: {len(campaign_data)}")
            self.log_manage(f"  • 飞书表格 Offer 总数: {len(feishu_data)}")
            self.log_manage(f"  • PartnerBoost 交易总数: {len(commission_data)}")
            self.log_manage(f"  • YeahPromos 交易总数: {len(yp_commission_data)}")
            self.log_manage(f"  • 本次更新的 Offer 数: {len(updates)}")
            
            self.log_manage("\n⚠️ 未匹配数据:")
            self.log_manage(f"  • 未匹配到飞书Offer的广告系列: {len(unmatched_campaigns)} 个")
            if unmatched_campaigns:
                # 按(ASIN, 国家)分组显示
                unmatched_by_key = {}
                for c in unmatched_campaigns:
                    key = (c['asin'], c['country'])
                    if key not in unmatched_by_key:
                        unmatched_by_key[key] = []
                    unmatched_by_key[key].append(c)
                
                for (asin, country), campaigns_list in list(unmatched_by_key.items())[:5]:
                    cost_sum = sum(c['cost_usd'] for c in campaigns_list)
                    self.log_manage(f"    - ASIN={asin}, 国家={country}, 广告系列数={len(campaigns_list)}, 花费=${cost_sum:.2f}")
                if len(unmatched_by_key) > 5:
                    self.log_manage(f"    ... 还有 {len(unmatched_by_key) - 5} 个(ASIN+国家)组合未显示")
            
            self.log_manage(f"  • 未匹配到飞书Offer的PB佣金(ASIN+国家): {len(unmatched_commission_asin_country)} 个")
            # 计算ASIN+国家未匹配的佣金总额
            total_unmatched_asin_country_comm = sum(unmatched_commission_asin_country.values()) if unmatched_commission_asin_country else 0
            if unmatched_commission_asin_country:
                self.log_manage(f"    未匹配佣金总额: ${total_unmatched_asin_country_comm:.2f}")
                for (asin, country), comm in list(unmatched_commission_asin_country.items())[:3]:
                    self.log_manage(f"    - ASIN={asin}, 国家={country}, 佣金=${comm:.2f}")
                if len(unmatched_commission_asin_country) > 3:
                    self.log_manage(f"    ... 还有 {len(unmatched_commission_asin_country) - 3} 条未显示")
            
            self.log_manage(f"  • 未匹配到飞书Offer的PB佣金(仅ASIN): {len(unmatched_commission_asin_only)} 个")
            # 计算仅ASIN未匹配的佣金总额
            total_unmatched_asin_only_comm = sum(unmatched_commission_asin_only.values()) if unmatched_commission_asin_only else 0
            if unmatched_commission_asin_only:
                self.log_manage(f"    未匹配佣金总额: ${total_unmatched_asin_only_comm:.2f}")
            
            # 计算未匹配佣金总额：直接用 PB总额(非Rejected) - 已匹配 来避免重复计算
            # （旧方法 asin_country + asin_only 相加会对"ASIN完全不在飞书中"的情况重复计算）
            total_pb_non_rejected_for_unmatched = sum(asin_country_commission.values())
            total_matched_commission = sum(
                comm for key, comm in asin_country_commission.items()
                if key in feishu_asin_country_set
            )
            total_unmatched_commission = total_pb_non_rejected_for_unmatched - total_matched_commission
            self.log_manage(f"  • 未匹配佣金合计: ${total_unmatched_commission:.2f}（= PB总额${total_pb_non_rejected_for_unmatched:.2f} - 已匹配${total_matched_commission:.2f}）")
            self.log_manage(f"  • Rejected佣金合计: ${rejected_commission_total + yp_rejected_commission_total:.2f}")
            
            self.log_manage("\n📈 更新统计:")
            if status_counts:
                for status, count in status_counts.items():
                    self.log_manage(f"  • {status}: {count} 个")
            else:
                self.log_manage(f"  • 无状态更新")
            
            self.log_manage(f"\n💰 本次更新涉及的金额:")
            self.log_manage(f"  • 广告花费: ${total_cost_updated:.2f}")
            self.log_manage(f"  • 佣金: ${total_commission_updated:.2f}")
            if commission_with_asterisk > 0:
                self.log_manage(f"  • 带*标记的佣金数 (可能不准确): {commission_with_asterisk} 个")
                self.log_manage(f"    (原因: PB数据无国家信息，且飞书有多个相同ASIN的Offer)")
            
            # 计算广告数据中无法提取信息的统计
            self.log_manage("\n🔍 数据质量:")
            self.log_manage(f"  • 无法提取ASIN的广告系列: {campaigns_without_asin} 个")
            self.log_manage(f"  • 无法提取国家的广告系列: {campaigns_without_country} 个")
            self.log_manage(f"  • PB交易中Rejected状态: {skipped_rejected} 条")
            self.log_manage(f"  • YP交易中缺少ASIN/品牌ID: {yp_missing_asin_brand} 条")
            self.log_manage(f"  • PB交易中无ASIN: {skipped_no_asin} 条")
            self.log_manage(f"  • PB佣金国家来源: customer_country={country_from_customer}, merchant_name={country_from_merchant}, mcid={country_from_mcid}, 默认US={country_default_us}")
            
            # 步骤5：更新"总计"行
            self.log_manage("\n【步骤5】更新飞书表格'总计'行")
            self.update_progress_manage("更新总计行...")
            
            try:
                # 直接从PB源数据计算已匹配佣金（不再从单元格求和，避免旧值导致重复计算）
                # 已匹配佣金 = PB中(ASIN+国家)在飞书表格中存在的佣金之和
                total_commission_sum = 0
                for key, comm in asin_country_commission.items():
                    if key in feishu_asin_country_set:
                        total_commission_sum += comm
                
                # 校验：PB总佣金（非Rejected） = 已匹配 + 未匹配
                total_pb_non_rejected = sum(asin_country_commission.values())
                self.log_manage(f"  PB佣金校验: PB总额(非Rejected)=${total_pb_non_rejected:.2f} = 已匹配${total_commission_sum:.2f} + 未匹配${total_unmatched_commission:.2f}")
                
                # 总花费：从MCC获取选定日期范围内的全部花费（包含已删除广告系列）
                self.log_manage(f"  从MCC获取全部花费（含已删除广告系列）...")
                mcc_total_cost = self.get_mcc_total_cost(start_date_str, end_date_str)
                if mcc_total_cost is not None:
                    total_cost_sum = mcc_total_cost
                else:
                    current_campaign_cost_sum = sum(c.get('cost_usd', 0) for c in campaign_data)
                    total_cost_sum = self.get_last_successful_summary_total_cost(min_cost=current_campaign_cost_sum)
                    if total_cost_sum is None:
                        self.log_manage(f"  ⚠ 无法获取MCC全部花费，本次保留总计行原有广告花费")
                    else:
                        self.log_manage(f"  ⚠ 无法获取MCC全部花费，恢复最近一次成功总计花费 ${total_cost_sum:.2f}")
                
                self.log_manage(f"  飞书表格所有Offer汇总:")
                combined_non_rejected_commission_total = non_rejected_commission_total + yp_non_rejected_commission_total
                combined_gross_commission_total = gross_commission_total + yp_gross_commission_total
                combined_rejected_commission_total = rejected_commission_total + yp_rejected_commission_total

                self.log_manage(f"    • 非Rejected佣金: ${combined_non_rejected_commission_total:.2f}")
                self.log_manage(f"    • 总佣金: ${combined_gross_commission_total:.2f}")
                self.log_manage(f"    • Rejected佣金: ${combined_rejected_commission_total:.2f}")
                if total_cost_sum is None:
                    self.log_manage(f"    • MCC总花费(USD): 保留原值")
                else:
                    self.log_manage(f"    • MCC总花费(USD): ${total_cost_sum:.2f}")
                
                # 更新第二行（总计行），传入未匹配佣金和Rejected佣金
                self.update_feishu_summary_row(
                    feishu_token,
                    combined_non_rejected_commission_total,
                    total_cost_sum,
                    combined_gross_commission_total,
                    combined_rejected_commission_total
                )
                
            except Exception as e:
                self.log_manage(f"  更新总计行时出错: {str(e)}")

            # 先统计最近N天新增数据，日志延迟到流程末尾输出，避免被步骤6/7刷上去。
            recent_increment_log_lines = []
            self.update_progress_manage("统计新增数据...")
            try:
                recent_increment_log_lines = self.build_recent_increment_log_lines(increment_days)
            except Exception as e:
                recent_increment_log_lines = [
                    "\n【步骤5.5】输出新增数据",
                    f"  统计新增数据时出错: {str(e)}",
                ]
            
            # 步骤6：更新"广告系列"表格
            self.log_manage("\n【步骤6】更新'广告系列'表格")
            self.update_progress_manage("更新广告系列表格...")
            
            try:
                # 构建 campaign_id → tracking_link 的精确映射（基于row_campaigns的匹配结果）
                campaign_id_to_tracking_link = {}
                if row_campaigns:
                    # 先建一个 row_index → tracking_link 的映射
                    row_index_to_link = {}
                    for row in feishu_data:
                        ri = row.get('row_index')
                        if ri in row_campaigns:
                            tl = row.get('投放链接', '')
                            if isinstance(tl, list) and len(tl) > 0 and isinstance(tl[0], dict):
                                tl = tl[0].get('link', '') or tl[0].get('text', '')
                            if isinstance(tl, str) and tl:
                                row_index_to_link[ri] = tl
                    # 然后 campaign_id → tracking_link
                    for ri, cams in row_campaigns.items():
                        link = row_index_to_link.get(ri, '')
                        if link:
                            for c in cams:
                                cid = c.get('campaign_id')
                                if cid:
                                    campaign_id_to_tracking_link[cid] = link
                    if campaign_id_to_tracking_link:
                        self.log_manage(f"  构建了 {len(campaign_id_to_tracking_link)} 个广告系列→投放链接的精确映射")
                
                campaign_sheet_report = self.update_campaigns_sheet(
                    feishu_token=feishu_token,
                    campaign_data=campaign_data,
                    commission_data=commission_data,
                    yp_commission_data=yp_commission_data,
                    feishu_data=feishu_data,
                    campaign_id_to_tracking_link=campaign_id_to_tracking_link,
                    offer_commission_context=offer_commission_context,
                    start_date_str=start_date_str,
                    end_date_str=end_date_str,
                    increment_days=increment_days
                )
            except Exception as e:
                self.log_manage(f"  更新广告系列表格时出错: {str(e)}")
                import traceback
                self.log_manage(traceback.format_exc())
                campaign_sheet_report = None

            if campaign_sheet_report:
                uid_conflicts = campaign_sheet_report.get('uid_conflicts', [])
                unmatched_uid_transactions = campaign_sheet_report.get('unmatched_uid_transactions', [])
                if uid_conflicts or unmatched_uid_transactions:
                    self.log_manage("\n【广告系列UID冲突报告】")
                    if uid_conflicts:
                        total_conflict_commission = sum(item.get('commission', 0.0) for item in uid_conflicts)
                        conflict_uids = {item.get('uid', '') for item in uid_conflicts if item.get('uid', '')}
                        conflict_campaign_keys = set()
                        for item in uid_conflicts:
                            for campaign in item.get('campaigns', []):
                                conflict_campaign_keys.add(
                                    (campaign.get('row_index', ''), campaign.get('campaign_name', ''))
                                )
                        self.log_manage(
                            f"  UID冲突: {len(conflict_uids)} 个UID，涉及 {len(conflict_campaign_keys)} 个广告系列，冲突佣金合计 ${total_conflict_commission:.2f}"
                        )
                        for item in uid_conflicts:
                            self.log_manage(
                                f"    - UID={item.get('uid', '')}, 佣金=${item.get('commission', 0.0):.2f}, "
                                f"ASIN={item.get('asin', '')}, 国家={item.get('country', '')}, "
                                f"关联广告系列={len(item.get('campaigns', []))} 个"
                            )
                            for campaign in item.get('campaigns', []):
                                self.log_manage(
                                    f"      row={campaign.get('row_index', '')}, 广告系列={campaign.get('campaign_name', '')}, 状态={campaign.get('status', '')}"
                                )
                    if unmatched_uid_transactions:
                        total_unmatched_commission = sum(item.get('commission', 0.0) for item in unmatched_uid_transactions)
                        self.log_manage(
                            f"  UID未匹配到广告系列: {len(unmatched_uid_transactions)} 条佣金，金额合计 ${total_unmatched_commission:.2f}"
                        )
                        for item in unmatched_uid_transactions:
                            self.log_manage(
                                f"    - UID={item.get('uid', '')}, 佣金=${item.get('commission', 0.0):.2f}, "
                                f"ASIN={item.get('asin', '')}, 国家={item.get('country', '')}, 原因={item.get('reason', '')}"
                            )

            # 步骤7：更新"ads | 品牌"表格
            self.log_manage("\n【步骤7】更新'ads | 品牌'表格")
            self.update_progress_manage("更新ads品牌表...")
            try:
                spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
                latest_feishu_data = self.get_feishu_sheet_data(feishu_token)
                if latest_feishu_data:
                    feishu_data = latest_feishu_data
                self.update_ads_brand_sheet(
                    feishu_token=feishu_token,
                    spreadsheet_token=spreadsheet_token,
                    feishu_data=feishu_data,
                    commission_data=commission_data,
                    yp_commission_data=yp_commission_data,
                    start_date_str=start_date_str,
                    end_date_str=end_date_str
                )
            except Exception as e:
                self.log_manage(f"  更新ads | 品牌表格时出错: {str(e)}")
                import traceback
                self.log_manage(traceback.format_exc())
            
            for line in recent_increment_log_lines:
                self.log_manage(line)

            self.log_manage("\n" + "=" * 50)
            self.log_manage("✅ 统计完成！")
            self.log_manage("=" * 50)
            
        except Exception as e:
            self.log_manage(f"\n错误: {str(e)}")
            import traceback
            self.log_manage(traceback.format_exc())
        finally:
            self.root.after(0, self._restore_manage_ui)
    
    def update_campaigns_sheet(self, feishu_token, campaign_data, commission_data, yp_commission_data, feishu_data, campaign_id_to_tracking_link=None, offer_commission_context=None, start_date_str=None, end_date_str=None, increment_days=None):
        """更新'广告系列'表格
        
        Args:
            feishu_token: 飞书访问令牌
            campaign_data: Google Ads广告系列数据列表
            commission_data: PB佣金数据列表
            yp_commission_data: YP佣金数据列表
            feishu_data: offer表格数据（用于获取广告系列名称-投放链接映射）
            offer_commission_context: offer行级佣金归因结果，确保广告系列与offer使用同口径
            start_date_str: 本次统计开始日期（用于汇总已移除广告系列历史指标）
            end_date_str: 本次统计结束日期（用于汇总已移除广告系列历史指标）
        """
        # 新表格的配置
        campaigns_spreadsheet_token = "KnJ1wphpBiVMrGkWl5ncUkMGnfe"
        campaigns_sheet_id = CAMPAIGNS_SHEET_ID
        report = {
            'uid_conflicts': [],
            'unmatched_uid_transactions': [],
        }
        
        # 步骤1：从offer表格构建 (ASIN, 国家) -> 投放链接 和 每单佣金 的映射
        self.log_manage("  构建(ASIN+国家)到投放链接和每单佣金的映射...")
        if campaign_id_to_tracking_link is None:
            campaign_id_to_tracking_link = {}

        try:
            latest_feishu_data = self.get_feishu_sheet_data(feishu_token)
            if latest_feishu_data:
                feishu_data = latest_feishu_data
        except Exception as e:
            self.log_manage(f"  读取最新Offer表失败，继续使用内存快照: {str(e)}")
        
        asin_country_to_tracking_links = {}  # {(asin, country): [tracking_link, ...]} 改为列表，支持多个链接
        asin_country_to_commission_per_order = {}  # {(asin, country): commission_per_order}
        product_link_to_tracking_link, duplicate_product_link_mappings = self.build_offer_product_link_to_tracking_link_map(feishu_data)
        
        rows_with_tracking_link = 0
        for row in feishu_data:
            if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                continue
            asin = row.get('ASIN', '')
            country = row.get('国家代码', '')
            tracking_link = row.get('投放链接', '')
            normalized_tracking_link = self.normalize_sheet_cell_value(tracking_link)
            # 直接读取折扣价和佣金列来计算每单佣金，而不是读取公式列
            discount_price = row.get('折扣价', '')
            commission_rate = row.get('佣金', '')
            
            if asin and country:
                key = (asin, self.normalize_country_code(country))
                if normalized_tracking_link:
                    rows_with_tracking_link += 1
                    if key not in asin_country_to_tracking_links:
                        asin_country_to_tracking_links[key] = []
                    asin_country_to_tracking_links[key].append(normalized_tracking_link)
                # 计算每单佣金 = 折扣价 * 佣金（而不是直接读取公式列）
                if discount_price and commission_rate:
                    try:
                        # 处理折扣价（可能带$符号）
                        price_str = str(discount_price).replace('$', '').replace(',', '').strip()
                        price_val = float(price_str) if price_str else 0
                        
                        # 处理佣金（可能是百分比如"8%"或小数如"0.08"）
                        comm_str = str(commission_rate).strip()
                        if '%' in comm_str:
                            comm_val = float(comm_str.replace('%', '')) / 100
                        else:
                            comm_val = float(comm_str)
                            # 如果大于1，认为是百分比形式
                            if comm_val > 1:
                                comm_val = comm_val / 100
                        
                        # 计算每单佣金
                        commission_per_order = price_val * comm_val
                        if commission_per_order > 0:
                            asin_country_to_commission_per_order[key] = commission_per_order
                    except (ValueError, TypeError):
                        pass  # 无法计算，跳过
        
        self.log_manage(f"    找到 {len(asin_country_to_tracking_links)} 个(ASIN+国家)到投放链接的映射")
        self.log_manage(f"    找到 {len(product_link_to_tracking_link)} 个产品链接到投放链接的映射")
        if duplicate_product_link_mappings:
            self.log_manage(f"    ⚠ 产品链接对应多个投放链接，跳过冲突映射 {len(duplicate_product_link_mappings)} 个")
        self.log_manage(f"    找到 {len(asin_country_to_commission_per_order)} 个(ASIN+国家)到每单佣金的映射")
        
        # 调试日志
        self.log_debug(f"=== (ASIN+国家)到投放链接映射 ===")
        for key, links in list(asin_country_to_tracking_links.items())[:20]:
            self.log_debug(f"  {key[0]}_{key[1]} -> {len(links)}个链接")
        
        if offer_commission_context is None:
            offer_commission_context = {}

        offer_row_commissions = offer_commission_context.get('row_commissions', {}) or {}
        yp_offer_context = self.build_yp_offer_group_context(
            feishu_data,
            offer_row_commissions=offer_row_commissions
        )
        offer_group_summary_by_key = yp_offer_context.get('offer_group_summary_by_key', {})

        def normalize_link_value(value):
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                return value[0].get('link', '') or value[0].get('text', '')
            return value if isinstance(value, str) else ''

        def normalize_country_value(value):
            return self.normalize_country_code(value)
        
        # 步骤3：构建MCC中所有广告系列的信息
        self.log_manage("  处理MCC广告系列数据...")
        mcc_campaigns = {}  # {campaign_name: campaign_info}
        
        for campaign in campaign_data:
            campaign_name = campaign.get('campaign_name', '')
            if not campaign_name:
                continue
            
            if campaign_name not in mcc_campaigns:
                mcc_campaigns[campaign_name] = {
                    'account_id': campaign.get('account_id', ''),
                    'cost_usd': 0,
                    'status': campaign.get('status', 'UNKNOWN'),
                    'campaign_id': campaign.get('campaign_id', ''),
                    'asin': campaign.get('asin', ''),
                    'country': campaign.get('country', ''),
                    'final_urls': campaign.get('final_urls', []),
                    'final_url_suffix': campaign.get('final_url_suffix', ''),
                    'clicks': 0,
                    'avg_cpc_usd': 0.0,
                    'preset_cpc_usd': float(campaign.get('preset_cpc_usd', 0.0) or 0.0),
                    '_cpc_weighted_cost': 0.0,
                    '_cpc_weighted_clicks': 0,
                    '_traffic_source': 'pb',
                    'enabled_keyword_ctr_weighted_numerator': 0.0,
                    'enabled_keyword_ctr_impressions': 0
                }
            
            # 累加花费
            mcc_campaigns[campaign_name]['cost_usd'] += campaign.get('cost_usd', 0)
            mcc_campaigns[campaign_name]['enabled_keyword_ctr_weighted_numerator'] += campaign.get(
                'enabled_keyword_ctr_weighted_numerator', 0.0
            )
            mcc_campaigns[campaign_name]['enabled_keyword_ctr_impressions'] += campaign.get(
                'enabled_keyword_ctr_impressions', 0
            )
            campaign_clicks = int(campaign.get('clicks', 0) or 0)
            campaign_avg_cpc = float(campaign.get('avg_cpc_usd', 0.0) or 0.0)
            campaign_preset_cpc = float(campaign.get('preset_cpc_usd', 0.0) or 0.0)
            mcc_campaigns[campaign_name]['clicks'] += campaign_clicks
            if campaign_clicks > 0:
                mcc_campaigns[campaign_name]['_cpc_weighted_cost'] += campaign_avg_cpc * campaign_clicks
                mcc_campaigns[campaign_name]['_cpc_weighted_clicks'] += campaign_clicks
                total_clicks_for_cpc = mcc_campaigns[campaign_name]['_cpc_weighted_clicks']
                if total_clicks_for_cpc > 0:
                    mcc_campaigns[campaign_name]['avg_cpc_usd'] = (
                        mcc_campaigns[campaign_name]['_cpc_weighted_cost'] / total_clicks_for_cpc
                    )
            if campaign_preset_cpc > mcc_campaigns[campaign_name].get('preset_cpc_usd', 0.0):
                mcc_campaigns[campaign_name]['preset_cpc_usd'] = campaign_preset_cpc
            
            # 如果之前没有ASIN/国家，尝试从当前campaign获取
            if not mcc_campaigns[campaign_name]['asin'] and campaign.get('asin'):
                mcc_campaigns[campaign_name]['asin'] = campaign.get('asin')
            if not mcc_campaigns[campaign_name]['country'] and campaign.get('country'):
                mcc_campaigns[campaign_name]['country'] = campaign.get('country')
        
        self.log_manage(f"    MCC中共有 {len(mcc_campaigns)} 个广告系列")
        
        # 步骤4：读取现有的"广告系列"表格数据
        self.log_manage("  读取现有广告系列表格...")
        existing_campaigns_data = self.read_campaigns_sheet(feishu_token, campaigns_spreadsheet_token, campaigns_sheet_id)
        
        if existing_campaigns_data is None:
            self.log_manage("  ✗ 无法读取广告系列表格")
            return
        
        existing_rows, column_map, first_empty_row = existing_campaigns_data
        self.log_manage(f"    读取到 {len(existing_rows)} 行现有数据，第一个空行: {first_empty_row}")
        campaign_commission_baseline = self.build_campaign_commission_baseline_snapshot(existing_rows)
        
        # 构建现有广告系列名称到行索引、投放链接、上次花费/佣金的映射
        existing_campaign_names = {}  # {campaign_name: row_index}
        existing_campaign_links = {}  # {campaign_name: tracking_link} 已有的投放链接
        existing_campaign_cost = {}   # {campaign_name: float} 上次运行时的花费
        existing_campaign_commission = {}  # {campaign_name: float} 上次运行时的佣金
        existing_campaign_status = {}  # {campaign_name: str} 上次运行时的状态
        existing_campaign_clicks = {}  # {campaign_name: str}
        existing_campaign_cpc = {}  # {campaign_name: str}
        existing_campaign_preset_cpc = {}  # {campaign_name: str}
        for row_data in existing_rows:
            name = row_data.get('广告系列名称', '')
            row_idx = row_data.get('row_index')
            status = str(row_data.get('状态', '') or '').strip()
            if status == SUMMARY_STATUS_TEXT:
                continue
            if name and row_idx:
                existing_campaign_names[name] = row_idx
                existing_link = row_data.get('投放链接', '')
                if isinstance(existing_link, list) and len(existing_link) > 0 and isinstance(existing_link[0], dict):
                    existing_link = existing_link[0].get('link', '') or existing_link[0].get('text', '')
                if existing_link:
                    existing_campaign_links[name] = existing_link
                # 记录上次的花费和佣金（用于计算新增量）
                existing_campaign_cost[name] = self.parse_commission_value(row_data.get('广告系列总花费', ''))
                existing_campaign_commission[name] = self.parse_commission_value(row_data.get('总佣金', ''))
                # 记录上次的状态（用于检测状态变更）
                existing_campaign_status[name] = str(row_data.get('状态', '') or '').strip()
                existing_campaign_clicks[name] = str(row_data.get('总点击数', '') or '').strip()
                existing_campaign_cpc[name] = str(row_data.get('CPC', '') or '').strip()
                existing_campaign_preset_cpc[name] = str(row_data.get('预设CPC', '') or '').strip()

        existing_campaign_row_info = {}
        for row_data in existing_rows:
            name = row_data.get('广告系列名称', '')
            row_idx = row_data.get('row_index')
            status = str(row_data.get('状态', '') or '').strip()
            if status == SUMMARY_STATUS_TEXT:
                continue
            if name and row_idx:
                existing_campaign_row_info[name] = {
                    'row_index': row_idx,
                    'campaign_name': name,
                    'status': status,
                    'tracking_link': row_data.get('投放链接', ''),
                    'product_link': row_data.get('产品链接', ''),
                }

        def extract_tracking_uid(link):
            link = normalize_link_value(link)
            if not link:
                return ''
            uid = self.extract_tracking_uid_from_link(link)
            if uid:
                return uid
            return link[-7:] if len(link) >= 7 else link

        current_campaign_id_to_name = {}
        for campaign_name, campaign_info in mcc_campaigns.items():
            campaign_id = str(campaign_info.get('campaign_id', '') or '').strip()
            if campaign_id:
                current_campaign_id_to_name[campaign_id] = campaign_name
        
        # 调试日志
        self.log_debug(f"广告系列表格列映射: {column_map}")
        self.log_debug(f"现有表格中有 {len(existing_campaign_names)} 个广告系列名称")
        for name in list(mcc_campaigns.keys()):
            asin = mcc_campaigns[name].get('asin', 'None')
            country = mcc_campaigns[name].get('country', 'None')
            in_existing = "已存在" if name in existing_campaign_names else "新增"
            self.log_debug(f"  MCC广告系列: {name} | ASIN={asin} | 国家={country} | [{in_existing}]")
        
        # 步骤5：准备更新数据
        self.log_manage("  准备更新数据...")
        updates = []  # 更新现有行
        new_rows = []  # 新增行
        preserved_campaign_cost_rows = []

        increment_start_date_str = None
        increment_end_date_str = None
        increment_campaign_cost_by_name = {}
        increment_pb_commission_by_campaign = {}
        increment_pb_commission_by_offer_group = {}
        pb_commission_by_offer_group = {}
        increment_yp_commission_by_group = {}
        if increment_days and end_date_str:
            try:
                increment_end = datetime.strptime(end_date_str, '%Y-%m-%d')
                increment_start = increment_end - timedelta(days=int(increment_days) - 1)
                increment_start_date_str = increment_start.strftime('%Y-%m-%d')
                increment_end_date_str = increment_end.strftime('%Y-%m-%d')
            except Exception as e:
                self.log_manage(f"    新增窗口日期计算失败，将回退为空新增值: {str(e)}")

        # 基于uid构建广告系列佣金映射
        campaign_commission_totals = {}  # {campaign_name: float}
        campaign_commission_asins = {}   # {campaign_name: set()}
        unmatched_uid_commission_total = 0.0
        unmatched_uid_commission_count = 0

        uid_to_unique_campaign = {}
        ambiguous_uid_to_campaigns = {}
        uid_conflict_details = {}

        def remember_uid_campaign(uid, campaign_name, status=''):
            if not uid or not campaign_name:
                return
            ambiguous_uid_to_campaigns.setdefault(uid, set()).add(campaign_name)
            uid_conflict_details.setdefault(uid, {})
            if campaign_name not in uid_conflict_details[uid]:
                row_info = existing_campaign_row_info.get(campaign_name, {})
                uid_conflict_details[uid][campaign_name] = {
                    'row_index': row_info.get('row_index', existing_campaign_names.get(campaign_name, '')),
                    'campaign_name': campaign_name,
                    'status': status or row_info.get('status', existing_campaign_status.get(campaign_name, '')),
                }

        for campaign_id, link in campaign_id_to_tracking_link.items():
            uid = extract_tracking_uid(link)
            campaign_name = current_campaign_id_to_name.get(str(campaign_id).strip())
            if uid and campaign_name and campaign_name in mcc_campaigns:
                remember_uid_campaign(uid, campaign_name, mcc_campaigns.get(campaign_name, {}).get('status', ''))

        for campaign_name, link in existing_campaign_links.items():
            uid = extract_tracking_uid(link)
            if uid and campaign_name:
                remember_uid_campaign(uid, campaign_name, existing_campaign_status.get(campaign_name, ''))

        for uid, campaign_names in ambiguous_uid_to_campaigns.items():
            if len(campaign_names) == 1:
                uid_to_unique_campaign[uid] = next(iter(campaign_names))

        for trans in commission_data:
            if trans.get('status') == 'Rejected':
                continue
            uid = str(trans.get('uid', '') or '').strip()
            if not uid:
                continue

            campaign_name = uid_to_unique_campaign.get(uid)
            comm = float(trans.get('sale_comm', 0) or 0)
            asin = str(trans.get('prod_id', '') or '').strip()
            country = normalize_country_value(
                trans.get('customer_country', '') or trans.get('geo', '') or trans.get('country', '')
            )
            if not country and trans.get('merchant_name', ''):
                country = normalize_country_value(self.extract_country_from_merchant_name(trans.get('merchant_name', '')))
            if not country and trans.get('mcid', ''):
                country = normalize_country_value(self.extract_country_from_mcid(trans.get('mcid', '')))
            if not country:
                country = 'US'

            if campaign_name:
                campaign_commission_totals[campaign_name] = campaign_commission_totals.get(campaign_name, 0.0) + comm
                if campaign_name not in campaign_commission_asins:
                    campaign_commission_asins[campaign_name] = set()
                if asin:
                    campaign_commission_asins[campaign_name].add(f"{asin}_{country}")
            else:
                unmatched_uid_commission_total += comm
                unmatched_uid_commission_count += 1
                related_campaigns = []
                for related in sorted(uid_conflict_details.get(uid, {}).values(), key=lambda item: (item.get('row_index') or 10**9, item.get('campaign_name', ''))):
                    related_campaigns.append({
                        'row_index': related.get('row_index', ''),
                        'campaign_name': related.get('campaign_name', ''),
                        'status': related.get('status', ''),
                    })
                if related_campaigns:
                    report['uid_conflicts'].append({
                        'uid': uid,
                        'commission': comm,
                        'asin': asin,
                        'country': country,
                        'campaigns': related_campaigns,
                    })
                else:
                    report['unmatched_uid_transactions'].append({
                        'uid': uid,
                        'commission': comm,
                        'asin': asin,
                        'country': country,
                        'reason': '广告系列表中未找到该UID',
                    })

        self.log_manage(
            f"  基于uid构建广告系列佣金: {len(campaign_commission_totals)} 个广告系列, "
            f"合计=${sum(campaign_commission_totals.values()):.2f}"
        )
        if unmatched_uid_commission_total > 0:
            self.log_manage(
                f"    仍无法归到唯一广告系列的uid佣金: {unmatched_uid_commission_count} 条, "
                f"金额=${unmatched_uid_commission_total:.2f}"
            )
        
        # 处理MCC中的所有广告系列
        for campaign_name, campaign_info in mcc_campaigns.items():
            # 通过产品链接精确获取投放链接；每单佣金仍按(ASIN+国家)获取
            asin = campaign_info.get('asin', '')
            country = campaign_info.get('country', '')
            tracking_link = ''
            commission_per_order = ''
            
            total_commission = round(campaign_commission_totals.get(campaign_name, 0.0), 2)
            commission_asins = campaign_commission_asins.get(campaign_name, set())
            if total_commission > 0:
                self.log_debug(
                    f"  广告系列佣金按UID归因: {campaign_name} | 佣金=${total_commission:.2f} | "
                    f"ASINs={sorted(commission_asins)}"
                )

            # 计算ROI
            cost = campaign_info['cost_usd']
            
            # 确定状态
            status_raw = campaign_info['status']
            prev_status = existing_campaign_status.get(campaign_name, '')
            if status_raw == 'ENABLED':
                status = '投放中'
                status_color = 'green'
            elif status_raw == 'PAUSED':
                # 检查是否是从其他状态变为暂停
                if prev_status.startswith('广告系列暂停中'):
                    # 已经是暂停状态，保留原来的日期
                    status = prev_status
                else:
                    # 从其他状态（投放中、新增等）变为暂停，附加变更日期（如 2026-2-15）
                    now = datetime.now()
                    today_str = f"{now.year}-{now.month}-{now.day}"
                    status = f'广告系列暂停中{today_str}'
                status_color = 'orange'
            else:
                status = '投放中'  # 默认
                status_color = 'green'
            
            # 格式化佣金ASIN列表
            asins_str = ', '.join(sorted(commission_asins)) if commission_asins else ''
            
            # 构建产品链接 = Amazon最终到达网址 + ? + 最终到达网址后缀
            product_link_url = ''
            final_urls = campaign_info.get('final_urls', [])
            suffix = campaign_info.get('final_url_suffix', '')
            if final_urls and suffix:
                base_url = final_urls[0]
                separator = '&' if '?' in base_url else '?'
                product_link_url = f"{base_url}{separator}{suffix}"

            product_link_key = self.normalize_product_link_key(product_link_url)
            if product_link_key:
                tracking_link = product_link_to_tracking_link.get(product_link_key, '')

            if asin and country:
                key = (asin, self.normalize_country_code(country))
                commission_per_order_value = asin_country_to_commission_per_order.get(key)
                if commission_per_order_value is not None:
                    if isinstance(commission_per_order_value, (int, float)):
                        commission_per_order = f"${commission_per_order_value:.2f}"
                    else:
                        commission_per_order = str(commission_per_order_value)
            
            # 计算新增花费；新增佣金稍后统一按“本次总佣金 - 运行前总佣金”覆盖。
            prev_cost = existing_campaign_cost.get(campaign_name, 0)
            preserve_existing_cost = cost <= 0.0001 and prev_cost > 0.0001
            cost_increment = cost - prev_cost
            roi_cost = prev_cost if preserve_existing_cost else cost
            roi = round(total_commission / roi_cost, 1) if roi_cost > 0 else 0
            keyword_ctr_weighted_numerator = campaign_info.get('enabled_keyword_ctr_weighted_numerator', 0.0)
            keyword_ctr_impressions = campaign_info.get('enabled_keyword_ctr_impressions', 0)
            enabled_keyword_ctr = (
                keyword_ctr_weighted_numerator / keyword_ctr_impressions * 100
                if keyword_ctr_impressions > 0 else 0.0
            )
            total_clicks = int(campaign_info.get('clicks', 0) or 0)
            avg_cpc_usd = float(campaign_info.get('avg_cpc_usd', 0.0) or 0.0)
            preset_cpc_usd = float(campaign_info.get('preset_cpc_usd', 0.0) or 0.0)
            if preserve_existing_cost:
                preserved_campaign_cost_rows.append({
                    'campaign_name': campaign_name,
                    'previous_cost': prev_cost,
                    'row_index': existing_campaign_names.get(campaign_name)
                })
            
            # 未拿到产品链接映射时，优先保留广告系列表里已有的投放链接，避免空值覆盖
            tracking_link_to_write = tracking_link or existing_campaign_links.get(campaign_name) or None
            if not tracking_link and tracking_link_to_write:
                self.log_debug(
                    f"  产品链接未命中Offer映射，保留已有投放链接: {campaign_name}"
                )

            offer_group_key, offer_group_summary = self.resolve_yp_campaign_offer_group(
                campaign_name,
                {
                    **campaign_info,
                    'product_link': product_link_url,
                },
                yp_offer_context,
                tracking_link_to_write
            )
            if offer_group_key:
                campaign_info['_offer_group_key'] = offer_group_key
                campaign_info['_offer_group_summary'] = offer_group_summary
                campaign_info['_traffic_source'] = 'yp' if offer_group_summary.get('is_yp') else 'pb'

            # 准备更新数据
            update_data = {
                'campaign_name': campaign_name,
                '状态': status,
                '广告系列名称': campaign_name,
                '投放中的ads': campaign_info['account_id'].replace('-', ''),
                '投放链接': tracking_link_to_write,
                '品牌名': self.extract_brand_and_country_from_campaign_name(campaign_name)[0] or '',
                '品牌名称': self.extract_brand_and_country_from_campaign_name(campaign_name)[0] or '',
                '国家代码': self.normalize_country_code(
                    campaign_info.get('country', '') or self.extract_brand_and_country_from_campaign_name(campaign_name)[1] or ''
                ),
                '广告系列总花费': None if preserve_existing_cost else f"${cost:.2f}",
                '总佣金': f"${total_commission:.2f}",
                'ROI': f"{roi}",
                '佣金ASIN': asins_str,
                '每单佣金': commission_per_order,
                '产品链接': product_link_url,
                '已启用关键字CTR': f"{enabled_keyword_ctr:.2f}%",
                '总点击数': total_clicks,
                'CPC': f"${avg_cpc_usd:.2f}" if total_clicks > 0 else "$0.00",
                '预设CPC': f"${preset_cpc_usd:.2f}" if preset_cpc_usd > 0 else "$0.00",
                '新增广告系列花费': None if preserve_existing_cost else f"${cost_increment:.2f}",
                '新增佣金': '',
                'status_color': status_color,
                '_offer_group_key': offer_group_key,
                '_offer_group_summary': offer_group_summary,
            }
            
            if campaign_name in existing_campaign_names:
                # 更新现有行
                update_data['row_index'] = existing_campaign_names[campaign_name]
                updates.append(update_data)
            else:
                # 新增行（上次不存在，新增量就是本次的全部值）
                new_rows.append(update_data)
        
        if preserved_campaign_cost_rows:
            self.log_manage(
                f"    花费保护: 广告系列表保留 {len(preserved_campaign_cost_rows)} 行原花费，避免本次$0.00覆盖"
            )
            for item in preserved_campaign_cost_rows[:3]:
                self.log_manage(
                    f"      - row={item['row_index']}, 广告系列={item['campaign_name']}, "
                    f"保留原花费=${item['previous_cost']:.2f}"
                )
            if len(preserved_campaign_cost_rows) > 3:
                self.log_manage(f"      ... 还有 {len(preserved_campaign_cost_rows) - 3} 行未显示")

        removed_campaign_metrics = {}
        removed_campaign_sources = {}
        if start_date_str and end_date_str:
            removed_candidates = sorted(
                name for name in existing_campaign_names.keys()
                if name and name not in mcc_campaigns
            )
            if removed_candidates:
                for campaign_name in removed_candidates:
                    existing_tracking_link = existing_campaign_links.get(campaign_name, '')
                    _, offer_summary = self.resolve_yp_campaign_offer_group(
                        campaign_name,
                        {
                            'campaign_name': campaign_name,
                            'asin': '',
                            'country': '',
                            'campaign_id': '',
                            'tracking_link': existing_tracking_link,
                            'product_link': existing_campaign_row_info.get(campaign_name, {}).get('product_link', ''),
                        },
                        yp_offer_context,
                        existing_tracking_link
                    )
                    removed_campaign_sources[campaign_name] = 'yp' if offer_summary.get('is_yp') else 'pb'
                removed_campaign_metrics = self.get_removed_campaign_metrics(
                    start_date_str=start_date_str,
                    end_date_str=end_date_str,
                    campaign_names=removed_candidates
                )

        # 处理已删除的广告系列（在表格中存在但MCC中不存在）
        for campaign_name, row_idx in existing_campaign_names.items():
            if campaign_name not in mcc_campaigns:
                ended_commission = round(campaign_commission_totals.get(campaign_name, 0.0), 2)
                ended_asins = ', '.join(sorted(campaign_commission_asins.get(campaign_name, set())))
                existing_tracking_link = existing_campaign_links.get(campaign_name, '')
                ended_offer_group_key, ended_offer_group_summary = self.resolve_yp_campaign_offer_group(
                    campaign_name,
                    {
                        'campaign_name': campaign_name,
                        'asin': '',
                        'country': '',
                        'campaign_id': '',
                        'tracking_link': existing_tracking_link,
                        'product_link': existing_campaign_row_info.get(campaign_name, {}).get('product_link', ''),
                    },
                    yp_offer_context,
                    existing_tracking_link
                )
                ended_update = {
                    'campaign_name': campaign_name,
                    'row_index': row_idx,
                    '状态': '投放已结束',
                    '总佣金': f"${ended_commission:.2f}",
                    '佣金ASIN': ended_asins,
                    '已启用关键字CTR': '0.00%',
                    '新增佣金': '',
                    'status_color': 'black',
                    '_offer_group_key': ended_offer_group_key,
                    '_offer_group_summary': ended_offer_group_summary,
                }
                removed_metrics = removed_campaign_metrics.get(campaign_name, {})
                if not existing_campaign_clicks.get(campaign_name, '') and ('clicks' in removed_metrics):
                    ended_update['总点击数'] = int(removed_metrics.get('clicks', 0) or 0)
                if not existing_campaign_cpc.get(campaign_name, '') and ('avg_cpc_usd' in removed_metrics):
                    clicks = int(removed_metrics.get('clicks', 0) or 0)
                    ended_update['CPC'] = f"${float(removed_metrics.get('avg_cpc_usd', 0.0) or 0.0):.2f}" if clicks > 0 else "$0.00"
                if not existing_campaign_preset_cpc.get(campaign_name, '') and ('preset_cpc_usd' in removed_metrics):
                    ended_update['预设CPC'] = f"${float(removed_metrics.get('preset_cpc_usd', 0.0) or 0.0):.2f}" if float(removed_metrics.get('preset_cpc_usd', 0.0) or 0.0) > 0 else "$0.00"
                updates.append(ended_update)
        
        final_campaign_rows = {}
        for row_data in existing_rows:
            status = str(row_data.get('状态', '') or '').strip()
            campaign_name = str(row_data.get('广告系列名称', '') or '').strip()
            row_idx = row_data.get('row_index')
            if status == SUMMARY_STATUS_TEXT or not campaign_name or not row_idx:
                continue
            final_campaign_rows[campaign_name] = {
                'campaign_name': campaign_name,
                'row_index': row_idx,
                'status': status,
                'total_commission': self.parse_commission_value(row_data.get('总佣金', '')),
                'total_clicks': int(self.parse_commission_value(row_data.get('总点击数', '')) or 0),
            }

        for update in updates:
            campaign_name = str(update.get('campaign_name', '') or '').strip()
            row_idx = update.get('row_index')
            if not campaign_name or not row_idx:
                continue
            final_campaign_rows[campaign_name] = {
                'campaign_name': campaign_name,
                'row_index': row_idx,
                'status': str(update.get('状态', '') or final_campaign_rows.get(campaign_name, {}).get('status', '')).strip(),
                'total_commission': self.parse_commission_value(
                    update.get('总佣金', final_campaign_rows.get(campaign_name, {}).get('total_commission', 0.0))
                ),
                'total_clicks': int(
                    self.parse_commission_value(
                        update.get('总点击数', final_campaign_rows.get(campaign_name, {}).get('total_clicks', 0))
                    ) or 0
                ),
            }

        for new_row in new_rows:
            campaign_name = str(new_row.get('campaign_name', '') or '').strip()
            if not campaign_name:
                continue
            synthetic_row_index = f"new:{campaign_name}"
            final_campaign_rows[campaign_name] = {
                'campaign_name': campaign_name,
                'row_index': synthetic_row_index,
                'status': str(new_row.get('状态', '') or '').strip(),
                'total_commission': self.parse_commission_value(new_row.get('总佣金', '')),
                'total_clicks': int(self.parse_commission_value(new_row.get('总点击数', '')) or 0),
            }

        for row_info in offer_row_commissions.values():
            commission = self.get_pb_offer_row_campaign_commission(row_info)
            if commission <= 0:
                continue
            campaign_names = []
            for campaign_name in row_info.get('campaign_names', []) or []:
                campaign_name = str(campaign_name or '').strip()
                if campaign_name and campaign_name not in campaign_names:
                    campaign_names.append(campaign_name)
            if len(campaign_names) <= 1:
                continue
            offer_key = (
                str(row_info.get('asin', '') or '').strip().upper(),
                str(row_info.get('brand_id', '') or '').strip(),
                self.normalize_country_code(row_info.get('country', '') or '')
            )
            if offer_key[0] and offer_key[2]:
                pb_commission_by_offer_group[offer_key] = (
                    pb_commission_by_offer_group.get(offer_key, 0.0) + commission
                )

        if increment_start_date_str and increment_end_date_str:
            try:
                increment_pb_data = self.get_all_commissions(increment_start_date_str, increment_end_date_str)
                _, increment_offer_context = self.calculate_updates(
                    feishu_data=feishu_data,
                    asin_country_campaigns={},
                    asin_country_commission={},
                    asin_country_uid_commission={},
                    asin_country_no_uid_commission=self.build_pb_no_uid_commission_map(increment_pb_data),
                    row_campaigns={},
                    yp_asin_brand_commission={}
                )
                for row_info in (increment_offer_context.get('row_commissions', {}) or {}).values():
                    commission = self.get_pb_offer_row_campaign_commission(row_info)
                    if commission <= 0:
                        continue

                    campaign_names = []
                    for campaign_name in row_info.get('campaign_names', []) or []:
                        campaign_name = str(campaign_name or '').strip()
                        if campaign_name and campaign_name not in campaign_names:
                            campaign_names.append(campaign_name)
                    if len(campaign_names) == 1:
                        campaign_name = campaign_names[0]
                        increment_pb_commission_by_campaign[campaign_name] = (
                            increment_pb_commission_by_campaign.get(campaign_name, 0.0) + commission
                        )

                    brand_id = str(row_info.get('brand_id', '') or '').strip()
                    offer_key = (
                        str(row_info.get('asin', '') or '').strip().upper(),
                        brand_id,
                        self.normalize_country_code(row_info.get('country', '') or '')
                    )
                    if offer_key[0] and offer_key[2] and len(campaign_names) > 1:
                        increment_pb_commission_by_offer_group[offer_key] = (
                            increment_pb_commission_by_offer_group.get(offer_key, 0.0) + commission
                        )
            except Exception as e:
                self.log_manage(f"    PB新增佣金归因失败，将回退为空新增值: {str(e)}")

            try:
                increment_yp_data = self.get_all_yp_commissions(increment_start_date_str, increment_end_date_str)
                increment_yp_commission_by_group = self.calculate_yp_increment_commission_by_group(
                    feishu_data,
                    increment_yp_data
                )
            except Exception as e:
                self.log_manage(f"    YP新增佣金归因失败，将回退为空新增值: {str(e)}")

        yp_row_apply_result = self.apply_yp_group_commissions_to_campaign_rows(
            updates=updates,
            new_rows=new_rows,
            offer_group_summary_by_key=offer_group_summary_by_key,
            existing_campaign_commission=existing_campaign_commission,
            existing_campaign_cost=existing_campaign_cost,
            increment_yp_commission_by_group=increment_yp_commission_by_group,
        )
        if yp_row_apply_result.get('applied_rows', 0) > 0 or yp_row_apply_result.get('skipped_groups', 0) > 0:
            self.log_manage(
                f"    YP佣金补写到广告系列行: 成功 {yp_row_apply_result.get('applied_rows', 0)} 组, "
                f"多广告系列分组保留给统计行 {yp_row_apply_result.get('skipped_groups', 0)} 组"
            )

        pb_offer_apply_result = self.apply_pb_offer_commissions_to_campaign_rows(
            updates=updates,
            new_rows=new_rows,
            offer_row_commissions=offer_row_commissions,
            existing_campaign_cost=existing_campaign_cost,
            increment_pb_commission_by_campaign=increment_pb_commission_by_campaign,
        )
        if pb_offer_apply_result.get('applied_rows', 0) > 0 or pb_offer_apply_result.get('skipped_groups', 0) > 0:
            self.log_manage(
                f"    PB无UID佣金补写到广告系列行: 成功 {pb_offer_apply_result.get('applied_rows', 0)} 行, "
                f"多广告系列Offer保留给统计行 {pb_offer_apply_result.get('skipped_groups', 0)} 组"
            )

        grouped_offer_keys = {
            offer_key
            for offer_key, count in Counter(
                item.get('_offer_group_key')
                for item in (updates + new_rows)
                if item.get('_offer_group_key')
            ).items()
            if count > 1
        }
        self.apply_campaign_increment_commission_delta(
            updates + new_rows,
            campaign_commission_baseline,
            grouped_offer_keys=grouped_offer_keys
        )
        self.log_manage("    广告系列新增佣金口径: 本次总佣金 - 运行前总佣金")
        if grouped_offer_keys:
            self.log_manage(f"    多广告系列分组子行新增佣金已留空，交由 {len(grouped_offer_keys)} 个统计行汇总")

        brand_country_totals, brand_country_offer_index = self.calculate_break_even_brand_country_commissions(
            feishu_data=feishu_data,
            commission_data=commission_data,
            yp_commission_data=yp_commission_data
        )
        brand_country_clicks = {}
        if start_date_str and end_date_str:
            brand_country_clicks = self.get_brand_country_clicks(
                start_date_str=start_date_str,
                end_date_str=end_date_str,
                offer_index=brand_country_offer_index
            )
            for key, click_info in brand_country_clicks.items():
                aggregate = brand_country_totals.setdefault(key, {
                    'brand': click_info.get('brand', ''),
                    'country': click_info.get('country', ''),
                    'commission': 0.0,
                    'clicks': 0,
                })
                aggregate['clicks'] = int(click_info.get('clicks', 0) or 0)
        if start_date_str and end_date_str:
            self.log_manage(
                f"    品牌收支平衡CPC口径: 仅统计 {start_date_str} 至 {end_date_str} 的佣金与点击"
            )

        break_even_updates = self.apply_brand_break_even_cpc(
            updates=updates,
            new_rows=new_rows,
            mcc_campaigns=mcc_campaigns,
            brand_country_totals=brand_country_totals
        )

        if break_even_updates > 0:
            self.log_manage(f"    品牌收支平衡CPC已更新: {break_even_updates} 行")
        else:
            self.log_manage("    品牌收支平衡CPC: 本次没有可更新的投放中/暂停中广告系列")

        self.log_manage(f"    需要更新 {len(updates)} 行，新增 {len(new_rows)} 行")
        if increment_days and end_date_str:
            try:
                summary_candidate_names = set()
                for item in updates + new_rows:
                    name = str(item.get('campaign_name', '') or '').strip()
                    if name:
                        summary_candidate_names.add(name)
                if summary_candidate_names:
                    increment_campaign_cost_by_name = self.get_campaign_increment_costs_by_name(
                        increment_start_date_str,
                        increment_end_date_str,
                        summary_candidate_names
                    )
                self.log_manage(
                    f"    统计行新增口径: {increment_start_date_str} 至 {increment_end_date_str}, "
                    f"广告系列花费命中 {len(increment_campaign_cost_by_name)} 个, "
                    f"PB广告系列佣金 {len(increment_pb_commission_by_campaign)} 个, "
                    f"YP分组佣金 {len(increment_yp_commission_by_group)} 个"
                )
            except Exception as e:
                self.log_manage(f"    统计行新增窗口计算失败，回退为空新增值: {str(e)}")
        
        # 步骤6：执行更新
        if updates or new_rows:
            self.apply_campaigns_sheet_updates(
                feishu_token,
                campaigns_spreadsheet_token,
                campaigns_sheet_id,
                updates,
                new_rows,
                column_map,
                first_empty_row
            )
        else:
            self.log_manage("  没有需要更新的内容")

        try:
            self.consolidate_campaign_sheet_same_offer_rows(
                feishu_token,
                campaigns_spreadsheet_token,
                campaigns_sheet_id,
                mcc_campaigns,
                yp_offer_context,
                increment_start_date_str=increment_start_date_str,
                increment_end_date_str=increment_end_date_str,
                feishu_data=feishu_data,
                current_campaign_rows=updates + new_rows,
                campaign_commission_baseline=campaign_commission_baseline,
                increment_yp_commission_by_group=increment_yp_commission_by_group,
                increment_campaign_cost_by_name=increment_campaign_cost_by_name,
                increment_pb_commission_by_campaign=increment_pb_commission_by_campaign,
                pb_commission_by_offer_group=pb_commission_by_offer_group,
                increment_pb_commission_by_offer_group=increment_pb_commission_by_offer_group
            )
        except Exception as e:
            self.log_manage(f"  整理广告系列相同offer统计行时出错: {str(e)}")
        
        return report

    def backfill_ended_campaign_metrics_once(self, start_date_str, end_date_str):
        """一次性回填广告系列表中“投放已结束”行的总点击数、CPC和预设CPC。"""
        campaigns_spreadsheet_token = "KnJ1wphpBiVMrGkWl5ncUkMGnfe"
        campaigns_sheet_id = CAMPAIGNS_SHEET_ID

        self.log_manage("=" * 50)
        self.log_manage("开始一次性回填“投放已结束”广告系列点击数/CPC...")
        self.log_manage(f"统计日期范围: {start_date_str} 至 {end_date_str}")
        self.log_manage("=" * 50)

        feishu_token = self.get_feishu_token()
        if not feishu_token:
            self.log_manage("  ✗ 获取飞书Token失败")
            return {'updated_rows': 0, 'candidate_rows': 0, 'matched_metrics': 0}

        self.log_manage("  读取现有广告系列表格...")
        existing_campaigns_data = self.read_campaigns_sheet(feishu_token, campaigns_spreadsheet_token, campaigns_sheet_id)
        if existing_campaigns_data is None:
            self.log_manage("  ✗ 无法读取广告系列表格")
            return {'updated_rows': 0, 'candidate_rows': 0, 'matched_metrics': 0}

        existing_rows, column_map, first_empty_row = existing_campaigns_data

        ended_candidates = []
        ended_row_index_by_name = {}
        existing_clicks = {}
        existing_cpc = {}
        existing_preset_cpc = {}
        for row_data in existing_rows:
            status = str(row_data.get('状态', '') or '').strip()
            if status != '投放已结束':
                continue

            campaign_name = str(row_data.get('广告系列名称', '') or '').strip()
            row_idx = row_data.get('row_index')
            if not campaign_name or not row_idx:
                continue

            clicks_value = str(row_data.get('总点击数', '') or '').strip()
            cpc_value = str(row_data.get('CPC', '') or '').strip()
            preset_cpc_value = str(row_data.get('预设CPC', '') or '').strip()
            needs_clicks_backfill = not clicks_value
            needs_cpc_backfill = not cpc_value
            needs_preset_cpc_backfill = not preset_cpc_value
            if not (needs_clicks_backfill or needs_cpc_backfill or needs_preset_cpc_backfill):
                continue

            ended_candidates.append(campaign_name)
            ended_row_index_by_name[campaign_name] = row_idx
            existing_clicks[campaign_name] = clicks_value
            existing_cpc[campaign_name] = cpc_value
            existing_preset_cpc[campaign_name] = preset_cpc_value

        if not ended_candidates:
            self.log_manage("  没有需要补录点击数/CPC的“投放已结束”广告系列")
            return {'updated_rows': 0, 'candidate_rows': 0, 'matched_metrics': 0}

        self.log_manage(f"  待补录候选: {len(ended_candidates)} 个广告系列")
        removed_campaign_metrics = self.get_removed_campaign_metrics(
            start_date_str,
            end_date_str,
            ended_candidates
        )

        updates = []
        matched_metrics = 0
        for campaign_name in ended_candidates:
            removed_metrics = removed_campaign_metrics.get(campaign_name)
            if not removed_metrics:
                continue

            update = {
                'campaign_name': campaign_name,
                'row_index': ended_row_index_by_name[campaign_name],
            }
            wrote_any = False

            if not existing_clicks.get(campaign_name, '') and ('clicks' in removed_metrics):
                update['总点击数'] = int(removed_metrics.get('clicks', 0) or 0)
                wrote_any = True

            if not existing_cpc.get(campaign_name, '') and ('avg_cpc_usd' in removed_metrics):
                clicks = int(removed_metrics.get('clicks', 0) or 0)
                update['CPC'] = f"${float(removed_metrics.get('avg_cpc_usd', 0.0) or 0.0):.2f}" if clicks > 0 else "$0.00"
                wrote_any = True

            if not existing_preset_cpc.get(campaign_name, '') and ('preset_cpc_usd' in removed_metrics):
                preset_cpc_usd = float(removed_metrics.get('preset_cpc_usd', 0.0) or 0.0)
                update['预设CPC'] = f"${preset_cpc_usd:.2f}" if preset_cpc_usd > 0 else "$0.00"
                wrote_any = True

            if wrote_any:
                updates.append(update)
                matched_metrics += 1

        if not updates:
            self.log_manage("  Google Ads中未命中需要补录的已移除广告系列")
            return {'updated_rows': 0, 'candidate_rows': len(ended_candidates), 'matched_metrics': 0}

        self.log_manage(f"  准备写回 {len(updates)} 行“投放已结束”广告系列")
        self.apply_campaigns_sheet_updates(
            feishu_token,
            campaigns_spreadsheet_token,
            campaigns_sheet_id,
            updates,
            [],
            column_map,
            first_empty_row
        )
        self.log_manage("  ✅ 一次性回填完成")
        return {
            'updated_rows': len(updates),
            'candidate_rows': len(ended_candidates),
            'matched_metrics': matched_metrics
        }

    def backfill_ended_campaign_break_even_cpc_once(self, start_date_str, end_date_str):
        """一次性回填广告系列表中“投放已结束”且Google Ads中能命中的广告系列的品牌收支平衡CPC。"""
        campaigns_spreadsheet_token = "KnJ1wphpBiVMrGkWl5ncUkMGnfe"
        campaigns_sheet_id = CAMPAIGNS_SHEET_ID

        self.log_manage("=" * 50)
        self.log_manage("开始一次性回填“投放已结束”广告系列品牌收支平衡CPC...")
        self.log_manage(f"统计日期范围: {start_date_str} 至 {end_date_str}")
        self.log_manage("=" * 50)

        feishu_token = self.get_feishu_token()
        if not feishu_token:
            self.log_manage("  ✗ 获取飞书Token失败")
            return {'updated_rows': 0, 'candidate_rows': 0, 'matched_metrics': 0, 'break_even_updates': 0}

        self.log_manage("  读取Offer表格...")
        offer_rows = self.get_feishu_sheet_data(feishu_token)
        if not offer_rows:
            self.log_manage("  ✗ 无法读取Offer表格")
            return {'updated_rows': 0, 'candidate_rows': 0, 'matched_metrics': 0, 'break_even_updates': 0}

        self.log_manage("  读取广告系列表格...")
        existing_campaigns_data = self.read_campaigns_sheet(feishu_token, campaigns_spreadsheet_token, campaigns_sheet_id)
        if existing_campaigns_data is None:
            self.log_manage("  ✗ 无法读取广告系列表格")
            return {'updated_rows': 0, 'candidate_rows': 0, 'matched_metrics': 0, 'break_even_updates': 0}

        existing_rows, column_map, first_empty_row = existing_campaigns_data

        ended_candidates = []
        updates = []
        for row_data in existing_rows:
            status = str(row_data.get('状态', '') or '').strip()
            if status != '投放已结束':
                continue

            campaign_name = str(row_data.get('广告系列名称', '') or '').strip()
            row_idx = row_data.get('row_index')
            if not campaign_name or not row_idx:
                continue

            ended_candidates.append(campaign_name)
            updates.append({
                'campaign_name': campaign_name,
                'row_index': row_idx,
                '状态': status,
            })

        if not ended_candidates:
            self.log_manage("  没有“投放已结束”广告系列需要回填品牌收支平衡CPC")
            return {'updated_rows': 0, 'candidate_rows': 0, 'matched_metrics': 0, 'break_even_updates': 0}

        self.log_manage(f"  待处理候选: {len(ended_candidates)} 个广告系列")

        removed_campaign_metrics = self.get_removed_campaign_metrics(
            start_date_str=start_date_str,
            end_date_str=end_date_str,
            campaign_names=ended_candidates
        )
        matched_campaign_names = {
            str(name).strip()
            for name, metrics in (removed_campaign_metrics or {}).items()
            if metrics is not None
        }

        applicable_updates = [item for item in updates if str(item.get('campaign_name', '') or '').strip() in matched_campaign_names]
        if not applicable_updates:
            self.log_manage("  Google Ads中未命中任何“投放已结束”广告系列，无法回填品牌收支平衡CPC")
            return {
                'updated_rows': 0,
                'candidate_rows': len(ended_candidates),
                'matched_metrics': len(matched_campaign_names),
                'break_even_updates': 0
            }

        commission_data = self.get_all_commissions(start_date_str, end_date_str)
        yp_commission_data = self.get_all_yp_commissions(start_date_str, end_date_str)
        brand_country_totals, brand_country_offer_index = self.calculate_break_even_brand_country_commissions(
            feishu_data=offer_rows,
            commission_data=commission_data,
            yp_commission_data=yp_commission_data
        )
        brand_country_clicks = self.get_brand_country_clicks(
            start_date_str=start_date_str,
            end_date_str=end_date_str,
            offer_index=brand_country_offer_index
        )
        for key, click_info in brand_country_clicks.items():
            aggregate = brand_country_totals.setdefault(key, {
                'brand': click_info.get('brand', ''),
                'country': click_info.get('country', ''),
                'commission': 0.0,
                'clicks': 0,
            })
            aggregate['clicks'] = int(click_info.get('clicks', 0) or 0)

        break_even_updates = self.apply_brand_break_even_cpc_to_ended_rows(
            updates=applicable_updates,
            brand_country_totals=brand_country_totals
        )
        if break_even_updates <= 0:
            self.log_manage("  本次没有可回填品牌收支平衡CPC的已结束广告系列")
            return {
                'updated_rows': 0,
                'candidate_rows': len(ended_candidates),
                'matched_metrics': len(matched_campaign_names),
                'break_even_updates': 0
            }

        self.log_manage(f"  准备写回 {break_even_updates} 行“投放已结束”广告系列的品牌收支平衡CPC")
        self.apply_campaigns_sheet_updates(
            feishu_token,
            campaigns_spreadsheet_token,
            campaigns_sheet_id,
            applicable_updates,
            [],
            column_map,
            first_empty_row
        )
        self.log_manage("  ✅ 已结束广告系列品牌收支平衡CPC回填完成")
        return {
            'updated_rows': break_even_updates,
            'candidate_rows': len(ended_candidates),
            'matched_metrics': len(matched_campaign_names),
            'break_even_updates': break_even_updates
        }
    
    def read_campaigns_sheet(self, token, spreadsheet_token, sheet_id):
        """读取广告系列表格数据
        
        Returns:
            (rows_data, column_map, first_empty_row) 或 None
        """
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        # 读取表格数据（扩大范围到AZ列，支持更多列和更多广告系列行）
        url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/values/{sheet_id}!A1:AZ5000"
        
        try:
            response = requests.get(url, headers=headers, timeout=30)
            data = response.json()
            
            if data.get('code') != 0:
                self.log_manage(f"    读取失败: {data.get('msg', 'Unknown error')}")
                return None
            
            values = data.get('data', {}).get('valueRange', {}).get('values', [])
            
            if not values:
                self.log_manage("    表格为空")
                return [], {}, 1
            
            # 解析表头（第一行）
            header_row = values[0] if values else []
            column_map = {}  # {列名: 列字母}
            
            for i, header in enumerate(header_row):
                if header:
                    col_letter = self.index_to_column_letter(i)
                    column_map[str(header).strip()] = col_letter
            
            # 解析数据行，同时找到第一个空行。
            # 飞书 values API 在读取大范围时可能返回尾部空白行；不能把这些空白行当成
            # 已有数据，否则新增广告系列会写到当前网格行数之外而不可见。
            rows_data = []
            first_empty_row = 2  # 默认从第2行开始（第1行是表头）
            campaign_name_col_idx = None
            
            # 找到"广告系列名称"列的索引
            for i, header in enumerate(header_row):
                if header and str(header).strip() == '广告系列名称':
                    campaign_name_col_idx = i
                    break
            
            for row_idx, row in enumerate(values[1:], start=2):  # 从第2行开始（跳过表头）
                row_data = {'row_index': row_idx}
                has_campaign_name = False
                has_value = False
                
                for i, cell in enumerate(row):
                    if i < len(header_row) and header_row[i]:
                        row_data[str(header_row[i]).strip()] = cell if cell else ''
                        if cell:
                            has_value = True
                    
                    # 检查是否有广告系列名称
                    if campaign_name_col_idx is not None and i == campaign_name_col_idx and cell:
                        has_campaign_name = True
                
                if has_value:
                    rows_data.append(row_data)
                
                # 更新第一个空行位置
                if has_campaign_name:
                    first_empty_row = row_idx + 1
            
            # 返回第一个空行位置，而不是总行数
            return rows_data, column_map, first_empty_row
            
        except Exception as e:
            self.log_manage(f"    读取异常: {str(e)}")
            return None

    def read_generic_sheet(self, token, spreadsheet_token, sheet_id, max_range="A1:AZ2000"):
        """读取任意飞书工作表，第一行作为表头。"""
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/values/{sheet_id}!{max_range}"

        try:
            response = requests.get(url, headers=headers, timeout=30)
            data = response.json()
            if data.get('code') != 0:
                self.log_manage(f"    读取工作表失败: {data.get('msg', 'Unknown error')}")
                return None

            values = data.get('data', {}).get('valueRange', {}).get('values', []) or []
            if not values:
                return [], {}, []

            header_row = values[0]
            column_map = {}
            for i, header in enumerate(header_row):
                if header:
                    column_map[str(header).strip()] = self.index_to_column_letter(i)

            rows_data = []
            for row_idx, row in enumerate(values[1:], start=2):
                row_data = {'row_index': row_idx}
                has_value = False
                for i, cell in enumerate(row):
                    if i < len(header_row) and header_row[i]:
                        value = cell if cell else ''
                        row_data[str(header_row[i]).strip()] = value
                        if value:
                            has_value = True
                if has_value:
                    rows_data.append(row_data)

            return rows_data, column_map, header_row
        except Exception as e:
            self.log_manage(f"    读取工作表异常: {str(e)}")
            return None

    def clear_sheet_data_rows(self, token, spreadsheet_token, sheet_id, column_count, max_rows=2000):
        """清空工作表表头以下的数据区域。"""
        if column_count <= 0:
            return True

        end_col = self.index_to_column_letter(column_count - 1)
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
        blank_row = ['' for _ in range(column_count)]
        value_ranges = []
        for row_idx in range(2, max_rows + 1):
            value_ranges.append({
                "range": f"{sheet_id}!A{row_idx}:{end_col}{row_idx}",
                "values": [blank_row]
            })

        try:
            batch_size = 100
            for i in range(0, len(value_ranges), batch_size):
                body = {"valueRanges": value_ranges[i:i + batch_size]}
                response = requests.post(url, headers=headers, json=body, timeout=30)
                data = response.json()
                if data.get('code') != 0:
                    self.log_manage(f"    清空工作表失败: {data.get('msg', 'Unknown error')}")
                    return False
            return True
        except Exception as e:
            self.log_manage(f"    清空工作表异常: {str(e)}")
            return False

    def overwrite_sheet_rows(self, token, spreadsheet_token, sheet_id, header_row, rows, max_clear_rows=2000):
        """保留表头，清空旧数据后从第2行覆盖写入。"""
        column_count = len(header_row or [])
        if column_count <= 0:
            self.log_manage("    目标工作表缺少表头，无法写入")
            return False

        self.clear_sheet_data_rows(token, spreadsheet_token, sheet_id, column_count, max_rows=max_clear_rows)
        if not rows:
            self.log_manage("    已清空旧数据，本次没有新数据")
            return True

        end_col = self.index_to_column_letter(column_count - 1)
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
        value_ranges = []
        for offset, row in enumerate(rows, start=2):
            value_ranges.append({
                "range": f"{sheet_id}!A{offset}:{end_col}{offset}",
                "values": [[row.get(str(header).strip(), '') if header else '' for header in header_row]]
            })

        batch_size = 100
        for i in range(0, len(value_ranges), batch_size):
            batch = value_ranges[i:i + batch_size]
            body = {"valueRanges": batch}
            response = requests.post(url, headers=headers, json=body, timeout=30)
            data = response.json()
            if data.get('code') != 0:
                self.log_manage(f"    写入工作表失败: {data.get('msg', 'Unknown error')}")
                return False
        return True

    def apply_sheet_background_styles(self, token, spreadsheet_token, sheet_id, style_updates):
        """批量更新单元格背景色。"""
        valid_updates = []
        for item in style_updates or []:
            row_index = item.get('row_index')
            col_letter = item.get('column')
            color = str(item.get('background_color', '') or '').strip()
            if not row_index or not col_letter or not re.match(r'^#[0-9A-Fa-f]{6}$', color):
                continue
            valid_updates.append({
                'ranges': f"{sheet_id}!{col_letter}{row_index}:{col_letter}{row_index}",
                'style': {'backColor': color.upper()}
            })

        if not valid_updates:
            return True

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/styles_batch_update"

        success = True
        batch_size = 100
        for i in range(0, len(valid_updates), batch_size):
            body = {"data": valid_updates[i:i + batch_size]}
            try:
                response = requests.put(url, headers=headers, json=body, timeout=30)
                data = response.json()
                if data.get('code') != 0:
                    success = False
                    self.log_manage(f"    背景色更新失败: {data.get('msg', 'Unknown error')}")
            except Exception as e:
                success = False
                self.log_manage(f"    背景色更新异常: {str(e)}")

        if success:
            self.log_manage(f"    背景色更新成功: {len(valid_updates)} 个单元格")
        return success

    def get_campaign_metric_backgrounds(self):
        """广告系列表关键指标列背景色。"""
        return {
            '预设CPC': '#F8F9FA',
            'CPC': '#FAF1D1',
            '品牌收支平衡CPC': '#FFF258',
            '广告系列总花费': '#FBBFBC',
            '总佣金': '#D9F5D6',
        }

    def build_campaign_metric_background_style_updates(self, rows, column_map):
        """为广告系列表数据行生成关键指标背景色更新。"""
        backgrounds = self.get_campaign_metric_backgrounds()
        style_updates = []
        for row in rows or []:
            row_index = row.get('row_index')
            if not row_index:
                continue
            status = str(row.get('状态', '') or '').strip()
            campaign_name = str(row.get('广告系列名称', '') or '').strip()
            if not campaign_name or campaign_name == SUMMARY_PLACEHOLDER_TEXT:
                continue
            for header, color in backgrounds.items():
                col_letter = (column_map or {}).get(header)
                if col_letter:
                    style_updates.append({
                        'row_index': row_index,
                        'column': col_letter,
                        'background_color': color,
                    })
        return style_updates
    
    def apply_campaigns_sheet_updates(self, token, spreadsheet_token, sheet_id, updates, new_rows, column_map, first_empty_row):
        """应用广告系列表格更新"""
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        # 定义字段到列的映射（完全基于表头文字匹配，不使用硬编码列位置）
        all_fields = ['状态', '广告系列名称', '投放中的ads', '投放链接', '品牌名', '品牌名称', '国家代码', '广告系列总花费', 
                      '总佣金', 'ROI', '佣金ASIN', '每单佣金', '产品链接', '已启用关键字CTR',
                      '总点击数', 'CPC', '预设CPC', '品牌收支平衡CPC',
                      '新增广告系列花费', '新增佣金']
        field_to_column = {}
        missing_fields = []
        for f in all_fields:
            if f in column_map:
                field_to_column[f] = column_map[f]
            else:
                missing_fields.append(f)
        if missing_fields:
            self.log_manage(f"    ⚠ 广告系列表中未找到以下列: {missing_fields}")
            self.log_manage(f"    表格现有列: {list(column_map.keys())}")
        self.log_debug(f"  广告系列表列映射: {field_to_column}")
        
        value_ranges = []
        style_updates = []
        new_row_backgrounds = self.get_campaign_metric_backgrounds()
        
        # 处理更新
        for update in updates:
            row_idx = update.get('row_index')
            if not row_idx:
                continue
            
            for field, col in field_to_column.items():
                if field in update and update[field] is not None:
                    if field == '投放链接' and str(update[field]).strip() == '':
                        continue
                    value_ranges.append({
                        'range': f"{sheet_id}!{col}{row_idx}:{col}{row_idx}",
                        'values': [[update[field]]]
                    })
            
            # 收集样式更新
            if 'status_color' in update:
                style_updates.append({
                    'row_index': row_idx,
                    'color': update['status_color'],
                    'column': field_to_column['状态']
                })
        
        # 处理新增行
        if new_rows:
            self.log_manage(f"    新增广告系列将从第 {first_empty_row} 行开始写入")
            for offset, new_row in enumerate(new_rows[:5]):
                self.log_manage(
                    f"      新增行预览 row={first_empty_row + offset}: "
                    f"{new_row.get('广告系列名称', new_row.get('campaign_name', ''))}"
                )
            if len(new_rows) > 5:
                self.log_manage(f"      ... 还有 {len(new_rows) - 5} 行新增广告系列")
            self._ensure_sheet_rows(
                token,
                spreadsheet_token,
                sheet_id,
                first_empty_row + len(new_rows) - 1
            )

        current_row = first_empty_row
        for new_row in new_rows:
            for field, col in field_to_column.items():
                if field in new_row and new_row[field] is not None:
                    if field == '投放链接' and str(new_row[field]).strip() == '':
                        continue
                    value_ranges.append({
                        'range': f"{sheet_id}!{col}{current_row}:{col}{current_row}",
                        'values': [[new_row[field]]]
                    })
            
            # 收集样式更新
            if 'status_color' in new_row:
                style_updates.append({
                    'row_index': current_row,
                    'color': new_row['status_color'],
                    'column': field_to_column['状态']
                })
            for field, bg_color in new_row_backgrounds.items():
                if field in field_to_column:
                    style_updates.append({
                        'row_index': current_row,
                        'background_color': bg_color,
                        'column': field_to_column[field]
                    })
            
            current_row += 1
        
        # 批量更新值
        if value_ranges:
            # 分批处理，每批最多100个range
            batch_size = 100
            for i in range(0, len(value_ranges), batch_size):
                batch = value_ranges[i:i+batch_size]
                url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
                body = {"valueRanges": batch}
                
                try:
                    response = requests.post(url, headers=headers, json=body, timeout=30)
                    data = response.json()
                    if data.get('code') == 0:
                        self.log_manage(f"    值更新成功: {len(batch)} 个单元格")
                    else:
                        self.log_manage(f"    值更新失败: {data.get('msg', 'Unknown error')}")
                except Exception as e:
                    self.log_manage(f"    值更新异常: {str(e)}")
        
        # 批量更新样式
        if style_updates:
            self.apply_campaigns_style_updates(token, spreadsheet_token, sheet_id, style_updates)
        
        self.log_manage(f"  ✅ 广告系列表格更新完成")

    def build_ads_brand_campaign_sheet_maps(self, campaign_sheet_rows):
        """从广告系列表构建广告系列名称到投放链接、品牌收支平衡CPC的映射。"""
        maps = {
            'tracking_link_by_name': {},
            'break_even_cpc_by_name': {},
            'platforms_by_brand_country': {},
            'break_even_cpc_by_brand_country': {},
        }
        for row in campaign_sheet_rows or []:
            campaign_name = str(row.get('广告系列名称', '') or '').strip()
            if not campaign_name or campaign_name == SUMMARY_PLACEHOLDER_TEXT:
                continue
            status = str(row.get('状态', '') or '').strip()
            tracking_link = str(self.normalize_sheet_cell_value(row.get('投放链接', '')) or '').strip()
            break_even_cpc = ''
            if status != '投放已结束':
                break_even_cpc = str(self.normalize_sheet_cell_value(row.get('品牌收支平衡CPC', '')) or '').strip()
            if tracking_link and campaign_name not in maps['tracking_link_by_name']:
                maps['tracking_link_by_name'][campaign_name] = tracking_link
            if break_even_cpc and campaign_name not in maps['break_even_cpc_by_name']:
                maps['break_even_cpc_by_name'][campaign_name] = break_even_cpc

            brand, country = self.extract_brand_and_country_from_campaign_name(campaign_name)
            brand_key = self.normalize_brand_key(brand or '')
            country = self.normalize_country_code(country or '')
            if not brand_key or not country:
                continue

            platform = self.detect_platform_from_tracking_link(tracking_link)
            if platform:
                platform_values = maps['platforms_by_brand_country'].setdefault((brand_key, country), [])
                if platform not in platform_values:
                    platform_values.append(platform)
            if break_even_cpc:
                cpc_values = maps['break_even_cpc_by_brand_country'].setdefault((brand_key, country), [])
                if break_even_cpc not in cpc_values:
                    cpc_values.append(break_even_cpc)
        return maps

    def build_offer_tracking_link_maps_for_ads_brand(self, feishu_data):
        """从Offer表构建广告系列名称和品牌国家到平台列表的映射。"""
        maps = {
            'tracking_link_by_campaign_name': {},
            'platforms_by_brand_country': {},
        }
        for row in feishu_data or []:
            if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                continue
            tracking_link = str(self.normalize_sheet_cell_value(row.get('投放链接', '')) or '').strip()
            platform = self.detect_platform_from_tracking_link(tracking_link)
            if not platform:
                continue

            campaign_names = [
                name.strip()
                for name in str(row.get('广告系列名称', '') or '').split(',')
                if name.strip()
            ]
            for campaign_name in campaign_names:
                maps['tracking_link_by_campaign_name'].setdefault(campaign_name, tracking_link)

            brand_key = self.normalize_brand_key(row.get('品牌名称', ''))
            country = self.normalize_country_code(row.get('国家代码', '') or '')
            if brand_key and country:
                platform_values = maps['platforms_by_brand_country'].setdefault((brand_key, country), [])
                if platform not in platform_values:
                    platform_values.append(platform)
        return maps

    def detect_platform_from_tracking_link(self, tracking_link):
        link = str(tracking_link or '').lower()
        if 'pboost' in link:
            return 'pb'
        if 'yeahpromos' in link:
            return 'yp'
        return ''

    def infer_campaign_platform(self, campaign_name='', tracking_link='', campaign_info=None):
        """推断广告系列平台；只使用明确链接或已归因来源，不按命名默认PB。"""
        platform = self.detect_platform_from_tracking_link(tracking_link)
        if platform:
            return platform
        campaign_info = campaign_info or {}
        source = str(campaign_info.get('_traffic_source', '') or '').strip().lower()
        if source in ('pb', 'yp'):
            return source
        return ''

    def get_ads_brand_platforms_for_campaigns(self, campaign_names, tracking_link_by_name):
        platforms = []
        seen = set()
        for campaign_name in campaign_names or []:
            platform = self.infer_campaign_platform(campaign_name, tracking_link_by_name.get(campaign_name, ''))
            if platform and platform not in seen:
                platforms.append(platform)
                seen.add(platform)
        return ', '.join(platforms)

    def build_ads_brand_sheet_rows(self, campaigns, campaign_sheet_rows, feishu_data, commission_data, yp_commission_data, start_date_str, end_date_str, accounts=None):
        """构建 ads | 品牌 表数据行。"""
        sheet_maps = self.build_ads_brand_campaign_sheet_maps(campaign_sheet_rows)
        offer_link_maps = self.build_offer_tracking_link_maps_for_ads_brand(feishu_data)
        tracking_link_by_name = sheet_maps.get('tracking_link_by_name', {})
        offer_tracking_link_by_name = offer_link_maps.get('tracking_link_by_campaign_name', {})
        break_even_cpc_by_name = sheet_maps.get('break_even_cpc_by_name', {})
        platforms_by_brand_country = sheet_maps.get('platforms_by_brand_country', {})
        offer_platforms_by_brand_country = offer_link_maps.get('platforms_by_brand_country', {})
        break_even_cpc_by_brand_country = sheet_maps.get('break_even_cpc_by_brand_country', {})

        brand_country_totals = {}
        for c in campaigns or []:
            brand_key = c.get('brand_key') or self.normalize_brand_key(c.get('brand', ''))
            country = self.normalize_country_code(c.get('country', '') or '')
            if not brand_key or not country:
                continue
            key = (brand_key, country)
            item = brand_country_totals.setdefault(key, {
                'brand': c.get('brand', ''),
                'country': country,
                'cost': 0.0,
            })
            item['cost'] += float(c.get('cost_usd', 0.0) or 0.0)

        brand_country_commissions, _ = self.calculate_break_even_brand_country_commissions(
            feishu_data=feishu_data,
            commission_data=commission_data,
            yp_commission_data=yp_commission_data
        )

        account_rows = {}
        for account in accounts or []:
            account_id = str(account.get('account_id', account.get('id', '')) or '').strip()
            if not account_id:
                continue
            account_rows.setdefault(account_id, {
                'account_id': account_id,
                'account_name': account.get('account_name', account.get('name', '')),
                'mcc_id': account.get('mcc_id', ''),
                'mcc_name': account.get('mcc_name', ''),
                'campaigns': []
            })

        for c in campaigns or []:
            account_id = str(c.get('account_id', '') or '').strip()
            if not account_id:
                continue
            account = account_rows.setdefault(account_id, {
                'account_id': account_id,
                'account_name': c.get('account_name', ''),
                'mcc_id': c.get('mcc_id', ''),
                'mcc_name': c.get('mcc_name', ''),
                'campaigns': []
            })
            account['campaigns'].append(c)

        rows = []
        for account in account_rows.values():
            has_active_or_paused = any(
                c.get('status') in ('ENABLED', 'PAUSED')
                for c in account.get('campaigns', [])
            )
            if not has_active_or_paused:
                rows.append({
                    'ads': account.get('account_name', ''),
                    'ID': account.get('account_id', ''),
                    '所属mcc': account.get('mcc_name', ''),
                    'mccID': account.get('mcc_id', ''),
                    '_sort_country': '~~~~',
                    '_sort_brand': '',
                })
                continue

            enabled_country_brands = {}
            for c in account.get('campaigns', []):
                if c.get('status') != 'ENABLED':
                    continue
                brand_key = c.get('brand_key') or self.normalize_brand_key(c.get('brand', ''))
                country = self.normalize_country_code(c.get('country', '') or '')
                if not brand_key or not country:
                    continue
                item = enabled_country_brands.setdefault(country, {
                    'country': country,
                    'brands': {},
                })
                item['brands'].setdefault(brand_key, c.get('brand', '') or brand_key)

            for country, active_info in sorted(enabled_country_brands.items(), key=lambda kv: kv[0]):
                active_brand_keys = set(active_info.get('brands', {}).keys())
                matched_campaigns = []
                enabled_campaigns = []
                paused_campaigns = []
                removed_campaigns = []
                campaign_names = []

                for c in account.get('campaigns', []):
                    campaign_brand_key = c.get('brand_key') or self.normalize_brand_key(c.get('brand', ''))
                    campaign_country = self.normalize_country_code(c.get('country', '') or '')
                    if campaign_country != country or campaign_brand_key not in active_brand_keys:
                        continue
                    matched_campaigns.append(c)
                    campaign_name = c.get('campaign_name', '')
                    if campaign_name:
                        campaign_names.append(campaign_name)
                    status = c.get('status')
                    if status == 'ENABLED':
                        enabled_campaigns.append(c)
                    elif status == 'PAUSED':
                        paused_campaigns.append(c)
                    elif status == 'REMOVED':
                        removed_campaigns.append(c)

                if not matched_campaigns:
                    continue

                brand_names = [active_info['brands'][key] for key in sorted(active_brand_keys)]
                total_clicks = sum(int(c.get('clicks', 0) or 0) for c in matched_campaigns)
                total_cost = sum(float(c.get('cost_usd', 0.0) or 0.0) for c in matched_campaigns)
                recent_5_day_total_cost = sum(float(c.get('recent_5_day_cost_usd', 0.0) or 0.0) for c in matched_campaigns)

                enabled_preset_values = [float(c.get('preset_cpc_usd', 0.0) or 0.0) for c in enabled_campaigns if float(c.get('preset_cpc_usd', 0.0) or 0.0) > 0]
                avg_preset_cpc = sum(enabled_preset_values) / len(enabled_preset_values) if enabled_preset_values else 0.0
                enabled_clicks = sum(int(c.get('clicks', 0) or 0) for c in enabled_campaigns)
                enabled_cost = sum(float(c.get('cost_usd', 0.0) or 0.0) for c in enabled_campaigns)
                avg_cpc = (enabled_cost / enabled_clicks) if enabled_clicks > 0 else 0.0

                break_even_values = []
                seen_break_even = set()
                for c in enabled_campaigns:
                    value = break_even_cpc_by_name.get(c.get('campaign_name', ''))
                    if value and value not in seen_break_even:
                        break_even_values.append(value)
                        seen_break_even.add(value)
                for brand_key in sorted(active_brand_keys):
                    for value in break_even_cpc_by_brand_country.get((brand_key, country), []):
                        if value and value not in seen_break_even:
                            break_even_values.append(value)
                            seen_break_even.add(value)

                brand_total_cost = 0.0
                brand_total_commission = 0.0
                for brand_key in active_brand_keys:
                    total_key = (brand_key, country)
                    brand_total_cost += float(brand_country_totals.get(total_key, {}).get('cost', 0.0) or 0.0)
                    brand_total_commission += float(brand_country_commissions.get(total_key, {}).get('commission', 0.0) or 0.0)

                roi = (brand_total_commission / brand_total_cost) if brand_total_cost > 0 else 0.0
                platform_values = []
                seen_platforms = set()
                for c in matched_campaigns:
                    campaign_name = c.get('campaign_name', '')
                    platform = self.infer_campaign_platform(
                        campaign_name,
                        offer_tracking_link_by_name.get(campaign_name, '') or tracking_link_by_name.get(campaign_name, ''),
                        c
                    )
                    if platform and platform not in seen_platforms:
                        platform_values.append(platform)
                        seen_platforms.add(platform)
                for brand_key in sorted(active_brand_keys):
                    for platform in offer_platforms_by_brand_country.get((brand_key, country), []):
                        if platform and platform not in seen_platforms:
                            platform_values.append(platform)
                            seen_platforms.add(platform)
                    for platform in platforms_by_brand_country.get((brand_key, country), []):
                        if platform and platform not in seen_platforms:
                            platform_values.append(platform)
                            seen_platforms.add(platform)

                rows.append({
                    'ads': account.get('account_name', ''),
                    'ID': account.get('account_id', ''),
                    '所属mcc': account.get('mcc_name', ''),
                    'mccID': account.get('mcc_id', ''),
                    '在投品牌名': ', '.join(brand_names),
                    '国家': country,
                    '平台': ', '.join(platform_values),
                    '统计开始时间': start_date_str,
                    '统计结束时间': end_date_str,
                    '广告系列数量（投放中|暂停|结束）': f"{len(enabled_campaigns)} | {len(paused_campaigns)} | {len(removed_campaigns)}",
                    '总点击数': total_clicks,
                    '在投广告系列平均预设CPC': f"${avg_preset_cpc:.2f}",
                    '在投广告系列平均CPC': f"${avg_cpc:.2f}",
                    '品牌收支平衡CPC': ', '.join(break_even_values),
                    '广告系列总花费': f"${total_cost:.2f}",
                    '近5日总花费': f"${recent_5_day_total_cost:.2f}",
                    '品牌总花费': f"${brand_total_cost:.2f}",
                    '品牌总佣金': f"${brand_total_commission:.2f}",
                    'ROI': f"{roi:.1f}",
                    '_sort_country': country,
                    '_sort_brand': ', '.join(brand_names).lower(),
                })

        rows.sort(key=lambda row: (row.get('_sort_country', ''), row.get('_sort_brand', ''), row.get('ads', '')))
        for row in rows:
            row.pop('_sort_country', None)
            row.pop('_sort_brand', None)
        return rows

    def get_recent_window_for_ads_brand_cost(self, end_date_str=None, days=5):
        """获取 ads | 品牌 近N日花费统计窗口，包含结束日。"""
        try:
            end_date = datetime.strptime(str(end_date_str or '').strip(), '%Y-%m-%d').date()
        except Exception:
            end_date = datetime.now().date()
        start_date = end_date - timedelta(days=int(days) - 1)
        return start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')

    def update_ads_brand_sheet(self, feishu_token, spreadsheet_token, feishu_data, commission_data, yp_commission_data, start_date_str, end_date_str):
        """更新 ads | 品牌 工作表。"""
        sheet_id = self.get_feishu_sheet_id_by_title(feishu_token, spreadsheet_token, ADS_BRAND_SHEET_TITLE)
        if not sheet_id:
            self.log_manage(f"  ✗ 未找到工作表: {ADS_BRAND_SHEET_TITLE}")
            return False

        self.log_manage(f"  找到“{ADS_BRAND_SHEET_TITLE}” sheet ID: {sheet_id}")
        target_sheet = self.read_generic_sheet(feishu_token, spreadsheet_token, sheet_id, max_range="A1:AZ2000")
        if target_sheet is None:
            return False
        _, _, header_row = target_sheet
        if not header_row:
            self.log_manage("  ✗ ads | 品牌 表缺少表头")
            return False

        campaigns_sheet_data = self.read_campaigns_sheet(feishu_token, spreadsheet_token, CAMPAIGNS_SHEET_ID)
        campaign_sheet_rows = campaigns_sheet_data[0] if campaigns_sheet_data else []

        recent_cost_start_str, recent_cost_end_str = self.get_recent_window_for_ads_brand_cost(end_date_str, days=5)
        self.log_manage(f"  ads品牌表近5日总花费口径: {recent_cost_start_str} 至 {recent_cost_end_str}")

        campaigns, succeeded_account_ids, ads_accounts = self.get_all_campaigns_for_ads_brand_stats(
            start_date_str,
            end_date_str,
            recent_cost_start_str=recent_cost_start_str,
            recent_cost_end_str=recent_cost_end_str
        )
        self.log_manage(f"  ads品牌表获取到 {len(campaigns)} 个广告系列（来自 {len(succeeded_account_ids)} 个账户）")

        rows = self.build_ads_brand_sheet_rows(
            campaigns=campaigns,
            campaign_sheet_rows=campaign_sheet_rows,
            feishu_data=feishu_data,
            commission_data=commission_data,
            yp_commission_data=yp_commission_data,
            start_date_str=start_date_str,
            end_date_str=end_date_str,
            accounts=ads_accounts
        )
        self.log_manage(f"  ads品牌表准备写入 {len(rows)} 行")
        success = self.overwrite_sheet_rows(feishu_token, spreadsheet_token, sheet_id, header_row, rows)
        if success:
            column_map = {
                str(header).strip(): self.index_to_column_letter(index)
                for index, header in enumerate(header_row or [])
                if header
            }
            background_map = {
                '在投广告系列平均预设CPC': '#F8F9FA',
                '在投广告系列平均CPC': '#FAF1D1',
                '广告系列总花费': '#F8F9FA',
                '近5日总花费': '#FAF1D1',
                '品牌收支平衡CPC': '#FFF258',
                '品牌总花费': '#FBBFBC',
                '品牌总佣金': '#D9F5D6',
            }
            style_updates = []
            for offset, _ in enumerate(rows, start=2):
                for header, color in background_map.items():
                    col_letter = column_map.get(header)
                    if col_letter:
                        style_updates.append({
                            'row_index': offset,
                            'column': col_letter,
                            'background_color': color,
                        })
            if style_updates:
                self.apply_sheet_background_styles(feishu_token, spreadsheet_token, sheet_id, style_updates)
            self.log_manage("  ✅ ads | 品牌 表更新完成")
        return success
    
    def apply_campaigns_style_updates(self, token, spreadsheet_token, sheet_id, style_updates):
        """应用广告系列表格样式更新"""
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        color_map = {
            'green': {'red': 0, 'green': 128, 'blue': 0},
            'orange': {'red': 255, 'green': 165, 'blue': 0},
            'black': {'red': 0, 'green': 0, 'blue': 0},
            'red': {'red': 255, 'green': 0, 'blue': 0},
            'deep_pink': {'red': 199, 'green': 21, 'blue': 133}
        }
        
        # 准备样式更新请求
        style_ranges = []
        for update in style_updates:
            row_idx = update['row_index']
            col = update['column']
            cell_style = {}
            if update.get('background_color'):
                bg_color = str(update.get('background_color')).strip()
                if re.match(r'^#[0-9A-Fa-f]{6}$', bg_color):
                    cell_style['backColor'] = bg_color.upper()
            if update.get('color'):
                color = update['color']
                rgb = color_map.get(color, color_map['black'])
                cell_style.update({
                    'font': {
                        'bold': True
                    },
                    'foreColor': f"#{rgb['red']:02X}{rgb['green']:02X}{rgb['blue']:02X}"
                })

            if not cell_style:
                continue

            style_ranges.append({
                'ranges': f"{sheet_id}!{col}{row_idx}:{col}{row_idx}",
                'style': cell_style
            })
        
        if style_ranges:
            url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/styles_batch_update"
            body = {"data": style_ranges}
            
            try:
                response = requests.put(url, headers=headers, json=body, timeout=30)
                data = response.json()
                if data.get('code') == 0:
                    self.log_manage(f"    样式更新成功: {len(style_ranges)} 个单元格")
                else:
                    self.log_manage(f"    样式更新失败: {data.get('msg', 'Unknown error')}")
            except Exception as e:
                self.log_manage(f"    样式更新异常: {str(e)}")

    def get_campaign_status_style_color(self, status_text):
        """按广告系列表状态文本返回状态列文字颜色类型。"""
        status_text = str(status_text or '').strip()
        if status_text == SUMMARY_STATUS_TEXT:
            return 'deep_pink'
        if status_text.startswith('投放中'):
            return 'green'
        if status_text.startswith('广告系列暂停中') or status_text.startswith('暂停'):
            return 'orange'
        if status_text.startswith('投放已结束'):
            return 'black'
        return ''

    def apply_campaign_status_column_styles(self, token, spreadsheet_token, sheet_id, rows, column_map):
        """统一重刷广告系列表状态列文字颜色，避免行移动/重建后样式错位。"""
        status_col = (column_map or {}).get('状态')
        if not status_col:
            return

        style_updates = []
        for row in rows or []:
            row_index = row.get('row_index')
            color = self.get_campaign_status_style_color(row.get('状态', ''))
            if not row_index or not color:
                continue
            style_updates.append({
                'row_index': row_index,
                'color': color,
                'column': status_col,
            })

        if style_updates:
            self.apply_campaigns_style_updates(token, spreadsheet_token, sheet_id, style_updates)

    def get_campaign_increment_costs_by_name(self, start_date_str, end_date_str, campaign_names):
        """按广告系列名称汇总新增窗口内花费。"""
        target_names = {str(name or '').strip() for name in (campaign_names or []) if str(name or '').strip()}
        if not target_names or not start_date_str or not end_date_str:
            return {}

        costs = {}
        try:
            sub_accounts = self.iter_google_ads_mcc_accounts()

            CNY_TO_USD_RATE = 0.14
            client_by_mcc = {}
            for account in sub_accounts:
                if self.stop_flag:
                    break
                try:
                    mcc_id = account.get('mcc_id', '')
                    client = client_by_mcc.get(mcc_id)
                    if client is None:
                        client = self.get_google_ads_client(mcc_id)
                        client_by_mcc[mcc_id] = client
                    if not client:
                        continue
                    ga_service = client.get_service('GoogleAdsService')
                    account_currency = account.get('currency', 'USD')
                    query = f"""
                        SELECT
                            campaign.name,
                            metrics.cost_micros
                        FROM campaign
                        WHERE segments.date >= '{start_date_str}'
                            AND segments.date <= '{end_date_str}'
                    """
                    result = ga_service.search(customer_id=account['id'], query=query)
                    for row in result:
                        campaign_name = str(row.campaign.name or '').strip()
                        if campaign_name not in target_names:
                            continue
                        cost_original = row.metrics.cost_micros / 1_000_000 if row.metrics.cost_micros else 0.0
                        cost_usd = cost_original * CNY_TO_USD_RATE if account_currency == 'CNY' else cost_original
                        costs[campaign_name] = costs.get(campaign_name, 0.0) + cost_usd
                except Exception as e:
                    self.log_manage(f"  ⚠ 账户 {account.get('name', '')}({account.get('id', '')}) 查询统计行新增花费失败: {str(e)[:100]}")
        except Exception as e:
            self.log_manage(f"  汇总统计行新增花费失败: {e}")
        return costs

    def calculate_yp_increment_commission_by_group(self, feishu_data, yp_commission_data):
        """按YP offer统计行分组汇总新增窗口佣金。"""
        context = self.build_yp_offer_group_context(feishu_data)
        lookup = self.build_brand_country_lookup_maps(self.build_break_even_brand_country_offer_index(feishu_data))
        result = {}

        for trans in yp_commission_data or []:
            if str(trans.get('status', '') or '').strip().lower() == 'rejected':
                continue
            entry = self.resolve_yp_brand_country_entry(trans, lookup)
            if not entry:
                continue

            asin = str(trans.get('asin', '') or trans.get('prod_id', '') or '').strip().upper()
            if not asin:
                asin = self.extract_asin_from_yp_order_id(trans.get('id', ''))
            brand_ids = sorted(entry.get('brand_ids', set()) or [])
            brand_id = self.build_yp_transaction_brand_key(trans, entry)
            country = entry.get('country', '')
            if not asin or not brand_id or not country:
                continue

            candidates = [
                key for key in context.get('offer_group_summary_by_key', {})
                if key[0] == asin and key[1] == brand_id and key[2] == country
            ]
            if len(candidates) == 1:
                group_key = self.get_yp_campaign_group_key(candidates[0], context['offer_group_summary_by_key'].get(candidates[0], {}))
                result[group_key] = result.get(group_key, 0.0) + float(trans.get('sale_comm', 0.0) or 0.0)

        return result

    def build_offer_group_lookup_for_campaign_summary(self, feishu_data):
        lookup = {
            'by_product_link': {},
            'by_tracking_link': {},
            'summary_by_key': {},
            'brand_name_by_asin_brand_country': {},
            'brand_name_by_asin_country': {},
            'brand_name_by_brand_country': {},
        }
        for row in feishu_data or []:
            if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                continue
            asin = str(row.get('ASIN', '') or '').strip().upper()
            brand_id = self.build_yp_row_brand_key(row)
            country = self.normalize_country_code(row.get('国家代码', '') or '')
            if not asin or not brand_id or not country:
                continue
            key = (asin, brand_id, country)
            summary = lookup['summary_by_key'].setdefault(key, {
                'asin': asin,
                'brand_id': brand_id,
                'brand_name': str(row.get('品牌名称', '') or '').strip(),
                'country': country,
            })
            if not summary.get('brand_name') and row.get('品牌名称'):
                summary['brand_name'] = str(row.get('品牌名称', '') or '').strip()
            brand_name = str(row.get('品牌名称', '') or '').strip()
            if brand_name:
                lookup['brand_name_by_asin_brand_country'].setdefault(key, brand_name)
                lookup['brand_name_by_asin_country'].setdefault((asin, country), brand_name)
                lookup['brand_name_by_brand_country'].setdefault((brand_id, country), brand_name)
            product_link_key = self.normalize_product_link_key(row.get('产品链接', ''))
            tracking_link = str(self.normalize_sheet_cell_value(row.get('投放链接', '')) or '').strip()
            if product_link_key:
                lookup['by_product_link'][product_link_key] = key
            if tracking_link:
                lookup['by_tracking_link'][tracking_link] = key
        return lookup

    def resolve_campaign_summary_brand_name(self, offer_key, offer_summary, yp_offer_context, offer_group_lookup):
        """为广告系列统计行解析品牌名，兼容PB/YP和旧统计行分组。"""
        offer_summary = offer_summary or {}
        brand_name = str(offer_summary.get('brand_name', '') or '').strip()
        if brand_name:
            return brand_name

        yp_summary = (yp_offer_context or {}).get('offer_group_summary_by_key', {}).get(offer_key, {}) or {}
        brand_name = str(yp_summary.get('brand_name', '') or '').strip()
        if brand_name:
            return brand_name

        if not offer_key or len(offer_key) < 3:
            return ''

        asin = str(offer_key[0] or '').strip().upper()
        brand_id = str(offer_key[1] or '').strip()
        country = self.normalize_country_code(offer_key[2] or '')

        lookup = offer_group_lookup or {}
        for map_name, key in (
            ('brand_name_by_asin_brand_country', (asin, brand_id, country)),
            ('summary_by_key', (asin, brand_id, country)),
            ('brand_name_by_asin_country', (asin, country)),
            ('brand_name_by_brand_country', (brand_id, country)),
        ):
            value = (lookup.get(map_name, {}) or {}).get(key, '')
            if isinstance(value, dict):
                value = value.get('brand_name', '')
            value = str(value or '').strip()
            if value:
                return value

        return ''

    def consolidate_campaign_sheet_same_offer_rows(self, token, spreadsheet_token, sheet_id, mcc_campaigns, yp_offer_context=None, increment_start_date_str=None, increment_end_date_str=None, feishu_data=None, current_campaign_rows=None, campaign_commission_baseline=None, increment_yp_commission_by_group=None, increment_campaign_cost_by_name=None, increment_pb_commission_by_campaign=None, pb_commission_by_offer_group=None, increment_pb_commission_by_offer_group=None):
        """在广告系列表中为同一个offer的多个广告系列插入统计行。"""
        result = self.read_campaigns_sheet(token, spreadsheet_token, sheet_id)
        if result is None:
            self.log_manage("  无法读取广告系列表，跳过相同offer统计行整理")
            return

        existing_rows, column_map, _ = result
        campaign_name_col = column_map.get('广告系列名称')
        if not campaign_name_col:
            self.log_manage("  广告系列表缺少“广告系列名称”列，跳过相同offer统计行整理")
            return

        yp_offer_context = yp_offer_context or {}
        increment_yp_commission_by_group = increment_yp_commission_by_group or {}
        increment_campaign_cost_by_name = increment_campaign_cost_by_name or {}
        increment_pb_commission_by_campaign = increment_pb_commission_by_campaign or {}
        pb_commission_by_offer_group = pb_commission_by_offer_group or {}
        increment_pb_commission_by_offer_group = increment_pb_commission_by_offer_group or {}
        campaign_commission_baseline = campaign_commission_baseline or self.build_campaign_commission_baseline_snapshot(existing_rows)
        offer_group_lookup = self.build_offer_group_lookup_for_campaign_summary(feishu_data or [])
        current_campaign_rows_by_name = {
            str(row.get('campaign_name') or row.get('广告系列名称') or '').strip(): row
            for row in (current_campaign_rows or [])
            if str(row.get('campaign_name') or row.get('广告系列名称') or '').strip()
        }

        def get_increment_commission_for_offer_group(offer_key):
            if not offer_key:
                return None
            campaign_group_key = self.get_yp_campaign_group_key(offer_key, {}) or offer_key
            base_key = offer_key[:3] if len(offer_key) >= 3 else offer_key
            total = 0.0
            found = False
            for source in (
                increment_yp_commission_by_group,
                increment_pb_commission_by_offer_group,
            ):
                for key in (offer_key, campaign_group_key, base_key):
                    if key in source:
                        total += float(source.get(key, 0.0) or 0.0)
                        found = True
                        break
            return total if found else None

        def build_campaign_offer_group_map():
            rows_now, _, _ = self.read_campaigns_sheet(token, spreadsheet_token, sheet_id)
            group_to_rows = {}
            summary_rows = {}
            summary_rows_by_key = {}
            row_by_index = {}
            ordered_rows = []
            current_rows_by_name = {
                str(row.get('campaign_name') or row.get('广告系列名称') or '').strip(): row
                for row in (current_campaign_rows or [])
                if str(row.get('campaign_name') or row.get('广告系列名称') or '').strip()
            }

            for row in rows_now:
                row_index = row.get('row_index')
                if not row_index or row_index < 2:
                    continue
                row_by_index[row_index] = row
                ordered_rows.append(row_index)

                status = str(row.get('状态', '') or '').strip()
                campaign_name = str(row.get('广告系列名称', '') or '').strip()
                if not campaign_name:
                    continue

                if status == SUMMARY_STATUS_TEXT:
                    parsed_summary = self.parse_campaign_summary_label(campaign_name)
                    if parsed_summary:
                        asin = parsed_summary.get('asin', '')
                        brand_id = parsed_summary.get('brand_id', '')
                        country = parsed_summary.get('country', '')
                        yp_marker = parsed_summary.get('marker', '')
                        if asin and country:
                            summary_key = (asin, brand_id, country)
                            if yp_marker.startswith('adg:'):
                                matched_summary_key = None
                                for source_key, source_summary in yp_offer_context.get('offer_group_summary_by_key', {}).items():
                                    if source_key == summary_key:
                                        matched_summary_key = self.get_yp_campaign_group_key(source_key, source_summary)
                                        break
                                summary_key = matched_summary_key or summary_key
                            summary_rows_by_key.setdefault(summary_key, []).append(row_index)
                            summary_rows[summary_key] = row_index
                    continue

                campaign_info = mcc_campaigns.get(campaign_name)
                offer_key = None
                current_row = current_rows_by_name.get(campaign_name, {}) or {}
                if current_row:
                    offer_key = current_row.get('_offer_group_key')
                if campaign_info:
                    if not offer_key:
                        offer_key = campaign_info.get('_offer_group_key')
                if not offer_key:
                    tracking_link = row.get('投放链接', '')
                    offer_key, _ = self.resolve_yp_campaign_offer_group(
                        campaign_name,
                        {
                            'campaign_name': campaign_name,
                            'asin': '',
                            'country': '',
                            'campaign_id': '',
                            'tracking_link': tracking_link,
                            'product_link': row.get('产品链接', ''),
                        },
                        yp_offer_context,
                        tracking_link
                    )

                if not offer_key:
                    product_link_key = self.normalize_product_link_key(row.get('产品链接', ''))
                    tracking_link = str(self.normalize_sheet_cell_value(row.get('投放链接', '')) or '').strip()
                    offer_key = offer_group_lookup.get('by_product_link', {}).get(product_link_key)
                    if not offer_key and tracking_link:
                        offer_key = offer_group_lookup.get('by_tracking_link', {}).get(tracking_link)

                if not offer_key:
                    continue

                offer_summary = yp_offer_context.get('offer_group_summary_by_key', {}).get(offer_key, {}) or {}
                if not offer_summary:
                    offer_summary = offer_group_lookup.get('summary_by_key', {}).get((offer_key[0], offer_key[1], offer_key[2]), {}) or {}
                campaign_group_key = self.get_yp_campaign_group_key(offer_key, offer_summary)
                if not campaign_group_key:
                    continue
                group_to_rows.setdefault(campaign_group_key, []).append(row_index)

            return rows_now, row_by_index, ordered_rows, group_to_rows, summary_rows, summary_rows_by_key

        rows_now, row_by_index, ordered_rows, group_to_rows, summary_rows, summary_rows_by_key = build_campaign_offer_group_map()

        def get_visible_child_rows(summary_row_index, expected_rows=None):
            if not summary_row_index:
                return []
            expected_row_set = set(expected_rows or [])
            visible_rows = []
            for row_index in sorted(row_by_index):
                if row_index <= summary_row_index:
                    continue
                row_data = row_by_index.get(row_index, {}) or {}
                status_text = str(row_data.get('状态', '') or '').strip()
                if status_text == SUMMARY_STATUS_TEXT:
                    break
                if expected_row_set and row_index not in expected_row_set:
                    continue
                campaign_name = str(row_data.get('广告系列名称', '') or '').strip()
                if campaign_name and campaign_name != SUMMARY_PLACEHOLDER_TEXT:
                    visible_rows.append(row_index)
            return visible_rows

        duplicate_summary_rows_to_delete = []
        for offer_key, summary_row_indices in summary_rows_by_key.items():
            if len(summary_row_indices) <= 1:
                continue
            group_rows = sorted(group_to_rows.get(offer_key, []))
            if group_rows:
                keep_row = min(summary_row_indices, key=lambda ri: (abs(ri - group_rows[0]), ri))
            else:
                keep_row = min(summary_row_indices)
            duplicate_summary_rows_to_delete.extend(ri for ri in summary_row_indices if ri != keep_row)

        summary_rows_by_base_key = {}
        for summary_key, row_indices in summary_rows_by_key.items():
            if len(summary_key) < 3:
                continue
            base_key = summary_key[:3]
            for row_index in row_indices:
                summary_rows_by_base_key.setdefault(base_key, []).append((summary_key, row_index))

        for base_key, summary_items in summary_rows_by_base_key.items():
            has_marked_summary = any(len(summary_key) > 3 and summary_key[3] for summary_key, _ in summary_items)
            if not has_marked_summary:
                continue
            for summary_key, row_index in summary_items:
                if len(summary_key) <= 3 or not summary_key[3]:
                    duplicate_summary_rows_to_delete.append(row_index)

        data_column_names = [
            '状态', '广告系列名称', '投放中的ads', '投放链接', '品牌名', '品牌名称',
            '国家代码', '广告系列总花费', '总佣金', 'ROI', '佣金ASIN', '每单佣金',
            '产品链接', '已启用关键字CTR', '总点击数', 'CPC', '预设CPC',
            '品牌收支平衡CPC', '新增广告系列花费', '新增佣金'
        ]
        empty_gap_rows_to_delete = []
        all_data_rows = sorted(row_by_index)
        last_meaningful_row = 1
        for row_index in all_data_rows:
            row_data = row_by_index.get(row_index, {}) or {}
            if any(str(self.normalize_sheet_cell_value(row_data.get(name, '')) or '').strip() for name in data_column_names):
                last_meaningful_row = row_index
        for row_index in all_data_rows:
            if row_index >= last_meaningful_row:
                continue
            row_data = row_by_index.get(row_index, {}) or {}
            if any(str(self.normalize_sheet_cell_value(row_data.get(name, '')) or '').strip() for name in data_column_names):
                continue
            empty_gap_rows_to_delete.append(row_index)

        if empty_gap_rows_to_delete:
            duplicate_summary_rows_to_delete.extend(empty_gap_rows_to_delete)

        duplicate_summary_rows_to_delete = sorted(set(duplicate_summary_rows_to_delete))
        if duplicate_summary_rows_to_delete:
            if not self.delete_sheet_rows(token, spreadsheet_token, sheet_id, duplicate_summary_rows_to_delete):
                raise RuntimeError("无法删除重复的广告系列相同offer统计行")
            if empty_gap_rows_to_delete:
                self.log_manage(f"  已清理 {len(empty_gap_rows_to_delete)} 条广告系列表空白夹缝行")
            summary_deleted_count = len(set(duplicate_summary_rows_to_delete) - set(empty_gap_rows_to_delete))
            if summary_deleted_count:
                self.log_manage(f"  已清理 {summary_deleted_count} 条重复的广告系列相同offer统计行")
            rows_now, row_by_index, ordered_rows, group_to_rows, summary_rows, summary_rows_by_key = build_campaign_offer_group_map()

        duplicate_groups = {k: v for k, v in group_to_rows.items() if len(v) > 1}
        if not duplicate_groups:
            self.log_manage("  未发现需要写入统计行的YP广告系列相同offer分组")
            self.apply_campaign_status_column_styles(token, spreadsheet_token, sheet_id, rows_now, column_map)
            metric_style_updates = self.build_campaign_metric_background_style_updates(rows_now, column_map)
            if metric_style_updates:
                self.apply_campaigns_style_updates(token, spreadsheet_token, sheet_id, metric_style_updates)
            return

        self.log_manage(f"  发现 {len(duplicate_groups)} 组YP广告系列相同offer，开始整理")

        group_order = sorted(
            duplicate_groups.keys(),
            key=lambda key: min(group_to_rows.get(key, [10**9]))
        )

        def build_summary_row_updates(offer_key, rows, summary_row_index, row_by_index):
            asin, brand_id, country = offer_key[:3]
            yp_marker = offer_key[3] if len(offer_key) > 3 else ''
            offer_summary = {}
            for source_key, source_summary in yp_offer_context.get('offer_group_summary_by_key', {}).items():
                if self.get_yp_campaign_group_key(source_key, source_summary) == offer_key:
                    offer_summary = source_summary or {}
                    break

            effective_rows = get_visible_child_rows(summary_row_index, rows) or rows
            metrics = self.summarize_campaign_group_rows(row_by_index, effective_rows)
            total_cost = metrics.get('total_cost', 0.0)
            total_commission = metrics.get('total_commission', 0.0)
            if total_commission <= 0 and offer_summary:
                total_commission = float(offer_summary.get('commission', 0.0) or 0.0)
            total_commission += float(pb_commission_by_offer_group.get(offer_key[:3], 0.0) or 0.0)
            total_clicks = metrics.get('total_clicks', 0)

            child_status_counts = {'enabled': 0, 'paused': 0, 'ended': 0}
            platforms = set()
            child_campaign_names = []
            for row_index in effective_rows:
                row_data = row_by_index.get(row_index, {}) or {}
                child_campaign_name = str(row_data.get('广告系列名称', '') or '').strip()
                if child_campaign_name:
                    child_campaign_names.append(child_campaign_name)
                current_row = current_campaign_rows_by_name.get(child_campaign_name, {})
                status_text = str(current_row.get('状态', '') or row_data.get('状态', '') or '').strip()
                if status_text.startswith('投放中'):
                    child_status_counts['enabled'] += 1
                elif status_text.startswith('广告系列暂停中') or status_text.startswith('暂停'):
                    child_status_counts['paused'] += 1
                elif status_text.startswith('投放已结束'):
                    child_status_counts['ended'] += 1
                tracking_link = str(self.normalize_sheet_cell_value(row_data.get('投放链接', '')) or '').strip()
                platform = self.detect_platform_from_tracking_link(tracking_link)
                if platform:
                    platforms.add(platform)
            if yp_marker:
                platforms.add('yp')

            status_counts_text = f"{child_status_counts['enabled']} | {child_status_counts['paused']} | {child_status_counts['ended']}"
            brand_name = (
                self.resolve_campaign_summary_brand_name(
                    offer_key,
                    offer_summary,
                    yp_offer_context,
                    offer_group_lookup,
                )
                or '-'
            )
            summary_label = self.build_campaign_summary_label(
                brand_name=brand_name,
                asin=asin,
                brand_id=self.display_offer_brand_id(brand_id),
                country=country,
                platforms=sorted(platforms),
                status_counts=status_counts_text,
                marker=yp_marker,
            )

            increment_cost = sum(
                float((increment_campaign_cost_by_name or {}).get(name, 0.0) or 0.0)
                for name in child_campaign_names
            )
            previous_commission = self.get_campaign_summary_previous_commission(
                offer_key,
                child_campaign_names,
                campaign_commission_baseline
            )
            if total_commission <= 0 and previous_commission > 0:
                total_commission = previous_commission
            increment_commission = total_commission - previous_commission
            roi_value = round(total_commission / total_cost, 1) if total_cost > 0 else 0

            updates = []
            style_updates = []
            summary_values = {
                '状态': SUMMARY_STATUS_TEXT,
                '广告系列名称': summary_label,
                '品牌名': brand_name if brand_name != '-' else '',
                '品牌名称': brand_name if brand_name != '-' else '',
                '国家代码': country,
                '广告系列总花费': f"${total_cost:.2f}" if total_cost else '',
                '总佣金': f"${total_commission:.2f}" if total_commission else '',
                'ROI': f"{roi_value}",
                '总点击数': total_clicks if total_clicks else '',
                '佣金ASIN': ', '.join(metrics.get('commission_asins', [])),
                '品牌收支平衡CPC': metrics.get('brand_break_even_cpc', ''),
                '新增广告系列花费': f"${increment_cost:.2f}" if increment_cost else '',
                '新增佣金': f"${increment_commission:.2f}",
            }
            preserved_headers = set(summary_values.keys())

            for header, col_letter in column_map.items():
                col_idx = self.column_letter_to_index(col_letter)
                if col_idx is None:
                    continue
                value = summary_values.get(header, SUMMARY_PLACEHOLDER_TEXT if header not in preserved_headers else '')
                updates.append((summary_row_index, col_idx, value))

            if '状态' in column_map:
                style_updates.append({
                    'row_index': summary_row_index,
                    'color': 'deep_pink',
                    'column': column_map.get('状态', 'A')
                })

            commission_col_letter = column_map.get('总佣金')
            commission_col_idx = self.column_letter_to_index(commission_col_letter) if commission_col_letter else None
            increment_commission_col_letter = column_map.get('新增佣金')
            increment_commission_col_idx = self.column_letter_to_index(increment_commission_col_letter) if increment_commission_col_letter else None
            increment_cost_col_letter = column_map.get('新增广告系列花费')
            increment_cost_col_idx = self.column_letter_to_index(increment_cost_col_letter) if increment_cost_col_letter else None
            if commission_col_idx is not None:
                for row_index in effective_rows:
                    updates.append((row_index, commission_col_idx, '↑'))
                    if increment_commission_col_idx is not None:
                        updates.append((row_index, increment_commission_col_idx, '↑'))
                    if increment_cost_col_idx is not None:
                        updates.append((row_index, increment_cost_col_idx, '↑'))

            return updates, style_updates

        fast_path_ready = True
        for offer_key in group_order:
            rows = sorted(group_to_rows.get(offer_key, []))
            if len(rows) <= 1:
                continue
            target = rows[0]
            summary_row_index = summary_rows.get(offer_key)
            if summary_row_index != target - 1 or rows != list(range(target, target + len(rows))):
                fast_path_ready = False
                break

        if fast_path_ready and not getattr(self, '_force_rebuild_campaign_summary_rows', False):
            all_updates = []
            all_style_updates = []
            fast_group_count = 0
            for offer_key in group_order:
                rows = sorted(group_to_rows.get(offer_key, []))
                if len(rows) <= 1:
                    continue
                summary_row_index = summary_rows.get(offer_key)
                updates, style_updates = build_summary_row_updates(offer_key, rows, summary_row_index, row_by_index)
                all_updates.extend(updates)
                all_style_updates.extend(style_updates)
                fast_group_count += 1

            if all_updates:
                self._batch_update_sheet_cells(token, spreadsheet_token, all_updates, sheet_id=sheet_id)
            if all_style_updates:
                self.apply_campaigns_style_updates(token, spreadsheet_token, sheet_id, all_style_updates)
            self.apply_campaign_status_column_styles(token, spreadsheet_token, sheet_id, rows_now, column_map)
            metric_style_updates = self.build_campaign_metric_background_style_updates(rows_now, column_map)
            if metric_style_updates:
                self.apply_campaigns_style_updates(token, spreadsheet_token, sheet_id, metric_style_updates)
            self.log_manage(f"  统计行快速更新完成: {fast_group_count} 组，{len(all_updates)} 个单元格")
            return

        self.log_manage("  统计行结构需要整理，使用移动/插入慢速路径")

        header_by_col_idx = {}
        ordered_headers = []
        for header, col_letter in sorted(
            column_map.items(),
            key=lambda item: self.column_letter_to_index(item[1]) if self.column_letter_to_index(item[1]) is not None else 10**9
        ):
            col_idx = self.column_letter_to_index(col_letter)
            if col_idx is None:
                continue
            header_by_col_idx[col_idx] = header
            ordered_headers.append(header)

        group_by_child_row = {}
        grouped_child_rows = set()
        for offer_key, rows in duplicate_groups.items():
            for row_index in rows:
                group_by_child_row[row_index] = offer_key
                grouped_child_rows.add(row_index)

        summary_rows_to_skip = {
            row.get('row_index')
            for row in rows_now
            if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT
        }

        normal_output_rows = []
        grouped_output_rows = []
        summary_style_updates = []
        emitted_groups = set()

        for row_index in sorted(row_by_index):
            if row_index in summary_rows_to_skip:
                continue
            row_data = row_by_index.get(row_index, {}) or {}
            row_status = str(row_data.get('状态', '') or '').strip()
            row_campaign_name = str(row_data.get('广告系列名称', '') or '').strip()
            if not row_status and (not row_campaign_name or row_campaign_name == SUMMARY_PLACEHOLDER_TEXT):
                continue
            if not any(
                str(self.normalize_sheet_cell_value(row_data.get(header, '')) or '').strip()
                for header in ordered_headers
            ):
                continue

            offer_key = group_by_child_row.get(row_index)
            if offer_key:
                if offer_key in emitted_groups:
                    continue
                group_rows = sorted(group_to_rows.get(offer_key, []))
                if len(group_rows) <= 1:
                    for child_row_index in group_rows:
                        child = dict(row_by_index.get(child_row_index, {}) or {})
                        child.pop('row_index', None)
                        grouped_output_rows.append(child)
                    emitted_groups.add(offer_key)
                    continue

                summary_updates, style_updates = build_summary_row_updates(
                    offer_key,
                    group_rows,
                    group_rows[0] - 1,
                    row_by_index
                )

                summary_row_values = {header: '' for header in ordered_headers}
                child_value_updates = {}
                for update_row, col_idx, value in summary_updates:
                    header = header_by_col_idx.get(col_idx)
                    if not header:
                        continue
                    if update_row == group_rows[0] - 1:
                        summary_row_values[header] = value
                    elif update_row in group_rows:
                        child_value_updates.setdefault(update_row, {})[header] = value

                grouped_output_rows.append(summary_row_values)
                for style in style_updates:
                    style_copy = dict(style)
                    style_copy['_pending_group_offset'] = len(grouped_output_rows) - 1
                    summary_style_updates.append(style_copy)

                for child_row_index in group_rows:
                    child = dict(row_by_index.get(child_row_index, {}) or {})
                    child.update(child_value_updates.get(child_row_index, {}))
                    child.pop('row_index', None)
                    grouped_output_rows.append(child)

                emitted_groups.add(offer_key)
                asin, brand_id, country = offer_key[:3]
                yp_marker = offer_key[3] if len(offer_key) > 3 else ''
                self.log_manage(
                f"    已整理广告系列统计行: ASIN={asin}, 品牌ID={self.display_offer_brand_id(brand_id) or '-'}, 国家={country}, 分组={yp_marker or '-'}, 汇总{len(group_rows)}个广告系列"
                )
                continue

            if row_index in grouped_child_rows:
                continue

            normal_row = dict(row_data)
            normal_row.pop('row_index', None)
            normal_output_rows.append(normal_row)

        ordered_output_rows = normal_output_rows + grouped_output_rows
        first_group_row = 2 + len(normal_output_rows)
        for style in summary_style_updates:
            pending_offset = style.pop('_pending_group_offset', None)
            if pending_offset is not None:
                style['row_index'] = first_group_row + int(pending_offset)

        if not ordered_output_rows:
            self.log_manage("  统计行重建跳过：没有可写回的广告系列数据")
            return

        old_last_row = max(row_by_index.keys()) if row_by_index else 2
        self.clear_sheet_data_rows(
            token,
            spreadsheet_token,
            sheet_id,
            column_count=len(ordered_headers),
            max_rows=max(5000, old_last_row + 100)
        )

        end_col = self.index_to_column_letter(len(ordered_headers) - 1)
        value_ranges = []
        for offset, row in enumerate(ordered_output_rows, start=2):
            row['row_index'] = offset
            value_ranges.append({
                "range": f"{sheet_id}!A{offset}:{end_col}{offset}",
                "values": [[row.get(header, '') for header in ordered_headers]]
            })

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
        batch_size = 100
        for i in range(0, len(value_ranges), batch_size):
            body = {"valueRanges": value_ranges[i:i + batch_size]}
            response = requests.post(url, headers=headers, json=body, timeout=30)
            data = response.json()
            if data.get('code') != 0:
                raise RuntimeError(f"广告系列统计行重建写回失败: {data.get('msg', 'Unknown error')}")

        if summary_style_updates:
            self.apply_campaigns_style_updates(token, spreadsheet_token, sheet_id, summary_style_updates)
        self.apply_campaign_status_column_styles(token, spreadsheet_token, sheet_id, ordered_output_rows, column_map)
        metric_style_updates = self.build_campaign_metric_background_style_updates(ordered_output_rows, column_map)
        if metric_style_updates:
            self.apply_campaigns_style_updates(token, spreadsheet_token, sheet_id, metric_style_updates)
        self.log_manage(f"  统计行结构重建完成: 写回 {len(ordered_output_rows)} 行，统计行 {len(summary_style_updates)} 组")
        return

        slow_path_updates = []
        slow_path_style_updates = []
        for offer_key in group_order:
            rows_now, row_by_index, ordered_rows, group_to_rows, summary_rows, summary_rows_by_key = build_campaign_offer_group_map()
            rows = sorted(group_to_rows.get(offer_key, []))
            if len(rows) <= 1:
                continue

            target = rows[0]
            summary_row_index = summary_rows.get(offer_key)
            if summary_row_index is not None and summary_row_index != target - 1:
                src_0based = summary_row_index - 1
                dst_0based = target - 1
                if summary_row_index < target:
                    dst_0based = target - 2
                if not self.feishu_move_dimension(token, spreadsheet_token, sheet_id, src_0based, src_0based, dst_0based):
                    raise RuntimeError(f"无法移动广告系列统计行 {offer_key}")
                rows_now, row_by_index, ordered_rows, group_to_rows, summary_rows, summary_rows_by_key = build_campaign_offer_group_map()
                rows = sorted(group_to_rows.get(offer_key, rows))
                target = rows[0]

            desired = list(range(target, target + len(rows)))
            if rows != desired:
                working_order = list(ordered_rows)
                data_start_0based = 1
                for offset, row_index in enumerate(rows):
                    current_pos = working_order.index(row_index)
                    desired_pos = working_order.index(target) + offset
                    if current_pos == desired_pos:
                        continue
                    src_0based = data_start_0based + current_pos
                    dst_0based = data_start_0based + desired_pos
                    if not self.feishu_move_dimension(token, spreadsheet_token, sheet_id, src_0based, src_0based, dst_0based):
                        raise RuntimeError(f"无法整理广告系列分组 {offer_key}")
                    moved = working_order.pop(current_pos)
                    working_order.insert(desired_pos, moved)

                rows_now, row_by_index, ordered_rows, group_to_rows, summary_rows, summary_rows_by_key = build_campaign_offer_group_map()
                rows = sorted(group_to_rows.get(offer_key, rows))
                target = rows[0]

            summary_row_index = summary_rows.get(offer_key)
            if summary_row_index is None or summary_row_index != target - 1:
                summary_row_index = target
                if not self.insert_sheet_rows(token, spreadsheet_token, sheet_id, start_index_0based=target - 1, count=1):
                    raise RuntimeError(f"无法为广告系列分组 {offer_key} 插入统计行")
                rows = [ri + 1 if ri >= summary_row_index else ri for ri in rows]
                rows_now, row_by_index, ordered_rows, group_to_rows, summary_rows, summary_rows_by_key = build_campaign_offer_group_map()
                rows = sorted(group_to_rows.get(offer_key, rows))

            asin, brand_id, country = offer_key[:3]
            yp_marker = offer_key[3] if len(offer_key) > 3 else ''
            offer_summary = {}
            for source_key, source_summary in yp_offer_context.get('offer_group_summary_by_key', {}).items():
                if self.get_yp_campaign_group_key(source_key, source_summary) == offer_key:
                    offer_summary = source_summary or {}
                    break
            effective_rows = get_visible_child_rows(summary_row_index, rows) or rows
            metrics = self.summarize_campaign_group_rows(row_by_index, effective_rows)
            total_cost = metrics.get('total_cost', 0.0)
            total_commission = metrics.get('total_commission', 0.0)
            if total_commission <= 0 and offer_summary:
                total_commission = float(offer_summary.get('commission', 0.0) or 0.0)
            total_commission += float(pb_commission_by_offer_group.get(offer_key[:3], 0.0) or 0.0)
            total_clicks = metrics.get('total_clicks', 0)
            child_status_counts = {'enabled': 0, 'paused': 0, 'ended': 0}
            platforms = set()
            child_campaign_names = []
            for row_index in effective_rows:
                row_data = row_by_index.get(row_index, {}) or {}
                child_campaign_name = str(row_data.get('广告系列名称', '') or '').strip()
                if child_campaign_name:
                    child_campaign_names.append(child_campaign_name)
                current_row = current_campaign_rows_by_name.get(child_campaign_name, {})
                status_text = str(current_row.get('状态', '') or row_data.get('状态', '') or '').strip()
                if status_text.startswith('投放中'):
                    child_status_counts['enabled'] += 1
                elif status_text.startswith('广告系列暂停中') or status_text.startswith('暂停'):
                    child_status_counts['paused'] += 1
                elif status_text.startswith('投放已结束'):
                    child_status_counts['ended'] += 1
                tracking_link = str(self.normalize_sheet_cell_value(row_data.get('投放链接', '')) or '').strip()
                platform = self.detect_platform_from_tracking_link(tracking_link)
                if platform:
                    platforms.add(platform)
            if yp_marker:
                platforms.add('yp')
            status_counts_text = f"{child_status_counts['enabled']} | {child_status_counts['paused']} | {child_status_counts['ended']}"
            brand_name = (
                self.resolve_campaign_summary_brand_name(
                    offer_key,
                    offer_summary,
                    yp_offer_context,
                    offer_group_lookup,
                )
                or '-'
            )
            summary_label = self.build_campaign_summary_label(
                brand_name=brand_name,
                asin=asin,
                brand_id=self.display_offer_brand_id(brand_id),
                country=country,
                platforms=sorted(platforms),
                status_counts=status_counts_text,
                marker=yp_marker,
            )
            increment_cost = sum(
                float((increment_campaign_cost_by_name or {}).get(name, 0.0) or 0.0)
                for name in child_campaign_names
            )
            previous_commission = self.get_campaign_summary_previous_commission(
                offer_key,
                child_campaign_names,
                campaign_commission_baseline
            )
            if total_commission <= 0 and previous_commission > 0:
                total_commission = previous_commission
            increment_commission = total_commission - previous_commission
            roi_value = round(total_commission / total_cost, 1) if total_cost > 0 else 0
            updates = []
            style_updates = []
            summary_values = {
                '状态': SUMMARY_STATUS_TEXT,
                '广告系列名称': summary_label,
                '品牌名': brand_name if brand_name != '-' else '',
                '品牌名称': brand_name if brand_name != '-' else '',
                '国家代码': country,
                '广告系列总花费': f"${total_cost:.2f}" if total_cost else '',
                '总佣金': f"${total_commission:.2f}" if total_commission else '',
                'ROI': f"{roi_value}",
                '总点击数': total_clicks if total_clicks else '',
                '佣金ASIN': ', '.join(metrics.get('commission_asins', [])),
                '品牌收支平衡CPC': metrics.get('brand_break_even_cpc', ''),
                '新增广告系列花费': f"${increment_cost:.2f}" if increment_cost else '',
                '新增佣金': f"${increment_commission:.2f}",
            }
            preserved_headers = set(summary_values.keys())

            for header, col_letter in column_map.items():
                col_idx = self.column_letter_to_index(col_letter)
                if col_idx is None:
                    continue
                value = summary_values.get(header, SUMMARY_PLACEHOLDER_TEXT if header not in preserved_headers else '')
                updates.append((summary_row_index, col_idx, value))

            if '状态' in column_map:
                style_updates.append({
                    'row_index': summary_row_index,
                    'color': 'deep_pink',
                    'column': column_map.get('状态', 'A')
                })

            commission_col_letter = column_map.get('总佣金')
            commission_col_idx = self.column_letter_to_index(commission_col_letter) if commission_col_letter else None
            increment_commission_col_letter = column_map.get('新增佣金')
            increment_commission_col_idx = self.column_letter_to_index(increment_commission_col_letter) if increment_commission_col_letter else None
            increment_cost_col_letter = column_map.get('新增广告系列花费')
            increment_cost_col_idx = self.column_letter_to_index(increment_cost_col_letter) if increment_cost_col_letter else None
            if commission_col_idx is not None:
                for row_index in effective_rows:
                    updates.append((row_index, commission_col_idx, '↑'))
                    if increment_commission_col_idx is not None:
                        updates.append((row_index, increment_commission_col_idx, '↑'))
                    if increment_cost_col_idx is not None:
                        updates.append((row_index, increment_cost_col_idx, '↑'))

            slow_path_updates.extend(updates)
            slow_path_style_updates.extend(style_updates)

            self.log_manage(
                f"    已整理广告系列统计行: ASIN={asin}, 品牌ID={self.display_offer_brand_id(brand_id) or '-'}, 国家={country}, 分组={yp_marker or '-'}, 汇总{len(rows)}个广告系列"
            )

        if slow_path_updates:
            self._batch_update_sheet_cells(token, spreadsheet_token, slow_path_updates, sheet_id=sheet_id)
        if slow_path_style_updates:
            self.apply_campaigns_style_updates(token, spreadsheet_token, sheet_id, slow_path_style_updates)
        if slow_path_updates:
            self.log_manage(f"  统计行慢速路径批量写回完成: {len(slow_path_updates)} 个单元格")
    
    def get_all_campaigns_with_asin(self, start_date_str=None, end_date_str=None):
        """获取所有广告系列及其ASIN（从广告层级的最终到达网址提取）
        
        返回:
            (campaigns, succeeded_account_ids): 广告系列列表和成功查询的账户ID集合
        """
        campaigns = []
        succeeded_account_ids = set()
        
        try:
            sub_accounts = self.iter_google_ads_mcc_accounts()
            self.log_manage(f"  多MCC合计找到 {len(sub_accounts)} 个子账户")
            
            # CNY到USD的汇率（可根据实际情况调整）
            CNY_TO_USD_RATE = 0.14  # 1 CNY ≈ 0.14 USD
            client_by_mcc = {}
            
            # 遍历每个子账户获取广告系列和广告
            for account in sub_accounts:
                if self.stop_flag:
                    break
                
                try:
                    mcc_id = account.get('mcc_id', '')
                    client = client_by_mcc.get(mcc_id)
                    if client is None:
                        client = self.get_google_ads_client(mcc_id)
                        client_by_mcc[mcc_id] = client
                    if not client:
                        continue
                    ga_service = client.get_service('GoogleAdsService')
                    # 获取账户货币类型
                    account_currency = account.get('currency', 'USD')

                    # 先查询广告系列基础信息，确保所有非REMOVED广告系列都能纳入结果。
                    campaign_query = """
                        SELECT
                            campaign.id,
                            campaign.name,
                            campaign.status,
                            campaign.final_url_suffix
                        FROM campaign
                        WHERE campaign.status != 'REMOVED'
                    """
                    campaign_response = ga_service.search(customer_id=account['id'], query=campaign_query)

                    campaign_info = {}  # {campaign_id: campaign_data}
                    for row in campaign_response:
                        c = row.campaign
                        campaign_id = str(c.id)
                        if campaign_id not in campaign_info:
                            country = self.extract_country_from_campaign_name(c.name)
                            campaign_info[campaign_id] = {
                                'account_id': account['id'],
                                'account_name': account['name'],
                                'mcc_id': account.get('mcc_id', ''),
                                'mcc_name': account.get('mcc_name', ''),
                                'campaign_id': campaign_id,
                                'campaign_name': c.name,
                                'status': c.status.name,
                                'cost_usd': 0,
                                'currency': account_currency,
                                'asin': None,
                                'country': country,
                                'final_urls': [],
                                'final_url_suffix': c.final_url_suffix if c.final_url_suffix else '',
                                'link_ids': set(),
                                'uids': set(),
                                'clicks': 0,
                                'avg_cpc_usd': 0.0,
                                'preset_cpc_usd': 0.0,
                                'enabled_keyword_ctr_weighted_numerator': 0.0,
                                'enabled_keyword_ctr_impressions': 0
                            }

                    # 查询广告层级，获取最终到达网址，补充ASIN / uid / link_id。
                    ad_query = """
                        SELECT
                            campaign.id,
                            campaign.name,
                            campaign.status,
                            campaign.final_url_suffix,
                            ad_group_ad.ad.final_urls
                        FROM ad_group_ad
                        WHERE campaign.status != 'REMOVED'
                            AND ad_group_ad.status != 'REMOVED'
                    """
                    
                    ad_response = ga_service.search(customer_id=account['id'], query=ad_query)

                    for row in ad_response:
                        c = row.campaign
                        ad = row.ad_group_ad.ad
                        campaign_id = str(c.id)

                        if campaign_id not in campaign_info:
                            country = self.extract_country_from_campaign_name(c.name)
                            campaign_info[campaign_id] = {
                                'account_id': account['id'],
                                'account_name': account['name'],
                                'mcc_id': account.get('mcc_id', ''),
                                'mcc_name': account.get('mcc_name', ''),
                                'campaign_id': campaign_id,
                                'campaign_name': c.name,
                                'status': c.status.name,
                                'cost_usd': 0,
                                'currency': account_currency,
                                'asin': None,
                                'country': country,
                                'final_urls': [],
                                'final_url_suffix': c.final_url_suffix if c.final_url_suffix else '',
                                'link_ids': set(),
                                'uids': set(),
                                'clicks': 0,
                                'avg_cpc_usd': 0.0,
                                'preset_cpc_usd': 0.0,
                                'enabled_keyword_ctr_weighted_numerator': 0.0,
                                'enabled_keyword_ctr_impressions': 0
                            }
                        
                        # 从广告的最终到达网址提取ASIN
                        asin = None
                        ad_link_id = None
                        ad_uid = None
                        if ad.final_urls:
                            for url in ad.final_urls:
                                if not asin:
                                    asin = self.extract_asin_from_url(url)
                                # 也尝试从final_url中提取uid和link_id（适用于pboost.me短链接）
                                if not ad_uid:
                                    uid_match = re.search(r'[?&]uid=([^&]+)', url)
                                    if uid_match:
                                        ad_uid = uid_match.group(1)
                                if not ad_link_id:
                                    lid_match = re.search(r'[?&]aa_adgroupid=([^&]+)', url)
                                    if lid_match:
                                        ad_link_id = lid_match.group(1)
                        
                        # 从campaign的final_url_suffix（跟踪后缀）中提取link_id和uid
                        suffix = c.final_url_suffix if c.final_url_suffix else ''
                        if suffix:
                            if not ad_link_id:
                                lid_match = re.search(r'(?:^|&)aa_adgroupid=([^&]+)', suffix)
                                if lid_match:
                                    ad_link_id = lid_match.group(1)
                            if not ad_uid:
                                uid_match = re.search(r'(?:^|&)uid=([^&]+)', suffix)
                                if uid_match:
                                    ad_uid = uid_match.group(1)

                        if not campaign_info[campaign_id]['final_urls'] and ad.final_urls:
                            campaign_info[campaign_id]['final_urls'] = list(ad.final_urls)

                        # 如果之前没有ASIN，尝试从这个广告获取
                        if not campaign_info[campaign_id]['asin'] and asin:
                            campaign_info[campaign_id]['asin'] = asin
                        
                        # 收集link_id和uid（用于精确匹配offer行）
                        if ad_link_id:
                            campaign_info[campaign_id]['link_ids'].add(ad_link_id)
                        if ad_uid:
                            campaign_info[campaign_id]['uids'].add(ad_uid)

                    # 查询广告组层级预设出价，优先取每个campaign中的最高值作为展示口径。
                    ad_group_bid_query = """
                        SELECT
                            campaign.id,
                            ad_group.cpc_bid_micros
                        FROM ad_group
                        WHERE campaign.status != 'REMOVED'
                            AND ad_group.status != 'REMOVED'
                    """
                    ad_group_bid_response = ga_service.search(customer_id=account['id'], query=ad_group_bid_query)
                    for row in ad_group_bid_response:
                        campaign_id = str(row.campaign.id)
                        if campaign_id not in campaign_info:
                            continue
                        cpc_bid_micros = int(row.ad_group.cpc_bid_micros or 0)
                        if cpc_bid_micros <= 0:
                            continue
                        preset_cpc_original = cpc_bid_micros / 1_000_000
                        preset_cpc_usd = preset_cpc_original * CNY_TO_USD_RATE if account_currency == 'CNY' else preset_cpc_original
                        if preset_cpc_usd > campaign_info[campaign_id].get('preset_cpc_usd', 0.0):
                            campaign_info[campaign_id]['preset_cpc_usd'] = preset_cpc_usd

                    # 用统计日期范围在campaign级别汇总花费，避免广告级查询把历史花费漏掉后写成0。
                    cost_query = """
                        SELECT
                            campaign.id,
                            campaign.target_spend.cpc_bid_ceiling_micros,
                            metrics.cost_micros,
                            metrics.clicks,
                            metrics.average_cpc
                        FROM campaign
                        WHERE campaign.status != 'REMOVED'
                    """
                    if start_date_str and end_date_str:
                        cost_query += f"""
                            AND segments.date >= '{start_date_str}'
                            AND segments.date <= '{end_date_str}'
                        """

                    cost_response = ga_service.search(customer_id=account['id'], query=cost_query)
                    for row in cost_response:
                        campaign_id = str(row.campaign.id)
                        if campaign_id not in campaign_info:
                            continue

                        cost_in_original = row.metrics.cost_micros / 1000000 if row.metrics.cost_micros else 0
                        preset_cpc_original = (row.campaign.target_spend.cpc_bid_ceiling_micros / 1000000) if row.campaign.target_spend.cpc_bid_ceiling_micros else 0
                        if account_currency == 'CNY':
                            cost_in_usd = cost_in_original * CNY_TO_USD_RATE
                            avg_cpc_in_usd = (row.metrics.average_cpc / 1000000) * CNY_TO_USD_RATE if row.metrics.average_cpc else 0
                            preset_cpc_usd = preset_cpc_original * CNY_TO_USD_RATE if preset_cpc_original else 0
                        else:
                            cost_in_usd = cost_in_original
                            avg_cpc_in_usd = (row.metrics.average_cpc / 1000000) if row.metrics.average_cpc else 0
                            preset_cpc_usd = preset_cpc_original
                        campaign_info[campaign_id]['cost_usd'] += cost_in_usd
                        campaign_info[campaign_id]['clicks'] += int(row.metrics.clicks or 0)
                        if avg_cpc_in_usd > 0:
                            campaign_info[campaign_id]['avg_cpc_usd'] = avg_cpc_in_usd
                        if preset_cpc_usd > campaign_info[campaign_id].get('preset_cpc_usd', 0.0):
                            campaign_info[campaign_id]['preset_cpc_usd'] = preset_cpc_usd

                    # 查询关键词层级指标，汇总“启用且有点击”的关键词CTR加权平均所需数据
                    keyword_query = """
                        SELECT
                            campaign.id,
                            metrics.impressions,
                            metrics.clicks,
                            metrics.ctr
                        FROM keyword_view
                        WHERE campaign.status != 'REMOVED'
                            AND ad_group.status != 'REMOVED'
                            AND ad_group_criterion.status = 'ENABLED'
                    """
                    if start_date_str and end_date_str:
                        keyword_query += f"""
                            AND segments.date >= '{start_date_str}'
                            AND segments.date <= '{end_date_str}'
                        """

                    keyword_response = ga_service.search(customer_id=account['id'], query=keyword_query)
                    for row in keyword_response:
                        campaign_id = str(row.campaign.id)
                        if campaign_id not in campaign_info:
                            continue

                        clicks = int(row.metrics.clicks or 0)
                        if clicks <= 0:
                            continue

                        impressions = int(row.metrics.impressions or 0)
                        ctr = float(row.metrics.ctr or 0)
                        campaign_info[campaign_id]['enabled_keyword_ctr_weighted_numerator'] += ctr * impressions
                        campaign_info[campaign_id]['enabled_keyword_ctr_impressions'] += impressions
                    
                    # 将set转为list方便后续处理
                    for cinfo in campaign_info.values():
                        cinfo['link_ids'] = list(cinfo.get('link_ids', set()))
                        cinfo['uids'] = list(cinfo.get('uids', set()))
                    campaigns.extend(campaign_info.values())
                    succeeded_account_ids.add(account['id'])
                        
                except Exception as e:
                    self.log_manage(f"  ⚠ 账户 {account['name']}({account['id']}) 获取失败: {str(e)[:100]}")
            
            if len(succeeded_account_ids) < len(sub_accounts):
                failed_count = len(sub_accounts) - len(succeeded_account_ids)
                self.log_manage(f"  ⚠ {failed_count} 个账户查询失败，这些账户的广告系列不会参与删除检测")
            
        except Exception as e:
            self.log_manage(f"  获取广告系列失败: {e}")
        
        return campaigns, succeeded_account_ids
    
    def _resolve_campaign_to_row_mapping(self, feishu_data, asin_country_campaigns, campaign_data):
        """将广告系列精确匹配到具体的offer行
        
        匹配策略（按优先级）：
        0. 产品链接匹配（最高优先级）：从offer行产品链接中提取aa_adgroupid，与广告系列的aa_adgroupid比较（一对一匹配）
        1. 混合平台分流：同ASIN+国家同时存在PB/YP时，未命中PB aa_adgroupid的campaign直接分配给YP行
        2. 直接uid匹配：从Google Ads final URL提取的uid与飞书投放链接的uid比较
        3. link_id匹配：通过PB API获取link_id，与Google Ads final URL中的aa_adgroupid比较
        
        返回:
            {row_index: [campaign_info, ...]} 精确匹配的映射
        """
        row_campaigns = {}
        
        # 1. 构建 uid → campaigns 和 link_id → campaigns 映射
        uid_to_campaigns = {}     # {uid: [campaign]}
        link_id_to_campaigns = {} # {link_id: [campaign]}
        has_uids = False
        has_link_ids = False
        
        for campaign in campaign_data:
            for uid in campaign.get('uids', []):
                if uid:
                    has_uids = True
                    if uid not in uid_to_campaigns:
                        uid_to_campaigns[uid] = []
                    uid_to_campaigns[uid].append(campaign)
            for lid in campaign.get('link_ids', []):
                if lid:
                    has_link_ids = True
                    if lid not in link_id_to_campaigns:
                        link_id_to_campaigns[lid] = []
                    link_id_to_campaigns[lid].append(campaign)
        
        if not has_uids and not has_link_ids:
            self.log_manage("    （广告系列中未找到uid或link_id，跳过精确匹配）")
            return row_campaigns
        
        self.log_manage(f"    广告系列标识: {len(uid_to_campaigns)}个uid, {len(link_id_to_campaigns)}个link_id")

        def normalize_link_cell(value):
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                return value[0].get('link', '') or value[0].get('text', '')
            return value if isinstance(value, str) else ''

        def extract_product_adgroupid(row):
            product_link = normalize_link_cell(row.get('产品链接', ''))
            if not product_link:
                return ''
            adgroupid_match = re.search(r'[?&]aa_adgroupid=([^&]+)', product_link)
            return adgroupid_match.group(1) if adgroupid_match else ''

        def extract_row_tracking_uid(row):
            tracking_link = normalize_link_cell(row.get('投放链接', ''))
            if not tracking_link:
                return ''
            row_uid = self.extract_tracking_uid_from_link(tracking_link)
            if row_uid:
                return row_uid
            return tracking_link[-7:] if len(tracking_link) >= 7 else ''

        def is_yp_offer_row(row):
            tracking_link = normalize_link_cell(row.get('投放链接', ''))
            if tracking_link and self._is_yp_tracking_link(tracking_link):
                return True
            product_link = normalize_link_cell(row.get('产品链接', ''))
            return not bool(product_link)
        
        # 2. 策略零（最高优先级）：通过产品链接中的aa_adgroupid直接匹配
        # 每个投放链接重定向后的产品链接包含唯一的aa_adgroupid参数，与广告系列后缀一致
        if has_link_ids:
            product_link_matched = 0
        for row in feishu_data:
            row_index = row.get('row_index')
            if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                continue
            if row_index in row_campaigns:
                continue

            product_link = row.get('产品链接', '')
            if isinstance(product_link, list) and len(product_link) > 0 and isinstance(product_link[0], dict):
                product_link = product_link[0].get('link', '') or product_link[0].get('text', '')
            if not isinstance(product_link, str) or not product_link:
                continue

            # 从产品链接URL中提取aa_adgroupid参数
            adgroupid_match = re.search(r'[?&]aa_adgroupid=([^&]+)', product_link)
            if not adgroupid_match:
                continue

            row_adgroupid = adgroupid_match.group(1)
            if row_adgroupid in link_id_to_campaigns:
                row_campaigns[row_index] = link_id_to_campaigns[row_adgroupid]
                product_link_matched += 1

        if has_link_ids and product_link_matched:
            self.log_manage(f"    ✓ 通过产品链接aa_adgroupid匹配: {product_link_matched} 行")
        
        # 3. 找出所有需要精确匹配的(ASIN, 国家)组合
        # 不仅处理多行offer的组合，还处理所有有广告系列的组合
        # 这样才能为广告系列表建立完整的 campaign_id -> tracking_link 映射
        asin_country_rows = {}
        for row in feishu_data:
            if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                continue
            asin = row.get('ASIN', '') or ''
            country = self.normalize_country_code(row.get('国家代码', ''))
            if not asin or not country:
                continue
            key = (asin, country)
            if key not in asin_country_rows:
                asin_country_rows[key] = []
            asin_country_rows[key].append(row)
        
        # 处理所有有广告系列的key，而不仅仅是多行的
        keys_to_match = {
            key for key, rows in asin_country_rows.items()
            if key in asin_country_campaigns
        }
        
        if not keys_to_match:
            return row_campaigns
        
        multi_row_count = sum(1 for k in keys_to_match if len(asin_country_rows[k]) > 1)
        self.log_manage(f"    需要精确匹配: {len(keys_to_match)} 个(ASIN+国家)组合 (其中{multi_row_count}个有多行offer)")

        # 4. 混合平台专用分流：
        # 对同ASIN+国家同时存在PB/YP的组合，先让PB通过产品链接aa_adgroupid命中；
        # 剩余未命中的campaign直接分给YP行，避免后续按(ASIN+国家)回退时被PB误吸收。
        mixed_platform_keys = set()
        mixed_platform_pb_exact_matches = 0
        mixed_platform_yp_assigned = 0
        for key in keys_to_match:
            rows = asin_country_rows[key]
            yp_rows = [row for row in rows if is_yp_offer_row(row)]
            pb_rows = [row for row in rows if not is_yp_offer_row(row)]
            if not yp_rows or not pb_rows:
                continue

            mixed_platform_keys.add(key)

            assigned_campaign_ids = set()
            for row in rows:
                for campaign in row_campaigns.get(row.get('row_index'), []):
                    cid = str(campaign.get('campaign_id', '') or '').strip()
                    if cid:
                        assigned_campaign_ids.add(cid)
            mixed_platform_pb_exact_matches += len(assigned_campaign_ids)

            yp_row_uid_map = {}
            for row in yp_rows:
                row_index = row.get('row_index')
                if not row_index:
                    continue
                row_uid = extract_row_tracking_uid(row)
                if row_uid:
                    yp_row_uid_map.setdefault(row_uid, []).append(row_index)

            empty_product_link_yp_rows = [
                row for row in yp_rows
                if not normalize_link_cell(row.get('产品链接', ''))
            ]

            for campaign in asin_country_campaigns.get(key, []):
                cid = str(campaign.get('campaign_id', '') or '').strip()
                if cid and cid in assigned_campaign_ids:
                    continue

                target_row_index = None
                for uid in campaign.get('uids', []) or []:
                    uid_rows = yp_row_uid_map.get(uid, [])
                    if len(uid_rows) == 1:
                        target_row_index = uid_rows[0]
                        break

                if target_row_index is None:
                    preferred_rows = empty_product_link_yp_rows or yp_rows
                    if preferred_rows:
                        target_row_index = preferred_rows[0].get('row_index')

                if not target_row_index:
                    continue

                row_campaigns.setdefault(target_row_index, []).append(campaign)
                if cid:
                    assigned_campaign_ids.add(cid)
                mixed_platform_yp_assigned += 1

        if mixed_platform_keys:
            self.log_manage(
                f"    ✓ 混合平台分流: {len(mixed_platform_keys)} 个组合, "
                f"PB aa_adgroupid命中 {mixed_platform_pb_exact_matches} 个广告系列, "
                f"未命中PB改分配到YP {mixed_platform_yp_assigned} 个广告系列"
            )

        # 5. 策略一：直接用uid匹配（最快，无需API调用）
        uid_matched = 0
        for key in keys_to_match:
            if key in mixed_platform_keys:
                continue
            rows = asin_country_rows[key]
            for row in rows:
                row_index = row.get('row_index')
                tracking_link = row.get('投放链接', '')
                
                # 提取投放链接中的uid
                if isinstance(tracking_link, list) and len(tracking_link) > 0 and isinstance(tracking_link[0], dict):
                    tracking_link = tracking_link[0].get('link', '') or tracking_link[0].get('text', '')
                if not isinstance(tracking_link, str) or not tracking_link:
                    continue
                
                row_uid = self.extract_tracking_uid_from_link(tracking_link)
                
                if not row_uid:
                    continue
                
                # 直接在uid_to_campaigns中查找
                if row_uid in uid_to_campaigns:
                    row_campaigns[row_index] = uid_to_campaigns[row_uid]
                    uid_matched += 1
        
        if uid_matched:
            self.log_manage(f"    ✓ 通过uid直接匹配: {uid_matched} 行")
        
        # 6. 策略二：对未匹配的行，通过PB API获取link_id匹配
        if has_link_ids:
            unmatched_keys = {
                key for key in keys_to_match
                if key not in mixed_platform_keys
                and any(row.get('row_index') not in row_campaigns for row in asin_country_rows[key])
            }
            
            if unmatched_keys:
                token = self.pb_token_var.get().strip()
                api_calls = 0
                link_matched = 0
                
                for key in unmatched_keys:
                    if self.stop_flag:
                        break
                    rows = asin_country_rows[key]
                    asin, country = key
                    
                    for row in rows:
                        row_index = row.get('row_index')
                        if row_index in row_campaigns:
                            continue  # 已经通过uid匹配了
                        
                        tracking_link = row.get('投放链接', '')
                        if isinstance(tracking_link, list) and len(tracking_link) > 0 and isinstance(tracking_link[0], dict):
                            tracking_link = tracking_link[0].get('link', '') or tracking_link[0].get('text', '')
                        if not isinstance(tracking_link, str) or not tracking_link:
                            continue
                        
                        row_uid = self.extract_tracking_uid_from_link(tracking_link)
                        if not row_uid and len(tracking_link) >= 7:
                            row_uid = tracking_link[-7:]
                        if not row_uid:
                            continue
                        
                        try:
                            api_url = f"{PB_API_BASE_URL}/api/datafeed/get_amazon_link_by_asin"
                            body = {
                                "token": token, "asins": asin,
                                "country_code": country, "uid": row_uid,
                                "return_partnerboost_link": 1
                            }
                            resp = requests.post(api_url, json=body, timeout=15)
                            api_calls += 1
                            data = resp.json()
                            
                            if data.get('status', {}).get('code') == 0:
                                link_data = data.get('data', [])
                                if link_data:
                                    link_id = link_data[0].get('link_id', '')
                                    if link_id and link_id in link_id_to_campaigns:
                                        row_campaigns[row_index] = link_id_to_campaigns[link_id]
                                        link_matched += 1
                            time.sleep(0.1)
                        except Exception as e:
                            self.log_debug(f"    解析link_id失败 ASIN={asin} uid={row_uid}: {str(e)[:80]}")
                
                if api_calls:
                    self.log_manage(f"    ✓ 通过link_id匹配: {link_matched} 行 ({api_calls}次API调用)")
        
        # 去重：确保每个广告系列只分配到一个offer行（一一对应原则）
        # 如果同一个广告系列被分配到多行（例如PB API对同ASIN+国家返回相同link_id），
        # 只保留第一个匹配的行，其余行移除该广告系列
        assigned_campaign_ids = {}  # {campaign_id: row_index} 记录已分配的广告系列
        rows_to_remove = []
        
        for row_index in sorted(row_campaigns.keys()):
            campaigns = row_campaigns[row_index]
            unassigned = []
            for c in campaigns:
                cid = c.get('campaign_id')
                if cid not in assigned_campaign_ids:
                    assigned_campaign_ids[cid] = row_index
                    unassigned.append(c)
            if unassigned:
                row_campaigns[row_index] = unassigned
            else:
                rows_to_remove.append(row_index)
        
        for ri in rows_to_remove:
            del row_campaigns[ri]
        
        if rows_to_remove:
            self.log_manage(f"    去重: 移除了 {len(rows_to_remove)} 个重复分配（确保一一对应）")
        
        total = len(row_campaigns)
        self.log_manage(f"    精确匹配总计: {total} 行")
        
        return row_campaigns

    def build_offer_product_link_to_tracking_link_map(self, feishu_data):
        """从Offer表构建产品链接到投放链接的一一映射。"""
        mapping = {}
        duplicates = {}
        for row in feishu_data or []:
            if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                continue
            product_link_key = self.normalize_product_link_key(row.get('产品链接', ''))
            tracking_link = str(self.normalize_sheet_cell_value(row.get('投放链接', '')) or '').strip()
            if not product_link_key or not tracking_link:
                continue
            if product_link_key in mapping and mapping[product_link_key] != tracking_link:
                duplicates.setdefault(product_link_key, set()).update([mapping[product_link_key], tracking_link])
                continue
            mapping[product_link_key] = tracking_link
        return mapping, duplicates

    def repair_campaign_tracking_links_from_product_links_once(self, token=None):
        """一次性按产品链接修复广告系列表“投放链接”。"""
        self.log_manage("\n【一次性修复】按产品链接修复广告系列表投放链接...")
        token = token or self.get_feishu_token()
        if not token:
            self.log_manage("  ✗ 获取飞书Token失败")
            return {'updated': 0, 'matched': 0, 'missing': 0, 'duplicates': 0}

        spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
        campaigns_sheet_id = CAMPAIGNS_SHEET_ID

        offer_rows = self.get_feishu_sheet_data(token)
        product_link_to_tracking_link, duplicates = self.build_offer_product_link_to_tracking_link_map(offer_rows)
        self.log_manage(f"  Offer表产品链接映射: {len(product_link_to_tracking_link)} 个")
        if duplicates:
            self.log_manage(f"  ⚠ 产品链接对应多个投放链接，已跳过冲突项: {len(duplicates)} 个")

        campaign_sheet_data = self.read_campaigns_sheet(token, spreadsheet_token, campaigns_sheet_id)
        if not campaign_sheet_data:
            self.log_manage("  ✗ 无法读取广告系列表")
            return {'updated': 0, 'matched': 0, 'missing': 0, 'duplicates': len(duplicates)}

        campaign_rows, column_map, _ = campaign_sheet_data
        tracking_col = column_map.get('投放链接')
        if not tracking_col:
            self.log_manage("  ✗ 广告系列表缺少“投放链接”列")
            return {'updated': 0, 'matched': 0, 'missing': 0, 'duplicates': len(duplicates)}

        tracking_col_idx = self.column_letter_to_index(tracking_col)
        updates = []
        matched = 0
        missing = 0

        for row in campaign_rows or []:
            status = str(row.get('状态', '') or '').strip()
            if status == SUMMARY_STATUS_TEXT:
                continue
            product_link_key = self.normalize_product_link_key(row.get('产品链接', ''))
            if not product_link_key:
                missing += 1
                continue
            tracking_link = product_link_to_tracking_link.get(product_link_key, '')
            if not tracking_link:
                missing += 1
                continue
            matched += 1
            current_link = str(self.normalize_sheet_cell_value(row.get('投放链接', '')) or '').strip()
            if current_link != tracking_link:
                updates.append((row.get('row_index'), tracking_col_idx, tracking_link))

        if updates:
            self._batch_update_sheet_cells(token, spreadsheet_token, updates, sheet_id=campaigns_sheet_id)
        self.log_manage(f"  匹配到 {matched} 行，需要修复 {len(updates)} 行，未命中 {missing} 行")
        self.log_manage("  ✅ 广告系列表投放链接一次性修复完成")
        return {
            'updated': len(updates),
            'matched': matched,
            'missing': missing,
            'duplicates': len(duplicates),
        }

    def get_removed_campaign_metrics(self, start_date_str=None, end_date_str=None, campaign_names=None):
        """一次性汇总已移除广告系列的点击数、真实CPC与预设CPC。"""
        removed_metrics = {}
        target_names = {str(name).strip() for name in (campaign_names or []) if str(name).strip()}
        if not target_names:
            return removed_metrics

        try:
            sub_accounts = self.iter_google_ads_mcc_accounts()

            CNY_TO_USD_RATE = 0.14
            matched_count = 0
            client_by_mcc = {}

            for account in sub_accounts:
                if self.stop_flag:
                    break

                try:
                    mcc_id = account.get('mcc_id', '')
                    client = client_by_mcc.get(mcc_id)
                    if client is None:
                        client = self.get_google_ads_client(mcc_id)
                        client_by_mcc[mcc_id] = client
                    if not client:
                        continue
                    ga_service = client.get_service('GoogleAdsService')
                    account_currency = account.get('currency', 'USD')
                    removed_query = """
                        SELECT
                            campaign.id,
                            campaign.name,
                            campaign.target_spend.cpc_bid_ceiling_micros,
                            metrics.clicks,
                            metrics.average_cpc
                        FROM campaign
                        WHERE campaign.status = 'REMOVED'
                    """
                    if start_date_str and end_date_str:
                        removed_query += f"""
                            AND segments.date >= '{start_date_str}'
                            AND segments.date <= '{end_date_str}'
                        """

                    removed_response = ga_service.search(customer_id=account['id'], query=removed_query)
                    for row in removed_response:
                        campaign_name = str(row.campaign.name or '').strip()
                        if not campaign_name or campaign_name not in target_names:
                            continue

                        clicks = int(row.metrics.clicks or 0)
                        avg_cpc_original = (row.metrics.average_cpc / 1000000) if row.metrics.average_cpc else 0.0
                        preset_cpc_original = (row.campaign.target_spend.cpc_bid_ceiling_micros / 1000000) if row.campaign.target_spend.cpc_bid_ceiling_micros else 0.0
                        avg_cpc_usd = avg_cpc_original * CNY_TO_USD_RATE if account_currency == 'CNY' else avg_cpc_original
                        preset_cpc_usd = preset_cpc_original * CNY_TO_USD_RATE if account_currency == 'CNY' else preset_cpc_original

                        if campaign_name not in removed_metrics:
                            removed_metrics[campaign_name] = {
                                'clicks': 0,
                                'cost_for_cpc': 0.0,
                                'preset_cpc_usd': 0.0,
                            }
                            matched_count += 1

                        removed_metrics[campaign_name]['clicks'] += clicks
                        removed_metrics[campaign_name]['cost_for_cpc'] += avg_cpc_usd * clicks
                        if preset_cpc_usd > removed_metrics[campaign_name].get('preset_cpc_usd', 0.0):
                            removed_metrics[campaign_name]['preset_cpc_usd'] = preset_cpc_usd
                except Exception as e:
                    self.log_manage(f"  ⚠ 账户 {account['name']}({account['id']}) 汇总已移除广告系列指标失败: {str(e)[:100]}")

            for campaign_name, item in removed_metrics.items():
                clicks = int(item.get('clicks', 0) or 0)
                total_cost = float(item.get('cost_for_cpc', 0.0) or 0.0)
                item['avg_cpc_usd'] = (total_cost / clicks) if clicks > 0 else 0.0

            self.log_manage(f"  已移除广告系列指标汇总完成: 命中 {matched_count} 个广告系列")
        except Exception as e:
            self.log_manage(f"  汇总已移除广告系列指标失败: {e}")

        return removed_metrics

    def get_all_campaigns_for_ads_brand_stats(self, start_date_str=None, end_date_str=None, recent_cost_start_str=None, recent_cost_end_str=None):
        """获取ads品牌表需要的所有广告系列，包含ENABLED/PAUSED/REMOVED。"""
        campaigns = []
        succeeded_account_ids = set()

        try:
            sub_accounts = self.iter_google_ads_mcc_accounts()

            CNY_TO_USD_RATE = 0.14
            self.log_manage(f"  ads品牌表扫描 {len(sub_accounts)} 个子账户")
            client_by_mcc = {}

            for account in sub_accounts:
                if self.stop_flag:
                    break

                try:
                    mcc_id = account.get('mcc_id', '')
                    client = client_by_mcc.get(mcc_id)
                    if client is None:
                        client = self.get_google_ads_client(mcc_id)
                        client_by_mcc[mcc_id] = client
                    if not client:
                        continue
                    ga_service = client.get_service('GoogleAdsService')
                    account_currency = account.get('currency', 'USD')
                    campaign_info = {}

                    base_campaign_query = """
                        SELECT
                            campaign.id,
                            campaign.name,
                            campaign.status,
                            campaign.target_spend.cpc_bid_ceiling_micros
                        FROM campaign
                    """
                    base_campaign_response = ga_service.search(customer_id=account['id'], query=base_campaign_query)
                    for row in base_campaign_response:
                        campaign_id = str(row.campaign.id)
                        campaign_name = str(row.campaign.name or '').strip()
                        brand, country = self.extract_brand_and_country_from_campaign_name(campaign_name)
                        preset_cpc_original = row.campaign.target_spend.cpc_bid_ceiling_micros / 1_000_000 if row.campaign.target_spend.cpc_bid_ceiling_micros else 0.0
                        preset_cpc_usd = preset_cpc_original * CNY_TO_USD_RATE if account_currency == 'CNY' else preset_cpc_original
                        campaign_info[campaign_id] = {
                            'mcc_id': mcc_id,
                            'mcc_name': account.get('mcc_name', ''),
                            'account_id': account['id'],
                            'account_name': account['name'],
                            'campaign_id': campaign_id,
                            'campaign_name': campaign_name,
                            'status': row.campaign.status.name,
                            'brand': brand or '',
                            'brand_key': self.normalize_brand_key(brand or ''),
                            'country': country or '',
                            'cost_usd': 0.0,
                            'recent_5_day_cost_usd': 0.0,
                            'clicks': 0,
                            'avg_cpc_usd': 0.0,
                            'preset_cpc_usd': preset_cpc_usd,
                        }

                    metric_query = """
                        SELECT
                            campaign.id,
                            campaign.target_spend.cpc_bid_ceiling_micros,
                            metrics.cost_micros,
                            metrics.clicks,
                            metrics.average_cpc
                        FROM campaign
                    """
                    if start_date_str and end_date_str:
                        metric_query += f"""
                            WHERE segments.date >= '{start_date_str}'
                                AND segments.date <= '{end_date_str}'
                        """

                    metric_response = ga_service.search(customer_id=account['id'], query=metric_query)
                    for row in metric_response:
                        campaign_id = str(row.campaign.id)
                        item = campaign_info.get(campaign_id)
                        if not item:
                            continue

                        cost_original = row.metrics.cost_micros / 1_000_000 if row.metrics.cost_micros else 0.0
                        avg_cpc_original = row.metrics.average_cpc / 1_000_000 if row.metrics.average_cpc else 0.0
                        preset_cpc_original = row.campaign.target_spend.cpc_bid_ceiling_micros / 1_000_000 if row.campaign.target_spend.cpc_bid_ceiling_micros else 0.0
                        if account_currency == 'CNY':
                            cost_usd = cost_original * CNY_TO_USD_RATE
                            avg_cpc_usd = avg_cpc_original * CNY_TO_USD_RATE
                            preset_cpc_usd = preset_cpc_original * CNY_TO_USD_RATE
                        else:
                            cost_usd = cost_original
                            avg_cpc_usd = avg_cpc_original
                            preset_cpc_usd = preset_cpc_original

                        item['cost_usd'] += cost_usd
                        item['clicks'] += int(row.metrics.clicks or 0)
                        if avg_cpc_usd > 0:
                            item['avg_cpc_usd'] = avg_cpc_usd
                        if preset_cpc_usd > item.get('preset_cpc_usd', 0.0):
                            item['preset_cpc_usd'] = preset_cpc_usd

                    if recent_cost_start_str and recent_cost_end_str:
                        recent_cost_query = f"""
                            SELECT
                                campaign.id,
                                metrics.cost_micros
                            FROM campaign
                            WHERE segments.date >= '{recent_cost_start_str}'
                                AND segments.date <= '{recent_cost_end_str}'
                        """
                        recent_cost_response = ga_service.search(customer_id=account['id'], query=recent_cost_query)
                        for row in recent_cost_response:
                            campaign_id = str(row.campaign.id)
                            item = campaign_info.get(campaign_id)
                            if not item:
                                continue

                            cost_original = row.metrics.cost_micros / 1_000_000 if row.metrics.cost_micros else 0.0
                            cost_usd = cost_original * CNY_TO_USD_RATE if account_currency == 'CNY' else cost_original
                            item['recent_5_day_cost_usd'] += cost_usd

                    campaigns.extend(campaign_info.values())
                    succeeded_account_ids.add(account['id'])
                except Exception as e:
                    self.log_manage(f"  ⚠ ads品牌表账户 {account['name']}({account['id']}) 查询失败: {str(e)[:100]}")
        except Exception as e:
            self.log_manage(f"  获取ads品牌表广告系列失败: {e}")
            return campaigns, succeeded_account_ids, []

        return campaigns, succeeded_account_ids, [
            {
                'account_id': account.get('id', ''),
                'account_name': account.get('name', ''),
                'mcc_id': account.get('mcc_id', ''),
                'mcc_name': account.get('mcc_name', ''),
            }
            for account in sub_accounts
        ]
    
    def extract_asin_from_url(self, url):
        """从最终到达网址中提取ASIN"""
        if not url:
            return None
        
        # ASIN格式：https://www.amazon.com/dp/B0F9NK4M7Q 或类似格式
        # 尝试多种匹配模式
        patterns = [
            r'/dp/([A-Z0-9]{10})',           # 标准格式 /dp/ASIN
            r'/gp/product/([A-Z0-9]{10})',   # 旧格式 /gp/product/ASIN
            r'/product/([A-Z0-9]{10})',      # 简化格式
            r'[?&]asin=([A-Z0-9]{10})',      # URL参数 ?asin=ASIN
            r'[?&]ASIN=([A-Z0-9]{10})',      # URL参数大写
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url, re.IGNORECASE)
            if match:
                return match.group(1).upper()
        
        return None
    
    def extract_country_from_campaign_name(self, campaign_name):
        """从广告系列名称中提取国家代码
        
        广告系列名称格式示例：2037-2589-anker-US-Search-20260122205857
        提取其中的国家代码（如US, DE, UK等）
        """
        if not campaign_name:
            return None
        
        # 常见国家代码列表
        country_codes = ['US', 'UK', 'GB', 'DE', 'FR', 'IT', 'ES', 'CA', 'JP', 'AU', 'NL', 'BE', 'MX', 'BR', 'IN', 'SG', 'AE', 'SA', 'PL', 'SE', 'TR', 'EG']
        
        # 尝试从广告系列名称中匹配国家代码
        # 格式：xxxx-xxxx-brand-COUNTRY-type-date
        parts = campaign_name.split('-')
        for part in parts:
            part_upper = part.upper()
            if part_upper in country_codes:
                return self.normalize_country_code(part_upper)
        
        # 也尝试用正则匹配，以防格式略有不同
        for code in country_codes:
            # 匹配被-包围的国家代码，不区分大小写
            pattern = rf'[-_]({code})[-_]'
            match = re.search(pattern, campaign_name, re.IGNORECASE)
            if match:
                return self.normalize_country_code(match.group(1).upper())
        
        return None

    def extract_brand_and_country_from_campaign_name(self, campaign_name):
        """从广告系列名称中提取品牌名和国家代码。

        示例:
            Miir_DE_5801_9711_20260425162956454 -> (Miir, DE)
            2037-2589-anker-US-Search-20260122205857 -> (anker, US)
        """
        if not campaign_name:
            return None, None

        country = self.extract_country_from_campaign_name(campaign_name)
        if not country:
            return None, None

        tokens = [token.strip() for token in re.split(r'[-_]', str(campaign_name)) if str(token).strip()]
        brand = None

        for idx, token in enumerate(tokens):
            if token.upper() != country:
                continue

            for left_idx in range(idx - 1, -1, -1):
                candidate = tokens[left_idx].strip()
                if not candidate:
                    continue
                if candidate.upper() == country:
                    continue
                if candidate.isdigit():
                    continue
                if not re.search(r'[A-Za-z]', candidate):
                    continue
                brand = candidate
                break
            if brand:
                break

        if not brand:
            for candidate in tokens:
                if candidate.upper() == country:
                    break
                if candidate.isdigit():
                    continue
                if not re.search(r'[A-Za-z]', candidate):
                    continue
                brand = candidate
                break

        return (brand, country) if brand and country else (None, country)
    
    def extract_country_from_merchant_name(self, merchant_name):
        """从品牌名称中提取国家代码
        
        品牌名称格式示例：DE-Anker, UK-Anker, Anker
        提取其中的国家代码（如 DE, UK 等）
        
        返回:
            提取到的国家代码（大写），如果没有则返回 None
        """
        if not merchant_name:
            return None
        
        # 常见国家代码列表
        country_codes = ['US', 'UK', 'DE', 'FR', 'IT', 'ES', 'CA', 'JP', 'AU', 'NL', 'BE', 'MX', 'BR', 'IN', 'SG', 'AE', 'SA', 'PL', 'SE', 'TR', 'EG']
        
        # 尝试匹配开头的国家代码（如 "DE-Anker" -> "DE"）
        # 格式：COUNTRY-BrandName
        parts = merchant_name.split('-')
        if len(parts) >= 2:
            first_part = parts[0].upper()
            if first_part in country_codes:
                return first_part
        
        # 也尝试匹配其他位置的国家代码
        for code in country_codes:
            # 匹配被-或_包围的国家代码，或在开头/结尾
            patterns = [
                rf'^({code})[-_]',      # 开头：DE-Anker
                rf'[-_]({code})$',      # 结尾：Anker-DE
                rf'[-_]({code})[-_]',   # 中间：Brand-DE-Name
            ]
            for pattern in patterns:
                match = re.search(pattern, merchant_name, re.IGNORECASE)
                if match:
                    return match.group(1).upper()
        
        return None

    def normalize_offer_brand_sort_key(self, brand_name):
        """Offer排序用品牌键：把“品牌名”和“品牌名+国家代码”视为同一品牌。"""
        raw_text = str(brand_name or '').strip()
        if not raw_text or raw_text.lower() in ('none', 'null'):
            return ''
        normalized = self.normalize_brand_key(brand_name)
        if normalized:
            return normalized

        text = raw_text.lower()
        text = re.sub(r'[\s\-_]+', ' ', text)
        return text.strip()

    def get_offer_sort_status_priority(self, status):
        """Offer顺序整理状态优先级：已结束、暂停中、投放中、新增、新复制。"""
        status_text = str(status or '').strip()
        if status_text.startswith('投放已结束'):
            return 0
        if status_text.startswith('暂停中') or status_text.startswith('暂停') or status_text.startswith('广告系列暂停中'):
            return 1
        if status_text.startswith('投放中'):
            return 2
        if status_text.startswith('新增'):
            return 3
        if status_text.startswith('新复制'):
            return 4
        return 5
    
    def extract_country_from_mcid(self, mcid):
        """从 mcid 中提取国家代码
        
        mcid 格式示例：amzdeanker, amzusanker, amzukanker
        其中 amz 是 Amazon 前缀，de/us/uk 是国家代码，anker 是品牌名
        
        返回:
            提取到的国家代码（大写），如果没有则返回 None
        """
        if not mcid:
            return None
        
        mcid_lower = mcid.lower()
        
        # 常见国家代码及其在 mcid 中的表示形式
        # mcid 格式：amz + 国家代码 + 品牌名
        country_patterns = {
            'de': 'DE',   # amzdeanker -> DE
            'us': 'US',   # amzusanker -> US
            'uk': 'UK',   # amzukanker -> UK
            'fr': 'FR',   # amzfranker -> FR
            'it': 'IT',   # amzitanker -> IT
            'es': 'ES',   # amzesanker -> ES
            'ca': 'CA',   # amzcanker -> CA (注意可能有歧义)
            'jp': 'JP',   # amzjpanker -> JP
            'au': 'AU',   # amzauanker -> AU
            'nl': 'NL',   # amznlanker -> NL
            'be': 'BE',   # amzbeanker -> BE
            'mx': 'MX',   # amzmxanker -> MX
            'br': 'BR',   # amzbranker -> BR
            'in': 'IN',   # amzinanker -> IN
            'sg': 'SG',   # amzsganker -> SG
            'ae': 'AE',   # amzaeanker -> AE
            'sa': 'SA',   # amzsaanker -> SA
            'pl': 'PL',   # amzplanker -> PL
            'se': 'SE',   # amzseanker -> SE
            'tr': 'TR',   # amztranker -> TR
            'eg': 'EG',   # amzeganker -> EG
        }
        
        # 检查 mcid 是否以 amz 开头
        if mcid_lower.startswith('amz'):
            # 尝试提取国家代码（amz 后面的 2 个字符）
            if len(mcid_lower) >= 5:
                potential_country = mcid_lower[3:5]
                if potential_country in country_patterns:
                    return country_patterns[potential_country]
        
        return None
    
    def get_feishu_sheet_data(self, token):
        """获取飞书表格数据"""
        spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
        sheet_id = self.feishu_sheet_id_var.get().strip()
        
        # 获取工作表元数据
        meta_url = f"{FEISHU_API_BASE_URL}/sheets/v3/spreadsheets/{spreadsheet_token}/sheets/{sheet_id}"
        headers = {"Authorization": f"Bearer {token}"}
        
        response = requests.get(meta_url, headers=headers, timeout=30)
        meta = response.json()
        row_count = meta.get("data", {}).get("sheet", {}).get("grid_properties", {}).get("row_count", 100)
        
        # 读取数据
        range_str = f"{sheet_id}!A1:Z{row_count}"  # 扩大范围以支持更多列
        data_url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/values/{range_str}"
        
        response = requests.get(data_url, headers=headers, timeout=30)
        data = response.json()
        
        values = data.get("data", {}).get("valueRange", {}).get("values", [])
        
        # 解析数据，跳过表头
        rows = []
        column_map = {}  # 表头名称 -> 列字母的映射
        
        if len(values) > 1:
            headers_row = values[0]
            
            # 构建列映射：表头名称 -> 列字母
            for j, header in enumerate(headers_row):
                if header:
                    col_letter = self.index_to_column_letter(j)
                    column_map[header] = col_letter
            
            for i, row in enumerate(values[1:], start=2):  # 从第2行开始
                row_data = {'row_index': i}
                for j, cell in enumerate(row):
                    if j < len(headers_row):
                        # 处理URL类型的单元格
                        if isinstance(cell, list) and len(cell) > 0 and isinstance(cell[0], dict):
                            cell = cell[0].get('link', '') or cell[0].get('text', '')
                        row_data[headers_row[j]] = cell
                rows.append(row_data)
        
        # 保存列映射供写入时使用
        self.feishu_column_map = column_map
        
        return rows
    
    def index_to_column_letter(self, index):
        """将列索引（0开始）转换为Excel列字母（A, B, ..., Z, AA, AB, ...）"""
        result = ""
        while index >= 0:
            result = chr(index % 26 + ord('A')) + result
            index = index // 26 - 1
        return result

    def column_letter_to_index(self, column_letter):
        """将Excel列字母转换为列索引（0开始）。"""
        if not column_letter:
            return None
        index = 0
        for ch in str(column_letter).strip().upper():
            if not ('A' <= ch <= 'Z'):
                return None
            index = index * 26 + (ord(ch) - ord('A') + 1)
        return index - 1
    
    def load_old_campaign_consume(self):
        """读取旧广告系列花费记录
        
        返回:
            dict: {(asin, country): {'old_cost': float, 'last_updated': str, 'brand_name': str}}
        """
        old_consume_data = {}
        
        if not os.path.exists(OLD_CAMPAIGN_CONSUME_FILE):
            return old_consume_data
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(OLD_CAMPAIGN_CONSUME_FILE)
            ws = wb.active
            
            # 跳过表头，从第2行开始读取
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # ASIN 和 国家代码 不为空
                    asin = str(row[0]).strip()
                    country = str(row[1]).strip().upper()
                    old_cost = float(row[2]) if row[2] else 0
                    brand_name = str(row[3]) if row[3] else ''
                    last_updated = str(row[4]) if row[4] else ''
                    
                    key = (asin, country)
                    old_consume_data[key] = {
                        'old_cost': old_cost,
                        'brand_name': brand_name,
                        'last_updated': last_updated
                    }
            
            wb.close()
        except Exception as e:
            self.log_manage(f"  读取旧广告系列花费文件失败: {e}")
        
        return old_consume_data
    
    def save_old_campaign_consume(self, old_consume_data):
        """保存旧广告系列花费记录
        
        参数:
            old_consume_data: dict, {(asin, country): {'old_cost': float, 'last_updated': str, 'brand_name': str}}
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "旧广告系列花费"
            
            # 写入表头
            headers = ['ASIN', '国家代码', '旧广告系列累计花费', '品牌名称', '最后更新时间']
            ws.append(headers)
            
            # 写入数据
            for (asin, country), data in old_consume_data.items():
                ws.append([
                    asin,
                    country,
                    data.get('old_cost', 0),
                    data.get('brand_name', ''),
                    data.get('last_updated', '')
                ])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            
            wb.save(OLD_CAMPAIGN_CONSUME_FILE)
            self.log_manage(f"  旧广告系列花费已保存到: {OLD_CAMPAIGN_CONSUME_FILE}")
        except Exception as e:
            self.log_manage(f"  保存旧广告系列花费文件失败: {e}")
    
    def add_old_campaign_cost(self, old_consume_data, asin, country, cost_to_add, brand_name=''):
        """添加或累加旧广告系列花费
        
        参数:
            old_consume_data: dict, 旧花费数据字典
            asin: str, ASIN码
            country: str, 国家代码
            cost_to_add: float, 要累加的花费
            brand_name: str, 品牌名称
        """
        key = (asin, self.normalize_country_code(country))
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if key in old_consume_data:
            # 累加花费
            old_consume_data[key]['old_cost'] += cost_to_add
            old_consume_data[key]['last_updated'] = current_time
            if brand_name:
                old_consume_data[key]['brand_name'] = brand_name
        else:
            # 新增记录
            old_consume_data[key] = {
                'old_cost': cost_to_add,
                'brand_name': brand_name,
                'last_updated': current_time
            }
    
    def load_campaign_cost_snapshot(self):
        """读取广告系列花费快照
        
        返回:
            dict: {campaign_id: {'asin': str, 'country': str, 'cost': float, 'campaign_name': str}}
        """
        snapshot = {}
        
        if not os.path.exists(CAMPAIGN_COST_SNAPSHOT_FILE):
            return snapshot
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(CAMPAIGN_COST_SNAPSHOT_FILE)
            ws = wb.active
            
            # 跳过表头，从第2行开始读取
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # campaign_id 不为空
                    campaign_id = str(row[0]).strip()
                    snapshot[campaign_id] = {
                        'asin': str(row[1]).strip() if row[1] else '',
                        'country': str(row[2]).strip().upper() if row[2] else '',
                        'cost': float(row[3]) if row[3] else 0,
                        'campaign_name': str(row[4]) if row[4] else '',
                        'account_id': str(row[5]) if row[5] else ''
                    }
            
            wb.close()
        except Exception as e:
            self.log_manage(f"  读取广告系列快照文件失败: {e}")
        
        return snapshot
    
    def save_campaign_cost_snapshot(self, snapshot):
        """保存广告系列花费快照
        
        参数:
            snapshot: dict, {campaign_id: {'asin': str, 'country': str, 'cost': float, 'campaign_name': str}}
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "广告系列花费快照"
            
            # 写入表头
            headers = ['广告系列ID', 'ASIN', '国家代码', '花费(USD)', '广告系列名称', '账户ID']
            ws.append(headers)
            
            # 写入数据
            for campaign_id, data in snapshot.items():
                ws.append([
                    campaign_id,
                    data.get('asin', ''),
                    data.get('country', ''),
                    data.get('cost', 0),
                    data.get('campaign_name', ''),
                    data.get('account_id', '')
                ])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 15
            
            wb.save(CAMPAIGN_COST_SNAPSHOT_FILE)
        except Exception as e:
            self.log_manage(f"  保存广告系列快照文件失败: {e}")
    
    def get_all_commissions(self, start_date_str, end_date_str):
        """获取指定日期范围内的佣金数据"""
        all_transactions = []
        url = f"{PB_API_BASE_URL}/api.php?mod=medium&op=transaction"
        token = self.pb_token_var.get().strip()
        
        start_from = datetime.strptime(start_date_str, '%Y-%m-%d')
        current_end = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        self.log_manage(f"  查询佣金数据，日期范围: {start_date_str} 到 {end_date_str}")
        
        api_calls = 0
        api_errors = 0
        
        while current_end >= start_from:
            if self.stop_flag:
                break
            
            current_begin = current_end - timedelta(days=60)
            if current_begin < start_from:
                current_begin = start_from
            
            begin_str = current_begin.strftime('%Y-%m-%d')
            end_str = current_end.strftime('%Y-%m-%d')
            
            page = 1
            chunk_actual_count = 0
            chunk_api_total_count = None
            chunk_api_total_page = None
            chunk_response_pages = 0
            while True:
                body = {
                    'token': token,
                    'begin_date': begin_str,
                    'end_date': end_str,
                    'type': 'json',
                    'status': 'All',
                    'limit': 2000,
                    'page': page
                }
                
                try:
                    response = requests.post(url, json=body, timeout=60)
                    data = response.json()
                    api_calls += 1
                    
                    status_code = data.get('status', {}).get('code')
                    if status_code != 0:
                        error_msg = data.get('status', {}).get('message', 'Unknown error')
                        self.log_manage(f"  API返回错误 ({begin_str}~{end_str}): code={status_code}, msg={error_msg}")
                        api_errors += 1
                        break
                    
                    transactions = data.get('data', {}).get('list', [])
                    total_page = int(data.get('data', {}).get('total_page', 0))
                    total_count = int(data.get('data', {}).get('total_count', 0))

                    if page == 1:
                        chunk_api_total_count = total_count
                        chunk_api_total_page = total_page if total_page > 0 else 1

                    chunk_actual_count += len(transactions)
                    chunk_response_pages += 1
                    all_transactions.extend(transactions)
                    
                    if page >= total_page or not transactions:
                        break
                    page += 1
                    
                except Exception as e:
                    self.log_manage(f"  API请求异常 ({begin_str}~{end_str}): {str(e)[:100]}")
                    api_errors += 1
                    break

            if chunk_response_pages > 0:
                reported_pages = chunk_api_total_page if chunk_api_total_page else chunk_response_pages
                if chunk_api_total_count is not None and chunk_api_total_count != chunk_actual_count:
                    self.log_manage(
                        f"  {begin_str}~{end_str}: API声称共{chunk_api_total_count}条，"
                        f"{reported_pages}页；实际获取{chunk_actual_count}条"
                    )
                else:
                    self.log_manage(
                        f"  {begin_str}~{end_str}: 实际获取{chunk_actual_count}条，{reported_pages}页"
                    )
            
            current_end = current_begin - timedelta(days=1)
        
        self.log_manage(f"  API调用统计: {api_calls}次请求, {api_errors}次错误")
        
        return all_transactions

    def get_all_yp_commissions(self, start_date_str, end_date_str):
        """获取指定日期范围内的YeahPromos佣金数据。"""
        yp_token = self.yp_token_var.get().strip()
        yp_site_id = self.yp_site_id_var.get().strip()
        if not yp_token or not yp_site_id:
            self.log_manage("  未配置YP Token或Site ID，跳过YP佣金")
            return []

        all_transactions = []
        page = 1
        total_pages = None
        api_calls = 0
        api_errors = 0
        url = f"{YP_API_BASE_URL}/index/Getorder/getorder"
        headers = {"token": yp_token}

        self.log_manage(f"  查询YP佣金数据，日期范围: {start_date_str} 到 {end_date_str}")

        while True:
            if self.stop_flag:
                break

            params = {
                "site_id": yp_site_id,
                "startDate": start_date_str,
                "endDate": end_date_str,
                "is_amazon": 1,
                "page": page,
                "limit": 1000,
            }
            try:
                response = requests.get(url, headers=headers, params=params, timeout=60)
                data = response.json()
                api_calls += 1

                if str(data.get("code", "")) != "100000":
                    msg = data.get("msg") or data.get("message") or data.get("status") or "Unknown error"
                    self.log_manage(f"  YP API返回错误: {msg}")
                    api_errors += 1
                    break

                payload = data.get('data')
                items = []
                if isinstance(payload, list):
                    items = payload
                elif isinstance(payload, dict):
                    for key in ('Data', 'data', 'list', 'orders', 'rows'):
                        value = payload.get(key)
                        if isinstance(value, list):
                            items = value
                            break
                    if total_pages is None:
                        total_pages = int(payload.get('PageTotal') or payload.get('pageTotal') or payload.get('total_page') or payload.get('totalPage') or 0)
                if total_pages is None:
                    total_pages = int(data.get('PageTotal') or data.get('pageTotal') or 0)

                all_transactions.extend(items)
                if not items:
                    break
                if total_pages and page >= total_pages:
                    break
                page += 1
            except Exception as e:
                self.log_manage(f"  YP API请求异常: {str(e)[:100]}")
                api_errors += 1
                break

        self.log_manage(f"  YP API调用统计: {api_calls}次请求, {api_errors}次错误")
        self.log_manage(f"  获取到 {len(all_transactions)} 条YP交易")
        return all_transactions
    
    def calculate_updates(self, feishu_data, asin_country_campaigns, asin_country_commission, asin_only_commission=None, asin_country_uid_commission=None, asin_country_no_uid_commission=None, row_campaigns=None, yp_asin_brand_commission=None):
        """计算需要更新的内容
        
        参数:
            feishu_data: 飞书表格数据
            asin_country_campaigns: {(asin, country): [campaign_info, ...]} 广告系列映射
            asin_country_commission: {(asin, country): total_commission} 所有佣金总映射（用于总计）
            asin_only_commission: {asin: total_commission} 无uid且国家默认US的佣金（ASIN-only备选匹配）
            asin_country_uid_commission: {(asin, country, uid): total_commission} 有uid的佣金（UID池）
            asin_country_no_uid_commission: {(asin, country): total_commission} 无uid的佣金（非UID池）
            row_campaigns: {row_index: [campaign_info, ...]} 通过uid/link_id精确匹配的行级广告系列映射
        
        返回:
            (updates, commission_context):
                updates: 更新列表
                commission_context: offer行级佣金归因结果，供广告系列表复用
        """
        if asin_only_commission is None:
            asin_only_commission = {}
        if asin_country_uid_commission is None:
            asin_country_uid_commission = {}
        if asin_country_no_uid_commission is None:
            asin_country_no_uid_commission = {}
        if row_campaigns is None:
            row_campaigns = {}
        if yp_asin_brand_commission is None:
            yp_asin_brand_commission = {}
        updates = []
        offer_row_commissions = {}
        
        # 构建已被精确匹配认领的广告系列ID集合
        # 用于在回退到(ASIN+国家)匹配时，排除已认领的广告系列
        claimed_campaign_ids = set()
        claimed_campaign_ids_by_key = {}
        for row_cams in row_campaigns.values():
            for c in row_cams:
                cid = c.get('campaign_id')
                if not cid:
                    continue
                claimed_campaign_ids.add(cid)
                campaign_key = (
                    str(c.get('asin', '') or '').strip(),
                    self.normalize_country_code(c.get('country', '') or '')
                )
                if campaign_key[0] and campaign_key[1]:
                    claimed_campaign_ids_by_key.setdefault(campaign_key, set()).add(cid)
        
        # 第一遍遍历：构建每个(ASIN, 国家)组合的状态分布
        # 用于判断佣金应该更新到哪一行
        # active_rows: 正在投放的行（投放中、广告系列暂停中等）
        # ended_rows: 投放已结束的行
        # untested_rows: 未测试的行
        asin_country_rows = {}  # {(asin, country): {'active_rows': [...], 'ended_rows': [...], 'untested_rows': [...]}}
        
        # 统计每个ASIN在飞书表格中出现的次数（用于判断是否有多个相同ASIN的offer）
        asin_row_count = {}  # {asin: count}
        # 记录每个ASIN有哪些国家的offer
        asin_countries = {}  # {asin: set(country1, country2, ...)}
        # 记录哪些ASIN有US offer
        asin_has_us_offer = set()
        def normalize_link_value(value):
            """标准化飞书中的链接单元格为纯字符串链接。"""
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                return value[0].get('link', '') or value[0].get('text', '')
            return value if isinstance(value, str) else ''

        def extract_tracking_uid(value):
            """优先提取链接中的uid参数，缺失时回退到末尾7位。"""
            link = normalize_link_value(value)
            if not link:
                return ''
            uid = self.extract_tracking_uid_from_link(link)
            if uid:
                return uid
            return link[-7:] if len(link) >= 7 else link

        # 记录每行的uid
        row_uid_map = {}  # {row_index: uid}
        row_is_yp_map = {}  # {row_index: bool}
        asin_country_platforms = {}  # {(asin, country): {'pb', 'yp'}}
        exact_uid_row_candidates = {}  # {(asin, country, uid): [row_index, ...]}
        # 记录每个(asin, country)有多少个不同的投放链接（用于判断是否有复制的offer）
        asin_country_uids = {}  # {(asin, country): [uid1, uid2, ...]}
        
        for row in feishu_data:
            asin = row.get('ASIN', '')
            country = row.get('国家代码', '')
            current_status = row.get('状态', '')
            row_index = row.get('row_index')
            tracking_link = row.get('投放链接', '')
            
            if not asin or str(current_status or '').strip() == SUMMARY_STATUS_TEXT:
                continue
            
            # 提取投放链接的uid
            row_uid = extract_tracking_uid(tracking_link)
            row_uid_map[row_index] = row_uid
            row_is_yp = self._is_yp_tracking_link(normalize_link_value(tracking_link))
            row_is_yp_map[row_index] = row_is_yp
            
            # 统计每个ASIN的出现次数
            if asin not in asin_row_count:
                asin_row_count[asin] = 0
            asin_row_count[asin] += 1
            
            # 标准化国家代码
            if country:
                country = self.normalize_country_code(country)
            
            # 记录每个ASIN有哪些国家
            if country:
                if asin not in asin_countries:
                    asin_countries[asin] = set()
                asin_countries[asin].add(country)
                
                # 记录有US offer的ASIN
                if country == 'US':
                    asin_has_us_offer.add(asin)
            
            key = (asin, country) if country else None
            if not key:
                continue

            asin_country_platforms.setdefault(key, set()).add('yp' if row_is_yp else 'pb')
            
            # 记录每个(asin, country)的uid列表
            if key not in asin_country_uids:
                asin_country_uids[key] = []
            if row_uid:
                asin_country_uids[key].append(row_uid)
                exact_uid_row_candidates.setdefault((asin, country, row_uid), []).append(row_index)
            
            if key not in asin_country_rows:
                asin_country_rows[key] = {'active_rows': [], 'ended_rows': [], 'untested_rows': []}
            
            if current_status == '投放已结束':
                asin_country_rows[key]['ended_rows'].append(row_index)
            elif current_status == '未测试' or current_status == '新复制' or not current_status:
                asin_country_rows[key]['untested_rows'].append(row_index)
            else:
                # 投放中、广告系列暂停中等活跃状态
                asin_country_rows[key]['active_rows'].append(row_index)
        
        # 记录UID精确匹配失败后，回退到(ASIN+国家)的分配情况
        uid_fallback_events = []
        unresolved_uid_commissions = []
        preserved_offer_cost_rows = []
        precomputed_row_commissions = {}

        mixed_platform_keys = {
            key for key, platforms in asin_country_platforms.items()
            if len(platforms) > 1
        }

        def get_fallback_row_for_key(key, preferred_platform=None):
            rows_info = asin_country_rows.get(key, {})
            for bucket in ('active_rows', 'ended_rows', 'untested_rows'):
                bucket_rows = rows_info.get(bucket, [])
                if preferred_platform:
                    platform_rows = [
                        row_index for row_index in bucket_rows
                        if ('yp' if row_is_yp_map.get(row_index) else 'pb') == preferred_platform
                    ]
                    if platform_rows:
                        return platform_rows[0]
                elif bucket_rows:
                    return bucket_rows[0]
            return None

        def get_or_create_row_commission(row_index, asin, country):
            if row_index not in precomputed_row_commissions:
                precomputed_row_commissions[row_index] = {
                    'row_index': row_index,
                    'asin': asin,
                    'country': country,
                    'total': 0.0,
                    'used_asin_only_match': False,
                    'uid_allocations': [],
                    'non_uid_commission': 0.0,
                    'asin_only_commission': 0.0,
                    'yp_commission': 0.0,
                    'yp_gross_commission': 0.0,
                    'yp_rejected_commission': 0.0,
                }
            return precomputed_row_commissions[row_index]

        for uid_key, commission_value in asin_country_uid_commission.items():
            asin, country, uid = uid_key
            exact_rows = exact_uid_row_candidates.get(uid_key, [])

            if len(exact_rows) == 1:
                target_row = exact_rows[0]
                result = get_or_create_row_commission(target_row, asin, country)
                result['total'] += commission_value
                result['uid_allocations'].append({
                    'uid': uid,
                    'commission': commission_value,
                    'match_type': 'uid_exact'
                })
                continue

            fallback_row = get_fallback_row_for_key((asin, country), preferred_platform='pb')
            if fallback_row is not None:
                result = get_or_create_row_commission(fallback_row, asin, country)
                result['total'] += commission_value
                result['uid_allocations'].append({
                    'uid': uid,
                    'commission': commission_value,
                    'match_type': 'asin_country_fallback'
                })
                uid_fallback_events.append({
                    'row_index': fallback_row,
                    'asin': asin,
                    'country': country,
                    'uids': [uid],
                    'commission': commission_value,
                    'exact_match_count': len(exact_rows)
                })
            else:
                unresolved_uid_commissions.append((uid_key, commission_value))

        allocated_no_uid_keys = set()
        for key, commission_value in asin_country_no_uid_commission.items():
            fallback_row = get_fallback_row_for_key(key, preferred_platform='pb')
            if fallback_row is None:
                continue
            result = get_or_create_row_commission(fallback_row, key[0], key[1])
            result['total'] += commission_value
            result['non_uid_commission'] += commission_value
            allocated_no_uid_keys.add(key)

        asin_candidate_keys = {}
        for key in asin_country_rows.keys():
            asin_candidate_keys.setdefault(key[0], []).append(key)
        for asin, commission_value in asin_only_commission.items():
            if asin in asin_has_us_offer or (asin, 'US') in allocated_no_uid_keys:
                continue
            target_key = None
            target_row = None
            for key in asin_candidate_keys.get(asin, []):
                row_index = get_fallback_row_for_key(key, preferred_platform='pb')
                if row_index is not None:
                    target_key = key
                    target_row = row_index
                    break
            if target_key is None or target_row is None:
                continue
            result = get_or_create_row_commission(target_row, target_key[0], target_key[1])
            result['total'] += commission_value
            result['asin_only_commission'] += commission_value
            result['used_asin_only_match'] = True

        yp_duplicate_groups = {}
        for row in feishu_data:
            row_index = row.get('row_index')
            if not row_index or row_index <= 2:
                continue
            status = str(row.get('状态', '') or '').strip()
            if status == SUMMARY_STATUS_TEXT:
                continue
            asin = str(row.get('ASIN', '') or '').strip()
            brand_id = self.build_yp_row_brand_key(row)
            country = self.normalize_country_code(row.get('国家代码', '') or '')
            tracking_link = str(normalize_link_value(row.get('投放链接', '')) or '').strip()
            if not asin or not brand_id or not self._is_yp_tracking_link(tracking_link):
                continue
            yp_marker = self.extract_yp_offer_marker(tracking_link, row.get('产品链接', ''))
            yp_duplicate_groups.setdefault((asin, brand_id, country, yp_marker), []).append(row_index)

        for yp_key, yp_info in yp_asin_brand_commission.items():
            row_indices = yp_duplicate_groups.get(yp_key, [])
            if not row_indices and len(yp_key) >= 2:
                fallback_matches = [
                    rows for key, rows in yp_duplicate_groups.items()
                    if key[0] == yp_key[0] and key[1] == yp_key[1]
                ]
                if len(fallback_matches) == 1:
                    row_indices = fallback_matches[0]
            if not row_indices:
                continue
            target_row = min(row_indices)
            target_row_data = next((r for r in feishu_data if r.get('row_index') == target_row), None)
            if not target_row_data:
                continue
            target_country = self.normalize_country_code(target_row_data.get('国家代码', '') or '')
            result = get_or_create_row_commission(target_row, yp_key[0], target_country)
            result['total'] += yp_info.get('non_rejected', 0.0)
            result['yp_commission'] += yp_info.get('non_rejected', 0.0)
            result['yp_gross_commission'] += yp_info.get('gross', 0.0)
            result['yp_rejected_commission'] += yp_info.get('rejected', 0.0)

        def get_commission_for_row(asin, country, row_index, is_first_offer):
            """获取某行的佣金归因结果。"""
            result = precomputed_row_commissions.get(row_index)
            if not result:
                return {
                    'total': None,
                    'used_asin_only_match': False,
                    'uid_allocations': [],
                    'non_uid_commission': 0.0,
                    'asin_only_commission': 0.0,
                    'yp_commission': 0.0,
                    'yp_gross_commission': 0.0,
                    'yp_rejected_commission': 0.0,
                }

            return {
                'total': result['total'],
                'used_asin_only_match': result['used_asin_only_match'],
                'uid_allocations': list(result.get('uid_allocations', [])),
                'non_uid_commission': result.get('non_uid_commission', 0.0) or 0.0,
                'asin_only_commission': result.get('asin_only_commission', 0.0) or 0.0,
                'yp_commission': result.get('yp_commission', 0.0) or 0.0,
                'yp_gross_commission': result.get('yp_gross_commission', 0.0) or 0.0,
                'yp_rejected_commission': result.get('yp_rejected_commission', 0.0) or 0.0,
            }

        def apply_mixed_platform_yp_cleanup(update):
            update['ads_ids'] = ''
            update['campaign_count'] = ''
            update['campaign_names'] = ''
            update['clear_total_cost'] = True

        def store_offer_row_commission(row_data, asin, country, row_index, commission_result, campaigns=None):
            """记录offer行级佣金归因结果，供广告系列表复用。"""
            commission_value = commission_result.get('total')

            campaign_names = []
            campaign_ids = []
            for campaign in campaigns or []:
                campaign_name = campaign.get('campaign_name', '')
                campaign_id = campaign.get('campaign_id', '')
                if campaign_name:
                    campaign_names.append(campaign_name)
                if campaign_id:
                    campaign_ids.append(campaign_id)

            # 对于没有当前广告系列匹配的行，保留offer表中原有的广告系列名称作为弱回退线索
            if not campaign_names:
                existing_campaign_names = row_data.get('广告系列名称', '')
                if isinstance(existing_campaign_names, str) and existing_campaign_names.strip():
                    campaign_names = [name.strip() for name in existing_campaign_names.split(',') if name.strip()]

            if commission_value is None and not campaign_names and not campaign_ids:
                return

            offer_row_commissions[row_index] = {
                'row_index': row_index,
                'asin': asin,
                'country': country,
                'brand_id': str(row_data.get('品牌ID', '') or '').strip(),
                'brand_name': str(row_data.get('品牌名称', '') or '').strip(),
                'is_yp': self._is_yp_tracking_link(normalize_link_value(row_data.get('投放链接', ''))),
                'commission': round(commission_value or 0.0, 2),
                'tracking_link': normalize_link_value(row_data.get('投放链接', '')),
                'product_link': normalize_link_value(row_data.get('产品链接', '')),
                'campaign_names': campaign_names,
                'campaign_ids': campaign_ids,
                'uid_allocations': list(commission_result.get('uid_allocations', [])),
                'non_uid_commission': commission_result.get('non_uid_commission', 0.0) or 0.0,
                'asin_only_commission': commission_result.get('asin_only_commission', 0.0) or 0.0,
                'used_asin_only_match': commission_result.get('used_asin_only_match', False),
                'yp_commission': commission_result.get('yp_commission', 0.0) or 0.0,
                'yp_gross_commission': commission_result.get('yp_gross_commission', 0.0) or 0.0,
                'yp_rejected_commission': commission_result.get('yp_rejected_commission', 0.0) or 0.0,
            }
        
        # 记录每个(asin, country)已处理的行数，用于判断是否是第一个offer
        asin_country_processed = {}  # {(asin, country): processed_count}
        
        # 调试日志：检查广告系列和offer表格的匹配情况
        self.log_debug("=== 广告系列与offer表格匹配检查 ===")
        for campaign_key in asin_country_campaigns.keys():
            if campaign_key in asin_country_rows:
                rows_info = asin_country_rows[campaign_key]
                self.log_debug(f"  ✓ {campaign_key[0]}_{campaign_key[1]}: 找到匹配 (active={len(rows_info['active_rows'])}, ended={len(rows_info['ended_rows'])}, untested={len(rows_info['untested_rows'])})")
            else:
                self.log_debug(f"  ✗ {campaign_key[0]}_{campaign_key[1]}: 未在offer表格中找到")
        
        mixed_platform_fallback_skips = 0

        # 第二遍遍历：处理更新
        for row in feishu_data:
            row_index = row.get('row_index')
            asin = row.get('ASIN', '')
            country = row.get('国家代码', '')
            current_status = row.get('状态', '')
            
            if not asin or str(current_status or '').strip() == SUMMARY_STATUS_TEXT:
                continue
            
            # 标准化国家代码
            if country:
                country = self.normalize_country_code(country)
            
            key = (asin, country) if country else None
            row_is_yp = row_is_yp_map.get(row_index, False)
            
            # 记录当前行是否是该(asin, country)的第一个offer
            if key:
                if key not in asin_country_processed:
                    asin_country_processed[key] = 0
                asin_country_processed[key] += 1
                is_first_offer = (asin_country_processed[key] == 1)
            else:
                is_first_offer = True
            
            # 处理"投放已结束"状态的offer
            if current_status == '投放已结束':
                # 首先检查是否有新的广告系列，如果有，则不跳过，继续后面的处理让状态变成"投放中"
                if key and key in asin_country_campaigns:
                    # 有广告系列，不跳过，继续后面的正常处理流程
                    pass
                else:
                    # 没有广告系列，只更新佣金
                    # 使用辅助函数获取佣金
                    commission_result = get_commission_for_row(asin, country, row_index, is_first_offer)
                    commission_value = commission_result.get('total')
                    used_asin_only_match = commission_result.get('used_asin_only_match', False)
                    needs_mixed_yp_cleanup = bool(row_is_yp and key in mixed_platform_keys)

                    if commission_value is not None:
                        commission_display = round(commission_value, 2)
                        if used_asin_only_match and len(asin_countries.get(asin, set())) > 1:
                            commission_display = f"{commission_display}*"
                        
                        update = {
                            'row_index': row_index,
                            'asin': asin,
                            'country': country,
                            'commission': commission_display
                        }
                        if needs_mixed_yp_cleanup:
                            apply_mixed_platform_yp_cleanup(update)
                        updates.append(update)
                        store_offer_row_commission(row, asin, country, row_index, commission_result, [])
                    elif needs_mixed_yp_cleanup:
                        update = {'row_index': row_index, 'asin': asin, 'country': country}
                        apply_mixed_platform_yp_cleanup(update)
                        updates.append(update)
                    # 没有广告系列，跳过其他处理
                    continue
            
            # 处理"未测试"或"新复制"状态的offer（没有广告系列的情况）
            if (current_status == '未测试' or current_status == '新复制' or not current_status):
                # 检查是否有广告系列
                if key and key in asin_country_campaigns:
                    # 有广告系列，继续后面的正常处理流程
                    pass
                else:
                    # 没有广告系列，检查是否需要更新佣金
                    # 使用辅助函数获取佣金
                    commission_result = get_commission_for_row(asin, country, row_index, is_first_offer)
                    commission_value = commission_result.get('total')
                    used_asin_only_match = commission_result.get('used_asin_only_match', False)
                    needs_mixed_yp_cleanup = bool(row_is_yp and key in mixed_platform_keys)

                    if commission_value is not None:
                        commission_display = round(commission_value, 2)
                        if used_asin_only_match and len(asin_countries.get(asin, set())) > 1:
                            commission_display = f"{commission_display}*"
                        
                        update = {
                            'row_index': row_index,
                            'asin': asin,
                            'country': country,
                            'commission': commission_display
                        }
                        if needs_mixed_yp_cleanup:
                            apply_mixed_platform_yp_cleanup(update)
                        updates.append(update)
                        store_offer_row_commission(row, asin, country, row_index, commission_result, [])
                    elif needs_mixed_yp_cleanup:
                        update = {'row_index': row_index, 'asin': asin, 'country': country}
                        apply_mixed_platform_yp_cleanup(update)
                        updates.append(update)
                    # 跳过其他处理（未测试且没有广告系列，不更新状态等）
                    continue
            
            update = {'row_index': row_index, 'asin': asin, 'country': country}
            skipped_by_mixed_platform_guard = False
            
            # 优先使用link_id精确匹配的广告系列，否则回退到(ASIN+国家)匹配
            has_precise_match = row_index in row_campaigns
            
            if has_precise_match:
                campaigns = row_campaigns[row_index]
            elif key and key in asin_country_campaigns:
                if key in mixed_platform_keys:
                    campaigns = []
                    if row_is_yp:
                        skipped_by_mixed_platform_guard = True
                        mixed_platform_fallback_skips += 1
                else:
                # 回退到(ASIN+国家)匹配，但排除已被其他行精确认领的广告系列
                    all_campaigns = asin_country_campaigns[key]
                    key_claimed_ids = claimed_campaign_ids_by_key.get(key, set())
                    if key_claimed_ids:
                        campaigns = [c for c in all_campaigns if c.get('campaign_id') not in key_claimed_ids]
                    else:
                        campaigns = list(all_campaigns)
            else:
                campaigns = []
            
            has_campaigns = len(campaigns) > 0
            
            if has_campaigns:
                total_campaigns = len(campaigns)
                enabled_campaigns = [c for c in campaigns if c['status'] == 'ENABLED']
                paused_campaigns = [c for c in campaigns if c['status'] == 'PAUSED']
                
                # 计算当前广告系列的花费（从Google Ads获取，即全部花费）
                total_cost = sum(c['cost_usd'] for c in campaigns)
                previous_total_cost = self.parse_commission_value(row.get('广告系列总花费', ''))
                
                # 获取所有账户ID（去除横杠）
                account_ids = list(set(c['account_id'].replace('-', '') for c in campaigns))
                
                # 获取所有广告系列名称
                campaign_names = [c['campaign_name'] for c in campaigns]
                
                # 确定状态
                if len(enabled_campaigns) == total_campaigns:
                    # 全部启用
                    update['status'] = '投放中'
                    update['status_color'] = 'green'
                elif len(enabled_campaigns) > 0:
                    # 部分启用
                    update['status'] = f'投放中({len(enabled_campaigns)}/{total_campaigns})'
                    update['status_color'] = 'green'
                else:
                    # 全部暂停
                    update['status'] = '广告系列暂停中'
                    update['status_color'] = 'orange'
                
                update['ads_ids'] = ','.join(account_ids)
                update['campaign_count'] = total_campaigns
                if total_cost <= 0.0001 and previous_total_cost > 0 and not has_precise_match:
                    update['total_cost'] = None
                    preserved_offer_cost_rows.append({
                        'row_index': row_index,
                        'asin': asin,
                        'country': country,
                        'previous_cost': previous_total_cost,
                    })
                else:
                    update['total_cost'] = 0.0 if total_cost <= 0.0001 else round(total_cost, 2)
                # 多个广告系列名称用逗号分隔
                update['campaign_names'] = ', '.join(campaign_names)
                if not has_precise_match and key:
                    claimed_ids = claimed_campaign_ids_by_key.setdefault(key, set())
                    for c in campaigns:
                        cid = c.get('campaign_id')
                        if cid:
                            claimed_ids.add(cid)
                            claimed_campaign_ids.add(cid)
                
            else:
                # 这个ASIN+国家组合没有广告系列
                if current_status and current_status not in ['未测试', '新复制']:
                    # 之前有状态，现在没有广告系列了
                    update['status'] = '投放已结束'
                    update['status_color'] = 'black'
                    update['ads_ids'] = ''
                    update['campaign_count'] = ''
                    update['campaign_names'] = ''  # 同时清空广告系列名称
                    if skipped_by_mixed_platform_guard:
                        update['clear_total_cost'] = True
                    else:
                        update['total_cost'] = None  # 保留原有花费数据
                else:
                    # 未测试或新复制状态，不更新
                    continue
            
            # 添加佣金数据 - 非"投放已结束"状态的行
            # 使用辅助函数获取佣金
            commission_result = get_commission_for_row(asin, country, row_index, is_first_offer)
            commission_value = commission_result.get('total')
            used_asin_only_match = commission_result.get('used_asin_only_match', False)

            if commission_value is not None:
                commission_display = round(commission_value, 2)
                if used_asin_only_match and len(asin_countries.get(asin, set())) > 1:
                    commission_display = f"{commission_display}*"
                update['commission'] = commission_display

            if campaigns:
                store_offer_row_commission(row, asin, country, row_index, commission_result, campaigns)
            elif commission_value is not None:
                store_offer_row_commission(row, asin, country, row_index, commission_result, campaigns)

            updates.append(update)
        
        if uid_fallback_events:
            fallback_total = sum(item['commission'] for item in uid_fallback_events)
            self.log_manage(
                f"  UID未精确命中，已按ASIN+国家回退到Offer: {len(uid_fallback_events)} 行，佣金 ${fallback_total:.2f}"
            )
            for item in uid_fallback_events[:3]:
                self.log_manage(
                    f"    - row={item['row_index']}, ASIN={item['asin']}, 国家={item['country']}, "
                    f"UIDs={','.join(item['uids'])}, 佣金=${item['commission']:.2f}"
                )
            if len(uid_fallback_events) > 3:
                self.log_manage(f"    ... 还有 {len(uid_fallback_events) - 3} 行未显示")
        
        if unresolved_uid_commissions:
            unresolved_total = sum(comm for _, comm in unresolved_uid_commissions)
            self.log_manage(
                f"  ⚠ 存在 {len(unresolved_uid_commissions)} 条UID佣金未落到具体offer行，金额 ${unresolved_total:.2f}"
            )
            for (asin, country, uid), comm in unresolved_uid_commissions[:3]:
                self.log_manage(
                    f"    - ASIN={asin}, 国家={country}, UID={uid}, 佣金=${comm:.2f}"
                )
            if len(unresolved_uid_commissions) > 3:
                self.log_manage(f"    ... 还有 {len(unresolved_uid_commissions) - 3} 条未显示")

        if preserved_offer_cost_rows:
            self.log_manage(
                f"  花费保护: Offer表保留 {len(preserved_offer_cost_rows)} 行原花费，避免本次$0.00覆盖"
            )
            for item in preserved_offer_cost_rows[:3]:
                self.log_manage(
                    f"    - row={item['row_index']}, ASIN={item['asin']}, 国家={item['country']}, "
                    f"保留原花费=${item['previous_cost']:.2f}"
                )
            if len(preserved_offer_cost_rows) > 3:
                self.log_manage(f"    ... 还有 {len(preserved_offer_cost_rows) - 3} 行未显示")

        if mixed_platform_fallback_skips:
            self.log_manage(
                f"  混合平台保护: 跳过 {mixed_platform_fallback_skips} 行 YP Offer 的(ASIN+国家)广告系列回退匹配，避免PB/YP串行"
            )
        
        commission_context = {
            'row_commissions': offer_row_commissions,
            'uid_fallback_events': uid_fallback_events,
            'unresolved_uid_commissions': unresolved_uid_commissions,
            'yp_duplicate_groups': {k: v for k, v in yp_duplicate_groups.items() if len(v) > 1},
        }
        return updates, commission_context
    
    def apply_feishu_updates(self, token, updates):
        """应用飞书表格更新"""
        spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
        sheet_id = self.feishu_sheet_id_var.get().strip()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        # 获取列映射（表头名称 -> 列字母）
        column_map = getattr(self, 'feishu_column_map', {})
        
        # 定义数据字段与表头名称的对应关系
        field_to_header = {
            'status': '状态',
            'ads_ids': '投放中的ads',
            'campaign_count': '广告系列数量',
            'campaign_names': '广告系列名称',
            'total_cost': '广告系列总花费',
            'commission': '总佣金'
        }
        
        # 检查必要的列是否存在
        missing_columns = []
        for field, header in field_to_header.items():
            if header not in column_map:
                missing_columns.append(header)
        
        if missing_columns:
            self.log_manage(f"  警告：以下列在表头中未找到: {missing_columns}")
            self.log_manage(f"  当前表头列: {list(column_map.keys())}")
        
        self.log_manage(f"  需要更新 {len(updates)} 行数据")
        self.log_manage(f"  列位置映射: 状态={column_map.get('状态', '?')}, 投放中的ads={column_map.get('投放中的ads', '?')}, 广告系列数量={column_map.get('广告系列数量', '?')}, 广告系列名称={column_map.get('广告系列名称', '?')}, 广告系列总花费={column_map.get('广告系列总花费', '?')}, 总佣金={column_map.get('总佣金', '?')}")
        
        # 分批处理，每批最多10条更新
        batch_size = 10
        total_updates = len(updates)
        
        for batch_start in range(0, total_updates, batch_size):
            if self.stop_flag:
                break
            
            batch_end = min(batch_start + batch_size, total_updates)
            batch_updates = updates[batch_start:batch_end]
            
            # 准备批量更新的数据
            value_ranges = []
            style_updates = []
            
            for update in batch_updates:
                row_index = update['row_index']
                
                # 状态列
                if 'status' in update and '状态' in column_map:
                    col = column_map['状态']
                    value_ranges.append({
                        'range': f"{sheet_id}!{col}{row_index}:{col}{row_index}",
                        'values': [[update['status']]]
                    })
                
                # 投放中的ads列
                if 'ads_ids' in update and '投放中的ads' in column_map:
                    col = column_map['投放中的ads']
                    value_ranges.append({
                        'range': f"{sheet_id}!{col}{row_index}:{col}{row_index}",
                        'values': [[update['ads_ids']]]
                    })
                
                # 广告系列数量列
                if 'campaign_count' in update and '广告系列数量' in column_map:
                    col = column_map['广告系列数量']
                    value_ranges.append({
                        'range': f"{sheet_id}!{col}{row_index}:{col}{row_index}",
                        'values': [[update['campaign_count']]]
                    })
                
                # 广告系列名称列
                if 'campaign_names' in update and '广告系列名称' in column_map:
                    col = column_map['广告系列名称']
                    value_ranges.append({
                        'range': f"{sheet_id}!{col}{row_index}:{col}{row_index}",
                        'values': [[update['campaign_names']]]
                    })
                
                # 广告系列总花费列
                if '广告系列总花费' in column_map:
                    col = column_map['广告系列总花费']
                    if update.get('clear_total_cost'):
                        value_ranges.append({
                            'range': f"{sheet_id}!{col}{row_index}:{col}{row_index}",
                            'values': [['']]
                        })
                    elif 'total_cost' in update and update['total_cost'] is not None:
                        value_ranges.append({
                            'range': f"{sheet_id}!{col}{row_index}:{col}{row_index}",
                            'values': [[f"${update['total_cost']}"]]
                        })
                
                # 总佣金列
                if 'commission' in update and '总佣金' in column_map:
                    col = column_map['总佣金']
                    value_ranges.append({
                        'range': f"{sheet_id}!{col}{row_index}:{col}{row_index}",
                        'values': [[f"${update['commission']}"]]
                    })
                
                # 收集样式更新（状态列的样式）
                if 'status_color' in update and '状态' in column_map:
                    style_updates.append({
                        'row_index': row_index,
                        'color': update['status_color'],
                        'column': column_map['状态']
                    })
            
            # 批量更新值 - 使用 values_batch_update API
            if value_ranges:
                try:
                    url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
                    body = {
                        "valueRanges": value_ranges
                    }
                    self.log_manage(f"    调用API: POST {url}")
                    self.log_manage(f"    更新范围数: {len(value_ranges)}")
                    response = requests.post(url, headers=headers, json=body, timeout=60)
                    result = response.json()
                    if result.get('code') != 0:
                        self.log_manage(f"  批量更新失败 (code={result.get('code')}): {result.get('msg', 'Unknown error')}")
                    else:
                        # 成功时记录更新的单元格数
                        responses = result.get('data', {}).get('responses', [])
                        total_cells = sum(r.get('updatedCells', 0) for r in responses)
                        self.log_manage(f"    成功更新 {total_cells} 个单元格")
                except Exception as e:
                    self.log_manage(f"  批量更新异常: {e}")
            
            # 批量更新样式
            if style_updates:
                self.batch_update_cell_styles(token, spreadsheet_token, sheet_id, style_updates)
            
            self.log_manage(f"  已更新 {batch_end}/{total_updates} 行")
            self.update_progress_manage(f"更新中 {batch_end}/{total_updates}")
            
            time.sleep(0.3)  # 避免请求过快
        
        self.log_manage(f"  更新完成")

    def repair_offer_total_costs_from_campaign_sheet(self, token):
        """一次性按广告系列表重算Offer表“广告系列总花费”，只更新这一列。"""
        spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
        offer_sheet_id = self.feishu_sheet_id_var.get().strip()
        campaigns_sheet_id = CAMPAIGNS_SHEET_ID

        self.log_manage("\n【一次性修复】按广告系列表重算Offer表广告系列总花费...")
        offer_rows = self.get_feishu_sheet_data(token)
        if not offer_rows:
            self.log_manage("  无法读取Offer表数据")
            return {'updated': 0, 'rows': 0}

        campaign_sheet_data = self.read_campaigns_sheet(token, spreadsheet_token, campaigns_sheet_id)
        if not campaign_sheet_data:
            self.log_manage("  无法读取广告系列表数据")
            return {'updated': 0, 'rows': 0}

        campaign_rows, _, _ = campaign_sheet_data
        cost_by_campaign_name = {}
        for row in campaign_rows or []:
            status = str(row.get('状态', '') or '').strip()
            if status == SUMMARY_STATUS_TEXT:
                continue
            campaign_name = str(row.get('广告系列名称', '') or '').strip()
            if not campaign_name:
                continue
            cost_by_campaign_name[campaign_name] = self.parse_commission_value(row.get('广告系列总花费', ''))

        cost_col_letter = self.feishu_column_map.get('广告系列总花费')
        cost_col_idx = self.column_letter_to_index(cost_col_letter)
        if cost_col_idx is None:
            self.log_manage("  Offer表缺少“广告系列总花费”列")
            return {'updated': 0, 'rows': 0}

        updates = []
        checked_rows = 0
        for row in offer_rows:
            row_index = row.get('row_index')
            if not row_index:
                continue
            status = str(row.get('状态', '') or '').strip()
            if status == SUMMARY_STATUS_TEXT:
                continue

            campaign_names_raw = str(row.get('广告系列名称', '') or '').strip()
            if not campaign_names_raw:
                continue

            campaign_names = [name.strip() for name in campaign_names_raw.split(',') if name.strip()]
            if not campaign_names:
                continue

            checked_rows += 1
            total_cost = 0.0
            found_campaign = False
            for campaign_name in campaign_names:
                if campaign_name in cost_by_campaign_name:
                    total_cost += float(cost_by_campaign_name.get(campaign_name, 0.0) or 0.0)
                    found_campaign = True

            if not found_campaign:
                continue

            normalized_total_cost = round(total_cost, 2)
            current_cost = self.parse_commission_value(row.get('广告系列总花费', ''))
            if abs(current_cost - normalized_total_cost) <= 0.0001:
                continue

            updates.append((row_index, cost_col_idx, f"${normalized_total_cost:.2f}"))

        self.log_manage(f"  广告系列表中读取到 {len(cost_by_campaign_name)} 个广告系列")
        self.log_manage(f"  检查了 {checked_rows} 行带广告系列名称的Offer")
        self.log_manage(f"  准备修复 {len(updates)} 行Offer的“广告系列总花费”")

        if not updates:
            self.log_manage("  无需修复")
            return {'updated': 0, 'rows': checked_rows}

        success = self._batch_update_sheet_cells(token, spreadsheet_token, updates, sheet_id=offer_sheet_id)
        if success:
            self.log_manage("  Offer表“广告系列总花费”修复完成")
        else:
            self.log_manage("  Offer表“广告系列总花费”修复存在部分失败")

        return {'updated': len(updates), 'rows': checked_rows}
    
    def batch_update_cell_styles(self, token, spreadsheet_token, sheet_id, style_updates):
        """批量更新单元格样式（字体颜色+加粗）"""
        # 字体颜色映射 - 使用十六进制颜色代码
        color_map = {
            'green': '#00AA00',    # 绿色
            'orange': '#FF8C00',   # 橙色
            'black': '#333333',    # 深灰/黑色
            'deep_pink': '#C71585' # 深粉色
        }
        
        # 按颜色分组
        color_groups = {}
        for style in style_updates:
            color = style['color']
            col = style.get('column', 'A')  # 使用动态列，默认为A
            if color not in color_groups:
                color_groups[color] = []
            color_groups[color].append(f"{sheet_id}!{col}{style['row_index']}:{col}{style['row_index']}")
        
        # 构建批量样式更新数据
        data = []
        for color, ranges in color_groups.items():
            font_color = color_map.get(color, '#000000')
            data.append({
                "ranges": ranges,
                "style": {
                    "font": {
                        "bold": True
                    },
                    "foreColor": font_color
                }
            })
        
        if not data:
            return
        
        try:
            url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/styles_batch_update"
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }
            body = {
                "data": data
            }
            response = requests.put(url, headers=headers, json=body, timeout=60)
            result = response.json()
            if result.get('code') != 0:
                self.log_manage(f"    样式更新失败 (code={result.get('code')}): {result.get('msg', 'Unknown error')}")
            else:
                total_cells = result.get('data', {}).get('totalUpdatedCells', 0)
                self.log_manage(f"    样式更新成功: {total_cells} 个单元格")
        except Exception as e:
            self.log_manage(f"    样式更新异常: {e}")
    
    def update_cell_style(self, token, spreadsheet_token, sheet_id, row_index, color, column='A'):
        """更新单个单元格样式（字体颜色+加粗）- 兼容旧接口"""
        self.batch_update_cell_styles(token, spreadsheet_token, sheet_id, [{'row_index': row_index, 'color': color, 'column': column}])
    
    def get_mcc_total_cost(self, start_date_str, end_date_str):
        """从MCC获取所有子账户在指定日期范围内的总花费（含已删除广告系列），转为USD"""
        try:
            sub_accounts = self.iter_google_ads_mcc_accounts()
            
            CNY_TO_USD_RATE = 0.14
            total_cost_usd = 0
            client_by_mcc = {}
            
            for account in sub_accounts:
                if self.stop_flag:
                    break
                try:
                    mcc_id = account.get('mcc_id', '')
                    client = client_by_mcc.get(mcc_id)
                    if client is None:
                        client = self.get_google_ads_client(mcc_id)
                        client_by_mcc[mcc_id] = client
                    if not client:
                        continue
                    ga_service = client.get_service('GoogleAdsService')
                    account_currency = account.get('currency', 'USD')
                    
                    # 查询该账户在指定日期范围内的总花费（包含所有广告系列，含已删除的）
                    cost_query = f"""
                        SELECT metrics.cost_micros
                        FROM customer
                        WHERE segments.date >= '{start_date_str}'
                            AND segments.date <= '{end_date_str}'
                    """
                    cost_response = ga_service.search(customer_id=account['id'], query=cost_query)
                    
                    for row in cost_response:
                        cost_original = row.metrics.cost_micros / 1000000 if row.metrics.cost_micros else 0
                        if account_currency == 'CNY':
                            total_cost_usd += cost_original * CNY_TO_USD_RATE
                        else:
                            total_cost_usd += cost_original
                            
                except Exception as e:
                    self.log_manage(f"  ⚠ 获取账户 {account['id']} 花费失败: {str(e)[:80]}")
            
            return total_cost_usd
            
        except Exception as e:
            self.log_manage(f"  获取MCC总花费失败: {e}")
            return None

    def get_last_successful_summary_total_cost(self, min_cost=None):
        """从日志中读取最近一次成功写入的总计行花费，用于MCC限流时恢复。"""
        try:
            if not os.path.exists(LOG_FILE):
                return None
            last_cost = None
            pattern = re.compile(r'总计行更新成功: .*总花费=\$([0-9,]+(?:\.[0-9]+)?)')
            with open(LOG_FILE, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    if '总花费保留原值' in line:
                        continue
                    match = pattern.search(line)
                    if match:
                        cost = float(match.group(1).replace(',', ''))
                        if min_cost is not None and cost <= float(min_cost or 0.0) + 0.01:
                            continue
                        last_cost = cost
            return last_cost
        except Exception as e:
            self.log_manage(f"  读取最近一次总计花费失败: {str(e)[:80]}")
            return None

    def get_mcc_daily_costs(self, start_date_str, end_date_str):
        """获取MCC在指定日期范围内的每日花费（含已删除广告系列），转为USD"""
        daily_costs = {}
        try:
            sub_accounts = self.iter_google_ads_mcc_accounts()

            CNY_TO_USD_RATE = 0.14
            client_by_mcc = {}

            for account in sub_accounts:
                if self.stop_flag:
                    break
                try:
                    mcc_id = account.get('mcc_id', '')
                    client = client_by_mcc.get(mcc_id)
                    if client is None:
                        client = self.get_google_ads_client(mcc_id)
                        client_by_mcc[mcc_id] = client
                    if not client:
                        continue
                    ga_service = client.get_service('GoogleAdsService')
                    account_currency = account.get('currency', 'USD')
                    cost_query = f"""
                        SELECT segments.date, metrics.cost_micros
                        FROM customer
                        WHERE segments.date >= '{start_date_str}'
                            AND segments.date <= '{end_date_str}'
                    """
                    cost_response = ga_service.search(customer_id=account['id'], query=cost_query)

                    for row in cost_response:
                        day = str(row.segments.date)
                        cost_original = row.metrics.cost_micros / 1000000 if row.metrics.cost_micros else 0
                        cost_usd = cost_original * CNY_TO_USD_RATE if account_currency == 'CNY' else cost_original
                        daily_costs[day] = daily_costs.get(day, 0) + cost_usd

                except Exception as e:
                    self.log_manage(f"  ⚠ 获取账户 {account['id']} 每日花费失败: {str(e)[:80]}")

        except Exception as e:
            self.log_manage(f"  获取MCC每日花费失败: {e}")

        return daily_costs

    def get_recent_daily_changes(self, days):
        """获取最近N天（包含今天）的每日新增花费/佣金/利润"""
        today = datetime.now().date()
        end_date = today
        start_date = today - timedelta(days=days - 1)

        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        daily_costs = self.get_mcc_daily_costs(start_date_str, end_date_str)
        commission_data = self.get_all_commissions(start_date_str, end_date_str)
        yp_commission_data = self.get_all_yp_commissions(start_date_str, end_date_str)

        daily_commissions = {}
        for trans in commission_data:
            if trans.get('status') == 'Rejected':
                continue

            order_time = trans.get('order_time', '')
            if str(order_time).isdigit():
                day = datetime.fromtimestamp(int(order_time)).strftime('%Y-%m-%d')
            else:
                continue

            comm = float(trans.get('sale_comm', 0) or 0)
            daily_commissions[day] = daily_commissions.get(day, 0) + comm

        for trans in yp_commission_data:
            if str(trans.get('status', '') or '').strip().lower() == 'rejected':
                continue

            order_time = trans.get('creationDate_time', '') or trans.get('creationDate', '')
            day = ''
            if isinstance(order_time, (int, float)) or str(order_time).isdigit():
                try:
                    day = datetime.fromtimestamp(int(order_time)).strftime('%Y-%m-%d')
                except Exception:
                    day = ''
            if not day and order_time:
                try:
                    day = datetime.strptime(str(order_time)[:19], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
                except Exception:
                    day = str(order_time)[:10] if len(str(order_time)) >= 10 else ''
            if not day:
                continue

            comm = float(trans.get('sale_comm', 0) or 0)
            daily_commissions[day] = daily_commissions.get(day, 0) + comm

        results = []
        current = start_date
        while current <= end_date:
            day = current.strftime('%Y-%m-%d')
            cost = daily_costs.get(day, 0)
            commission = daily_commissions.get(day, 0)
            profit = commission - cost
            results.append({
                'date': day,
                'cost': cost,
                'commission': commission,
                'profit': profit
            })
            current += timedelta(days=1)

        return results

    def build_recent_increment_log_lines(self, days):
        """生成最近N天新增数据日志，供主流程末尾统一输出。"""
        today = datetime.now().date()
        end_date = today
        start_date = today - timedelta(days=days - 1)
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        daily_changes = self.get_recent_daily_changes(days)
        lines = [
            "\n【步骤5.5】输出新增数据",
            f"  统计最近{days}天新增数据: {start_date_str} 至 {end_date_str}",
            "  新增数据:",
        ]
        for item in daily_changes:
            lines.append(
                f"  【{item['date']}】新增花费${item['cost']:.2f}；"
                f"新增佣金${item['commission']:.2f}；"
                f"利润${item['profit']:.2f}；"
            )
        return lines
    
    def update_feishu_summary_row(self, token, non_rejected_commission, total_cost=None, gross_commission=0, rejected_commission=0):
        """更新飞书表格第二行'总计'行的汇总数据
        
        参数:
            token: 飞书访问令牌
            non_rejected_commission: 非Rejected佣金
            total_cost: 总花费；None 表示保留原值
            gross_commission: 总佣金（含Rejected）
            rejected_commission: Rejected状态的佣金
        """
        spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
        sheet_id = self.feishu_sheet_id_var.get().strip()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        # 获取列映射
        column_map = getattr(self, 'feishu_column_map', {})
        
        # 准备更新数据
        value_ranges = []
        
        # 总佣金列 - 格式：$非Rejected佣金 ($总佣金-$Rejected佣金)
        if '总佣金' in column_map:
            col = column_map['总佣金']
            commission_display = f"${non_rejected_commission:.2f} (${gross_commission:.2f}-${rejected_commission:.2f})"
            value_ranges.append({
                'range': f"{sheet_id}!{col}2:{col}2",
                'values': [[commission_display]]
            })
        else:
            self.log_manage("  警告：未找到'总佣金'列")
        
        # 广告系列总花费列
        if total_cost is not None and '广告系列总花费' in column_map:
            col = column_map['广告系列总花费']
            value_ranges.append({
                'range': f"{sheet_id}!{col}2:{col}2",
                'values': [[f"${total_cost:.2f}"]]
            })
        elif total_cost is not None:
            self.log_manage("  警告：未找到'广告系列总花费'列")
        
        if not value_ranges:
            self.log_manage("  无法更新总计行：缺少必要的列")
            return
        
        try:
            url = f"{FEISHU_API_BASE_URL}/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
            body = {
                "valueRanges": value_ranges
            }
            response = requests.post(url, headers=headers, json=body, timeout=60)
            result = response.json()
            if result.get('code') != 0:
                self.log_manage(f"  更新总计行失败 (code={result.get('code')}): {result.get('msg', 'Unknown error')}")
            else:
                if total_cost is None:
                    self.log_manage(f"  ✅ 总计行更新成功: 总佣金=${non_rejected_commission:.2f} (${gross_commission:.2f}-${rejected_commission:.2f}), 总花费保留原值")
                else:
                    self.log_manage(f"  ✅ 总计行更新成功: 总佣金=${non_rejected_commission:.2f} (${gross_commission:.2f}-${rejected_commission:.2f}), 总花费=${total_cost:.2f}")
        except Exception as e:
            self.log_manage(f"  更新总计行异常: {e}")

    def consolidate_yp_duplicate_offer_rows(self, token, matched_keys=None):
        """整理并汇总YP重复offer行。"""
        spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
        sheet_id = self.feishu_sheet_id_var.get().strip()
        matched_keys = set(matched_keys or [])

        feishu_data = self.get_feishu_sheet_data(token)
        groups = self.build_offer_row_groups(feishu_data)
        duplicate_groups = {
            key: rows for key, rows in groups.get('yp_asin_brand_rows', {}).items()
            if len(rows) > 1 and (not matched_keys or key in matched_keys)
        }
        if not duplicate_groups:
            self.log_manage("  未发现需要汇总的YP重复offer")
            return

        self.log_manage(f"  发现 {len(duplicate_groups)} 组YP重复offer，开始整理")
        sorted_keys = [key for key, _ in sorted(duplicate_groups.items(), key=lambda item: min(item[1]))]

        for key in sorted_keys:
            feishu_data = self.get_feishu_sheet_data(token)
            raw_row_by_index = {row.get('row_index'): row for row in feishu_data if row.get('row_index')}
            groups = self.build_offer_row_groups(feishu_data)
            row_order = list(groups.get('row_sequence', []))
            row_by_index = groups.get('row_by_index', {})
            rows = sorted(groups.get('yp_asin_brand_rows', {}).get(key, []))
            if len(rows) <= 1:
                continue

            target = rows[0]
            desired = list(range(target, target + len(rows)))
            if rows != desired:
                working_order = list(row_order)
                data_start_0based = 2
                for offset, row_index in enumerate(rows):
                    current_pos = working_order.index(row_index)
                    desired_pos = working_order.index(target) + offset
                    if current_pos == desired_pos:
                        continue
                    src_0based = data_start_0based + current_pos
                    dst_0based = data_start_0based + desired_pos
                    if not self.feishu_move_dimension(token, spreadsheet_token, sheet_id, src_0based, src_0based, dst_0based):
                        raise RuntimeError(f"无法整理重复offer分组 {key}")
                    moved = working_order.pop(current_pos)
                    working_order.insert(desired_pos, moved)
                feishu_data = self.get_feishu_sheet_data(token)
                raw_row_by_index = {row.get('row_index'): row for row in feishu_data if row.get('row_index')}
                groups = self.build_offer_row_groups(feishu_data)
                row_by_index = groups.get('row_by_index', {})
                rows = sorted(groups.get('yp_asin_brand_rows', {}).get(key, rows))
                target = rows[0]

            existing_summary_row = raw_row_by_index.get(target - 1)
            if (
                existing_summary_row
                and str(existing_summary_row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT
                and str(existing_summary_row.get('ASIN', '') or '').strip() == key[0]
                and self.build_yp_row_brand_key(existing_summary_row) == str(key[1])
                and self.normalize_country_code(existing_summary_row.get('国家代码', '') or '') == (key[2] or '')
            ):
                summary_row_index = target - 1
            else:
                summary_row_index = target
                if not self.insert_sheet_rows(token, spreadsheet_token, sheet_id, start_index_0based=target - 1, count=1):
                    raise RuntimeError(f"无法为重复offer分组 {key} 插入统计行")
                feishu_data = self.get_feishu_sheet_data(token)
                groups = self.build_offer_row_groups(feishu_data)
                row_by_index = groups.get('row_by_index', {})
                rows = [ri + 1 if ri >= summary_row_index else ri for ri in rows]

            column_map = getattr(self, 'feishu_column_map', {})
            all_headers = [header for header, _ in sorted(column_map.items(), key=lambda item: self.get_column_index_by_header(item[0]) or 0)]
            excluded_headers = {
                '状态',
                '广告系列数量', '广告系列名称', '广告系列总花费', '总佣金',
                '国家代码', '品牌名称', 'ASIN', '品牌ID'
            }
            placeholder_headers = []
            for header in all_headers:
                if header in excluded_headers:
                    continue
                col_idx = self.get_column_index_by_header(header)
                currency_idx = self.get_column_index_by_header('货币')
                if col_idx is not None and currency_idx is not None and col_idx <= currency_idx:
                    placeholder_headers.append(header)

            updates = []
            style_updates = []
            summary_source = row_by_index.get(rows[0])
            if not summary_source:
                continue
            total_campaign_count = 0
            total_cost = 0.0
            total_commission = 0.0
            campaign_names = []

            for row_index in rows:
                row_data = row_by_index.get(row_index)
                if not row_data:
                    continue
                count_value = row_data.get('广告系列数量', '')
                try:
                    if str(count_value).strip():
                        total_campaign_count += int(float(str(count_value).replace(',', '').strip()))
                except Exception:
                    pass
                total_cost += self.parse_commission_value(row_data.get('广告系列总花费', ''))
                total_commission += self.parse_commission_value(row_data.get('总佣金', ''))
                names_value = str(row_data.get('广告系列名称', '') or '').strip()
                if names_value:
                    campaign_names.extend([name.strip() for name in names_value.split(',') if name.strip()])

            summary_values = {
                '状态': SUMMARY_STATUS_TEXT,
                '广告系列数量': total_campaign_count if total_campaign_count else '',
                '广告系列名称': ', '.join(campaign_names),
                '广告系列总花费': f"${total_cost:.2f}" if total_cost or campaign_names else '',
                '总佣金': f"${total_commission:.2f}" if total_commission or campaign_names else '',
                '国家代码': summary_source.get('国家代码', ''),
                '品牌名称': summary_source.get('品牌名称', ''),
                'ASIN': summary_source.get('ASIN', ''),
                '品牌ID': summary_source.get('品牌ID', ''),
            }

            for header, value in summary_values.items():
                col_idx = self.get_column_index_by_header(header)
                if col_idx is not None:
                    updates.append((summary_row_index, col_idx, value))

            for header in placeholder_headers:
                col_idx = self.get_column_index_by_header(header)
                if col_idx is not None:
                    updates.append((summary_row_index, col_idx, SUMMARY_PLACEHOLDER_TEXT))

            status_col_idx = self.get_column_index_by_header('状态')
            if status_col_idx is not None:
                style_updates.append({
                    'row_index': summary_row_index,
                    'color': 'deep_pink',
                    'column': self.feishu_column_map.get('状态', 'A')
                })

            for row_index in rows:
                for header in ('广告系列数量', '广告系列名称'):
                    col_idx = self.get_column_index_by_header(header)
                    if col_idx is not None:
                        updates.append((row_index, col_idx, ''))
                for header in ('广告系列总花费', '总佣金'):
                    col_idx = self.get_column_index_by_header(header)
                    if col_idx is not None:
                        updates.append((row_index, col_idx, '↑'))

            self._batch_update_sheet_cells(token, spreadsheet_token, updates)
            if style_updates:
                self.batch_update_cell_styles(token, spreadsheet_token, sheet_id, style_updates)

            self.log_manage(
                f"    已生成统计行: ASIN={key[0]}, 品牌ID={key[1]}, 汇总{len(rows)}行offer"
            )

    # ==================== 表格行排序功能 ====================
    
    def parse_commission_value(self, val):
        """从佣金字符串中提取数值
        
        支持格式: "$123.45", "*$123.45", "$4481.79 (+$133.91-$278.33)", "123.45", 数字等
        """
        if not val:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        val_str = str(val).strip()
        # 去除$、*、逗号
        val_str = val_str.replace('$', '').replace('*', '').replace(',', '')
        # 如果有空格+括号补充信息（如 "4481.79 (+133.91-278.33)"），只取第一部分
        if ' ' in val_str:
            val_str = val_str.split(' ')[0]
        # 如果有中文括号（如 "1143.82（35.63）"）
        if '（' in val_str:
            val_str = val_str.split('（')[0]
        try:
            return float(val_str) if val_str else 0.0
        except (ValueError, TypeError):
            return 0.0
    
    def feishu_move_dimension(self, token, spreadsheet_token, sheet_id, start_index, end_index, destination_index):
        """调用飞书API移动行
        
        参数:
            token: 飞书访问令牌
            spreadsheet_token: 电子表格token
            sheet_id: 工作表ID
            start_index: 要移动的起始行（0-based，包含）
            end_index: 要移动的结束行（0-based，包含）
            destination_index: 目标位置（0-based，移动后的位置，基于移除源行后的索引）
        
        返回:
            bool: 是否成功
        """
        if destination_index == start_index or destination_index == end_index + 1:
            return True

        url = f"{FEISHU_API_BASE_URL}/sheets/v3/spreadsheets/{spreadsheet_token}/sheets/{sheet_id}/move_dimension"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json; charset=utf-8"
        }
        body = {
            "source": {
                "major_dimension": "ROWS",
                "start_index": start_index,
                "end_index": end_index
            },
            "destination_index": destination_index
        }
        
        try:
            response = requests.post(url, headers=headers, json=body, timeout=30)
            result = response.json()
            if result.get('code') != 0:
                self.log_manage(f"    移动行失败 (src={start_index}, dst={destination_index}): code={result.get('code')}, {result.get('msg', '')}")
                return False
            return True
        except Exception as e:
            self.log_manage(f"    移动行异常: {e}")
            return False
    
    def sort_sheet_rows(self, token, spreadsheet_token, sheet_id, data_rows, data_start_0based, table_name=""):
        """对飞书表格的数据行进行排序（通用方法）
        
        参数:
            token: 飞书访问令牌
            spreadsheet_token: 电子表格token
            sheet_id: 工作表ID
            data_rows: 数据行列表，每项为 {'row_index': int, 'status': str, 'total_commission': float}，
                       必须按当前物理顺序排列
            data_start_0based: 第一个数据行的0-based索引
            table_name: 表格名称（用于日志）
        
        返回:
            int: 执行的移动次数
        """
        if len(data_rows) <= 1:
            return 0
        
        # 定义状态优先级：投放中(0) > 暂停中(1) > 投放已结束(2) > 未测试(3)
        status_priority = {
            SUMMARY_STATUS_TEXT: -1,
            '投放中': 0,
            '广告系列暂停中': 1,
            '暂停中': 1,
            '暂停': 1,
            '投放已结束': 2,
            '未测试': 3,
            '新复制': 3,
        }
        
        def get_status_priority(status_str):
            """获取状态优先级，支持带日期后缀的状态（如'广告系列暂停中2026-2-15'）"""
            if status_str in status_priority:
                return status_priority[status_str]
            # 前缀匹配（处理带日期的状态）
            for key, pri in status_priority.items():
                if status_str.startswith(key):
                    return pri
            return 99
        
        # 排序：先按状态优先级升序，再按总佣金降序
        sorted_rows = sorted(data_rows, key=lambda r: (
            get_status_priority(r['status']),
            -r['total_commission']
        ))
        
        # 统计各状态数量
        status_count = {}
        for r in sorted_rows:
            s = r['status'] if r['status'] else '(空)'
            status_count[s] = status_count.get(s, 0) + 1
        status_summary = ', '.join(f"{s}:{c}行" for s, c in status_count.items())
        self.log_manage(f"    {table_name}状态分布: {status_summary}")
        
        # 对比当前顺序和目标顺序
        current_order = [r['row_index'] for r in data_rows]
        desired_order = [r['row_index'] for r in sorted_rows]
        
        if current_order == desired_order:
            self.log_manage(f"    {table_name}已经是正确的排序顺序，无需移动")
            return 0
        
        # 预先计算实际需要移动的次数（模拟排序过程）
        temp_order = list(current_order)
        expected_moves = 0
        for i in range(len(temp_order)):
            target = desired_order[i]
            j = temp_order.index(target)
            if j != i:
                expected_moves += 1
                item = temp_order.pop(j)
                temp_order.insert(i, item)
        
        # 100次/分钟 => 每次调用间隔至少0.6秒
        MIN_INTERVAL = 0.6
        estimated_seconds = int(expected_moves * MIN_INTERVAL)
        self.log_manage(f"    需要执行 {expected_moves} 次移动，预计耗时约 {estimated_seconds // 60} 分 {estimated_seconds % 60} 秒")
        
        # 使用选择排序算法，从上往下依次将正确的行移到目标位置
        working_order = list(current_order)
        move_count = 0
        start_time = time.time()
        last_call_time = 0  # 上一次API调用的时间戳
        
        for i in range(len(working_order)):
            if self.stop_flag:
                self.log_manage(f"    排序被用户中断")
                break
            
            target = desired_order[i]
            j = working_order.index(target)
            if j == i:
                continue  # 已在正确位置
            
            # 精确控制API调用频率：确保两次调用之间间隔至少MIN_INTERVAL秒
            now = time.time()
            elapsed_since_last = now - last_call_time
            if elapsed_since_last < MIN_INTERVAL:
                time.sleep(MIN_INTERVAL - elapsed_since_last)
            
            # 计算飞书API所需的0-based行索引
            src_0based = data_start_0based + j
            dst_0based = data_start_0based + i
            
            last_call_time = time.time()
            success = self.feishu_move_dimension(
                token, spreadsheet_token, sheet_id,
                src_0based, src_0based, dst_0based
            )
            
            if success:
                # 更新工作顺序：模拟移动操作（pop后insert）
                item = working_order.pop(j)
                working_order.insert(i, item)
                move_count += 1
                
                # 每20次移动输出一次进度
                if move_count % 20 == 0:
                    elapsed = time.time() - start_time
                    remaining = (expected_moves - move_count) * (elapsed / move_count) if move_count > 0 else 0
                    self.log_manage(f"    进度: {move_count}/{expected_moves} ({move_count*100//expected_moves}%)，已用 {elapsed:.0f}s，剩余约 {remaining:.0f}s")
                    self.update_progress_manage(f"排序{table_name} {move_count}/{expected_moves}")
            else:
                self.log_manage(f"    排序中断：移动行失败")
                break
        
        elapsed_total = time.time() - start_time
        self.log_manage(f"    排序完成: {move_count}/{expected_moves} 次移动，耗时 {elapsed_total:.0f} 秒")
        return move_count

    def sort_pb_offer_rows_by_brand(self, token):
        """按品牌整理Offer表中的PB行，覆盖历史已新增但未归位的PB offer。"""
        self.log_manage("  按品牌整理PB offer行...")

        spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
        sheet_id = self.feishu_sheet_id_var.get().strip()
        feishu_data = self.get_feishu_sheet_data(token)
        column_map = getattr(self, 'feishu_column_map', {}) or {}
        if not feishu_data or not column_map:
            self.log_manage("    无法读取Offer表格数据")
            return 0

        row_by_index = {}
        last_data_ri = 0
        col_indices = {}
        for header, col_letter in column_map.items():
            col_idx = self.column_letter_to_index(col_letter)
            if col_idx is not None:
                col_indices[header] = col_idx

        for row in feishu_data:
            ri = row.get('row_index')
            if ri is None or ri < 3:
                continue
            has_content = (
                row.get('ASIN') or
                row.get('状态') or
                row.get('品牌名称') or
                row.get('投放链接')
            )
            if not has_content:
                continue
            row_by_index[ri] = row
            last_data_ri = max(last_data_ri, ri)

        if last_data_ri < 3:
            self.log_manage("    没有需要整理的PB offer行")
            return 0

        protected_rows = set()
        summary_rows = []
        for ri in range(3, last_data_ri + 1):
            row = row_by_index.get(ri, {})
            if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                summary_rows.append(ri)
                protected_rows.add(ri)

        for summary_ri in summary_rows:
            summary_row = row_by_index.get(summary_ri, {})
            summary_key = (
                str(summary_row.get('ASIN', '') or '').strip(),
                self.build_yp_row_brand_key(summary_row),
                self.normalize_country_code(summary_row.get('国家代码', '') or ''),
            )
            scan_ri = summary_ri + 1
            while scan_ri <= last_data_ri:
                row = row_by_index.get(scan_ri, {})
                if not row:
                    break
                if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                    break
                row_key = (
                    str(row.get('ASIN', '') or '').strip(),
                    str(row.get('品牌ID', '') or '').strip(),
                    self.normalize_country_code(row.get('国家代码', '') or ''),
                )
                is_summary_child = (
                    row_key == summary_key
                    and (
                        str(row.get('广告系列总花费', '') or '').strip() == '↑'
                        or str(row.get('总佣金', '') or '').strip() == '↑'
                    )
                )
                if not is_summary_child:
                    break
                protected_rows.add(scan_ri)
                scan_ri += 1

        if protected_rows:
            self.log_manage(
                f"    检测到 {len(summary_rows)} 条相同offer统计行，已保护 {len(protected_rows)} 行不参与品牌整理"
            )

        def build_sort_item(ri, segment_pos):
            row = row_by_index.get(ri, {})
            raw_row = []
            max_col_idx = max(col_indices.values(), default=-1)
            if max_col_idx >= 0:
                raw_row = [''] * (max_col_idx + 1)
                for header, col_idx in col_indices.items():
                    if col_idx < len(raw_row):
                        raw_row[col_idx] = row.get(header, '')

            is_pb_row = self.is_pb_offer_row(raw_row, col_indices)
            brand_key = self.normalize_offer_brand_sort_key(row.get('品牌名称', ''))
            asin = str(row.get('ASIN', '') or '').strip().upper()
            country = self.normalize_country_code(row.get('国家代码', '') or '')
            commission = self.parse_commission_value(row.get('总佣金', ''))
            return {
                'row_index': ri,
                'is_pb': is_pb_row,
                'brand_key': brand_key,
                'commission': commission,
                'asin': asin,
                'country': country,
                'original_pos': segment_pos,
            }

        segments = []
        current_segment = []
        for ri in range(3, last_data_ri + 1):
            if ri in protected_rows:
                if current_segment:
                    segments.append(current_segment)
                    current_segment = []
                continue
            if ri not in row_by_index:
                if current_segment:
                    segments.append(current_segment)
                    current_segment = []
                continue
            current_segment.append(ri)
        if current_segment:
            segments.append(current_segment)

        sortable_segments = []
        expected_moves = 0
        for segment in segments:
            if len(segment) <= 1:
                continue
            data_rows = [build_sort_item(ri, pos) for pos, ri in enumerate(segment)]
            sorted_rows = sorted(data_rows, key=lambda r: (
                0 if r['is_pb'] else 1,
                r['brand_key'] if r['is_pb'] else '',
                -r['commission'] if r['is_pb'] else 0,
                r['asin'] if r['is_pb'] else '',
                r['country'] if r['is_pb'] else '',
                r['original_pos'],
            ))
            current_order = [row['row_index'] for row in data_rows]
            desired_order = [row['row_index'] for row in sorted_rows]
            if current_order == desired_order:
                continue
            temp_order = list(current_order)
            segment_moves = 0
            for i, target in enumerate(desired_order):
                j = temp_order.index(target)
                if j != i:
                    segment_moves += 1
                    item = temp_order.pop(j)
                    temp_order.insert(i, item)
            expected_moves += segment_moves
            sortable_segments.append((segment[0], current_order, desired_order))

        if not sortable_segments:
            self.log_manage("    PB offer品牌顺序已经正确，无需移动")
            return 0

        self.log_manage(f"    需要移动 {expected_moves} 行来整理PB offer品牌顺序")
        move_count = 0

        for segment_start_ri, current_order, desired_order in sortable_segments:
            working_order = list(current_order)
            data_start_0based = segment_start_ri - 1

            for target_pos, target_row_index in enumerate(desired_order):
                if self.stop_flag:
                    self.log_manage("    PB offer品牌整理被停止")
                    break
                current_pos = working_order.index(target_row_index)
                if current_pos == target_pos:
                    continue
                if not self.feishu_move_dimension(
                    token,
                    spreadsheet_token,
                    sheet_id,
                    data_start_0based + current_pos,
                    data_start_0based + current_pos,
                    data_start_0based + target_pos
                ):
                    self.log_manage("    PB offer品牌整理中断：移动行失败")
                    break
                moved = working_order.pop(current_pos)
                working_order.insert(target_pos, moved)
                move_count += 1
                if move_count % 20 == 0:
                    self.log_manage(f"    PB offer品牌整理进度: {move_count}/{expected_moves}")
                time.sleep(0.6)
            if self.stop_flag:
                break

        self.log_manage(f"  ✅ PB offer品牌整理完成，共执行 {move_count} 次移动")
        return move_count
    
    def sort_offer_table(self, token):
        """对Offer表格进行排序
        
        排序规则：
        1. 表头行(第1行)、总计行(第2行)固定不动
        2. 保护“相同offer统计行”及其关联offer行，不打散汇总分组
        3. 其余PB/YP offer按品牌、国家、状态、ASIN整理
        """
        self.log_manage("  对Offer表格行排序整理...")
        
        spreadsheet_token = self.feishu_spreadsheet_var.get().strip()
        sheet_id = self.feishu_sheet_id_var.get().strip()
        
        # 重新读取最新数据
        feishu_data = self.get_feishu_sheet_data(token)
        
        if not feishu_data:
            self.log_manage("    无法读取Offer表格数据")
            return

        last_data_ri = 0
        row_by_index = {}
        for row in feishu_data:
            ri = row.get('row_index')
            if ri is None or ri < 3:
                continue
            has_content = (
                row.get('ASIN') or
                row.get('状态') or
                row.get('品牌名称') or
                row.get('投放链接') or
                row.get('总佣金')
            )
            if not has_content:
                continue
            row_by_index[ri] = row
            last_data_ri = max(last_data_ri, ri)
        
        if last_data_ri < 3:
            self.log_manage("    没有需要排序的数据行")
            return

        protected_rows = set()
        summary_rows = []
        for ri in range(3, last_data_ri + 1):
            row = row_by_index.get(ri, {})
            status = str(row.get('状态', '') or '').strip()
            if status == SUMMARY_STATUS_TEXT:
                summary_rows.append(ri)
                protected_rows.add(ri)

        for summary_ri in summary_rows:
            summary_row = row_by_index.get(summary_ri, {})
            summary_key = (
                str(summary_row.get('ASIN', '') or '').strip(),
                self.build_yp_row_brand_key(summary_row),
                self.normalize_country_code(summary_row.get('国家代码', '') or ''),
            )
            scan_ri = summary_ri + 1
            while scan_ri <= last_data_ri:
                row = row_by_index.get(scan_ri, {})
                if not row:
                    break
                if str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT:
                    break
                row_key = (
                    str(row.get('ASIN', '') or '').strip(),
                    str(row.get('品牌ID', '') or '').strip(),
                    self.normalize_country_code(row.get('国家代码', '') or ''),
                )
                is_summary_child = (
                    row_key == summary_key
                    and (
                        str(row.get('广告系列总花费', '') or '').strip() == '↑'
                        or str(row.get('总佣金', '') or '').strip() == '↑'
                    )
                )
                if not is_summary_child:
                    break
                protected_rows.add(scan_ri)
                scan_ri += 1

        if protected_rows:
            self.log_manage(
                f"    检测到 {len(summary_rows)} 条相同offer统计行，已保护 {len(protected_rows)} 行不参与顺序整理"
            )

        def build_sort_item(ri, segment_pos):
            row = row_by_index.get(ri, {})
            brand_key = self.normalize_offer_brand_sort_key(row.get('品牌名称', ''))
            country = self.normalize_country_code(row.get('国家代码', '') or '')
            status = str(row.get('状态', '') or '').strip()
            asin = str(row.get('ASIN', '') or '').strip().upper()
            return {
                'row_index': ri,
                'brand_key': brand_key,
                'country': country,
                'status_priority': self.get_offer_sort_status_priority(status),
                'asin': asin,
                'original_pos': segment_pos,
            }

        segments = []
        current_segment = []
        for ri in range(3, last_data_ri + 1):
            if ri in protected_rows:
                if current_segment:
                    segments.append(current_segment)
                    current_segment = []
                continue
            if ri not in row_by_index:
                if current_segment:
                    segments.append(current_segment)
                    current_segment = []
                continue
            current_segment.append(ri)
        if current_segment:
            segments.append(current_segment)

        sortable_segments = []
        expected_moves = 0
        sortable_row_count = 0
        for segment in segments:
            if len(segment) <= 1:
                continue
            data_rows = [build_sort_item(ri, pos) for pos, ri in enumerate(segment)]
            sortable_row_count += len(data_rows)
            sorted_rows = sorted(data_rows, key=lambda r: (
                1 if not r['brand_key'] else 0,
                r['brand_key'],
                r['country'],
                r['status_priority'],
                r['asin'],
                r['original_pos'],
            ))
            current_order = [row['row_index'] for row in data_rows]
            desired_order = [row['row_index'] for row in sorted_rows]
            if current_order == desired_order:
                continue
            temp_order = list(current_order)
            segment_moves = 0
            for i, target in enumerate(desired_order):
                j = temp_order.index(target)
                if j != i:
                    segment_moves += 1
                    item = temp_order.pop(j)
                    temp_order.insert(i, item)
            expected_moves += segment_moves
            sortable_segments.append((segment[0], current_order, desired_order))

        if not sortable_segments:
            self.log_manage("    Offer表品牌/状态/ASIN顺序已经正确，无需移动")
            return

        self.log_manage(f"    共 {sortable_row_count} 行可整理offer，需要执行 {expected_moves} 次移动")
        move_count = 0

        for segment_start_ri, current_order, desired_order in sortable_segments:
            working_order = list(current_order)
            data_start_0based = segment_start_ri - 1

            for target_pos, target_row_index in enumerate(desired_order):
                if self.stop_flag:
                    self.log_manage("    Offer顺序整理被停止")
                    break
                current_pos = working_order.index(target_row_index)
                if current_pos == target_pos:
                    continue
                if not self.feishu_move_dimension(
                    token,
                    spreadsheet_token,
                    sheet_id,
                    data_start_0based + current_pos,
                    data_start_0based + current_pos,
                    data_start_0based + target_pos
                ):
                    self.log_manage("    Offer顺序整理中断：移动行失败")
                    break
                moved = working_order.pop(current_pos)
                working_order.insert(target_pos, moved)
                move_count += 1
                if move_count % 20 == 0:
                    self.log_manage(f"    Offer顺序整理进度: {move_count}/{expected_moves}")
                    self.update_progress_manage(f"排序Offer表 {move_count}/{expected_moves}")
                time.sleep(0.6)
            if self.stop_flag:
                break

        self.log_manage(f"  ✅ Offer表格排序完成，共执行 {move_count} 次移动")
    
    def sort_campaigns_table(self, token):
        """对广告系列表格进行排序
        
        排序规则：
        1. 表头行(第1行)固定不动
        2. 数据行按状态排序：投放中 > 广告系列暂停中 > 投放已结束 > 未测试
        3. 同状态内按总佣金降序排列
        """
        self.log_manage("  对广告系列表格行排序...")
        
        campaigns_spreadsheet_token = "KnJ1wphpBiVMrGkWl5ncUkMGnfe"
        campaigns_sheet_id = CAMPAIGNS_SHEET_ID
        
        # 读取最新数据
        result = self.read_campaigns_sheet(token, campaigns_spreadsheet_token, campaigns_sheet_id)
        if result is None:
            self.log_manage("    无法读取广告系列表格数据")
            return
        
        existing_rows, column_map, first_empty_row = result
        
        if not existing_rows:
            self.log_manage("    广告系列表格无数据")
            return

        if any(str(row.get('状态', '') or '').strip() == SUMMARY_STATUS_TEXT for row in existing_rows):
            self.log_manage("    检测到“相同offer统计行”，已跳过广告系列表自动排序，避免打散汇总分组")
            return
        
        # 找到最后一个有内容的数据行
        last_data_ri = 0
        row_by_index = {}
        for row in existing_rows:
            ri = row.get('row_index')
            if ri is None:
                continue
            row_by_index[ri] = row
            campaign_name = row.get('广告系列名称', '')
            if campaign_name:
                last_data_ri = max(last_data_ri, ri)
        
        if last_data_ri < 2:
            self.log_manage("    没有需要排序的数据行")
            return
        
        # 构建连续的数据行列表（从第2行到最后有内容的行）
        data_rows = []
        for ri in range(2, last_data_ri + 1):
            row = row_by_index.get(ri, {})
            status = str(row.get('状态', '') or '').strip()
            commission = self.parse_commission_value(row.get('总佣金', ''))
            data_rows.append({
                'row_index': ri,
                'status': status,
                'total_commission': commission
            })
        
        self.log_manage(f"    共 {len(data_rows)} 行数据需要排序")
        
        # 广告系列表数据行从第2行开始（1-based），即0-based索引为1
        move_count = self.sort_sheet_rows(
            token, campaigns_spreadsheet_token, campaigns_sheet_id,
            data_rows, data_start_0based=1, table_name="广告系列表"
        )
        
        self.log_manage(f"  ✅ 广告系列表格排序完成，共执行 {move_count} 次移动")


def main():
    root = tk.Tk()
    app = OfferToolApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
